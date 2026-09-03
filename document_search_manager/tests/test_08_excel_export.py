# -*- coding: utf-8 -*-
"""Excel出力のハイパーリンク化

ネットワークには一切アクセスせず、最新バージョンの本体を読み込んで検証する。
実行: python tests/test_08_excel_export.py   （まとめて実行する場合は python tests/run_tests.py）
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _harness import dsm, DummyAuth  # noqa: E402

import io
import tempfile
from openpyxl import load_workbook

ok, ng = 0, 0
def check(label, cond, detail=""):
    global ok, ng
    if cond: ok += 1; print(f"  OK   {label}")
    else:    ng += 1; print(f"  NG   {label}  {detail}")

CFG = dict(dsm.DEFAULT_CFG)
SD = Path(tempfile.mkdtemp())
dsm.STATE_PATH = SD / "s.json"; dsm.EXPORT_DIR = SD / "e"; dsm.DOWNLOAD_DIR = SD / "d"
dsm._cfg = CFG; dsm._auth = DummyAuth()

BASE = "https://nexperia.sharepoint.com/sites/P020015_Caracal/Shared%20Documents/"
p = dsm.SharePointProvider(CFG, auth=None)
rows = [
    p._hit_to_result({"resource": {
        "webUrl": BASE + "40.%20Bench%20Validation/01.%20Validation_Plan/MRA2.0/%5B40VB%5D%20LDO.pptx",
        "lastModifiedDateTime": "2026-09-02T01:00:00Z",
        "createdBy": {"user": {"displayName": "Jake Smith;Takahiro Miyazaki"}}}}, 1),
    p._hit_to_result({"resource": {
        "webUrl": BASE + "12.%20Validation",
        "lastModifiedDateTime": "2026-07-31T02:00:00Z",
        "createdBy": {"user": {"displayName": "Jennifer Mitchell"}}}}, 2),
]
# リンクが一切無い行（リンクを張らないことの確認用）
rows.append(dsm.SearchResult(source="SharePoint", title="リンク無し", author="X",
                             last_modified="2026-01-01", doc_type="docx", rank=3))

mgr = dsm.SearchManager(CFG, DummyAuth())
mgr.providers[dsm.TARGET_SHAREPOINT].search = lambda kw, mx: {
    "results": rows, "total": len(rows), "note": ""}
dsm._manager = mgr
client = dsm.flask_app.test_client()
client.post("/api/search", json={"keyword": "validation", "target": "sharepoint"})

print("\n[E1] Excel出力の構成")
resp = client.get("/api/export?format=xlsx")
check("Excel出力が200", resp.status_code == 200, str(resp.status_code))
wb = load_workbook(io.BytesIO(resp.get_data()))
ws = wb.active
header = [c.value for c in ws[1]]
check("列は7列（URL列を廃止）", len(header) == 7, str(header))
check("列構成が期待どおり",
      header == ["ソース", "タイトル", "作成者", "最終更新日", "種別", "サイト", "フォルダ"],
      str(header))
for gone in ("ファイルリンク", "サイトリンク", "フォルダリンク"):
    check(f"「{gone}」列が無い", gone not in header)

print("\n[E2] ハイパーリンク")
title_cell = ws.cell(row=2, column=2)
site_cell  = ws.cell(row=2, column=6)
folder_cell = ws.cell(row=2, column=7)
check("タイトルセルがリンク", title_cell.hyperlink is not None)
check("タイトルのリンク先がファイル本体",
      title_cell.hyperlink and title_cell.hyperlink.target.endswith("LDO.pptx"),
      title_cell.hyperlink.target if title_cell.hyperlink else "なし")
check("サイトセルがリンク", site_cell.hyperlink is not None)
check("サイトのリンク先がサイトURL",
      site_cell.hyperlink and site_cell.hyperlink.target.endswith("P020015_Caracal"),
      site_cell.hyperlink.target if site_cell.hyperlink else "なし")
check("フォルダセルがリンク", folder_cell.hyperlink is not None)
check("フォルダのリンク先がフォルダビュー",
      folder_cell.hyperlink and "AllItems.aspx" in folder_cell.hyperlink.target,
      folder_cell.hyperlink.target if folder_cell.hyperlink else "なし")

check("リンクセルは青＋下線", title_cell.font.color.rgb.endswith("0563C1")
      and title_cell.font.underline == "single",
      f"{title_cell.font.color.rgb}/{title_cell.font.underline}")
check("リンクでない列は通常表示", ws.cell(row=2, column=3).font.underline in (None, "none"))

print("\n[E3] 値の内容")
check("タイトルは表示名のまま", title_cell.value == "[40VB] LDO.pptx", str(title_cell.value))
check("フォルダはフルパス",
      folder_cell.value == "Shared Documents/40. Bench Validation/01. Validation_Plan/MRA2.0",
      str(folder_cell.value))
check("フォルダ行の種別が「フォルダ」", ws.cell(row=3, column=5).value == "フォルダ",
      str(ws.cell(row=3, column=5).value))

print("\n[E4] リンクが無い行の扱い")
no_link = ws.cell(row=4, column=2)
check("URLが無ければリンクを張らない", no_link.hyperlink is None)
check("値は残る", no_link.value == "リンク無し", str(no_link.value))
check("空欄のサイトセルにもリンクを張らない", ws.cell(row=4, column=6).hyperlink is None)

print("\n[E5] 使い勝手")
check("見出し行を固定", ws.freeze_panes == "A2", str(ws.freeze_panes))
check("オートフィルタを設定", ws.auto_filter.ref == "A1:G4", str(ws.auto_filter.ref))
check("列幅を設定", ws.column_dimensions["B"].width == 56,
      str(ws.column_dimensions["B"].width))
check("見出しが濃色＋白文字", ws.cell(row=1, column=1).font.color.rgb.endswith("FFFFFF"))

print("\n[E6] CSVは従来どおりURL列を保持")
csv_body = client.get("/api/export?format=csv").get_data(as_text=True)
head = csv_body.splitlines()[0]
for col in ("ファイルリンク", "サイトリンク", "フォルダリンク"):
    check(f"CSVには「{col}」列が残る", col in head, head)
check("CSVにURLの値が入る", "AllItems.aspx" in csv_body)

print("\n[E7] 絞り込み後の行だけ出力（従来機能の維持）")
resp2 = client.get("/api/export?format=xlsx&idx=1")
ws2 = load_workbook(io.BytesIO(resp2.get_data())).active
check("選択した1行だけ出力", ws2.max_row == 2, str(ws2.max_row))
check("その行はフォルダ行", ws2.cell(row=2, column=5).value == "フォルダ",
      str(ws2.cell(row=2, column=5).value))

print(f"\n{'=' * 46}\n  成功 {ok} 件 / 失敗 {ng} 件\n{'=' * 46}")
sys.exit(1 if ng else 0)
