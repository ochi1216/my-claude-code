#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Document Search Manager  v20260903_02
=====================================
社内のドキュメント管理システムを、同一キーワードで横断検索するツール。

  0. All        … 有効な全系統を並列検索（既定）
  1. SharePoint … 社内SharePoint全社横断検索（本バージョンで実装）
  2. Nexus      … Shareflex品質文書サイト（Phase 2で実装予定）
  3. Enovia     … 3DEXPERIENCE/ENOVIA（Phase 3で実装予定）

【本バージョン(Phase 1)のスコープ】
  - SharePoint全社検索のみを実装する。Nexus / Enovia はUI上に枠だけ用意し、
    「未実装」と明示する（黙って0件を返さない）。

【認証方針】
  - 既存の po_database_organizer と同一のEntra IDアプリ登録を流用する。
  - 要求スコープは Sites.Read.All のみ。新規のGraph権限申請は一切発生させない。
  - Graphの /search/query が Sites.Read.All で通らなかった場合は、
    サイト単位の drive/root/search へ自動フォールバックする（権限追加は行わない）。

【パス方針】
  - 設定・キャッシュ・出力は全て「このスクリプトが置かれたフォルダ」基準で解決する。
    （カレントディレクトリに依存しないため、bat/ショートカットからでも安全に動く）
"""

import csv
import json
import sys
import threading
import webbrowser
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, asdict
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional, List, Dict, Any

try:
    import msal
    import requests as http_req
    from flask import Flask, jsonify, request as flask_req, Response, send_file
except ModuleNotFoundError as e:
    print(f"\n[エラー] 必要なライブラリ '{e.name}' が見つかりません。\n"
          "次のコマンドを実行してから再度実行してください:\n\n"
          "    pip install -r requirements.txt\n")
    sys.exit(1)


# ─────────────────────────────────────────────────────────────
# パス定義（すべて __file__ 基準。カレントディレクトリに依存しない）
# ─────────────────────────────────────────────────────────────
BASE_DIR         = Path(__file__).resolve().parent
CONFIG_PATH      = BASE_DIR / "config.json"
CONFIG_EXAMPLE   = BASE_DIR / "config.example.json"

# tenant_id / client_id の流用元候補（この順に探し、最初に揃ったものを使う）。
# いずれも「読むだけ」で、書き換えは一切行わない。
CREDENTIAL_SOURCES = [
    BASE_DIR.parent / "po_database_organizer"   / "config.json",
    BASE_DIR.parent / "onenote_report_generator" / "config.json",
]

TOKEN_CACHE_PATH = BASE_DIR / "token_cache.json"
CACHE_DIR        = BASE_DIR / "cache"
EXPORT_DIR       = BASE_DIR / "exports"

JST      = timezone(timedelta(hours=9))
GRAPH_V1 = "https://graph.microsoft.com/v1.0"

# 検索対象の識別子
TARGET_ALL        = "all"
TARGET_SHAREPOINT = "sharepoint"
TARGET_NEXUS      = "nexus"
TARGET_ENOVIA     = "enovia"


# ─────────────────────────────────────────────────────────────
# 設定ファイル読み込み
# ─────────────────────────────────────────────────────────────
DEFAULT_CFG: Dict[str, Any] = {
    "tenant_id": "",
    "client_id": "",
    "graph_scopes": ["https://graph.microsoft.com/Sites.Read.All"],
    "flask_port": 5020,
    "auto_open_browser": True,
    "page_size": 25,
    "default_max_results": 100,
    "hard_max_results": 500,
    "request_timeout_sec": 30,
    "provider_timeout_sec": 30,
    "sharepoint_host": "nexperia.sharepoint.com",
    "nexus_site_url": "https://nexperia.sharepoint.com/sites/SF_QualityDocumentsProd",
    "nexus_folder_path": "/sites/SF_QualityDocumentsProd/Documents",
    "nexus_list_id": "eb4ed9f8-81f8-4484-b8f6-64f22d30bfe0",
    "dedupe_nexus_from_sharepoint": False,
    "rewrite_host_to_mcas": False,
    "mcas_host": "nexperia.sharepoint.com.mcas.ms",
    "credentials_from": "",
    "fallback_site_urls": [],
}


def _read_credentials(path: Path) -> Dict[str, str]:
    """設定ファイルから tenant_id / client_id だけを取り出す。

    ツールによってキーの表記が異なる（po_database_organizer は小文字、
    onenote_report_generator は大文字）ため、両方の表記を受け付ける。
    値が未設定・プレースホルダ（<YOUR_...> 形式）の場合は採用しない。
    """
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    lowered = {str(k).lower(): v for k, v in data.items()}

    def valid(value) -> str:
        text = str(value or "").strip()
        if not text or text.startswith("<"):
            return ""
        return text

    return {
        "tenant_id": valid(lowered.get("tenant_id")),
        "client_id": valid(lowered.get("client_id")),
    }


def _load_config() -> Dict[str, Any]:
    """設定を読み込む。

    tenant_id / client_id が未設定の場合は、同一リポジトリ内の既存ツールの
    config.json から自動的に借用する（読むだけで、書き換えは一切行わない）。
    探索順は CREDENTIAL_SOURCES のとおり。config.json の "credentials_from" に
    パスを書けば、任意のファイルを優先的に参照させることもできる。
    """
    cfg = dict(DEFAULT_CFG)

    if CONFIG_PATH.exists():
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                cfg.update(json.load(f))
            print(f"⚙️  設定を読み込みました: {CONFIG_PATH}")
        except Exception as e:
            print(f"❌ config.json の読み込みに失敗しました: {e}")
            raise SystemExit(1)
    else:
        print(f"⚠️  config.json が見つかりません: {CONFIG_PATH}")
        print(f"    （{CONFIG_EXAMPLE.name} をコピーして作成できます）")

    # config.json のプレースホルダ（<YOUR_...>）は未設定として扱う
    for key in ("tenant_id", "client_id"):
        value = str(cfg.get(key) or "").strip()
        cfg[key] = "" if value.startswith("<") else value

    # ── tenant_id / client_id のフォールバック（既存ツールから借用） ──
    searched: List[Path] = []
    if cfg.get("credentials_from"):
        searched.append(Path(str(cfg["credentials_from"])).expanduser())
    searched.extend(CREDENTIAL_SOURCES)

    if not cfg.get("tenant_id") or not cfg.get("client_id"):
        for source in searched:
            if cfg.get("tenant_id") and cfg.get("client_id"):
                break
            if not source.exists():
                print(f"   ・流用元なし: {source}")
                continue
            try:
                found = _read_credentials(source)
            except Exception as e:
                print(f"   ・読み込み失敗: {source} ({e})")
                continue

            borrowed = []
            for key in ("tenant_id", "client_id"):
                if not cfg.get(key) and found.get(key):
                    cfg[key] = found[key]
                    borrowed.append(key)
            if borrowed:
                print(f"🔗 {source} から {' / '.join(borrowed)} を流用しました。")
            else:
                print(f"   ・該当項目なし: {source}")

    if not cfg.get("tenant_id") or not cfg.get("client_id"):
        print("\n❌ tenant_id / client_id が特定できませんでした。")
        print("   次の場所を探しましたが、いずれにも設定がありませんでした:")
        for source in searched:
            print(f"     - {source}")
        print("\n   対処方法（どちらか一方）:")
        print(f"     (1) {CONFIG_PATH.name} の tenant_id / client_id に、既存ツールと")
        print("         同じEntra IDアプリ登録の値を直接記入する")
        print("     (2) 既存ツールの config.json があるフォルダを、")
        print("         config.json の \"credentials_from\" にフルパスで指定する")
        print("         例: \"credentials_from\": "
              "\"C:\\\\Users\\\\xxxx\\\\po_database_organizer\\\\config.json\"\n")
        raise SystemExit(1)

    return cfg


# ─────────────────────────────────────────────────────────────
# 認証（po_database_organizer の GraphAuthManager を移植）
# ─────────────────────────────────────────────────────────────
class GraphAuthManager:
    """MSAL Device Code Flow で Graph API のアクセストークンを取得・保持する。"""

    def __init__(self, tenant_id: str, client_id: str, cache_path: Path):
        self.cache_path = Path(cache_path)
        self.cache = msal.SerializableTokenCache()
        if self.cache_path.exists():
            try:
                self.cache.deserialize(self.cache_path.read_text(encoding="utf-8"))
            except Exception as e:
                print(f"⚠️  トークンキャッシュの読み込みに失敗しました（無視して続行）: {e}")

        self.app = msal.PublicClientApplication(
            client_id,
            authority=f"https://login.microsoftonline.com/{tenant_id}",
            token_cache=self.cache,
        )

    def _save_cache(self) -> None:
        if self.cache.has_state_changed:
            try:
                self.cache_path.write_text(self.cache.serialize(), encoding="utf-8")
            except Exception as e:
                print(f"⚠️  トークンキャッシュの保存に失敗しました（無視して続行）: {e}")

    def get_token(self, scopes: List[str]) -> str:
        result = None
        accounts = self.app.get_accounts()
        if accounts:
            result = self.app.acquire_token_silent(scopes, account=accounts[0])

        if not result:
            flow = self.app.initiate_device_flow(scopes=scopes)
            if "user_code" not in flow:
                raise RuntimeError(f"デバイスコードフローの開始に失敗しました: {flow}")
            print("\n" + "=" * 70)
            print("🔑 サインインが必要です。以下の手順で認証してください。")
            print(flow["message"])
            print("=" * 70 + "\n")
            result = self.app.acquire_token_by_device_flow(flow)

        if "access_token" not in result:
            raise RuntimeError(
                f"アクセストークンの取得に失敗しました: "
                f"{result.get('error')} / {result.get('error_description')}"
            )

        self._save_cache()
        return result["access_token"]


# ─────────────────────────────────────────────────────────────
# 検索結果の正規化スキーマ
# ─────────────────────────────────────────────────────────────
@dataclass
class SearchResult:
    """3系統の検索結果を統一表現するためのスキーマ。

    Q-C（まずは一覧まで）の回答に従い、本文スニペットは保持しない。
    """
    source: str = ""             # SharePoint / Nexus / Enovia
    document_number: str = ""    # Nexus等の文書番号（SharePoint全社検索では通常空）
    title: str = ""
    author: str = ""
    last_modified: str = ""      # YYYY-MM-DD（JST）
    doc_type: str = ""           # 拡張子など
    site: str = ""               # 所属サイト名
    url: str = ""
    is_nexus_path: bool = False  # Nexusサイト配下の文書か（重複判定用）
    rank: int = 0


# ─────────────────────────────────────────────────────────────
# ユーティリティ
# ─────────────────────────────────────────────────────────────
def _pick(*candidates) -> str:
    """複数の候補値から、最初に見つかった非空の文字列を返す。"""
    for c in candidates:
        if c is None:
            continue
        s = str(c).strip()
        if s and s.lower() != "none":
            return s
    return ""


def _lower_keys(d: Any) -> Dict[str, Any]:
    """辞書のキーを小文字化した辞書を返す（Graphのfields名の揺れを吸収する）。"""
    if not isinstance(d, dict):
        return {}
    return {str(k).lower(): v for k, v in d.items()}


def _format_datetime(value: str) -> str:
    """ISO8601形式の日時をJSTの YYYY-MM-DD に整形する。失敗時は原文を返す。"""
    if not value:
        return ""
    text = str(value).strip()
    try:
        normalized = text.replace("Z", "+00:00")
        dt = datetime.fromisoformat(normalized)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(JST).strftime("%Y-%m-%d")
    except Exception:
        return text[:10]


def _extension_of(name: str) -> str:
    if not name or "." not in name:
        return ""
    return name.rsplit(".", 1)[-1].lower()


def _rewrite_url(url: str, cfg: Dict[str, Any]) -> str:
    """設定に応じてリンクのホストを .mcas.ms 経由に書き換える。"""
    if not url or not cfg.get("rewrite_host_to_mcas"):
        return url
    host = cfg.get("sharepoint_host", "")
    mcas = cfg.get("mcas_host", "")
    if host and mcas and host in url and mcas not in url:
        return url.replace(host, mcas, 1)
    return url


# ─────────────────────────────────────────────────────────────
# 検索プロバイダ（抽象）
# ─────────────────────────────────────────────────────────────
class SearchProvider:
    """検索対象1系統を表す抽象クラス。

    Nexus / Enovia を後から足すときに、この I/F だけを満たせばよい構造にする。
    """
    key = ""
    label = ""
    implemented = False

    def __init__(self, cfg: Dict[str, Any], auth: Optional[GraphAuthManager]):
        self.cfg = cfg
        self.auth = auth

    def probe(self) -> Dict[str, Any]:
        """疎通診断。{'ok': bool, 'mode': str, 'message': str} を返す。"""
        return {"ok": False, "mode": "未実装", "message": "この系統は未実装です。"}

    def search(self, keyword: str, max_results: int) -> Dict[str, Any]:
        """検索を実行し、{'results': [...], 'total': int, 'note': str} を返す。"""
        raise NotImplementedError


# ─────────────────────────────────────────────────────────────
# 1. SharePoint 全社横断検索プロバイダ
# ─────────────────────────────────────────────────────────────
class SharePointProvider(SearchProvider):
    """Microsoft Graph Search API による社内SharePoint全社横断検索。

    第一候補: POST /search/query  (entityTypes = listItem)
      - SharePointライブラリ上の文書は listItem として返るため、
        Sites.Read.All のみで全社検索が成立することを狙う。
    フォールバック: GET /sites/{host}:{path}:/drive/root/search(q='...')
      - config.json の fallback_site_urls に列挙したサイトを個別に検索する。
      - こちらは Sites.Read.All で確実に動作する。
    """
    key = TARGET_SHAREPOINT
    label = "SharePoint"
    implemented = True

    # 疎通診断の結果を保持（毎回判定し直さない）
    _mode: Optional[str] = None   # "search-api" / "site-drive" / None

    def _token(self) -> str:
        return self.auth.get_token(self.cfg["graph_scopes"])

    # ── 疎通診断 ────────────────────────────────────────────
    def probe(self) -> Dict[str, Any]:
        try:
            token = self._token()
        except Exception as e:
            return {"ok": False, "mode": "認証エラー", "message": f"認証に失敗しました: {e}"}

        status, _body, err = self._call_search_api(token, "test", frm=0, size=1)

        if status == 200:
            self._mode = "search-api"
            return {
                "ok": True,
                "mode": "search-api",
                "message": "全社横断検索が利用できます（/search/query・listItem方式）。",
            }

        if status in (401, 403):
            sites = self.cfg.get("fallback_site_urls") or []
            if sites:
                self._mode = "site-drive"
                return {
                    "ok": True,
                    "mode": "site-drive",
                    "message": (f"全社横断検索は権限不足のため、サイト単位検索に切り替えます"
                                f"（対象 {len(sites)} サイト）。"),
                }
            self._mode = None
            return {
                "ok": False,
                "mode": "権限不足",
                "message": (f"/search/query が HTTP {status} で拒否されました。"
                            f"config.json の fallback_site_urls に検索したいサイトURLを"
                            f"設定するとサイト単位検索に切り替えられます。"),
            }

        self._mode = None
        return {"ok": False, "mode": "エラー", "message": f"HTTP {status}: {err}"}

    # ── Graph Search API 呼び出し ──────────────────────────
    def _call_search_api(self, token: str, query_string: str, frm: int, size: int):
        """(status_code, body_dict, error_text) を返す。例外は投げない。"""
        body = {
            "requests": [{
                "entityTypes": ["listItem"],
                "query": {"queryString": query_string},
                "from": frm,
                "size": size,
            }]
        }
        try:
            resp = http_req.post(
                f"{GRAPH_V1}/search/query",
                headers={
                    "Authorization": f"Bearer {token}",
                    "Content-Type": "application/json",
                },
                json=body,
                timeout=self.cfg.get("request_timeout_sec", 30),
            )
        except Exception as e:
            return 0, None, str(e)

        if resp.status_code != 200:
            return resp.status_code, None, resp.text[:500]

        try:
            return 200, resp.json(), ""
        except Exception as e:
            return 200, None, f"レスポンスのJSON解析に失敗しました: {e}"

    # ── 検索本体 ────────────────────────────────────────────
    def search(self, keyword: str, max_results: int) -> Dict[str, Any]:
        token = self._token()

        if self._mode is None:
            self.probe()

        if self._mode == "site-drive":
            return self._search_by_site_drive(token, keyword, max_results)

        return self._search_by_search_api(token, keyword, max_results)

    def _search_by_search_api(self, token: str, keyword: str,
                              max_results: int) -> Dict[str, Any]:
        page_size = int(self.cfg.get("page_size", 25))
        results: List[SearchResult] = []
        total = 0
        frm = 0
        rank = 0

        while len(results) < max_results:
            size = min(page_size, max_results - len(results))
            status, body, err = self._call_search_api(token, keyword, frm, size)
            if status != 200 or body is None:
                if results:
                    break  # 途中まで取れていれば、それを返す
                raise RuntimeError(f"HTTP {status}: {err}")

            containers = (body.get("value") or [{}])[0].get("hitsContainers") or []
            if not containers:
                break

            container = containers[0]
            total = container.get("total", total) or total
            hits = container.get("hits") or []
            if not hits:
                break

            for hit in hits:
                rank += 1
                results.append(self._hit_to_result(hit, rank))

            if not container.get("moreResultsAvailable"):
                break
            frm += len(hits)

        return {"results": results, "total": total or len(results), "note": ""}

    def _hit_to_result(self, hit: Dict[str, Any], rank: int) -> SearchResult:
        """Graph Search の hit を SearchResult に変換する。

        Graphが返すフィールド名は環境により揺れるため、複数の候補を順に見る
        寛容なパーサとする（形が変わっても落ちないことを優先する）。
        """
        res = hit.get("resource") or {}
        fields = _lower_keys(res.get("fields") or {})

        name = _pick(fields.get("filename"), fields.get("name"), res.get("name"))
        title = _pick(fields.get("title"), name, "(タイトルなし)")

        url = _pick(res.get("webUrl"), fields.get("path"), fields.get("spwebur"))

        author = _pick(
            fields.get("author"),
            fields.get("createdby"),
            ((res.get("createdBy") or {}).get("user") or {}).get("displayName"),
            ((res.get("lastModifiedBy") or {}).get("user") or {}).get("displayName"),
        )

        last_modified = _format_datetime(_pick(
            fields.get("lastmodifiedtime"),
            fields.get("lastmodifieddatetime"),
            res.get("lastModifiedDateTime"),
        ))

        doc_type = _pick(
            fields.get("fileextension"),
            fields.get("filetype"),
            _extension_of(name),
        ).lower()

        document_number = _pick(
            fields.get("documentnumber"),
            fields.get("document_x0020_number"),
            fields.get("oldsystemidentifier"),
        )

        site = _pick(fields.get("sitetitle"), fields.get("spsiteurl"))

        nexus_url = str(self.cfg.get("nexus_site_url", "")).rstrip("/")
        is_nexus = bool(nexus_url) and url.lower().startswith(nexus_url.lower())

        return SearchResult(
            source=self.label,
            document_number=document_number,
            title=title,
            author=author,
            last_modified=last_modified,
            doc_type=doc_type,
            site=site,
            url=_rewrite_url(url, self.cfg),
            is_nexus_path=is_nexus,
            rank=rank,
        )

    # ── フォールバック（サイト単位検索） ───────────────────
    def _search_by_site_drive(self, token: str, keyword: str,
                              max_results: int) -> Dict[str, Any]:
        results: List[SearchResult] = []
        rank = 0
        errors: List[str] = []

        for site_url in (self.cfg.get("fallback_site_urls") or []):
            if len(results) >= max_results:
                break
            try:
                host, path = self._split_site_url(site_url)
            except Exception as e:
                errors.append(f"{site_url}: {e}")
                continue

            endpoint = (f"{GRAPH_V1}/sites/{host}:{path}:/drive/root/"
                        f"search(q='{keyword}')")
            try:
                resp = http_req.get(
                    endpoint,
                    headers={"Authorization": f"Bearer {token}"},
                    params={"$top": min(50, max_results - len(results))},
                    timeout=self.cfg.get("request_timeout_sec", 30),
                )
            except Exception as e:
                errors.append(f"{site_url}: {e}")
                continue

            if resp.status_code != 200:
                errors.append(f"{site_url}: HTTP {resp.status_code}")
                continue

            for item in (resp.json().get("value") or []):
                if len(results) >= max_results:
                    break
                rank += 1
                name = _pick(item.get("name"))
                url = _pick(item.get("webUrl"))
                nexus_url = str(self.cfg.get("nexus_site_url", "")).rstrip("/")
                results.append(SearchResult(
                    source=self.label,
                    document_number="",
                    title=_pick(name, "(タイトルなし)"),
                    author=_pick(((item.get("createdBy") or {}).get("user") or {})
                                 .get("displayName")),
                    last_modified=_format_datetime(item.get("lastModifiedDateTime", "")),
                    doc_type=_extension_of(name),
                    site=site_url,
                    url=_rewrite_url(url, self.cfg),
                    is_nexus_path=(bool(nexus_url)
                                   and url.lower().startswith(nexus_url.lower())),
                    rank=rank,
                ))

        note = "サイト単位検索モードで実行しました。"
        if errors:
            note += " 失敗したサイト: " + " / ".join(errors[:3])
        return {"results": results, "total": len(results), "note": note}

    @staticmethod
    def _split_site_url(site_url: str):
        """https://host/sites/xxx を (host, /sites/xxx) に分解する。"""
        text = site_url.strip().rstrip("/")
        if not text.startswith("http"):
            raise ValueError("URLは https:// から始めてください")
        without_scheme = text.split("://", 1)[1]
        if "/" not in without_scheme:
            raise ValueError("サイトパスが含まれていません")
        host, path = without_scheme.split("/", 1)
        return host, "/" + path


# ─────────────────────────────────────────────────────────────
# 2. Nexus プロバイダ（Phase 2で実装予定）
# ─────────────────────────────────────────────────────────────
class NexusProvider(SearchProvider):
    key = TARGET_NEXUS
    label = "Nexus"
    implemented = False

    def probe(self) -> Dict[str, Any]:
        return {
            "ok": False,
            "mode": "未実装",
            "message": ("Phase 2で実装予定です。Graph Search を "
                        + str(self.cfg.get("nexus_folder_path", ""))
                        + " に限定して検索する方式を予定しています。"),
        }

    def search(self, keyword: str, max_results: int) -> Dict[str, Any]:
        raise NotImplementedError("Nexus検索はPhase 2で実装予定です。")


# ─────────────────────────────────────────────────────────────
# 3. Enovia プロバイダ（Phase 3で実装予定）
# ─────────────────────────────────────────────────────────────
class EnoviaProvider(SearchProvider):
    key = TARGET_ENOVIA
    label = "Enovia"
    implemented = False

    def probe(self) -> Dict[str, Any]:
        return {
            "ok": False,
            "mode": "未実装",
            "message": "Phase 3で実装予定です（実装方式は調査後に確定）。",
        }

    def search(self, keyword: str, max_results: int) -> Dict[str, Any]:
        raise NotImplementedError("Enovia検索はPhase 3で実装予定です。")


# ─────────────────────────────────────────────────────────────
# 実行層（並列実行・部分成功）
# ─────────────────────────────────────────────────────────────
class SearchManager:
    """複数プロバイダを並列実行し、結果をマージする。

    1系統が失敗しても他系統の結果は必ず返す（部分成功方式）。
    """

    def __init__(self, cfg: Dict[str, Any], auth: GraphAuthManager):
        self.cfg = cfg
        self.providers: Dict[str, SearchProvider] = {
            TARGET_SHAREPOINT: SharePointProvider(cfg, auth),
            TARGET_NEXUS:      NexusProvider(cfg, auth),
            TARGET_ENOVIA:     EnoviaProvider(cfg, auth),
        }

    def probe_all(self) -> List[Dict[str, Any]]:
        out = []
        for key, provider in self.providers.items():
            try:
                result = provider.probe()
            except Exception as e:
                result = {"ok": False, "mode": "エラー", "message": str(e)}
            result.update({"key": key, "label": provider.label,
                           "implemented": provider.implemented})
            out.append(result)
        return out

    def _targets(self, target: str) -> List[SearchProvider]:
        if target == TARGET_ALL:
            return [p for p in self.providers.values() if p.implemented]
        provider = self.providers.get(target)
        # 未実装の系統は実行対象に含めない（呼び出し側で「未実装」として扱う）
        return [provider] if provider and provider.implemented else []

    def search(self, keyword: str, target: str, max_results: int) -> Dict[str, Any]:
        providers = self._targets(target)

        # 未実装の系統が明示的に選ばれた場合は、その旨をそのまま返す
        if not providers:
            provider = self.providers.get(target)
            if provider and not provider.implemented:
                return {
                    "results": [],
                    "statuses": [dict(provider.probe(), key=provider.key,
                                      label=provider.label, implemented=False,
                                      state="pending", count=0, total=0)],
                    "excluded_nexus": 0,
                }
            return {"results": [], "statuses": [], "excluded_nexus": 0}

        statuses: List[Dict[str, Any]] = []
        merged: List[SearchResult] = []
        timeout = self.cfg.get("provider_timeout_sec", 30)

        with ThreadPoolExecutor(max_workers=max(1, len(providers))) as pool:
            futures = {
                pool.submit(p.search, keyword, max_results): p for p in providers
            }
            for future in as_completed(futures, timeout=timeout * len(providers)):
                provider = futures[future]
                try:
                    outcome = future.result(timeout=timeout)
                    results = outcome.get("results") or []
                    merged.extend(results)
                    statuses.append({
                        "key": provider.key,
                        "label": provider.label,
                        "implemented": True,
                        "state": "ok" if results else "empty",
                        "count": len(results),
                        "total": outcome.get("total", len(results)),
                        "message": outcome.get("note", ""),
                    })
                except Exception as e:
                    statuses.append({
                        "key": provider.key,
                        "label": provider.label,
                        "implemented": True,
                        "state": "error",
                        "count": 0,
                        "total": 0,
                        "message": str(e),
                    })

        # 未実装の系統も「0. All」のときは状態として見せる（黙って隠さない）
        if target == TARGET_ALL:
            for provider in self.providers.values():
                if provider.implemented:
                    continue
                info = provider.probe()
                statuses.append({
                    "key": provider.key,
                    "label": provider.label,
                    "implemented": False,
                    "state": "pending",
                    "count": 0,
                    "total": 0,
                    "message": info.get("message", ""),
                })

        merged, excluded = self._dedupe(merged)
        merged.sort(key=lambda r: (r.source, r.rank))

        statuses.sort(key=lambda s: ["sharepoint", "nexus", "enovia"].index(s["key"])
                      if s["key"] in ("sharepoint", "nexus", "enovia") else 99)

        return {"results": merged, "statuses": statuses, "excluded_nexus": excluded}

    def _dedupe(self, results: List[SearchResult]):
        """Nexusサイト配下の文書をSharePoint結果から除外する。

        Phase 1 では Nexus プロバイダが未実装のため、既定では除外しない
        （除外するとNexus文書が一切見えなくなってしまうため）。
        Phase 2 で Nexus プロバイダを有効化するのと同時に、
        config.json の dedupe_nexus_from_sharepoint を true にする。
        """
        if not self.cfg.get("dedupe_nexus_from_sharepoint"):
            return results, 0

        kept, excluded = [], 0
        for r in results:
            if r.source == "SharePoint" and r.is_nexus_path:
                excluded += 1
                continue
            kept.append(r)
        return kept, excluded


# ─────────────────────────────────────────────────────────────
# 画面（ダークテーマ）
# ─────────────────────────────────────────────────────────────
# 絵文字はJS側の文字列定数として定義する（Python f-string の外で扱う）
INDEX_HTML = """<!doctype html>
<html lang="ja">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Document Search Manager</title>
<style>
  :root {
    --bg: #1a1a2e;
    --panel: #23233f;
    --panel2: #2c2c4d;
    --accent: #e94560;
    --text: #e8e8f0;
    --muted: #9a9ab5;
    --border: #3a3a5c;
  }
  * { box-sizing: border-box; }
  body {
    margin: 0; background: var(--bg); color: var(--text);
    font-family: "Segoe UI", "Meiryo", system-ui, sans-serif; font-size: 14px;
  }
  header {
    background: linear-gradient(90deg, #16213e, #1a1a2e);
    border-bottom: 2px solid var(--accent); padding: 16px 24px;
  }
  header h1 { margin: 0; font-size: 20px; letter-spacing: .04em; }
  header .sub { color: var(--muted); font-size: 12px; margin-top: 4px; }
  main { padding: 20px 24px 60px; max-width: 1500px; }
  .panel {
    background: var(--panel); border: 1px solid var(--border);
    border-radius: 8px; padding: 16px; margin-bottom: 16px;
  }
  .row { display: flex; gap: 12px; flex-wrap: wrap; align-items: center; }
  input[type=text], select {
    background: var(--panel2); color: var(--text);
    border: 1px solid var(--border); border-radius: 6px;
    padding: 10px 12px; font-size: 14px;
  }
  input[type=text] { flex: 1; min-width: 260px; }
  button {
    background: var(--accent); color: #fff; border: 0; border-radius: 6px;
    padding: 10px 18px; font-size: 14px; cursor: pointer; font-weight: 600;
  }
  button:hover { filter: brightness(1.12); }
  button:disabled { opacity: .5; cursor: not-allowed; }
  button.ghost { background: transparent; border: 1px solid var(--border); color: var(--text); }
  .targets { display: flex; gap: 18px; flex-wrap: wrap; margin-top: 14px; }
  .targets label { cursor: pointer; color: var(--muted); }
  .targets input { accent-color: var(--accent); margin-right: 6px; }
  .targets label.on { color: var(--text); font-weight: 600; }
  .status { display: flex; gap: 10px; flex-wrap: wrap; margin-top: 12px; }
  .chip {
    background: var(--panel2); border: 1px solid var(--border);
    border-radius: 999px; padding: 6px 14px; font-size: 12px;
  }
  .chip small { color: var(--muted); }
  table { width: 100%; border-collapse: collapse; margin-top: 4px; }
  th, td { text-align: left; padding: 10px 10px; border-bottom: 1px solid var(--border); vertical-align: top; }
  th { color: var(--muted); font-size: 12px; font-weight: 600; white-space: nowrap; }
  td a { color: #7fb4ff; text-decoration: none; }
  td a:hover { text-decoration: underline; }
  tr:hover td { background: rgba(233, 69, 96, .07); }
  .src { font-size: 11px; padding: 2px 8px; border-radius: 4px; background: var(--accent); color: #fff; white-space: nowrap; }
  .muted { color: var(--muted); }
  .log {
    background: #14142a; border: 1px solid var(--border); border-radius: 6px;
    padding: 12px; font-family: Consolas, monospace; font-size: 12px;
    max-height: 180px; overflow-y: auto; white-space: pre-wrap;
  }
  .tablewrap { overflow-x: auto; }
</style>
</head>
<body>
<header>
  <h1>Document Search Manager</h1>
  <div class="sub">社内ドキュメント横断検索 &nbsp;/&nbsp; v20260903_02 (Phase 1: SharePoint)</div>
</header>

<main>
  <div class="panel">
    <div class="row">
      <input type="text" id="keyword" placeholder="検索キーワードを入力（例: validation）" autofocus>
      <select id="maxResults">
        <option value="50">上位 50 件</option>
        <option value="100" selected>上位 100 件</option>
        <option value="200">上位 200 件</option>
        <option value="500">上位 500 件</option>
      </select>
      <button id="btnSearch">検索</button>
      <button id="btnProbe" class="ghost">疎通診断</button>
    </div>
    <div class="targets" id="targets">
      <label class="on"><input type="radio" name="target" value="all" checked>0. All</label>
      <label><input type="radio" name="target" value="sharepoint">1. SharePoint</label>
      <label><input type="radio" name="target" value="nexus">2. Nexus</label>
      <label><input type="radio" name="target" value="enovia">3. Enovia</label>
    </div>
    <div class="status" id="status"></div>
  </div>

  <div class="panel">
    <div class="row" style="justify-content: space-between;">
      <div><strong id="resultCount">検索結果: 0 件</strong> <span class="muted" id="excludedNote"></span></div>
      <div class="row">
        <button id="btnXlsx" class="ghost" disabled>Excel出力</button>
        <button id="btnCsv" class="ghost" disabled>CSV出力</button>
      </div>
    </div>
    <div class="tablewrap">
      <table id="resultTable">
        <thead>
          <tr>
            <th>ソース</th><th>Document Number</th><th>タイトル</th>
            <th>作成者</th><th>最終更新日</th><th>種別</th><th>リンク</th>
          </tr>
        </thead>
        <tbody id="resultBody">
          <tr><td colspan="7" class="muted">キーワードを入力して「検索」を押してください。</td></tr>
        </tbody>
      </table>
    </div>
  </div>

  <div class="panel">
    <div class="muted" style="margin-bottom:8px;">実行ログ</div>
    <div class="log" id="log"></div>
  </div>
</main>

<script>
var ICON_OK      = "\\u{1F7E2}";
var ICON_EMPTY   = "\\u{1F7E1}";
var ICON_ERROR   = "\\u{1F534}";
var ICON_PENDING = "\\u{26AA}";
var ICON_SEARCH  = "\\u{1F50D}";
var ICON_INFO    = "\\u{2139}";

var lastResults = [];

function log(msg) {
  var el = document.getElementById("log");
  var t = new Date().toLocaleTimeString("ja-JP");
  el.textContent += "[" + t + "] " + msg + "\\n";
  el.scrollTop = el.scrollHeight;
}

function stateIcon(state) {
  if (state === "ok") return ICON_OK;
  if (state === "empty") return ICON_EMPTY;
  if (state === "error") return ICON_ERROR;
  return ICON_PENDING;
}

function renderStatuses(statuses) {
  var box = document.getElementById("status");
  box.innerHTML = "";
  statuses.forEach(function (s) {
    var div = document.createElement("div");
    div.className = "chip";
    var head = stateIcon(s.state) + " " + s.label + " : " + s.count + " 件";
    div.textContent = head;
    if (s.message) {
      var small = document.createElement("small");
      small.textContent = "  " + s.message;
      div.appendChild(small);
    }
    box.appendChild(div);
  });
}

function renderResults(results) {
  var body = document.getElementById("resultBody");
  body.innerHTML = "";
  if (!results.length) {
    var tr = document.createElement("tr");
    var td = document.createElement("td");
    td.colSpan = 7; td.className = "muted";
    td.textContent = "該当する文書が見つかりませんでした。";
    tr.appendChild(td); body.appendChild(tr);
    return;
  }
  results.forEach(function (r) {
    var tr = document.createElement("tr");

    var tdSrc = document.createElement("td");
    var span = document.createElement("span");
    span.className = "src"; span.textContent = r.source;
    tdSrc.appendChild(span); tr.appendChild(tdSrc);

    ["document_number", "title", "author", "last_modified", "doc_type"].forEach(function (k) {
      var td = document.createElement("td");
      td.textContent = r[k] || "";
      tr.appendChild(td);
    });

    var tdLink = document.createElement("td");
    if (r.url) {
      var a = document.createElement("a");
      a.href = r.url; a.target = "_blank"; a.rel = "noopener";
      a.textContent = "開く";
      tdLink.appendChild(a);
    }
    tr.appendChild(tdLink);
    body.appendChild(tr);
  });
}

function setBusy(busy) {
  document.getElementById("btnSearch").disabled = busy;
  document.getElementById("btnProbe").disabled = busy;
}

function currentTarget() {
  var radios = document.getElementsByName("target");
  for (var i = 0; i < radios.length; i++) {
    if (radios[i].checked) return radios[i].value;
  }
  return "all";
}

document.getElementById("targets").addEventListener("change", function () {
  var labels = document.querySelectorAll("#targets label");
  labels.forEach(function (l) {
    l.className = l.querySelector("input").checked ? "on" : "";
  });
});

document.getElementById("btnSearch").addEventListener("click", function () {
  var keyword = document.getElementById("keyword").value.trim();
  if (!keyword) { log("キーワードが空です。"); return; }
  var target = currentTarget();
  var maxResults = parseInt(document.getElementById("maxResults").value, 10);

  setBusy(true);
  log(ICON_SEARCH + " 検索開始: \\"" + keyword + "\\" / 対象: " + target + " / 上限: " + maxResults + " 件");

  fetch("/api/search", {
    method: "POST",
    headers: { "Content-Type": "application/json" },
    body: JSON.stringify({ keyword: keyword, target: target, max_results: maxResults })
  })
  .then(function (r) { return r.json(); })
  .then(function (data) {
    if (data.error) { log(ICON_ERROR + " " + data.error); setBusy(false); return; }
    lastResults = data.results || [];
    renderStatuses(data.statuses || []);
    renderResults(lastResults);
    document.getElementById("resultCount").textContent = "検索結果: " + lastResults.length + " 件";
    var note = "";
    if (data.excluded_nexus > 0) {
      note = "（Nexus重複 " + data.excluded_nexus + " 件を除外）";
    }
    document.getElementById("excludedNote").textContent = note;
    document.getElementById("btnXlsx").disabled = (lastResults.length === 0);
    document.getElementById("btnCsv").disabled = (lastResults.length === 0);
    log(ICON_OK + " 検索完了: " + lastResults.length + " 件 (" + (data.elapsed_sec || 0) + " 秒)");
  })
  .catch(function (e) { log(ICON_ERROR + " 通信エラー: " + e); })
  .finally(function () { setBusy(false); });
});

document.getElementById("keyword").addEventListener("keydown", function (e) {
  if (e.key === "Enter") { document.getElementById("btnSearch").click(); }
});

document.getElementById("btnProbe").addEventListener("click", function () {
  setBusy(true);
  log(ICON_INFO + " 疎通診断を実行します...");
  fetch("/api/probe", { method: "POST" })
    .then(function (r) { return r.json(); })
    .then(function (data) {
      (data.probes || []).forEach(function (p) {
        var icon = p.ok ? ICON_OK : (p.implemented ? ICON_ERROR : ICON_PENDING);
        log(icon + " " + p.label + " [" + p.mode + "] " + p.message);
      });
    })
    .catch(function (e) { log(ICON_ERROR + " 通信エラー: " + e); })
    .finally(function () { setBusy(false); });
});

function download(fmt) {
  log(ICON_INFO + " " + fmt.toUpperCase() + " を出力しています...");
  window.location.href = "/api/export?format=" + fmt;
}
document.getElementById("btnXlsx").addEventListener("click", function () { download("xlsx"); });
document.getElementById("btnCsv").addEventListener("click", function () { download("csv"); });

log(ICON_INFO + " Document Search Manager を起動しました。");
</script>
</body>
</html>
"""


# ─────────────────────────────────────────────────────────────
# Flask アプリ
# ─────────────────────────────────────────────────────────────
flask_app = Flask(__name__)

_cfg: Dict[str, Any] = {}
_manager: Optional[SearchManager] = None
_last_results: List[SearchResult] = []
_last_keyword: str = ""
_lock = threading.Lock()


@flask_app.route("/")
def index() -> Response:
    return Response(INDEX_HTML, mimetype="text/html; charset=utf-8")


@flask_app.route("/api/probe", methods=["POST"])
def api_probe():
    try:
        return jsonify({"probes": _manager.probe_all()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@flask_app.route("/api/search", methods=["POST"])
def api_search():
    global _last_results, _last_keyword

    payload = flask_req.get_json(silent=True) or {}
    keyword = str(payload.get("keyword", "")).strip()
    target = str(payload.get("target", TARGET_ALL)).strip()
    max_results = int(payload.get("max_results", _cfg.get("default_max_results", 100)))
    max_results = max(1, min(max_results, int(_cfg.get("hard_max_results", 500))))

    if not keyword:
        return jsonify({"error": "キーワードが指定されていません。"}), 400

    started = datetime.now()
    try:
        outcome = _manager.search(keyword, target, max_results)
    except Exception as e:
        print(f"❌ 検索でエラーが発生しました: {e}")
        return jsonify({"error": f"検索でエラーが発生しました: {e}"}), 500

    elapsed = round((datetime.now() - started).total_seconds(), 1)
    results = outcome["results"]

    with _lock:
        _last_results = results
        _last_keyword = keyword

    print(f"🔍 検索: \"{keyword}\" / 対象={target} / {len(results)} 件 / {elapsed} 秒")

    return jsonify({
        "results": [asdict(r) for r in results],
        "statuses": outcome["statuses"],
        "excluded_nexus": outcome["excluded_nexus"],
        "elapsed_sec": elapsed,
    })


@flask_app.route("/api/export")
def api_export():
    fmt = (flask_req.args.get("format") or "xlsx").lower()

    with _lock:
        results = list(_last_results)
        keyword = _last_keyword

    if not results:
        return jsonify({"error": "出力対象の検索結果がありません。"}), 400

    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(JST).strftime("%Y%m%d_%H%M%S")
    safe_keyword = "".join(c for c in keyword if c.isalnum() or c in ("-", "_"))[:30]
    basename = f"document_search_{safe_keyword or 'result'}_{stamp}"

    headers = ["ソース", "Document Number", "タイトル", "作成者",
               "最終更新日", "種別", "サイト", "リンク"]

    def row_of(r: SearchResult):
        return [r.source, r.document_number, r.title, r.author,
                r.last_modified, r.doc_type, r.site, r.url]

    if fmt == "csv":
        path = EXPORT_DIR / (basename + ".csv")
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for r in results:
                writer.writerow(row_of(r))
        print(f"💾 CSVを出力しました: {path}")
        return send_file(str(path), as_attachment=True)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ModuleNotFoundError:
        return jsonify({"error": "openpyxl が未インストールです。CSV出力をご利用ください。"}), 500

    path = EXPORT_DIR / (basename + ".xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "検索結果"
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for r in results:
        ws.append(row_of(r))
    for column, width in zip("ABCDEFGH", (12, 18, 52, 20, 14, 10, 24, 60)):
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"💾 Excelを出力しました: {path}")
    return send_file(str(path), as_attachment=True)


# ─────────────────────────────────────────────────────────────
# エントリポイント
# ─────────────────────────────────────────────────────────────
def main() -> None:
    global _cfg, _manager

    print("=" * 70)
    print("📚 Document Search Manager  v20260903_02  (Phase 1: SharePoint)")
    print("=" * 70)
    print(f"📁 作業フォルダ: {BASE_DIR}")

    _cfg = _load_config()
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    auth = GraphAuthManager(_cfg["tenant_id"], _cfg["client_id"], TOKEN_CACHE_PATH)
    print("🔑 Graph APIの認証を行います（スコープ: "
          + ", ".join(_cfg["graph_scopes"]) + "）")
    auth.get_token(_cfg["graph_scopes"])
    print("✅ 認証に成功しました。")

    _manager = SearchManager(_cfg, auth)

    print("🩺 疎通診断を実行します...")
    for probe in _manager.probe_all():
        icon = "🟢" if probe.get("ok") else ("🔴" if probe.get("implemented") else "⚪")
        print(f"   {icon} {probe['label']:<11} [{probe['mode']}] {probe['message']}")

    port = int(_cfg.get("flask_port", 5020))
    url = f"http://127.0.0.1:{port}"
    print(f"\n🌐 ブラウザで次のURLを開いてください: {url}")
    print("   （終了するには、このウィンドウで Ctrl+C を押してください）\n")

    if _cfg.get("auto_open_browser", True):
        threading.Timer(1.2, lambda: webbrowser.open(url)).start()

    flask_app.run(host="127.0.0.1", port=port, debug=False, use_reloader=False)


if __name__ == "__main__":
    main()
