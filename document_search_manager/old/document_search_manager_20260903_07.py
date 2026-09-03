#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Document Search Manager  v20260903_04
=====================================
社内のドキュメント管理システムを、同一キーワードで横断検索するツール。

  0. All        … 有効な全系統を並列検索（既定）
  1. SharePoint … 社内SharePoint全社横断検索（本バージョンで実装）
  2. Nexus      … Shareflex品質文書サイト（Phase 2で実装予定）
  3. Enovia     … 3DEXPERIENCE/ENOVIA（Phase 3で実装予定）

【本バージョン(Phase 1)のスコープ】
  - SharePoint全社検索のみを実装する。Nexus / Enovia はUI上に枠だけ用意し、
    「未実装」と明示する（黙って0件を返さない）。

【認証方針】
  - 既存の po_database_organizer と同一のEntra IDアプリ登録を流用する。
  - 要求スコープは Sites.Read.All のみ。新規のGraph権限申請は一切発生させない。
  - Graphの /search/query が Sites.Read.All で通らなかった場合は、
    サイト単位の drive/root/search へ自動フォールバックする（権限追加は行わない）。

【パス方針】
  - 設定・キャッシュ・出力は全て「このスクリプトが置かれたフォルダ」基準で解決する。
    （カレントディレクトリに依存しないため、bat/ショートカットからでも安全に動く）
"""

import base64
import csv
import json
import re
import zipfile
import sys
import threading
import webbrowser
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, asdict
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional, List, Dict, Any
from urllib.parse import quote, unquote, urlparse

try:
    import msal
    import requests as http_req
    from flask import Flask, jsonify, request as flask_req, Response, send_file
except ModuleNotFoundError as e:
    print(f"\n[エラー] 必要なライブラリ '{e.name}' が見つかりません。\n"
          "次のコマンドを実行してから再度実行してください:\n\n"
          "    pip install -r requirements.txt\n")
    sys.exit(1)


# ─────────────────────────────────────────────────────────────
# パス定義（すべて __file__ 基準。カレントディレクトリに依存しない）
# ─────────────────────────────────────────────────────────────
BASE_DIR         = Path(__file__).resolve().parent
CONFIG_PATH      = BASE_DIR / "config.json"
CONFIG_EXAMPLE   = BASE_DIR / "config.example.json"

# tenant_id / client_id の流用元候補（この順に探し、最初に揃ったものを使う）。
# いずれも「読むだけ」で、書き換えは一切行わない。
CREDENTIAL_SOURCES = [
    BASE_DIR.parent / "po_database_organizer"   / "config.json",
    BASE_DIR.parent / "onenote_report_generator" / "config.json",
]

TOKEN_CACHE_PATH = BASE_DIR / "token_cache.json"
CACHE_DIR        = BASE_DIR / "cache"
EXPORT_DIR       = BASE_DIR / "exports"
DOWNLOAD_DIR     = BASE_DIR / "downloads"
STATE_PATH       = BASE_DIR / "session_state.json"

JST      = timezone(timedelta(hours=9))
GRAPH_V1 = "https://graph.microsoft.com/v1.0"

# 検索対象の識別子
TARGET_ALL        = "all"
TARGET_SHAREPOINT = "sharepoint"
TARGET_NEXUS      = "nexus"
TARGET_ENOVIA     = "enovia"

# 種別列でフォルダを示すラベル（種別フィルタでの絞り込みにも使う）
FOLDER_TYPE_LABEL = "フォルダ"


# ─────────────────────────────────────────────────────────────
# 設定ファイル読み込み
# ─────────────────────────────────────────────────────────────
DEFAULT_CFG: Dict[str, Any] = {
    "tenant_id": "",
    "client_id": "",
    "graph_scopes": ["https://graph.microsoft.com/Sites.Read.All"],
    "flask_port": 5020,
    "auto_open_browser": True,
    "page_size": 25,
    "default_max_results": 10,
    "hard_max_results": 500,
    "request_timeout_sec": 30,
    "provider_timeout_sec": 30,
    "sharepoint_host": "nexperia.sharepoint.com",
    "nexus_site_url": "https://nexperia.sharepoint.com/sites/SF_QualityDocumentsProd",
    "nexus_folder_path": "/sites/SF_QualityDocumentsProd/Documents",
    "nexus_list_id": "eb4ed9f8-81f8-4484-b8f6-64f22d30bfe0",
    "dedupe_nexus_from_sharepoint": False,
    "rewrite_host_to_mcas": False,
    "mcas_host": "nexperia.sharepoint.com.mcas.ms",
    "credentials_from": "",
    "exclude_folders": False,
    "max_download_files": 50,
    "download_timeout_sec": 120,
    "restore_last_search": True,
    "search_fields": [
        "title", "filename", "fileExtension", "path",
        "author", "lastModifiedTime", "siteTitle",
    ],
    "fallback_site_urls": [],
}


_VERSION_RE = re.compile(r"^document_search_manager_(\d{8})_(\d+)\.py$")


def _archive_old_versions() -> None:
    """自分より古いバージョンのスクリプトを old/ フォルダへ移動する。

    フォルダ直下には最新版だけが残る状態を保つ。判定は
    ファイル名の (日付, 連番) の大小で行い、自分自身と自分より新しい版は動かさない
    （古い版を起動したときに新しい版を退避させてしまわないため）。
    移動に失敗しても起動は止めない。
    """
    current = Path(__file__).name
    match = _VERSION_RE.match(current)
    if not match:
        return
    current_key = (match.group(1), int(match.group(2)))

    targets = []
    for path in BASE_DIR.glob("document_search_manager_*.py"):
        if path.name == current:
            continue
        m = _VERSION_RE.match(path.name)
        if not m:
            continue
        if (m.group(1), int(m.group(2))) < current_key:
            targets.append(path)

    if not targets:
        return

    old_dir = BASE_DIR / "old"
    try:
        old_dir.mkdir(exist_ok=True)
    except Exception as e:
        print(f"⚠️  old フォルダを作成できませんでした（移動をスキップします）: {e}")
        return

    for path in sorted(targets):
        destination = old_dir / path.name
        try:
            if destination.exists():
                # 既に退避済みの同名ファイルがある場合は上書きせず、手元の方を削除しない
                print(f"   ・old に同名あり（移動をスキップ）: {path.name}")
                continue
            path.replace(destination)
            print(f"📦 旧バージョンを old へ移動しました: {path.name}")
        except Exception as e:
            print(f"⚠️  {path.name} の移動に失敗しました（無視して続行）: {e}")


def _read_credentials(path: Path) -> Dict[str, str]:
    """設定ファイルから tenant_id / client_id だけを取り出す。

    ツールによってキーの表記が異なる（po_database_organizer は小文字、
    onenote_report_generator は大文字）ため、両方の表記を受け付ける。
    値が未設定・プレースホルダ（<YOUR_...> 形式）の場合は採用しない。
    """
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    lowered = {str(k).lower(): v for k, v in data.items()}

    def valid(value) -> str:
        text = str(value or "").strip()
        if not text or text.startswith("<"):
            return ""
        return text

    return {
        "tenant_id": valid(lowered.get("tenant_id")),
        "client_id": valid(lowered.get("client_id")),
    }


def _load_config() -> Dict[str, Any]:
    """設定を読み込む。

    tenant_id / client_id が未設定の場合は、同一リポジトリ内の既存ツールの
    config.json から自動的に借用する（読むだけで、書き換えは一切行わない）。
    探索順は CREDENTIAL_SOURCES のとおり。config.json の "credentials_from" に
    パスを書けば、任意のファイルを優先的に参照させることもできる。
    """
    cfg = dict(DEFAULT_CFG)

    if CONFIG_PATH.exists():
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                cfg.update(json.load(f))
            print(f"⚙️  設定を読み込みました: {CONFIG_PATH}")
        except Exception as e:
            print(f"❌ config.json の読み込みに失敗しました: {e}")
            raise SystemExit(1)
    else:
        print(f"⚠️  config.json が見つかりません: {CONFIG_PATH}")
        print(f"    （{CONFIG_EXAMPLE.name} をコピーして作成できます）")

    # config.json のプレースホルダ（<YOUR_...>）は未設定として扱う
    for key in ("tenant_id", "client_id"):
        value = str(cfg.get(key) or "").strip()
        cfg[key] = "" if value.startswith("<") else value

    # ── tenant_id / client_id のフォールバック（既存ツールから借用） ──
    searched: List[Path] = []
    if cfg.get("credentials_from"):
        searched.append(Path(str(cfg["credentials_from"])).expanduser())
    searched.extend(CREDENTIAL_SOURCES)

    if not cfg.get("tenant_id") or not cfg.get("client_id"):
        for source in searched:
            if cfg.get("tenant_id") and cfg.get("client_id"):
                break
            if not source.exists():
                print(f"   ・流用元なし: {source}")
                continue
            try:
                found = _read_credentials(source)
            except Exception as e:
                print(f"   ・読み込み失敗: {source} ({e})")
                continue

            borrowed = []
            for key in ("tenant_id", "client_id"):
                if not cfg.get(key) and found.get(key):
                    cfg[key] = found[key]
                    borrowed.append(key)
            if borrowed:
                print(f"🔗 {source} から {' / '.join(borrowed)} を流用しました。")
            else:
                print(f"   ・該当項目なし: {source}")

    if not cfg.get("tenant_id") or not cfg.get("client_id"):
        print("\n❌ tenant_id / client_id が特定できませんでした。")
        print("   次の場所を探しましたが、いずれにも設定がありませんでした:")
        for source in searched:
            print(f"     - {source}")
        print("\n   対処方法（どちらか一方）:")
        print(f"     (1) {CONFIG_PATH.name} の tenant_id / client_id に、既存ツールと")
        print("         同じEntra IDアプリ登録の値を直接記入する")
        print("     (2) 既存ツールの config.json があるフォルダを、")
        print("         config.json の \"credentials_from\" にフルパスで指定する")
        print("         例: \"credentials_from\": "
              "\"C:\\\\Users\\\\xxxx\\\\po_database_organizer\\\\config.json\"\n")
        raise SystemExit(1)

    return cfg


# ─────────────────────────────────────────────────────────────
# 認証（po_database_organizer の GraphAuthManager を移植）
# ─────────────────────────────────────────────────────────────
class GraphAuthManager:
    """MSAL Device Code Flow で Graph API のアクセストークンを取得・保持する。"""

    def __init__(self, tenant_id: str, client_id: str, cache_path: Path):
        self.cache_path = Path(cache_path)
        self.cache = msal.SerializableTokenCache()
        if self.cache_path.exists():
            try:
                self.cache.deserialize(self.cache_path.read_text(encoding="utf-8"))
            except Exception as e:
                print(f"⚠️  トークンキャッシュの読み込みに失敗しました（無視して続行）: {e}")

        self.app = msal.PublicClientApplication(
            client_id,
            authority=f"https://login.microsoftonline.com/{tenant_id}",
            token_cache=self.cache,
        )

    def _save_cache(self) -> None:
        if self.cache.has_state_changed:
            try:
                self.cache_path.write_text(self.cache.serialize(), encoding="utf-8")
            except Exception as e:
                print(f"⚠️  トークンキャッシュの保存に失敗しました（無視して続行）: {e}")

    def get_token(self, scopes: List[str]) -> str:
        result = None
        accounts = self.app.get_accounts()
        if accounts:
            result = self.app.acquire_token_silent(scopes, account=accounts[0])

        if not result:
            flow = self.app.initiate_device_flow(scopes=scopes)
            if "user_code" not in flow:
                raise RuntimeError(f"デバイスコードフローの開始に失敗しました: {flow}")
            print("\n" + "=" * 70)
            print("🔑 サインインが必要です。以下の手順で認証してください。")
            print(flow["message"])
            print("=" * 70 + "\n")
            result = self.app.acquire_token_by_device_flow(flow)

        if "access_token" not in result:
            raise RuntimeError(
                f"アクセストークンの取得に失敗しました: "
                f"{result.get('error')} / {result.get('error_description')}"
            )

        self._save_cache()
        return result["access_token"]


# ─────────────────────────────────────────────────────────────
# 検索結果の正規化スキーマ
# ─────────────────────────────────────────────────────────────
@dataclass
class SearchResult:
    """3系統の検索結果を統一表現するためのスキーマ。

    Q-C（まずは一覧まで）の回答に従い、本文スニペットは保持しない。
    """
    source: str = ""             # SharePoint / Nexus / Enovia
    document_number: str = ""    # Nexus等の文書番号（SharePoint全社検索では通常空）
    title: str = ""
    author: str = ""
    last_modified: str = ""      # YYYY-MM-DD（JST）
    doc_type: str = ""           # 拡張子など
    site: str = ""               # 所属サイト名（表示用）
    site_url: str = ""           # 所属サイトのトップURL
    folder: str = ""             # サイト内の保管フォルダパス（表示用・短縮）
    folder_full: str = ""        # 同上（ライブラリ名を含むフルパス）
    folder_url: str = ""         # 保管フォルダを開くURL
    is_folder: bool = False      # ヒットがフォルダ自体かどうか
    url: str = ""
    is_nexus_path: bool = False  # Nexusサイト配下の文書か（重複判定用）
    rank: int = 0


# ─────────────────────────────────────────────────────────────
# ユーティリティ
# ─────────────────────────────────────────────────────────────
def _pick(*candidates) -> str:
    """複数の候補値から、最初に見つかった非空の文字列を返す。"""
    for c in candidates:
        if c is None:
            continue
        s = str(c).strip()
        if s and s.lower() != "none":
            return s
    return ""


def _lower_keys(d: Any) -> Dict[str, Any]:
    """辞書のキーを小文字化した辞書を返す（Graphのfields名の揺れを吸収する）。"""
    if not isinstance(d, dict):
        return {}
    return {str(k).lower(): v for k, v in d.items()}


def _format_datetime(value: str) -> str:
    """ISO8601形式の日時をJSTの YYYY-MM-DD に整形する。失敗時は原文を返す。"""
    if not value:
        return ""
    text = str(value).strip()
    try:
        normalized = text.replace("Z", "+00:00")
        dt = datetime.fromisoformat(normalized)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(JST).strftime("%Y-%m-%d")
    except Exception:
        return text[:10]


def _name_from_url(url: str) -> str:
    """URLの末尾セグメントからファイル名を復元する。

    Graphの listItem はタイトルを返さない場合があるため、
    リンク先URLからファイル名を取り出して代替タイトルとして使う。
    """
    if not url:
        return ""
    try:
        path = urlparse(str(url)).path
    except Exception:
        return ""
    segment = path.rstrip("/").rsplit("/", 1)[-1]
    if not segment:
        return ""
    try:
        return unquote(segment)
    except Exception:
        return segment


def _site_url_from_url(url: str) -> str:
    """文書URLから、その文書が置かれているサイトのトップURLを導く。

    例: https://host/sites/JapanTE/Shared Documents/a.pptx
        → https://host/sites/JapanTE
    /sites/ や /teams/ を含まない場合は、ホストのトップURLを返す。
    """
    if not url:
        return ""
    try:
        parsed = urlparse(str(url))
    except Exception:
        return ""
    if not parsed.scheme or not parsed.netloc:
        return ""

    root = f"{parsed.scheme}://{parsed.netloc}"
    segments = [s for s in parsed.path.split("/") if s]
    for prefix in ("sites", "teams", "personal"):
        if prefix in segments:
            index = segments.index(prefix)
            if index + 1 < len(segments):
                return f"{root}/{segments[index]}/{segments[index + 1]}"
    return root


def _site_name_from_url(site_url: str) -> str:
    """サイトURLから表示用のサイト名（最後のセグメント）を取り出す。"""
    if not site_url:
        return ""
    try:
        parsed = urlparse(str(site_url))
    except Exception:
        return ""
    segments = [s for s in parsed.path.split("/") if s]
    if segments:
        return unquote(segments[-1])
    return parsed.netloc


def _folder_of_url(url: str, site_url: str, drop_last: bool = True):
    """URLから (フォルダの表示パス, フォルダを開くURL) を導く。

    drop_last=True  … ファイルURLを渡し、その「格納フォルダ」を得る
    drop_last=False … フォルダURLを渡し、「そのフォルダ自体」を得る

    例: https://host/sites/JapanTE/Shared%20Documents/2026/a.pptx
        表示パス : Shared Documents/2026
        フォルダURL:
          https://host/sites/JapanTE/Shared%20Documents/Forms/AllItems.aspx
          ?id=%2Fsites%2FJapanTE%2FShared%20Documents%2F2026
        （SharePointがフォルダを開くときに使う標準の形式）
    サイト直下にファイルがある等でフォルダを特定できない場合は ("", "") を返す。
    """
    if not url or not site_url:
        return "", ""
    try:
        parsed = urlparse(str(url))
        site = urlparse(str(site_url))
    except Exception:
        return "", ""
    if not parsed.path or not site.path:
        return "", ""

    site_path = site.path.rstrip("/")
    if not parsed.path.lower().startswith(site_path.lower() + "/"):
        return "", ""

    rest = parsed.path[len(site_path):].strip("/")
    segments = [s for s in rest.split("/") if s]

    folder_segments = segments[:-1] if drop_last else segments
    if not folder_segments:
        return "", ""          # フォルダ階層を特定できない

    library = folder_segments[0]             # ライブラリ名（符号化されたまま）

    display = unquote("/".join(folder_segments))
    server_relative = site_path + "/" + display

    folder_url = (f"{site.scheme}://{site.netloc}{site_path}/{library}"
                  f"/Forms/AllItems.aspx?id={quote(server_relative, safe='')}")
    return display, folder_url


# 「Shared Documents」のような既定のドキュメントライブラリ名は、どのサイトにも
# 存在して情報価値が低いため、表示上はライブラリのルートを示す "/" に置き換える。
# PO のような固有のライブラリ名は情報価値があるのでそのまま残す。
DEFAULT_LIBRARY_NAMES = {
    "shared documents", "documents", "共有ドキュメント", "ドキュメント",
}


def _shorten_folder(full_path: str) -> str:
    """フォルダのフルパスを表示用に短縮する。

      Shared Documents                                   → /
      Shared Documents/40. Bench Validation              → /40. Bench Validation
      PO/2026/Vendor                                     → PO/2026/Vendor（そのまま）
    """
    if not full_path:
        return ""
    segments = [s for s in str(full_path).split("/") if s]
    if not segments:
        return ""
    if segments[0].strip().lower() not in DEFAULT_LIBRARY_NAMES:
        return full_path
    return "/" + "/".join(segments[1:])


def _looks_like_folder(url: str, fields: Dict[str, Any]) -> bool:
    """ヒットがファイルではなくフォルダ自体かどうかを判定する。

    Graphの検索結果にはフォルダ自体も含まれるため、一覧をファイルだけに
    そろえる目的で使う。判定は次の順で行う。
      1. fields に isDocument / contentclass があればそれを使う
      2. 無ければ、URL末尾のセグメントに拡張子が無いものをフォルダとみなす
    """
    is_document = fields.get("isdocument")
    if is_document is not None:
        return str(is_document).strip().lower() in ("false", "0", "no")

    content_class = str(fields.get("contentclass") or "").lower()
    if content_class:
        return "folder" in content_class

    name = _name_from_url(url)
    if not name:
        return False
    # 拡張子と判断できる末尾を持たないものはフォルダとみなす
    return _extension_of(name) == ""


# 拡張子とみなす文字列の条件。
#   - 英字で始まる1〜10文字の英数字（docx / pptx / xlsm / pdf など）
#   - 例外として 7z を許容する
# 社内では「12. Validation」「40. Bench Validation」「MRA2.0」のように
# 番号やバージョンを含むフォルダ名が多く、最後のピリオド以降を無条件に拡張子と
# みなすと「. Validation」「0」などを拡張子と誤認してしまうため、条件を絞る。
_EXT_RE = re.compile(r"^(?:[A-Za-z][A-Za-z0-9]{0,9}|7[zZ])$")


def _decode_name(name: str) -> str:
    """URLエンコードが残っている名前を復号する（%20 → 空白 など）。"""
    text = str(name or "")
    if "%" not in text:
        return text
    try:
        return unquote(text)
    except Exception:
        return text


def _extension_of(name: str) -> str:
    """ファイル名から拡張子を取り出す。拡張子と判断できない場合は空を返す。"""
    text = _decode_name(name)
    if not text or "." not in text:
        return ""
    candidate = text.rsplit(".", 1)[-1]
    return candidate.lower() if _EXT_RE.match(candidate) else ""


def _rewrite_url(url: str, cfg: Dict[str, Any]) -> str:
    """設定に応じてリンクのホストを .mcas.ms 経由に書き換える。"""
    if not url or not cfg.get("rewrite_host_to_mcas"):
        return url
    host = cfg.get("sharepoint_host", "")
    mcas = cfg.get("mcas_host", "")
    if host and mcas and host in url and mcas not in url:
        return url.replace(host, mcas, 1)
    return url


# ─────────────────────────────────────────────────────────────
# 検索プロバイダ（抽象）
# ─────────────────────────────────────────────────────────────
class SearchProvider:
    """検索対象1系統を表す抽象クラス。

    Nexus / Enovia を後から足すときに、この I/F だけを満たせばよい構造にする。
    """
    key = ""
    label = ""
    implemented = False

    def __init__(self, cfg: Dict[str, Any], auth: Optional[GraphAuthManager]):
        self.cfg = cfg
        self.auth = auth

    def probe(self) -> Dict[str, Any]:
        """疎通診断。{'ok': bool, 'mode': str, 'message': str} を返す。"""
        return {"ok": False, "mode": "未実装", "message": "この系統は未実装です。"}

    def search(self, keyword: str, max_results: int) -> Dict[str, Any]:
        """検索を実行し、{'results': [...], 'total': int, 'note': str} を返す。"""
        raise NotImplementedError


# ─────────────────────────────────────────────────────────────
# 1. SharePoint 全社横断検索プロバイダ
# ─────────────────────────────────────────────────────────────
class SharePointProvider(SearchProvider):
    """Microsoft Graph Search API による社内SharePoint全社横断検索。

    第一候補: POST /search/query  (entityTypes = listItem)
      - SharePointライブラリ上の文書は listItem として返るため、
        Sites.Read.All のみで全社検索が成立することを狙う。
    フォールバック: GET /sites/{host}:{path}:/drive/root/search(q='...')
      - config.json の fallback_site_urls に列挙したサイトを個別に検索する。
      - こちらは Sites.Read.All で確実に動作する。
    """
    key = TARGET_SHAREPOINT
    label = "SharePoint"
    implemented = True

    # 疎通診断の結果を保持（毎回判定し直さない）
    _mode: Optional[str] = None   # "search-api" / "site-drive" / None
    # fields指定がテナントに拒否された場合に True（以後 fields を要求しない）
    _fields_disabled: bool = False

    def _token(self) -> str:
        return self.auth.get_token(self.cfg["graph_scopes"])

    # ── 疎通診断 ────────────────────────────────────────────
    def probe(self) -> Dict[str, Any]:
        try:
            token = self._token()
        except Exception as e:
            return {"ok": False, "mode": "認証エラー", "message": f"認証に失敗しました: {e}"}

        status, _body, err = self._call_search_api(token, "test", frm=0, size=1)

        if status == 200:
            self._mode = "search-api"
            return {
                "ok": True,
                "mode": "search-api",
                "message": "全社横断検索が利用できます（/search/query・listItem方式）。",
            }

        if status in (401, 403):
            sites = self.cfg.get("fallback_site_urls") or []
            if sites:
                self._mode = "site-drive"
                return {
                    "ok": True,
                    "mode": "site-drive",
                    "message": (f"全社横断検索は権限不足のため、サイト単位検索に切り替えます"
                                f"（対象 {len(sites)} サイト）。"),
                }
            self._mode = None
            return {
                "ok": False,
                "mode": "権限不足",
                "message": (f"/search/query が HTTP {status} で拒否されました。"
                            f"config.json の fallback_site_urls に検索したいサイトURLを"
                            f"設定するとサイト単位検索に切り替えられます。"),
            }

        self._mode = None
        return {"ok": False, "mode": "エラー", "message": f"HTTP {status}: {err}"}

    # ── Graph Search API 呼び出し ──────────────────────────
    def _call_search_api(self, token: str, query_string: str, frm: int, size: int):
        """(status_code, body_dict, error_text) を返す。例外は投げない。

        Graphの listItem は、既定ではタイトル等のメタデータを返さない。
        そのため fields（検索マネージドプロパティ名）を明示的に要求する。
        指定した名前がテナントに存在しない場合はHTTP 400になり得るため、
        400が返ったときは fields 無しで1度だけ再試行し、以後は要求しない。
        """
        request: Dict[str, Any] = {
            "entityTypes": ["listItem"],
            "query": {"queryString": query_string},
            "from": frm,
            "size": size,
        }

        search_fields = self.cfg.get("search_fields") or []
        use_fields = bool(search_fields) and not self._fields_disabled
        if use_fields:
            request["fields"] = list(search_fields)

        try:
            resp = http_req.post(
                f"{GRAPH_V1}/search/query",
                headers={
                    "Authorization": f"Bearer {token}",
                    "Content-Type": "application/json",
                },
                json={"requests": [request]},
                timeout=self.cfg.get("request_timeout_sec", 30),
            )
        except Exception as e:
            return 0, None, str(e)

        # fields が原因で拒否された場合は、fields 無しで再試行する
        if resp.status_code == 400 and use_fields:
            print("⚠️  fields指定が拒否されたため、fields無しで再試行します。"
                  "（タイトル等はURLから補完します）")
            self._fields_disabled = True
            return self._call_search_api(token, query_string, frm, size)

        if resp.status_code != 200:
            return resp.status_code, None, resp.text[:500]

        try:
            return 200, resp.json(), ""
        except Exception as e:
            return 200, None, f"レスポンスのJSON解析に失敗しました: {e}"

    # ── 検索本体 ────────────────────────────────────────────
    def search(self, keyword: str, max_results: int) -> Dict[str, Any]:
        token = self._token()

        if self._mode is None:
            self.probe()

        if self._mode == "site-drive":
            return self._search_by_site_drive(token, keyword, max_results)

        return self._search_by_search_api(token, keyword, max_results)

    def _search_by_search_api(self, token: str, keyword: str,
                              max_results: int) -> Dict[str, Any]:
        page_size = int(self.cfg.get("page_size", 25))
        results: List[SearchResult] = []
        total = 0
        frm = 0
        rank = 0

        while len(results) < max_results:
            size = min(page_size, max_results - len(results))
            status, body, err = self._call_search_api(token, keyword, frm, size)
            if status != 200 or body is None:
                if results:
                    break  # 途中まで取れていれば、それを返す
                raise RuntimeError(f"HTTP {status}: {err}")

            containers = (body.get("value") or [{}])[0].get("hitsContainers") or []
            if not containers:
                break

            container = containers[0]
            total = container.get("total", total) or total
            hits = container.get("hits") or []
            if not hits:
                break

            for hit in hits:
                result = self._hit_to_result(hit, rank + 1)
                # 既定ではフォルダも残す（案A: 区別して表示）。
                # exclude_folders を true にした場合のみ、一覧から取り除く。
                if result.is_folder and self.cfg.get("exclude_folders"):
                    continue
                rank += 1
                results.append(result)

            if not container.get("moreResultsAvailable"):
                break
            frm += len(hits)

        return {"results": results, "total": total or len(results), "note": ""}

    def _hit_to_result(self, hit: Dict[str, Any], rank: int) -> SearchResult:
        """Graph Search の hit を SearchResult に変換する。

        Graphが返すフィールド名は環境により揺れるため、複数の候補を順に見る
        寛容なパーサとする（形が変わっても落ちないことを優先する）。
        """
        res = hit.get("resource") or {}
        fields = _lower_keys(res.get("fields") or {})

        url = _pick(res.get("webUrl"), fields.get("path"), fields.get("spwebur"))

        # listItem は既定でタイトルを返さないことがあるため、
        # fields → resource.name → URL末尾のファイル名、の順に補完する
        name = _decode_name(_pick(fields.get("filename"), fields.get("name"),
                                  res.get("name"), _name_from_url(url)))
        title = _pick(fields.get("title"), name, "(タイトルなし)")

        author = _pick(
            fields.get("author"),
            fields.get("createdby"),
            ((res.get("createdBy") or {}).get("user") or {}).get("displayName"),
            ((res.get("lastModifiedBy") or {}).get("user") or {}).get("displayName"),
        )

        last_modified = _format_datetime(_pick(
            fields.get("lastmodifiedtime"),
            fields.get("lastmodifieddatetime"),
            res.get("lastModifiedDateTime"),
        ))

        # ヒットがフォルダ自体か、ファイルかを判定する。
        # フォルダは除外せず「種別=フォルダ」として区別する
        # （フォルダ名だけがキーワードに一致し、中のファイルには一致しない
        #   ケースでは、フォルダの存在そのものが手がかりになるため）。
        is_folder = _looks_like_folder(url, fields)

        if is_folder:
            doc_type = FOLDER_TYPE_LABEL
        else:
            raw_ext = _decode_name(_pick(fields.get("fileextension"),
                                         fields.get("filetype"))).lower()
            if raw_ext and not _EXT_RE.match(raw_ext):
                raw_ext = ""      # fields側が壊れた値を返した場合は採用しない
            doc_type = _pick(raw_ext, _extension_of(name))

        document_number = _pick(
            fields.get("documentnumber"),
            fields.get("document_x0020_number"),
            fields.get("oldsystemidentifier"),
        )

        # 文書が置かれているサイト。fields.sitetitle があればそれを表示名にし、
        # 無ければ文書URLからサイトURLを導いて、その最後のセグメントを使う。
        site_url = _pick(fields.get("spsiteurl"), _site_url_from_url(url))
        site = _pick(fields.get("sitetitle"), _site_name_from_url(site_url))

        # ファイルなら「格納フォルダ」、フォルダ自体なら「その親フォルダ」を示す
        folder_full, folder_url = _folder_of_url(url, site_url)
        folder = _shorten_folder(folder_full)

        # フォルダ行は、タイトルのリンク先をそのフォルダを開くURLにする
        link_url = url
        if is_folder:
            _, self_url = _folder_of_url(url, site_url, drop_last=False)
            link_url = _pick(self_url, url)

        nexus_url = str(self.cfg.get("nexus_site_url", "")).rstrip("/")
        is_nexus = bool(nexus_url) and url.lower().startswith(nexus_url.lower())

        return SearchResult(
            source=self.label,
            document_number=document_number,
            title=title,
            author=author,
            last_modified=last_modified,
            doc_type=doc_type,
            site=site,
            site_url=_rewrite_url(site_url, self.cfg),
            folder=folder,
            folder_full=folder_full,
            folder_url=_rewrite_url(folder_url, self.cfg),
            is_folder=is_folder,
            url=_rewrite_url(link_url, self.cfg),
            is_nexus_path=is_nexus,
            rank=rank,
        )

    # ── フォールバック（サイト単位検索） ───────────────────
    def _search_by_site_drive(self, token: str, keyword: str,
                              max_results: int) -> Dict[str, Any]:
        results: List[SearchResult] = []
        rank = 0
        errors: List[str] = []

        for site_url in (self.cfg.get("fallback_site_urls") or []):
            if len(results) >= max_results:
                break
            try:
                host, path = self._split_site_url(site_url)
            except Exception as e:
                errors.append(f"{site_url}: {e}")
                continue

            endpoint = (f"{GRAPH_V1}/sites/{host}:{path}:/drive/root/"
                        f"search(q='{keyword}')")
            try:
                resp = http_req.get(
                    endpoint,
                    headers={"Authorization": f"Bearer {token}"},
                    params={"$top": min(50, max_results - len(results))},
                    timeout=self.cfg.get("request_timeout_sec", 30),
                )
            except Exception as e:
                errors.append(f"{site_url}: {e}")
                continue

            if resp.status_code != 200:
                errors.append(f"{site_url}: HTTP {resp.status_code}")
                continue

            for item in (resp.json().get("value") or []):
                if len(results) >= max_results:
                    break
                rank += 1
                name = _pick(item.get("name"))
                url = _pick(item.get("webUrl"))
                nexus_url = str(self.cfg.get("nexus_site_url", "")).rstrip("/")
                item_site_url = _pick(_site_url_from_url(url), site_url)
                item_is_folder = ("folder" in item) or _looks_like_folder(url, {})
                item_folder_full, item_folder_url = _folder_of_url(url, item_site_url)
                item_folder = _shorten_folder(item_folder_full)
                if item_is_folder and self.cfg.get("exclude_folders"):
                    continue
                results.append(SearchResult(
                    source=self.label,
                    document_number="",
                    title=_pick(name, _name_from_url(url), "(タイトルなし)"),
                    author=_pick(((item.get("createdBy") or {}).get("user") or {})
                                 .get("displayName")),
                    last_modified=_format_datetime(item.get("lastModifiedDateTime", "")),
                    doc_type=(FOLDER_TYPE_LABEL if item_is_folder else
                              _pick(_extension_of(name),
                                    _extension_of(_name_from_url(url)))),
                    site=_site_name_from_url(item_site_url),
                    site_url=_rewrite_url(item_site_url, self.cfg),
                    folder=item_folder,
                    folder_full=item_folder_full,
                    folder_url=_rewrite_url(item_folder_url, self.cfg),
                    is_folder=item_is_folder,
                    url=_rewrite_url(url, self.cfg),
                    is_nexus_path=(bool(nexus_url)
                                   and url.lower().startswith(nexus_url.lower())),
                    rank=rank,
                ))

        note = "サイト単位検索モードで実行しました。"
        if errors:
            note += " 失敗したサイト: " + " / ".join(errors[:3])
        return {"results": results, "total": len(results), "note": note}

    @staticmethod
    def _split_site_url(site_url: str):
        """https://host/sites/xxx を (host, /sites/xxx) に分解する。"""
        text = site_url.strip().rstrip("/")
        if not text.startswith("http"):
            raise ValueError("URLは https:// から始めてください")
        without_scheme = text.split("://", 1)[1]
        if "/" not in without_scheme:
            raise ValueError("サイトパスが含まれていません")
        host, path = without_scheme.split("/", 1)
        return host, "/" + path


# ─────────────────────────────────────────────────────────────
# 2. Nexus プロバイダ（Phase 2で実装予定）
# ─────────────────────────────────────────────────────────────
class NexusProvider(SearchProvider):
    key = TARGET_NEXUS
    label = "Nexus"
    implemented = False

    def probe(self) -> Dict[str, Any]:
        return {
            "ok": False,
            "mode": "未実装",
            "message": ("Phase 2で実装予定です。Graph Search を "
                        + str(self.cfg.get("nexus_folder_path", ""))
                        + " に限定して検索する方式を予定しています。"),
        }

    def search(self, keyword: str, max_results: int) -> Dict[str, Any]:
        raise NotImplementedError("Nexus検索はPhase 2で実装予定です。")


# ─────────────────────────────────────────────────────────────
# 3. Enovia プロバイダ（Phase 3で実装予定）
# ─────────────────────────────────────────────────────────────
class EnoviaProvider(SearchProvider):
    key = TARGET_ENOVIA
    label = "Enovia"
    implemented = False

    def probe(self) -> Dict[str, Any]:
        return {
            "ok": False,
            "mode": "未実装",
            "message": "Phase 3で実装予定です（実装方式は調査後に確定）。",
        }

    def search(self, keyword: str, max_results: int) -> Dict[str, Any]:
        raise NotImplementedError("Enovia検索はPhase 3で実装予定です。")


# ─────────────────────────────────────────────────────────────
# 実行層（並列実行・部分成功）
# ─────────────────────────────────────────────────────────────
class SearchManager:
    """複数プロバイダを並列実行し、結果をマージする。

    1系統が失敗しても他系統の結果は必ず返す（部分成功方式）。
    """

    def __init__(self, cfg: Dict[str, Any], auth: GraphAuthManager):
        self.cfg = cfg
        self.providers: Dict[str, SearchProvider] = {
            TARGET_SHAREPOINT: SharePointProvider(cfg, auth),
            TARGET_NEXUS:      NexusProvider(cfg, auth),
            TARGET_ENOVIA:     EnoviaProvider(cfg, auth),
        }

    def probe_all(self) -> List[Dict[str, Any]]:
        out = []
        for key, provider in self.providers.items():
            try:
                result = provider.probe()
            except Exception as e:
                result = {"ok": False, "mode": "エラー", "message": str(e)}
            result.update({"key": key, "label": provider.label,
                           "implemented": provider.implemented})
            out.append(result)
        return out

    def _targets(self, target: str) -> List[SearchProvider]:
        if target == TARGET_ALL:
            return [p for p in self.providers.values() if p.implemented]
        provider = self.providers.get(target)
        # 未実装の系統は実行対象に含めない（呼び出し側で「未実装」として扱う）
        return [provider] if provider and provider.implemented else []

    def search(self, keyword: str, target: str, max_results: int) -> Dict[str, Any]:
        providers = self._targets(target)

        # 未実装の系統が明示的に選ばれた場合は、その旨をそのまま返す
        if not providers:
            provider = self.providers.get(target)
            if provider and not provider.implemented:
                return {
                    "results": [],
                    "statuses": [dict(provider.probe(), key=provider.key,
                                      label=provider.label, implemented=False,
                                      state="pending", count=0, total=0)],
                    "excluded_nexus": 0,
                }
            return {"results": [], "statuses": [], "excluded_nexus": 0}

        statuses: List[Dict[str, Any]] = []
        merged: List[SearchResult] = []
        timeout = self.cfg.get("provider_timeout_sec", 30)

        with ThreadPoolExecutor(max_workers=max(1, len(providers))) as pool:
            futures = {
                pool.submit(p.search, keyword, max_results): p for p in providers
            }
            for future in as_completed(futures, timeout=timeout * len(providers)):
                provider = futures[future]
                try:
                    outcome = future.result(timeout=timeout)
                    results = outcome.get("results") or []
                    merged.extend(results)
                    statuses.append({
                        "key": provider.key,
                        "label": provider.label,
                        "implemented": True,
                        "state": "ok" if results else "empty",
                        "count": len(results),
                        "total": outcome.get("total", len(results)),
                        "message": outcome.get("note", ""),
                    })
                except Exception as e:
                    statuses.append({
                        "key": provider.key,
                        "label": provider.label,
                        "implemented": True,
                        "state": "error",
                        "count": 0,
                        "total": 0,
                        "message": str(e),
                    })

        # 未実装の系統も「0. All」のときは状態として見せる（黙って隠さない）
        if target == TARGET_ALL:
            for provider in self.providers.values():
                if provider.implemented:
                    continue
                info = provider.probe()
                statuses.append({
                    "key": provider.key,
                    "label": provider.label,
                    "implemented": False,
                    "state": "pending",
                    "count": 0,
                    "total": 0,
                    "message": info.get("message", ""),
                })

        merged, excluded = self._dedupe(merged)
        merged.sort(key=lambda r: (r.source, r.rank))

        statuses.sort(key=lambda s: ["sharepoint", "nexus", "enovia"].index(s["key"])
                      if s["key"] in ("sharepoint", "nexus", "enovia") else 99)

        return {"results": merged, "statuses": statuses, "excluded_nexus": excluded}

    def _dedupe(self, results: List[SearchResult]):
        """Nexusサイト配下の文書をSharePoint結果から除外する。

        Phase 1 では Nexus プロバイダが未実装のため、既定では除外しない
        （除外するとNexus文書が一切見えなくなってしまうため）。
        Phase 2 で Nexus プロバイダを有効化するのと同時に、
        config.json の dedupe_nexus_from_sharepoint を true にする。
        """
        if not self.cfg.get("dedupe_nexus_from_sharepoint"):
            return results, 0

        kept, excluded = [], 0
        for r in results:
            if r.source == "SharePoint" and r.is_nexus_path:
                excluded += 1
                continue
            kept.append(r)
        return kept, excluded


# ─────────────────────────────────────────────────────────────
# 画面（ダークテーマ）
# ─────────────────────────────────────────────────────────────
# 絵文字はJS側の文字列定数として定義する（Python f-string の外で扱う）
INDEX_HTML = r"""<!doctype html>
<html lang="ja">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Document Search Manager</title>
<style>
  :root {
    --bg: #1a1a2e;
    --panel: #23233f;
    --panel2: #2c2c4d;
    --accent: #e94560;
    --text: #e8e8f0;
    --muted: #9a9ab5;
    --border: #3a3a5c;
  }
  * { box-sizing: border-box; }
  body {
    margin: 0; background: var(--bg); color: var(--text);
    font-family: "Segoe UI", "Meiryo", system-ui, sans-serif; font-size: 14px;
  }
  header {
    background: linear-gradient(90deg, #16213e, #1a1a2e);
    border-bottom: 2px solid var(--accent); padding: 16px 24px;
  }
  header h1 { margin: 0; font-size: 20px; letter-spacing: .04em; }
  header .sub { color: var(--muted); font-size: 12px; margin-top: 4px; }
  main { padding: 20px 24px 60px; max-width: 1600px; }
  .panel {
    background: var(--panel); border: 1px solid var(--border);
    border-radius: 8px; padding: 16px; margin-bottom: 16px;
  }
  .row { display: flex; gap: 12px; flex-wrap: wrap; align-items: center; }
  input[type=text], input[type=date], select {
    background: var(--panel2); color: var(--text);
    border: 1px solid var(--border); border-radius: 6px;
    padding: 10px 12px; font-size: 14px;
  }
  input[type=text].wide { flex: 1; min-width: 260px; }
  button {
    background: var(--accent); color: #fff; border: 0; border-radius: 6px;
    padding: 10px 18px; font-size: 14px; cursor: pointer; font-weight: 600;
  }
  button:hover { filter: brightness(1.12); }
  button:disabled { opacity: .5; cursor: not-allowed; }
  button.ghost { background: transparent; border: 1px solid var(--border); color: var(--text); }
  button.mini {
    padding: 4px 10px; font-size: 12px; font-weight: 500;
    background: transparent; border: 1px solid var(--border); color: var(--text);
  }
  .targets { display: flex; gap: 18px; flex-wrap: wrap; margin-top: 14px; }
  .targets label { cursor: pointer; color: var(--muted); }
  .targets input { accent-color: var(--accent); margin-right: 6px; }
  .targets label.on { color: var(--text); font-weight: 600; }
  .status { display: flex; gap: 10px; flex-wrap: wrap; margin-top: 12px; }
  .chip {
    background: var(--panel2); border: 1px solid var(--border);
    border-radius: 999px; padding: 6px 14px; font-size: 12px;
  }
  .chip small { color: var(--muted); }

  table { width: 100%; border-collapse: collapse; margin-top: 4px; }
  th, td { text-align: left; padding: 10px 10px; border-bottom: 1px solid var(--border); vertical-align: top; }
  th { color: var(--muted); font-size: 12px; font-weight: 600; white-space: nowrap; position: relative; }
  th .hdr { cursor: pointer; user-select: none; }
  th .hdr:hover { color: var(--text); }
  th .arrow { color: var(--accent); margin-left: 4px; }
  th .filt {
    display: inline-block; margin-left: 6px; padding: 0 5px; cursor: pointer;
    border: 1px solid var(--border); border-radius: 4px; color: var(--muted);
    background: transparent; font-size: 12px; line-height: 18px;
  }
  th .filt:hover { color: var(--text); }
  th .filt.on { color: #fff; background: var(--accent); border-color: var(--accent); }
  td a { color: #7fb4ff; text-decoration: none; }
  td a:hover { text-decoration: underline; }
  tr:hover td { background: rgba(233, 69, 96, .07); }
  .src { font-size: 11px; padding: 2px 8px; border-radius: 4px; background: var(--accent); color: #fff; white-space: nowrap; }
  .muted { color: var(--muted); }
  td.nowrap { white-space: nowrap; }
  td.title { min-width: 280px; }
  th.selcol, td.selcol { width: 46px; text-align: center; }
  th.selcol input, td.selcol input { accent-color: var(--accent); cursor: pointer; }
  td.title a { font-weight: 600; }

  .fpanel {
    position: fixed; z-index: 200;
    background: var(--panel2); border: 1px solid var(--border); border-radius: 6px;
    padding: 10px; min-width: 240px; max-width: 340px;
    box-shadow: 0 8px 24px rgba(0, 0, 0, .45); white-space: normal;
  }
  .fpanel .frow { display: flex; gap: 6px; margin-bottom: 8px; align-items: center; }
  .fpanel input[type=text], .fpanel input[type=date] { padding: 5px 8px; font-size: 12px; width: 100%; }
  .fpanel .flist { max-height: 240px; overflow-y: auto; }
  .fpanel label {
    display: flex; gap: 7px; align-items: flex-start; padding: 3px 2px;
    font-size: 12px; font-weight: 400; color: var(--text); cursor: pointer;
  }
  .fpanel label:hover { background: rgba(233, 69, 96, .12); }
  .fpanel input[type=checkbox] { accent-color: var(--accent); margin-top: 2px; }
  .fpanel .cnt { color: var(--muted); margin-left: auto; padding-left: 8px; }
  .fpanel .hint { color: var(--muted); font-size: 11px; margin-bottom: 6px; }

  .log {
    background: #14142a; border: 1px solid var(--border); border-radius: 6px;
    padding: 12px; font-family: Consolas, monospace; font-size: 12px;
    max-height: 180px; overflow-y: auto; white-space: pre-wrap;
  }
  .tablewrap { overflow-x: auto; }
</style>
</head>
<body>
<header>
  <h1>Document Search Manager</h1>
  <div class="sub">社内ドキュメント横断検索 &nbsp;/&nbsp; v20260903_07 (Phase 1: SharePoint)</div>
</header>

<main>
  <div class="panel">
    <div class="row">
      <input type="text" class="wide" id="keyword" placeholder="検索キーワードを入力（例: validation）" autofocus>
      <select id="maxResults">
        <option value="10" selected>上位 10 件</option>
        <option value="25">上位 25 件</option>
        <option value="50">上位 50 件</option>
        <option value="100">上位 100 件</option>
        <option value="200">上位 200 件</option>
        <option value="500">上位 500 件</option>
      </select>
      <button id="btnSearch">検索</button>
      <button id="btnProbe" class="ghost">疎通診断</button>
    </div>
    <div class="targets" id="targets">
      <label class="on"><input type="radio" name="target" value="all" checked>0. All</label>
      <label><input type="radio" name="target" value="sharepoint">1. SharePoint</label>
      <label><input type="radio" name="target" value="nexus">2. Nexus</label>
      <label><input type="radio" name="target" value="enovia">3. Enovia</label>
    </div>
    <div class="status" id="status"></div>
  </div>

  <div class="panel">
    <div class="row" style="justify-content: space-between;">
      <div>
        <strong id="resultCount">検索結果: 0 件</strong>
        <span class="muted" id="excludedNote"></span>
      </div>
      <div class="row">
        <button id="btnClearFilter" class="mini">フィルタ・ソート解除</button>
        <button id="btnDownload" disabled>選択ファイルをZIPで取得</button>
        <button id="btnXlsx" class="ghost" disabled>Excel出力</button>
        <button id="btnCsv" class="ghost" disabled>CSV出力</button>
      </div>
    </div>
    <div class="tablewrap">
      <table id="resultTable">
        <thead><tr id="resultHead"></tr></thead>
        <tbody id="resultBody">
          <tr><td class="muted">キーワードを入力して「検索」を押してください。</td></tr>
        </tbody>
      </table>
    </div>
  </div>

  <div class="panel">
    <div class="muted" style="margin-bottom:8px;">実行ログ</div>
    <div class="log" id="log"></div>
  </div>
</main>

<script>
var ICON_OK      = "\u{1F7E2}";
var ICON_EMPTY   = "\u{1F7E1}";
var ICON_ERROR   = "\u{1F534}";
var ICON_PENDING = "\u{26AA}";
var ICON_SEARCH  = "\u{1F50D}";
var ICON_INFO    = "\u{2139}";
var ICON_FILTER  = "\u{22EE}";
var ARROW_ASC    = "\u{25B2}";
var ARROW_DESC   = "\u{25BC}";

var EMPTY_LABEL = "(空欄)";

// 列定義。type が "date" の列は日付範囲フィルタ、それ以外はチェックボックス複数選択。
// Document Number は Enovia でのみ使うため、現段階では列に出さない
// （データ自体は保持しており、Phase 3 で復活させる想定）。
var COLUMNS = [
  { key: "source",        label: "ソース",      type: "set"    },
  { key: "title",         label: "タイトル",     type: "set"    },
  { key: "__select",      label: "選択",        type: "select" },
  { key: "author",        label: "作成者",       type: "set"    },
  { key: "last_modified", label: "最終更新日",   type: "date"   },
  { key: "doc_type",      label: "種別",        type: "set"    },
  { key: "site",          label: "サイト",       type: "set"    },
  { key: "folder",        label: "フォルダ",     type: "set"    }
];

var allResults = [];      // サーバーから受け取った全件
var viewResults = [];     // フィルタ・ソート適用後
var filters = {};         // {key: {values:[...]}} または {key: {from:"", to:""}}
var sortState = { key: null, dir: 0 };   // dir: 1=昇順 / -1=降順 / 0=解除
var selected = {};        // 一括ダウンロード用の選択状態 {_idx: true}
var pendingState = null;  // 起動時に復元する状態（検索結果の到着後に適用する）
var restoring = false;    // 復元中は状態を保存し直さない

function log(msg) {
  var el = document.getElementById("log");
  var t = new Date().toLocaleTimeString("ja-JP");
  el.textContent += "[" + t + "] " + msg + "\n";
  el.scrollTop = el.scrollHeight;
}

function stateIcon(state) {
  if (state === "ok") return ICON_OK;
  if (state === "empty") return ICON_EMPTY;
  if (state === "error") return ICON_ERROR;
  return ICON_PENDING;
}

function renderStatuses(statuses) {
  var box = document.getElementById("status");
  box.innerHTML = "";
  statuses.forEach(function (s) {
    var div = document.createElement("div");
    div.className = "chip";
    div.textContent = stateIcon(s.state) + " " + s.label + " : " + s.count + " 件";
    if (s.message) {
      var small = document.createElement("small");
      small.textContent = "  " + s.message;
      div.appendChild(small);
    }
    box.appendChild(div);
  });
}

// ── フィルタ判定 ─────────────────────────────────────────
function valueOf(row, key) {
  var v = row[key];
  return (v === null || v === undefined) ? "" : String(v);
}

function isFiltered(key) {
  var f = filters[key];
  if (!f) return false;
  if (f.values) return true;
  return !!(f.from || f.to);
}

function rowPasses(row) {
  for (var key in filters) {
    var f = filters[key];
    if (!f) continue;
    if (f.values) {
      if (f.values.indexOf(valueOf(row, key)) < 0) return false;
    } else {
      var v = valueOf(row, key);
      if (f.from) { if (!v || v < f.from) return false; }
      if (f.to)   { if (!v || v > f.to)   return false; }
    }
  }
  return true;
}

function compareRows(a, b) {
  var key = sortState.key;
  var x = valueOf(a, key), y = valueOf(b, key);
  // 空欄は常に末尾へ寄せる
  if (x === "" && y !== "") return 1;
  if (y === "" && x !== "") return -1;
  var c;
  if (key === "last_modified") { c = (x < y) ? -1 : (x > y) ? 1 : 0; }
  else { c = x.localeCompare(y, "ja"); }
  if (c !== 0) return c * sortState.dir;
  return a._idx - b._idx;
}

function applyView() {
  viewResults = allResults.filter(rowPasses);
  if (sortState.key && sortState.dir !== 0) {
    viewResults.sort(compareRows);
  } else {
    viewResults.sort(function (a, b) { return a._idx - b._idx; });
  }
  // フィルタパネルを開いたまま複数の値をチェックできるよう、
  // パネル表示中はヘッダーを作り直さず、表示状態だけを更新する。
  if (document.querySelector(".fpanel")) { refreshHeadIndicators(); }
  else { renderHead(); }
  renderBody();
  updateCounts();
  if (!restoring) { saveState(); }
}


function refreshHeadIndicators() {
  var ths = document.getElementById("resultHead").children;
  for (var i = 0; i < COLUMNS.length; i++) {
    var th = ths[i];
    if (!th) { continue; }
    var btn = th.querySelector(".filt");
    if (btn) { btn.className = "filt" + (isFiltered(COLUMNS[i].key) ? " on" : ""); }
  }
  refreshSelectAll();
}


function refreshSelectAll() {
  var box = document.querySelector("#resultHead th.selcol input");
  if (!box) { return; }
  box.checked = selectableRows().length > 0 && selectableRows().every(function (r) {
    return selected[r._idx];
  });
}


function selectableRows() {
  // フォルダとリンクの無い行は一括ダウンロードの対象外
  return viewResults.filter(function (r) { return r.url && !r.is_folder; });
}


// ── 検索・並び順・絞り込みの状態をサーバーに保存する ─────
function saveState() {
  var body = {
    keyword: document.getElementById("keyword").value,
    target: currentTarget(),
    max_results: parseInt(document.getElementById("maxResults").value, 10),
    sort: sortState,
    filters: filters
  };
  fetch("/api/state", {
    method: "POST",
    headers: { "Content-Type": "application/json" },
    body: JSON.stringify(body)
  }).catch(function () { /* 保存できなくても操作は続行する */ });
}

// ── ヘッダー描画（ソート＋フィルタUI） ───────────────────
function renderHead() {
  var head = document.getElementById("resultHead");
  head.innerHTML = "";
  COLUMNS.forEach(function (col) {
    var th = document.createElement("th");

    if (col.type === "select") {
      th.className = "selcol";
      var all = document.createElement("input");
      all.type = "checkbox";
      all.title = "表示中の行をすべて選択／解除";
      all.checked = selectableRows().length > 0 && selectableRows().every(function (r) {
        return selected[r._idx];
      });
      all.addEventListener("change", function () {
        viewResults.forEach(function (r) {
          if (r.is_folder || !r.url) { return; }   // フォルダは対象外
          if (all.checked) { selected[r._idx] = true; } else { delete selected[r._idx]; }
        });
        renderBody(); updateCounts();
      });
      th.appendChild(all);
      head.appendChild(th);
      return;
    }

    var span = document.createElement("span");
    span.className = "hdr";
    span.textContent = col.label;
    if (sortState.key === col.key && sortState.dir !== 0) {
      var arrow = document.createElement("span");
      arrow.className = "arrow";
      arrow.textContent = (sortState.dir === 1) ? ARROW_ASC : ARROW_DESC;
      span.appendChild(arrow);
    }
    span.addEventListener("click", function () { toggleSort(col.key); });
    th.appendChild(span);

    var btn = document.createElement("span");
    btn.className = "filt" + (isFiltered(col.key) ? " on" : "");
    btn.textContent = ICON_FILTER;
    btn.title = col.label + " で絞り込む";
    btn.addEventListener("click", function (e) {
      e.stopPropagation();
      openFilterPanel(btn, col);
    });
    th.appendChild(btn);

    head.appendChild(th);
  });
}

function toggleSort(key) {
  if (sortState.key !== key) { sortState = { key: key, dir: 1 }; }
  else if (sortState.dir === 1) { sortState.dir = -1; }
  else if (sortState.dir === -1) { sortState = { key: null, dir: 0 }; }
  else { sortState = { key: key, dir: 1 }; }
  applyView();
}

// ── フィルタパネル ───────────────────────────────────────
var openPanelKey = null;

function closeAllPanels() {
  var panels = document.querySelectorAll(".fpanel");
  for (var i = 0; i < panels.length; i++) { panels[i].parentNode.removeChild(panels[i]); }
  openPanelKey = null;
}

document.addEventListener("click", function () { closeAllPanels(); });
window.addEventListener("resize", function () { closeAllPanels(); });
window.addEventListener("scroll", function () { closeAllPanels(); }, true);

function openFilterPanel(btn, col) {
  var wasOpen = (openPanelKey === col.key);
  closeAllPanels();
  if (wasOpen) { return; }   // 同じボタンをもう一度押したら閉じる

  var panel = document.createElement("div");
  panel.className = "fpanel";
  panel.addEventListener("click", function (e) { e.stopPropagation(); });

  if (col.type === "date") { buildDatePanel(panel, col); }
  else { buildSetPanel(panel, col); }

  // 表の横スクロール領域に切り取られないよう body 直下に置き、
  // ボタンの位置から座標を決める。右端で画面外にはみ出す場合は左へ寄せる。
  document.body.appendChild(panel);
  var rect = btn.getBoundingClientRect();
  panel.style.top = (rect.bottom + 4) + "px";
  var left = rect.left;
  if (left + panel.offsetWidth > window.innerWidth - 8) {
    left = window.innerWidth - panel.offsetWidth - 8;
  }
  panel.style.left = Math.max(8, left) + "px";
  openPanelKey = col.key;
}

function buildDatePanel(panel, col) {
  var current = filters[col.key] || {};

  var hint = document.createElement("div");
  hint.className = "hint";
  hint.textContent = "指定した日を含めて絞り込みます。片方だけの指定も可能です。";
  panel.appendChild(hint);

  var rowFrom = document.createElement("div");
  rowFrom.className = "frow";
  var labFrom = document.createElement("span");
  labFrom.textContent = "以降";
  labFrom.style.minWidth = "34px";
  var inFrom = document.createElement("input");
  inFrom.type = "date";
  inFrom.value = current.from || "";
  rowFrom.appendChild(labFrom); rowFrom.appendChild(inFrom);
  panel.appendChild(rowFrom);

  var rowTo = document.createElement("div");
  rowTo.className = "frow";
  var labTo = document.createElement("span");
  labTo.textContent = "以前";
  labTo.style.minWidth = "34px";
  var inTo = document.createElement("input");
  inTo.type = "date";
  inTo.value = current.to || "";
  rowTo.appendChild(labTo); rowTo.appendChild(inTo);
  panel.appendChild(rowTo);

  function commit() {
    var from = inFrom.value, to = inTo.value;
    if (!from && !to) { delete filters[col.key]; }
    else { filters[col.key] = { from: from, to: to }; }
    applyView();
  }
  inFrom.addEventListener("change", commit);
  inTo.addEventListener("change", commit);

  var actions = document.createElement("div");
  actions.className = "frow";
  var btnClear = document.createElement("button");
  btnClear.className = "mini";
  btnClear.textContent = "この列の指定を解除";
  btnClear.addEventListener("click", function () {
    inFrom.value = ""; inTo.value = "";
    delete filters[col.key];
    applyView();
    closeAllPanels();
  });
  actions.appendChild(btnClear);
  panel.appendChild(actions);
}

function buildSetPanel(panel, col) {
  // 候補値と件数を全件から集計する
  var counts = {};
  allResults.forEach(function (r) {
    var v = valueOf(r, col.key);
    counts[v] = (counts[v] || 0) + 1;
  });
  var values = Object.keys(counts).sort(function (a, b) {
    if (a === "") return 1;
    if (b === "") return -1;
    return a.localeCompare(b, "ja");
  });

  var selected = filters[col.key] ? filters[col.key].values : null;   // null=全選択

  var searchRow = document.createElement("div");
  searchRow.className = "frow";
  var search = document.createElement("input");
  search.type = "text";
  search.placeholder = "候補を絞り込む";
  searchRow.appendChild(search);
  panel.appendChild(searchRow);

  var actions = document.createElement("div");
  actions.className = "frow";
  var btnAll = document.createElement("button");
  btnAll.className = "mini"; btnAll.textContent = "すべて選択";
  var btnNone = document.createElement("button");
  btnNone.className = "mini"; btnNone.textContent = "すべて解除";
  actions.appendChild(btnAll); actions.appendChild(btnNone);
  panel.appendChild(actions);

  var list = document.createElement("div");
  list.className = "flist";
  panel.appendChild(list);

  var boxes = [];

  function commit() {
    var picked = [];
    boxes.forEach(function (b) { if (b.checked) picked.push(b.value); });
    if (picked.length === values.length) { delete filters[col.key]; }
    else { filters[col.key] = { values: picked }; }
    applyView();
  }

  values.forEach(function (v) {
    var label = document.createElement("label");
    var box = document.createElement("input");
    box.type = "checkbox";
    box.value = v;
    box.checked = (selected === null) ? true : (selected.indexOf(v) >= 0);
    box.addEventListener("change", commit);
    boxes.push(box);

    var text = document.createElement("span");
    text.textContent = (v === "") ? EMPTY_LABEL : v;

    var cnt = document.createElement("span");
    cnt.className = "cnt";
    cnt.textContent = counts[v];

    label.appendChild(box); label.appendChild(text); label.appendChild(cnt);
    label.dataset.value = (v === "") ? EMPTY_LABEL : v;
    list.appendChild(label);
  });

  btnAll.addEventListener("click", function () {
    boxes.forEach(function (b) {
      if (b.parentNode.style.display !== "none") { b.checked = true; }
    });
    commit();
  });
  btnNone.addEventListener("click", function () {
    boxes.forEach(function (b) {
      if (b.parentNode.style.display !== "none") { b.checked = false; }
    });
    commit();
  });

  search.addEventListener("input", function () {
    var q = search.value.toLowerCase();
    var labels = list.querySelectorAll("label");
    for (var i = 0; i < labels.length; i++) {
      var hit = labels[i].dataset.value.toLowerCase().indexOf(q) >= 0;
      labels[i].style.display = hit ? "" : "none";
    }
  });
}

// ── 明細描画 ─────────────────────────────────────────────
function renderBody() {
  var body = document.getElementById("resultBody");
  body.innerHTML = "";

  if (!allResults.length) {
    var tr0 = document.createElement("tr");
    var td0 = document.createElement("td");
    td0.colSpan = COLUMNS.length; td0.className = "muted";
    td0.textContent = "該当する文書が見つかりませんでした。";
    tr0.appendChild(td0); body.appendChild(tr0);
    return;
  }
  if (!viewResults.length) {
    var tr1 = document.createElement("tr");
    var td1 = document.createElement("td");
    td1.colSpan = COLUMNS.length; td1.className = "muted";
    td1.textContent = "フィルタ条件に一致する行がありません。「フィルタ・ソート解除」で戻せます。";
    tr1.appendChild(td1); body.appendChild(tr1);
    return;
  }

  viewResults.forEach(function (r) {
    var tr = document.createElement("tr");
    COLUMNS.forEach(function (col) { tr.appendChild(buildCell(r, col)); });
    body.appendChild(tr);
  });
}

function buildCell(r, col) {
  var td = document.createElement("td");

  if (col.key === "source") {
    var span = document.createElement("span");
    span.className = "src"; span.textContent = r.source;
    td.appendChild(span);
    return td;
  }

  if (col.key === "title") {
    td.className = "title";
    if (r.url) {
      var aFile = document.createElement("a");
      aFile.href = r.url; aFile.target = "_blank"; aFile.rel = "noopener";
      aFile.title = "ファイルを開く: " + r.url;
      aFile.textContent = r.title || "(タイトルなし)";
      td.appendChild(aFile);
    } else {
      td.textContent = r.title || "(タイトルなし)";
    }
    return td;
  }

  if (col.key === "__select") {
    td.className = "selcol";
    var box = document.createElement("input");
    box.type = "checkbox";
    box.checked = !!selected[r._idx];
    // フォルダ自体はファイル本体を持たないため、一括ダウンロードの対象外にする
    box.disabled = (!r.url || r.is_folder);
    box.title = r.is_folder ? "フォルダはダウンロードできません"
              : (r.url ? "一括ダウンロードの対象にする" : "リンクが無いため選択できません");
    box.addEventListener("change", function () {
      if (box.checked) { selected[r._idx] = true; } else { delete selected[r._idx]; }
      refreshSelectAll(); updateCounts();
    });
    td.appendChild(box);
    return td;
  }

  if (col.key === "folder") {
    if (r.folder_url) {
      var aFolder = document.createElement("a");
      aFolder.href = r.folder_url; aFolder.target = "_blank"; aFolder.rel = "noopener";
      aFolder.title = "保管フォルダを開く: " + (r.folder_full || r.folder);
      aFolder.textContent = r.folder;
      td.appendChild(aFolder);
    } else {
      td.textContent = r.folder || "";
    }
    return td;
  }

  if (col.key === "site") {
    td.className = "nowrap";
    if (r.site_url) {
      var aSite = document.createElement("a");
      aSite.href = r.site_url; aSite.target = "_blank"; aSite.rel = "noopener";
      aSite.title = "保管先サイトを開く: " + r.site_url;
      aSite.textContent = r.site || "サイトを開く";
      td.appendChild(aSite);
    } else {
      td.textContent = r.site || "";
    }
    return td;
  }

  if (col.key === "last_modified" || col.key === "doc_type") { td.className = "nowrap"; }
  td.textContent = r[col.key] || "";
  return td;
}


function selectedIndices() {
  // 画面の並び順を保ったまま、選択されている行の索引を返す
  return viewResults.filter(function (r) { return selected[r._idx]; })
                    .map(function (r) { return r._idx; });
}


function updateCounts() {
  var text = "検索結果: " + allResults.length + " 件";
  if (viewResults.length !== allResults.length) {
    text += "（絞り込み後 " + viewResults.length + " 件）";
  }
  document.getElementById("resultCount").textContent = text;
  var enabled = (viewResults.length > 0);
  document.getElementById("btnXlsx").disabled = !enabled;
  document.getElementById("btnCsv").disabled = !enabled;

  var picked = selectedIndices().length;
  var btnDl = document.getElementById("btnDownload");
  btnDl.disabled = (picked === 0);
  btnDl.textContent = "選択ファイルをZIPで取得" + (picked ? " (" + picked + ")" : "");
}

// ── 検索 ─────────────────────────────────────────────────
function setBusy(busy) {
  document.getElementById("btnSearch").disabled = busy;
  document.getElementById("btnProbe").disabled = busy;
}

function currentTarget() {
  var radios = document.getElementsByName("target");
  for (var i = 0; i < radios.length; i++) {
    if (radios[i].checked) return radios[i].value;
  }
  return "all";
}

document.getElementById("targets").addEventListener("change", function () {
  var labels = document.querySelectorAll("#targets label");
  for (var i = 0; i < labels.length; i++) {
    labels[i].className = labels[i].querySelector("input").checked ? "on" : "";
  }
});

document.getElementById("btnSearch").addEventListener("click", function () {
  var keyword = document.getElementById("keyword").value.trim();
  if (!keyword) { log("キーワードが空です。"); return; }
  var target = currentTarget();
  var maxResults = parseInt(document.getElementById("maxResults").value, 10);

  setBusy(true);
  log(ICON_SEARCH + " 検索開始: \"" + keyword + "\" / 対象: " + target + " / 上限: " + maxResults + " 件");

  fetch("/api/search", {
    method: "POST",
    headers: { "Content-Type": "application/json" },
    body: JSON.stringify({ keyword: keyword, target: target, max_results: maxResults })
  })
  .then(function (r) { return r.json(); })
  .then(function (data) {
    if (data.error) { log(ICON_ERROR + " " + data.error); setBusy(false); return; }
    allResults = data.results || [];
    allResults.forEach(function (r, i) { r._idx = i; });
    filters = {};
    sortState = { key: null, dir: 0 };
    selected = {};
    if (pendingState) {
      filters = pendingState.filters || {};
      sortState = pendingState.sort || { key: null, dir: 0 };
      pendingState = null;
      restoring = true;
      log(ICON_INFO + " 前回の並び順と絞り込みを復元しました。");
    }
    renderStatuses(data.statuses || []);
    applyView();
    restoring = false;
    var note = "";
    if (data.excluded_nexus > 0) { note = "（Nexus重複 " + data.excluded_nexus + " 件を除外）"; }
    document.getElementById("excludedNote").textContent = note;
    log(ICON_OK + " 検索完了: " + allResults.length + " 件 (" + (data.elapsed_sec || 0) + " 秒)");
  })
  .catch(function (e) { log(ICON_ERROR + " 通信エラー: " + e); })
  .finally(function () { setBusy(false); });
});

document.getElementById("keyword").addEventListener("keydown", function (e) {
  if (e.key === "Enter") { document.getElementById("btnSearch").click(); }
});

document.getElementById("btnProbe").addEventListener("click", function () {
  setBusy(true);
  log(ICON_INFO + " 疎通診断を実行します...");
  fetch("/api/probe", { method: "POST" })
    .then(function (r) { return r.json(); })
    .then(function (data) {
      (data.probes || []).forEach(function (p) {
        var icon = p.ok ? ICON_OK : (p.implemented ? ICON_ERROR : ICON_PENDING);
        log(icon + " " + p.label + " [" + p.mode + "] " + p.message);
      });
    })
    .catch(function (e) { log(ICON_ERROR + " 通信エラー: " + e); })
    .finally(function () { setBusy(false); });
});

document.getElementById("btnClearFilter").addEventListener("click", function () {
  filters = {};
  sortState = { key: null, dir: 0 };
  closeAllPanels();
  applyView();
  log(ICON_INFO + " フィルタとソートを解除しました。");
});

// ── 出力（画面に表示されている行・並び順で出力する） ─────
function download(fmt) {
  if (!viewResults.length) { return; }
  var idx = viewResults.map(function (r) { return r._idx; }).join(",");
  log(ICON_INFO + " " + fmt.toUpperCase() + " を出力します（" + viewResults.length + " 件）。");
  window.location.href = "/api/export?format=" + fmt + "&idx=" + encodeURIComponent(idx);
}
document.getElementById("btnDownload").addEventListener("click", function () {
  var idx = selectedIndices();
  if (!idx.length) { return; }
  log(ICON_INFO + " " + idx.length + " 件のファイルをZIPにまとめています。"
      + "件数が多いと時間がかかります。");
  window.location.href = "/api/download?idx=" + encodeURIComponent(idx.join(","));
});

document.getElementById("btnXlsx").addEventListener("click", function () { download("xlsx"); });
document.getElementById("btnCsv").addEventListener("click", function () { download("csv"); });

// ── 前回の検索状態を復元する ─────────────────────────────
function restoreState() {
  return fetch("/api/state")
    .then(function (r) { return r.json(); })
    .then(function (s) {
      if (!s || !s.keyword) { return; }

      document.getElementById("keyword").value = s.keyword;

      var radios = document.getElementsByName("target");
      for (var i = 0; i < radios.length; i++) {
        radios[i].checked = (radios[i].value === s.target);
      }
      var labels = document.querySelectorAll("#targets label");
      for (var j = 0; j < labels.length; j++) {
        labels[j].className = labels[j].querySelector("input").checked ? "on" : "";
      }

      var sel = document.getElementById("maxResults");
      for (var k = 0; k < sel.options.length; k++) {
        if (sel.options[k].value === String(s.max_results)) { sel.selectedIndex = k; }
      }

      pendingState = s;
      log(ICON_INFO + " 前回の検索状態を復元しました: \"" + s.keyword + "\"（再検索します）");
      document.getElementById("btnSearch").click();
    })
    .catch(function () { /* 復元できなくても通常起動する */ });
}

renderHead();
log(ICON_INFO + " Document Search Manager を起動しました。");
// 件数の初期値は「上位 10 件」。前回の作業状態があればそちらを復元する。
restoreState();
</script>
</body>
</html>
"""


# ─────────────────────────────────────────────────────────────
# Flask アプリ
# ─────────────────────────────────────────────────────────────
flask_app = Flask(__name__)

_cfg: Dict[str, Any] = {}
_auth: Optional[GraphAuthManager] = None
_manager: Optional[SearchManager] = None
_last_results: List[SearchResult] = []
_last_keyword: str = ""
_lock = threading.Lock()


@flask_app.route("/")
def index() -> Response:
    return Response(INDEX_HTML, mimetype="text/html; charset=utf-8")


@flask_app.route("/api/config")
def api_config():
    """画面表示に必要な設定のみを返す（tenant_id / client_id は返さない）。"""
    return jsonify({
        "default_max_results": _cfg.get("default_max_results", 100),
        "hard_max_results": _cfg.get("hard_max_results", 500),
    })


@flask_app.route("/api/state", methods=["GET", "POST"])
def api_state():
    """検索キーワード・対象・件数・並び順・絞り込みを保存／復元する。

    ツールを終了して再起動しても、前回の作業状態から再開できるようにするため、
    フォルダ内の session_state.json に保存する。
    検索結果そのものは保存しない（古い結果を見せないため、再起動時に再検索する）。
    """
    if flask_req.method == "GET":
        if not _cfg.get("restore_last_search", True) or not STATE_PATH.exists():
            return jsonify({})
        try:
            with open(STATE_PATH, "r", encoding="utf-8") as f:
                return jsonify(json.load(f))
        except Exception as e:
            print(f"⚠️  前回の状態を読み込めませんでした（無視して続行）: {e}")
            return jsonify({})

    payload = flask_req.get_json(silent=True) or {}
    state = {
        "keyword": str(payload.get("keyword", ""))[:200],
        "target": str(payload.get("target", TARGET_ALL)),
        "max_results": int(payload.get("max_results") or
                           _cfg.get("default_max_results", 10)),
        "sort": payload.get("sort") or {},
        "filters": payload.get("filters") or {},
        "saved_at": datetime.now(JST).strftime("%Y-%m-%d %H:%M:%S"),
    }
    try:
        with open(STATE_PATH, "w", encoding="utf-8") as f:
            json.dump(state, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"⚠️  状態を保存できませんでした（無視して続行）: {e}")
        return jsonify({"saved": False})
    return jsonify({"saved": True})


@flask_app.route("/api/probe", methods=["POST"])
def api_probe():
    try:
        return jsonify({"probes": _manager.probe_all()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@flask_app.route("/api/search", methods=["POST"])
def api_search():
    global _last_results, _last_keyword

    payload = flask_req.get_json(silent=True) or {}
    keyword = str(payload.get("keyword", "")).strip()
    target = str(payload.get("target", TARGET_ALL)).strip()
    max_results = int(payload.get("max_results", _cfg.get("default_max_results", 100)))
    max_results = max(1, min(max_results, int(_cfg.get("hard_max_results", 500))))

    if not keyword:
        return jsonify({"error": "キーワードが指定されていません。"}), 400

    started = datetime.now()
    try:
        outcome = _manager.search(keyword, target, max_results)
    except Exception as e:
        print(f"❌ 検索でエラーが発生しました: {e}")
        return jsonify({"error": f"検索でエラーが発生しました: {e}"}), 500

    elapsed = round((datetime.now() - started).total_seconds(), 1)
    results = outcome["results"]

    with _lock:
        _last_results = results
        _last_keyword = keyword

    print(f"🔍 検索: \"{keyword}\" / 対象={target} / {len(results)} 件 / {elapsed} 秒")

    return jsonify({
        "results": [asdict(r) for r in results],
        "statuses": outcome["statuses"],
        "excluded_nexus": outcome["excluded_nexus"],
        "elapsed_sec": elapsed,
    })


def _pick_results(idx_param: Optional[str]) -> List[SearchResult]:
    """idx（カンマ区切りの索引）で、画面に表示されている行・並び順を再現する。

    idx が無い場合は全件を対象にする。範囲外・不正な値は無視する。
    """
    with _lock:
        results = list(_last_results)

    if not idx_param:
        return results

    picked: List[SearchResult] = []
    for token in str(idx_param).split(","):
        token = token.strip()
        if not token.isdigit():
            continue
        index = int(token)
        if 0 <= index < len(results):
            picked.append(results[index])
    return picked


@flask_app.route("/api/export")
def api_export():
    fmt = (flask_req.args.get("format") or "xlsx").lower()

    results = _pick_results(flask_req.args.get("idx"))
    with _lock:
        keyword = _last_keyword

    if not results:
        return jsonify({"error": "出力対象の検索結果がありません。"}), 400

    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(JST).strftime("%Y%m%d_%H%M%S")
    safe_keyword = "".join(c for c in keyword if c.isalnum() or c in ("-", "_"))[:30]
    basename = f"document_search_{safe_keyword or 'result'}_{stamp}"

    # Document Number は Enovia でのみ使うため、現段階では出力しない
    headers = ["ソース", "タイトル", "作成者", "最終更新日", "種別",
               "サイト", "フォルダ", "ファイルリンク", "サイトリンク", "フォルダリンク"]

    def row_of(r: SearchResult):
        return [r.source, r.title, r.author, r.last_modified, r.doc_type,
                r.site, (r.folder_full or r.folder), r.url, r.site_url, r.folder_url]

    if fmt == "csv":
        path = EXPORT_DIR / (basename + ".csv")
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for r in results:
                writer.writerow(row_of(r))
        print(f"💾 CSVを出力しました: {path}")
        return send_file(str(path), as_attachment=True)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ModuleNotFoundError:
        return jsonify({"error": "openpyxl が未インストールです。CSV出力をご利用ください。"}), 500

    path = EXPORT_DIR / (basename + ".xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "検索結果"
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for r in results:
        ws.append(row_of(r))
    for column, width in zip("ABCDEFGHIJ",
                             (12, 52, 22, 14, 10, 20, 34, 60, 44, 60)):
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"💾 Excelを出力しました: {path}")
    return send_file(str(path), as_attachment=True)


def _share_token(url: str) -> str:
    """共有URLを Graph の /shares で使える識別子に変換する。

    先頭に "u!" を付け、URLをBase64URL（パディング除去）で符号化する形式。
    """
    encoded = base64.urlsafe_b64encode(url.encode("utf-8")).decode("ascii")
    return "u!" + encoded.rstrip("=")


def _download_one(token: str, result: SearchResult, timeout: int):
    """1ファイルを取得して (ファイル名, バイト列) を返す。失敗時は例外を送出する。"""
    endpoint = f"{GRAPH_V1}/shares/{_share_token(result.url)}/driveItem/content"
    resp = http_req.get(
        endpoint,
        headers={"Authorization": f"Bearer {token}"},
        timeout=timeout,
        allow_redirects=True,
    )
    if resp.status_code != 200:
        raise RuntimeError(f"HTTP {resp.status_code}")

    name = _pick(_name_from_url(result.url), result.title, "document")
    return name, resp.content


@flask_app.route("/api/download")
def api_download():
    """選択されたファイルをまとめてZIPで返す。

    1件でも失敗した場合は、ZIP内に「_ダウンロード失敗一覧.txt」を同梱し、
    取得できた分はそのまま返す（全体を失敗させない）。
    """
    # ダウンロードは必ず明示的な選択（idx）を要求する。
    # 出力(export)と違い、未指定を「全件」と解釈すると、選択していない
    # ファイルまで一括取得してしまうため。
    idx_param = flask_req.args.get("idx")
    if not idx_param:
        return jsonify({"error": "ダウンロード対象が選択されていません。"}), 400

    # フォルダはファイル本体を持たないため対象から外す
    results = [r for r in _pick_results(idx_param) if r.url and not r.is_folder]
    if not results:
        return jsonify({"error": "ダウンロード対象が選択されていません。"}), 400

    limit = int(_cfg.get("max_download_files", 50))
    skipped_by_limit = []
    if len(results) > limit:
        skipped_by_limit = results[limit:]
        results = results[:limit]

    try:
        token = _auth.get_token(_cfg["graph_scopes"])
    except Exception as e:
        return jsonify({"error": f"認証に失敗しました: {e}"}), 500

    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(JST).strftime("%Y%m%d_%H%M%S")
    with _lock:
        keyword = _last_keyword
    safe_keyword = "".join(c for c in keyword if c.isalnum() or c in ("-", "_"))[:30]
    zip_path = DOWNLOAD_DIR / f"documents_{safe_keyword or 'selected'}_{stamp}.zip"

    timeout = int(_cfg.get("download_timeout_sec", 120))
    failures: List[str] = []
    used_names: Dict[str, int] = {}
    succeeded = 0

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for result in results:
            try:
                name, content = _download_one(token, result, timeout)
            except Exception as e:
                failures.append(f"{result.title}\t{result.url}\t{e}")
                print(f"⚠️  取得失敗: {result.title} ({e})")
                continue

            # ZIP内でファイル名が重複しないようにする
            if name in used_names:
                used_names[name] += 1
                stem, dot, ext = name.rpartition(".")
                suffix = f"_{used_names[name]}"
                name = (stem + suffix + dot + ext) if dot else (name + suffix)
            else:
                used_names[name] = 0

            archive.writestr(name, content)
            succeeded += 1

        for result in skipped_by_limit:
            failures.append(f"{result.title}\t{result.url}\t"
                            f"1回の上限({limit}件)を超えたため対象外")

        if failures:
            header = ("以下のファイルは取得できませんでした。\n"
                      "タイトル\tURL\t理由\n")
            archive.writestr("_ダウンロード失敗一覧.txt", header + "\n".join(failures))

    print(f"📥 一括ダウンロード: 成功 {succeeded} 件 / 失敗 {len(failures)} 件 → {zip_path}")

    if succeeded == 0:
        return jsonify({
            "error": ("すべてのファイルの取得に失敗しました。"
                      "ファイルへのアクセス権限をご確認ください。")
        }), 502

    return send_file(str(zip_path), as_attachment=True)


# ─────────────────────────────────────────────────────────────
# エントリポイント
# ─────────────────────────────────────────────────────────────
def main() -> None:
    global _cfg, _auth, _manager

    print("=" * 70)
    print("📚 Document Search Manager  v20260903_04  (Phase 1: SharePoint)")
    print("=" * 70)
    print(f"📁 作業フォルダ: {BASE_DIR}")

    _archive_old_versions()

    _cfg = _load_config()
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

    auth = GraphAuthManager(_cfg["tenant_id"], _cfg["client_id"], TOKEN_CACHE_PATH)
    _auth = auth
    print("🔑 Graph APIの認証を行います（スコープ: "
          + ", ".join(_cfg["graph_scopes"]) + "）")
    auth.get_token(_cfg["graph_scopes"])
    print("✅ 認証に成功しました。")

    _manager = SearchManager(_cfg, auth)

    print("🩺 疎通診断を実行します...")
    for probe in _manager.probe_all():
        icon = "🟢" if probe.get("ok") else ("🔴" if probe.get("implemented") else "⚪")
        print(f"   {icon} {probe['label']:<11} [{probe['mode']}] {probe['message']}")

    print(f"🔢 件数の初期値: 上位 {_cfg.get('default_max_results', 10)} 件"
          f"（前回の作業状態があればそちらを優先します）")

    port = int(_cfg.get("flask_port", 5020))
    url = f"http://127.0.0.1:{port}"
    print(f"\n🌐 ブラウザで次のURLを開いてください: {url}")
    print("   （終了するには、このウィンドウで Ctrl+C を押してください）\n")

    if _cfg.get("auto_open_browser", True):
        threading.Timer(1.2, lambda: webbrowser.open(url)).start()

    flask_app.run(host="127.0.0.1", port=port, debug=False, use_reloader=False)


if __name__ == "__main__":
    main()
