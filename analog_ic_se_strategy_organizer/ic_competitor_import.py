# -*- coding: utf-8 -*-
"""競合企業Excel(config/source_data/*.xlsx)からcompetitors_db.jsonを生成するインポートスクリプト。

越智さんが競合企業Excel（地域別シート: US/Europe/Japan/Asia）を更新した際、
本スクリプトを再実行するだけで config/competitors_db.json を最新化できる。
JSONを手動編集する運用は想定しない（Excelをマスターデータとする）。

使い方:
    python3 ic_competitor_import.py
    python3 ic_competitor_import.py --xlsx path/to/other.xlsx --out path/to/other.json
"""

import os
import json
import argparse
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_XLSX = os.path.join(BASE_DIR, "config", "source_data",
                             "analog_power_semiconductor_companies_global_2026.xlsx")
DEFAULT_OUT = os.path.join(BASE_DIR, "config", "competitors_db.json")

# シート名 -> region値（英語キー）。Excel側の「地域」列（米国/欧州/日本/アジア）とも対応する。
SHEET_TO_REGION = {"US": "US", "Europe": "Europe", "Japan": "Japan", "Asia": "Asia"}

# ヘッダー行の列順（0始まり index）と competitors_db.json のキー名対応。
# 8列目(index 8)から16列目(index 16)までが9つの製品カテゴリ列。
CATEGORY_COLUMNS = [
    ("dc_dc_pmic", 8),
    ("ldo", 9),
    ("led_driver", 10),
    ("ac_dc", 11),
    ("gate_driver", 12),
    ("load_switch_efuse", 13),
    ("ideal_diode_oring", 14),
    ("gan_power", 15),
    ("power_discrete_module", 16),
]

MARK_MAP = {"●": "primary", "△": "limited", "—": "none", "-": "none", None: "none"}

LEGEND = {
    "primary": "●: 公式カタログで独立した製品カテゴリ、複数シリーズ、または主要戦略製品として明確に確認できる",
    "limited": "△: 製品数が限定、特定用途への統合、モジュール／リファレンス中心、または主力カテゴリではない",
    "none": "—: 今回確認した公式公開カタログでは明確な独立製品群を確認できない（「技術的に不可能」を意味しない）",
}


def _mark(value):
    return MARK_MAP.get(str(value).strip() if value is not None else None, "none")


def _str(value):
    return str(value).strip() if value is not None else ""


def _iso_date(value):
    """Excelのdatetime/文字列いずれの確認日もYYYY-MM-DD文字列に揃える"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _row_to_company(sheet_name, row):
    region = SHEET_TO_REGION.get(sheet_name, sheet_name)
    categories = {key: _mark(row[idx]) for key, idx in CATEGORY_COLUMNS}
    return {
        "name": _str(row[2]),
        "region": region,
        "country": _str(row[1]),
        "parent_or_brand_status": _str(row[3]),
        "company_url": _str(row[4]),
        "product_url": _str(row[5]),
        "company_type": _str(row[6]),
        "automotive_capable": _mark(row[7]),
        "categories": categories,
        "breadth_score": row[17] if isinstance(row[17], (int, float)) else None,
        "product_overview": _str(row[18]),
        "market_positioning": _str(row[19]),
        "source_url": _str(row[20]),
        "verified_at": _iso_date(row[21]),
        "active": True,
    }


def import_competitors(xlsx_path=None, out_path=None):
    """Excelを読み込みcompetitors_db.jsonを生成して書き出す。生成したdictを返す。"""
    import openpyxl

    xlsx_path = xlsx_path or DEFAULT_XLSX
    out_path = out_path or DEFAULT_OUT

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    companies = []
    for sheet_name in SHEET_TO_REGION:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        # rows[0]=タイトル行, rows[1]=凡例行, rows[2]=ヘッダー行, rows[3:]=データ行
        for row in rows[3:]:
            if not row or not row[2]:
                continue
            companies.append(_row_to_company(sheet_name, row))

    db = {
        "generated_at": datetime.now().isoformat(),
        "source_file": os.path.basename(xlsx_path),
        "as_of": companies[0]["verified_at"] if companies else "",
        "legend": LEGEND,
        "categories": [key for key, _ in CATEGORY_COLUMNS],
        "companies": companies,
    }

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False, indent=1)

    return db


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="入力Excelパス")
    parser.add_argument("--out", default=DEFAULT_OUT, help="出力JSONパス")
    args = parser.parse_args()

    result = import_competitors(args.xlsx, args.out)
    by_region = {}
    for c in result["companies"]:
        by_region[c["region"]] = by_region.get(c["region"], 0) + 1
    print(f"取り込み完了: {len(result['companies'])}社 -> {args.out}")
    for region, count in by_region.items():
        print(f"  {region}: {count}社")
