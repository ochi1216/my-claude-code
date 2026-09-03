VERSION = "20260812_01"



import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import os
import sys

# ============================================================
# Gemini 共通クライアント(gemini_client.py)への互換シム
# ============================================================
# 会社PCからGemini APIへの直接アクセスが遮断される事象(2026-08-10頃)を受け、
# rtocs_organizer / analog_ic_se_strategy_organizer / outlook_total_organizer と
# 同様に、共通モジュール gemini_client.py の generate_advanced() 経由
# (直接呼び出しが失敗したら自宅PCプロキシへ自動フォールバック)へ移行した。
#
# 本ツールは旧SDK(google.generativeai)の genai.GenerativeModel(...).generate_content()
# を1箇所で使っていただけだが、他ツールと実装を揃えるため、同じ形の薄い互換シム
# (_CommonGeminiClient)を用意し、そこを経由する方式にした。
# 旧SDK(google-generativeai)への依存は本バージョンで無くなった。
#
# 必要な環境変数(会社PC):
#   GEMINI_API_KEY   … 直接呼び出し用(gemini_client.py 側が読む)
#   GEMINI_PROXY_URL … 自宅PCプロキシのURL(直接呼び出し失敗時のフォールバック先)
#   GEMINI_COMMON_DIR… gemini_client.py の置き場所を明示したい場合のみ(任意)

_GEMINI_COMMON_DIR_ENV = os.environ.get("GEMINI_COMMON_DIR")
if _GEMINI_COMMON_DIR_ENV:
    _COMMON_DIR_CANDIDATES = [_GEMINI_COMMON_DIR_ENV]
else:
    # 会社PCでは本スクリプトが PythonScripts\excel\excel_transrate\ に、
    # gemini_client.py が PythonScripts\common\ に置かれるため、正解は「1つ上」では
    # なく「2つ上」の common になる。他ツール(1つ上が common)と同じ配置に置かれた
    # 場合でも動くよう、上位ディレクトリを順に探して最初に見つかったものを使う。
    _SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    _COMMON_DIR_CANDIDATES = [
        os.path.abspath(os.path.join(_SCRIPT_DIR, *([os.pardir] * _n + ["common"])))
        for _n in (1, 2, 3)
    ]

_COMMON_DIR = next(
    (_d for _d in _COMMON_DIR_CANDIDATES
     if os.path.isfile(os.path.join(_d, "gemini_client.py"))),
    _COMMON_DIR_CANDIDATES[0])

if _COMMON_DIR not in sys.path:
    sys.path.insert(0, _COMMON_DIR)

# gemini_client のインポートはここで試みるが、失敗しても import 時点では落とさない
# (原因が分かるメッセージを起動時チェック・AI呼び出し時に出すため)。
try:
    from gemini_client import generate_advanced as _generate_advanced
    _GEMINI_CLIENT_IMPORT_ERROR = None
except Exception as _e:
    _generate_advanced = None
    _GEMINI_CLIENT_IMPORT_ERROR = _e

# 本ツールは全機能が翻訳(AI呼び出し)のため、共通モジュールを読み込めたかどうかが
# そのまま「Gemini が使えるか」になる。
HAS_GEMINI = _generate_advanced is not None
if not HAS_GEMINI:
    print(f"警告: Gemini共通モジュール(gemini_client.py)を読み込めませんでした: "
          f"{_GEMINI_CLIENT_IMPORT_ERROR}")


def _gemini_common_module_error_message():
    """共通モジュールを読み込めなかったときの、原因が分かる案内文を組み立てる。"""
    return ("Gemini共通モジュール(gemini_client.py)を読み込めませんでした。\n"
            f"探索したパス: {' / '.join(_COMMON_DIR_CANDIDATES)}\n"
            f"元のエラー: {_GEMINI_CLIENT_IMPORT_ERROR}\n\n"
            "gemini-common-tools を配置し、必要なら環境変数 GEMINI_COMMON_DIR で\n"
            "gemini_client.py のあるフォルダを指定してください。")


def _schema_to_jsonable(value):
    """REST APIのpayloadへそのまま載せられる素のdict/listへ変換する。
    本ツールは response_schema を使っていないため通常は出番がないが、
    他ツールのシムと実装を揃えるために残してある(pydanticモデルが渡された
    場合にJSON化できなくなるのを防ぐ保険)。"""
    if value is None or isinstance(value, (dict, list, str, int, float, bool)):
        return value
    # pydantic v2 (model_dump) / v1 (dict) の両方に対応。REST APIのフィールド名は
    # camelCase なので by_alias=True で別名を使う。
    for attr, kwargs in (("model_dump", {"mode": "json", "exclude_none": True, "by_alias": True}),
                         ("dict", {"exclude_none": True, "by_alias": True})):
        fn = getattr(value, attr, None)
        if callable(fn):
            try:
                return fn(**kwargs)
            except Exception:
                try:
                    return fn()
                except Exception:
                    pass
    return value


def _system_instruction_to_jsonable(system_instruction):
    """system_instruction を REST API の systemInstruction フィールド形式へ変換する。
    本ツールは文字列で渡す。旧SDKでは genai.GenerativeModel(system_instruction=...)
    に渡していたもので、REST では payload の最上位フィールドとして送る必要がある
    (ここを落とすと「翻訳結果だけを出力する」という指示が silent に消える)。"""
    if system_instruction is None:
        return None
    if isinstance(system_instruction, str):
        return {"parts": [{"text": system_instruction}]}
    if isinstance(system_instruction, dict):
        return system_instruction
    return _schema_to_jsonable(system_instruction)


class _GeminiGenerateConfig:
    """google.genai の types.GenerateContentConfig(...) 相当の設定オブジェクト。
    本ツールは旧SDKしか使っておらず google.genai を import していないため、
    新しい依存を増やさないよう同等の入れ物をここに置く。シム側は getattr で
    属性を読むだけなので、将来 types.GenerateContentConfig を渡すようにしても
    そのまま動く。"""
    def __init__(self, temperature=None, system_instruction=None,
                 response_mime_type=None, response_schema=None):
        self.temperature = temperature
        self.system_instruction = system_instruction
        self.response_mime_type = response_mime_type
        self.response_schema = response_schema


class _CommonUsageMetadata:
    """response.usage_metadata 互換(トークン計測用)。本ツールは現時点で参照して
    いないが、他ツールのシムと契約を揃えておく。"""
    def __init__(self, usage):
        usage = usage if isinstance(usage, dict) else {}
        self.prompt_token_count = usage.get("promptTokenCount", 0)
        self.candidates_token_count = usage.get("candidatesTokenCount", 0)


class _CommonGeminiResponse:
    """client.models.generate_content(...) の戻り値互換。
    レスポンスが空・想定外の形でも例外を投げず、text は空文字にする
    (呼び出し側で原文フォールバックできるようにするため)。"""
    def __init__(self, raw):
        try:
            self.text = raw["candidates"][0]["content"]["parts"][0]["text"]
        except (KeyError, IndexError, TypeError):
            self.text = ""
        self.usage_metadata = _CommonUsageMetadata(
            raw.get("usageMetadata", {}) if isinstance(raw, dict) else {})


class _CommonGeminiModels:
    """client.models 互換。"""
    def generate_content(self, model=None, contents=None, config=None):
        if _generate_advanced is None:
            raise RuntimeError(_gemini_common_module_error_message())

        payload = {"contents": [{"parts": [{"text": contents}]}]}
        if config is not None:
            system_instruction = _system_instruction_to_jsonable(
                getattr(config, "system_instruction", None))
            if system_instruction:
                payload["systemInstruction"] = system_instruction

            gen_cfg = {}
            mime = getattr(config, "response_mime_type", None)
            if mime:
                gen_cfg["responseMimeType"] = mime
            schema = getattr(config, "response_schema", None)
            if schema is not None:
                gen_cfg["responseSchema"] = _schema_to_jsonable(schema)
            temp = getattr(config, "temperature", None)
            if temp is not None:
                gen_cfg["temperature"] = temp
            if gen_cfg:
                payload["generationConfig"] = gen_cfg

        # model は明示的に渡す(共通モジュール側の既定モデルへ勝手にフォールバック
        # されると、意図したモデルと実際に使われるモデルが食い違うため)。
        raw = _generate_advanced(payload, model=model)
        return _CommonGeminiResponse(raw)


class _CommonGeminiClient:
    """genai.Client(api_key=...) 相当。api_key は gemini_client.py 側が環境変数
    GEMINI_API_KEY から読むため、ここでは互換性のために受け取るだけで使用しない。"""
    def __init__(self, api_key=None):
        self.models = _CommonGeminiModels()


def gemini_credentials_available():
    """AI呼び出しが行える見込みがあるかどうかの事前チェック。
    直接呼び出しが遮断されていてもプロキシ経由なら成功しうるため、
    GEMINI_API_KEY / GEMINI_PROXY_URL のどちらか一方でも設定されていれば通す
    (プロキシ専用構成を誤って弾かないため)。"""
    return bool(os.environ.get("GEMINI_API_KEY") or os.environ.get("GEMINI_PROXY_URL"))


# 必要なライブラリの存在確認
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border
    from openpyxl.utils import get_column_letter, column_index_from_string
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("警告: openpyxlライブラリがインストールされていません")

import shutil
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import time

def check_dependencies():
    """起動時の依存関係チェック"""
    missing_libs = []

    if not HAS_OPENPYXL:
        missing_libs.append("openpyxl")

    if missing_libs:
        error_msg = f"以下のライブラリがインストールされていません:\n"
        for lib in missing_libs:
            error_msg += f"- {lib}\n"
        error_msg += f"\n以下のコマンドでインストールしてください:\n"
        error_msg += f"pip install {' '.join(missing_libs)}"

        messagebox.showerror("依存関係エラー", error_msg)
        return False

    # 本ツールは全機能が翻訳(AI呼び出し)のため、共通モジュールが読めない場合は
    # 起動を続けても何もできない。原因が分かる形で案内して終了する。
    if not HAS_GEMINI:
        messagebox.showerror("依存関係エラー", _gemini_common_module_error_message())
        return False

    return True


# 旧版にあった genai.configure(api_key=..., transport='rest') は削除した。
# 認証情報(GEMINI_API_KEY / GEMINI_PROXY_URL)は共通モジュール gemini_client.py が
# 環境変数から読むため、本スクリプト側での設定は不要になった。


def get_merged_cell_info(worksheet):
    """結合セル情報を取得"""
    merged_ranges = {}
    for merged_range in worksheet.merged_cells.ranges:
        min_row, min_col = merged_range.min_row, merged_range.min_col
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_ranges[(row, col)] = (min_row, min_col)
    return merged_ranges

def parse_excel_range(range_str, max_row, max_col):
    """Excel範囲文字列を行列番号に変換"""
    if not range_str or not range_str.strip():
        return 1, 1, max_row, max_col
    
    range_str = range_str.strip().upper()
    
    try:
        # A1:C10 形式
        if ':' in range_str:
            start_cell, end_cell = range_str.split(':')
            
            # 列全体指定 (A:C)
            if start_cell.isalpha() and end_cell.isalpha():
                start_col = column_index_from_string(start_cell)
                end_col = column_index_from_string(end_cell)
                return 1, start_col, max_row, end_col
            
            # 行全体指定 (1:5)
            elif start_cell.isdigit() and end_cell.isdigit():
                start_row = int(start_cell)
                end_row = int(end_cell)
                return start_row, 1, end_row, max_col
            
            # 通常の範囲指定 (A1:C10)
            else:
                # 開始セル解析
                start_match = re.match(r'([A-Z]+)(\d+)', start_cell)
                if start_match:
                    start_col = column_index_from_string(start_match.group(1))
                    start_row = int(start_match.group(2))
                else:
                    raise ValueError("開始セルの形式が正しくありません")
                
                # 終了セル解析
                end_match = re.match(r'([A-Z]+)(\d+)', end_cell)
                if end_match:
                    end_col = column_index_from_string(end_match.group(1))
                    end_row = int(end_match.group(2))
                else:
                    raise ValueError("終了セルの形式が正しくありません")
                
                return start_row, start_col, end_row, end_col
        
        else:
            raise ValueError("範囲指定には':'が必要です")
    
    except Exception as e:
        raise ValueError(f"範囲の解析に失敗しました: {str(e)}")

def validate_range(range_str):
    """範囲指定の形式をチェック"""
    if not range_str or not range_str.strip():
        return True, ""
    
    try:
        # 基本的な形式チェック
        range_pattern = r'^[A-Z]+[0-9]*:[A-Z]+[0-9]*$|^[A-Z]+:[A-Z]+$|^[0-9]+:[0-9]+$'
        if re.match(range_pattern, range_str.strip().upper()):
            return True, ""
        else:
            return False, "範囲の形式が正しくありません（例: A1:C10, A:C, 1:5）"
    except:
        return False, "範囲の形式が正しくありません"

def is_translatable(text):
    """翻訳が必要なテキストかどうかを判定"""
    if not text or pd.isna(text):
        return False
    
    text_str = str(text).strip()
    
    # 空文字、記号のみ、数値のみは翻訳不要
    if text_str == "" or text_str in ["#", "-", "N/A", "NULL"]:
        return False
    
    # 純粋な数値は翻訳不要
    if re.match(r'^-?\d+\.?\d*$', text_str):
        return False
    
    # 日付形式は翻訳不要
    if re.match(r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}$', text_str):
        return False
    
    # 短い記号や英数字のみは翻訳不要
    if len(text_str) <= 2 and re.match(r'^[A-Za-z0-9\-_]+$', text_str):
        return False
    
    return True

def translate_batch_parallel(batch_items, target_language="English", batch_id=0):
    """並列処理用のバッチ翻訳関数"""
    if not HAS_GEMINI:
        print("Gemini共通モジュール(gemini_client.py)が利用できません")
        return batch_items

    try:
        if not batch_items:
            return []

        # バッチテキストを作成
        batch_text = ""
        for i, item in enumerate(batch_items, 1):
            batch_text += f"{i}. {item}\n"

        system_instruction = "You are a professional translator. Output only the translated text in the specified numbered list format."

        prompt = f"""Translate the following numbered texts to {target_language}. 
Keep the same numbering format and maintain the original meaning.
Do not use Markdown formatting. Do not bold numbers.
Output exactly in the format: "1. Translated Text"

{batch_text}
"""

        # 旧: genai.GenerativeModel(model_name=..., system_instruction=...,
        #         generation_config={"temperature": 0}).generate_content(prompt)
        # 新: 共通モジュール(gemini_client.py)経由の互換シムで同じ内容を送る。
        client = _CommonGeminiClient()
        config = _GeminiGenerateConfig(
            temperature=0,
            system_instruction=system_instruction
        )

        response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt,
            config=config
        )

        # レスポンスを行ベースで解析（番号境界問題を根本解消）
        response_text = response.text.strip()
        if not response_text:
            print(f"バッチ {batch_id} 警告: 応答が空でした（原文のまま出力します）")
        translated_items = []
        parsed = {}
        for line in response_text.splitlines():
            m = re.match(r'^(\d+)\.\s*(.+)', line.strip())
            if m:
                parsed[int(m.group(1))] = m.group(2).strip()
        for i in range(1, len(batch_items) + 1):
            translated_items.append(parsed.get(i, batch_items[i - 1]))

        print(f"バッチ {batch_id} 完了 ({len(batch_items)} 項目)")
        return translated_items

    except Exception as e:
        # エラー時はスキップ（原文を返す）
        print(f"バッチ {batch_id} エラー: {e}")
        return batch_items

class SimpleProgressWindow:
    """シンプルな進捗表示ウィンドウ"""
    def __init__(self, parent, total_batches, total_items):
        self.window = tk.Toplevel(parent)
        self.window.title("翻訳進捗")
        self.window.geometry("400x180")
        self.window.resizable(False, False)

        try:
            self.window.transient(parent)
            self.window.grab_set()
        except:
            pass

        # シート進捗表示（複数シート対応）
        self.sheet_label = tk.Label(self.window, text="", font=("Arial", 10, "bold"))
        self.sheet_label.pack(pady=5)

        # バッチ進捗ラベル
        self.progress_label = tk.Label(self.window, text="翻訳を開始しています...",
                                      font=("Arial", 10))
        self.progress_label.pack(pady=5)

        # 詳細ラベル
        self.detail_label = tk.Label(self.window, text="", font=("Arial", 9))
        self.detail_label.pack(pady=5)

        # 時間表示
        self.time_label = tk.Label(self.window, text="", font=("Arial", 8), fg="blue")
        self.time_label.pack(pady=5)

        self.total_batches = total_batches
        self.total_items = total_items
        self.total_sheets = 1
        self.current_sheet = 1
        self.completed_batches = 0
        self.start_time = time.time()
        self.lock = threading.Lock()

    def update_progress(self, completed_batches, current_sheet=1, total_sheets=1, sheet_name=""):
        """進捗を更新"""
        try:
            with self.lock:
                self.completed_batches = completed_batches
                percentage = int((completed_batches / self.total_batches) * 100) if self.total_batches > 0 else 0

                elapsed_time = time.time() - self.start_time

                self.sheet_label.config(
                    text=f"シート {current_sheet}/{total_sheets}: {sheet_name}")
                self.progress_label.config(
                    text=f"翻訳進捗: {completed_batches}/{self.total_batches} バッチ ({percentage}%)")
                self.detail_label.config(
                    text=f"完了項目数: 約{completed_batches * 8}/{self.total_items}")
                self.time_label.config(
                    text=f"経過時間: {elapsed_time:.1f}秒")

                self.window.update()
        except:
            pass

    def close(self):
        """ウィンドウを閉じる"""
        try:
            self.window.destroy()
        except:
            pass

def copy_cell_format(source_cell, target_cell):
    """セルの書式をコピー（エラーハンドリング強化）"""
    if not HAS_OPENPYXL:
        return
    
    try:
        if source_cell.font:
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                color=source_cell.font.color
            )
    except:
        pass
    
    try:
        if source_cell.fill and source_cell.fill.fill_type:
            target_cell.fill = PatternFill(
                fill_type=source_cell.fill.fill_type,
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color
            )
    except:
        pass
    
    try:
        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                wrap_text=source_cell.alignment.wrap_text
            )
    except:
        pass

def preserve_data_type(original_value, translated_value):
    """元のデータ型を保持"""
    if pd.isna(original_value) or original_value == "":
        return translated_value
    
    if isinstance(original_value, (int, float)):
        try:
            if str(translated_value).replace('.', '').replace('-', '').isdigit():
                return float(translated_value) if '.' in str(translated_value) else int(translated_value)
        except:
            pass
    
    return translated_value


def translate_excel_parallel(file_path, sheet_names, target_language="English", max_workers=3, translation_range=None):
    """複数シートを翻訳して1ファイルに出力する（1ファイル出力版）"""
    try:
        if not HAS_OPENPYXL:
            messagebox.showerror("エラー", "openpyxlライブラリが必要です。\npip install openpyxl")
            return
        if not HAS_GEMINI:
            messagebox.showerror("エラー", _gemini_common_module_error_message())
            return

        file_ext = os.path.splitext(file_path)[1].lower()
        is_macro_file = file_ext == '.xlsm'
        is_legacy_file = file_ext == '.xls'

        if file_ext not in ['.xlsx', '.xlsm', '.xls']:
            messagebox.showerror("エラー", f"サポートされていないファイル形式: {file_ext}")
            return

        if is_legacy_file:
            result = messagebox.askwarning("警告",
                ".xlsファイルは古い形式です。\n翻訳後は.xlsxとして保存されます。\n続行しますか？")
            if not result:
                return

        # ── 出力ファイル名を1回だけ生成 ──
        lang_abbr_map = {
            "English":             "en",
            "Chinese Simplified":  "zh",
            "Chinese Traditional": "zt",
            "Japanese":            "jp",
        }
        lang_abbr = lang_abbr_map.get(target_language, target_language[:2].lower())
        timestamp  = time.strftime("%Y%m%d_%H%M%S")
        base_name_raw = os.path.splitext(file_path)[0]
        base_name = re.sub(r'_(en|zh|zt|jp)_\d{8}_\d{6}$', '', base_name_raw)
        ext = ".xlsm" if is_macro_file else ".xlsx"
        output_path = f"{base_name}_{lang_abbr}_{timestamp}{ext}"

        # ── 元ファイルを出力先にコピー（書式・VBA保持）──
        load_params = {'data_only': False}
        if is_macro_file:
            load_params['keep_vba'] = True

        if is_legacy_file:
            # .xls は pandas 経由で新規ワークブックを作成
            try:
                df_dict = pd.read_excel(file_path, sheet_name=None, header=None)
                original_wb = openpyxl.Workbook()
                original_wb.remove(original_wb.active)
                for sname, df in df_dict.items():
                    ws = original_wb.create_sheet(title=sname)
                    for r_idx, row in df.iterrows():
                        for c_idx, val in enumerate(row):
                            if pd.notna(val):
                                ws.cell(row=r_idx+1, column=c_idx+1, value=val)
                work_wb = openpyxl.Workbook()
                work_wb.remove(work_wb.active)
                for sname in original_wb.sheetnames:
                    work_wb.create_sheet(title=sname)
            except Exception as e:
                messagebox.showerror("エラー", f".xlsファイルの読み込みに失敗:\n{e}")
                return
        else:
            try:
                shutil.copy2(file_path, output_path)
                original_wb = openpyxl.load_workbook(file_path, **load_params)
                work_wb    = openpyxl.load_workbook(output_path, **load_params)
            except Exception as e:
                messagebox.showerror("エラー", f"ファイルの読み込みに失敗:\n{e}")
                return

        # ── 各シートを順次翻訳 ──
        total_sheets = len(sheet_names)
        for sheet_index, sheet_name in enumerate(sheet_names, 1):
            if sheet_name not in original_wb.sheetnames:
                messagebox.showwarning("警告", f"シート '{sheet_name}' が見つかりません。スキップします。")
                continue
            _translate_sheet_content(
                original_wb, work_wb, sheet_name,
                target_language, max_workers, translation_range,
                sheet_index, total_sheets, is_legacy_file)

        # ── 全シート完了後に1回だけ保存 ──
        try:
            work_wb.save(output_path)
            work_wb.close()
            original_wb.close()
        except Exception as e:
            messagebox.showerror("エラー", f"ファイルの保存に失敗:\n{e}")
            try:
                work_wb.close()
                original_wb.close()
            except:
                pass
            return

        sheet_list = "、".join(sheet_names)
        messagebox.showinfo("完了",
            f"翻訳完了！\n"
            f"翻訳シート: {sheet_list}\n"
            f"保存先: {output_path}")

    except Exception as e:
        try:
            if 'original_wb' in locals(): original_wb.close()
            if 'work_wb'    in locals(): work_wb.close()
        except:
            pass
        messagebox.showerror("エラー", f"翻訳処理中にエラーが発生しました:\n{e}")


def _translate_sheet_content(original_wb, work_wb, sheet_name,
                              target_language, max_workers, translation_range,
                              sheet_index, total_sheets, is_legacy_file):
    """単一シートの翻訳内容書き込み（ファイルI/Oなし）"""
    try:
        original_ws = original_wb[sheet_name]
        work_ws     = work_wb[sheet_name]

        # 結合セル情報をコピー
        merged_ranges = {}
        if not is_legacy_file:
            try:
                merged_ranges = get_merged_cell_info(original_ws)
                for merged_range in original_ws.merged_cells.ranges:
                    try:
                        work_ws.merge_cells(str(merged_range))
                    except:
                        pass
            except Exception as e:
                print(f"結合セル情報の取得に失敗: {e}")

        # 列幅コピー
        if not is_legacy_file:
            try:
                for col_letter in original_ws.column_dimensions:
                    if original_ws.column_dimensions[col_letter].width:
                        work_ws.column_dimensions[col_letter].width = \
                            original_ws.column_dimensions[col_letter].width
            except Exception as e:
                print(f"列幅のコピーに失敗: {e}")

        # 翻訳範囲を決定
        if translation_range:
            try:
                start_row, start_col, end_row, end_col = parse_excel_range(
                    translation_range, original_ws.max_row, original_ws.max_column)
            except ValueError as e:
                messagebox.showerror("エラー", f"範囲指定エラー: {e}")
                return
        else:
            start_row, start_col = 1, 1
            end_row, end_col = original_ws.max_row, original_ws.max_column

        # 翻訳対象セルを収集
        translatable_items = []
        cell_positions     = []
        processed_merged   = set()

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                try:
                    if (row, col) in merged_ranges:
                        master_pos = merged_ranges[(row, col)]
                        if master_pos in processed_merged:
                            continue
                        processed_merged.add(master_pos)
                        cell = original_ws.cell(row=master_pos[0], column=master_pos[1])
                        pos  = master_pos
                    else:
                        cell = original_ws.cell(row=row, column=col)
                        pos  = (row, col)

                    if is_translatable(cell.value):
                        translatable_items.append(str(cell.value))
                        cell_positions.append(pos)
                except Exception as e:
                    print(f"セル ({row},{col}) 収集エラー: {e}")

        if not translatable_items:
            print(f"[{sheet_name}] 翻訳対象なし")
            return

        print(f"[{sheet_name}] 翻訳対象: {len(translatable_items)} 項目")

        # バッチ分割
        batch_size     = 8
        batches        = [translatable_items[i:i+batch_size]
                          for i in range(0, len(translatable_items), batch_size)]
        batch_positions = [cell_positions[i:i+batch_size]
                           for i in range(0, len(cell_positions), batch_size)]

        # 進捗ウィンドウ
        progress_window = SimpleProgressWindow(root, len(batches), len(translatable_items))
        progress_window.update_progress(0, sheet_index, total_sheets, sheet_name)

        translated_batches = [None] * len(batches)
        completed_count    = 0

        def update_progress_callback():
            nonlocal completed_count
            completed_count += 1
            progress_window.update_progress(completed_count, sheet_index, total_sheets, sheet_name)

        safe_workers = min(max_workers, len(batches), 4)

        try:
            with ThreadPoolExecutor(max_workers=safe_workers) as executor:
                future_to_index = {
                    executor.submit(translate_batch_parallel, batch, target_language, i+1): i
                    for i, batch in enumerate(batches)
                }
                for future in as_completed(future_to_index):
                    idx = future_to_index[future]
                    try:
                        translated_batches[idx] = future.result()
                    except Exception as e:
                        print(f"バッチ {idx+1} 例外: {e}")
                        translated_batches[idx] = batches[idx]
                    update_progress_callback()
        except Exception as e:
            progress_window.close()
            messagebox.showerror("エラー", f"翻訳処理中にエラー:\n{e}")
            return

        # 翻訳結果をwork_wsに書き込み
        for translated_batch, positions_batch in zip(translated_batches, batch_positions):
            if translated_batch is None:
                continue
            for translated_text, (row, col) in zip(translated_batch, positions_batch):
                try:
                    original_cell = original_ws.cell(row=row, column=col)
                    work_cell     = work_ws.cell(row=row, column=col)
                    work_cell.value = preserve_data_type(original_cell.value, translated_text)
                    if not is_legacy_file:
                        copy_cell_format(original_cell, work_cell)
                except Exception as e:
                    print(f"セル ({row},{col}) 書き込みエラー: {e}")

        # 翻訳範囲外セルをコピー（値・書式）
        translated_set = set(cell_positions)
        for row in range(1, original_ws.max_row + 1):
            for col in range(1, original_ws.max_column + 1):
                if (row, col) not in translated_set:
                    try:
                        original_cell = original_ws.cell(row=row, column=col)
                        work_cell     = work_ws.cell(row=row, column=col)
                        if original_cell.value is not None:
                            work_cell.value = original_cell.value
                        if not is_legacy_file:
                            copy_cell_format(original_cell, work_cell)
                    except Exception as e:
                        print(f"セル ({row},{col}) コピーエラー: {e}")

        progress_window.close()
        print(f"[{sheet_name}] 翻訳完了")

    except Exception as e:
        if 'progress_window' in locals():
            try: progress_window.close()
            except: pass
        messagebox.showerror("エラー", f"[{sheet_name}] 処理中にエラー:\n{e}")     
        
def select_file():
    """ファイル選択とシート選択の処理（複数シート対応版）"""
    path = filedialog.askopenfilename(
        title="翻訳するExcelファイルを選択してください",
        filetypes=[
            ("Excel files", "*.xlsx *.xlsm *.xls"),
            ("Excel Workbook", "*.xlsx"),
            ("Excel Macro-Enabled Workbook", "*.xlsm"),
            ("Excel 97-2003 Workbook", "*.xls"),
            ("All files", "*.*")
        ]
    )

    if not path:
        return

    file_ext = os.path.splitext(path)[1].lower()
    if file_ext not in ['.xlsx', '.xlsm', '.xls']:
        messagebox.showerror("エラー", f"サポートされていないファイル形式です: {file_ext}\n対応形式: .xlsx, .xlsm, .xls")
        return

    try:
        if file_ext == '.xls':
            excel_file = pd.ExcelFile(path)
            sheets = excel_file.sheet_names
            excel_file.close()
        else:
            load_params = {'data_only': True}
            if file_ext == '.xlsm':
                load_params['keep_vba'] = True
            wb = openpyxl.load_workbook(path, **load_params)
            sheets = wb.sheetnames
            wb.close()

        sheet_win = tk.Toplevel(root)
        sheet_win.title("翻訳設定")
        sheet_win.geometry("500x680")
        sheet_win.resizable(False, False)

        try:
            sheet_win.transient(root)
            sheet_win.grab_set()
        except:
            pass

        main_frame = tk.Frame(sheet_win)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # シート選択セクション（複数選択対応）
        sheet_frame = tk.Frame(main_frame)
        sheet_frame.pack(fill="x", pady=(0, 20))

        tk.Label(sheet_frame, text="翻訳対象シートを選択してください",
                font=("Arial", 12, "bold")).pack(anchor="w")

        sheet_listbox = tk.Listbox(sheet_frame, selectmode=tk.MULTIPLE,
                                   font=("Arial", 10), height=min(len(sheets), 6),
                                   exportselection=False)
        for s in sheets:
            sheet_listbox.insert(tk.END, s)
        sheet_listbox.selection_set(0)  # デフォルト: 先頭シートを選択
        sheet_listbox.pack(anchor="w", pady=(5, 0), fill="x")
        tk.Label(sheet_frame, text="※ Ctrl+クリックで複数選択",
                 font=("Arial", 8), fg="gray").pack(anchor="w")

        # 範囲選択セクション
        range_frame = tk.LabelFrame(main_frame, text="翻訳範囲設定",
                                   font=("Arial", 11, "bold"), padx=15, pady=15)
        range_frame.pack(fill="x", pady=(0, 20))

        translate_all_var = tk.BooleanVar(value=True)
        range_var = tk.StringVar()

        all_checkbox = tk.Checkbutton(range_frame,
                                     text="シート全体を翻訳",
                                     variable=translate_all_var,
                                     font=("Arial", 10))
        all_checkbox.pack(anchor="w")

        range_input_frame = tk.Frame(range_frame)
        range_input_frame.pack(fill="x", pady=(10, 0))

        range_checkbox_var = tk.BooleanVar(value=False)
        range_checkbox = tk.Checkbutton(range_input_frame,
                                       text="範囲指定:",
                                       variable=range_checkbox_var,
                                       font=("Arial", 10))
        range_checkbox.pack(side="left")

        range_entry = tk.Entry(range_input_frame, textvariable=range_var,
                              font=("Arial", 10), width=15)
        range_entry.pack(side="left", padx=(5, 0))

        example_label = tk.Label(range_frame,
                               text="例: A1:C10, B2:D20, A:C (列全体), 1:5 (行全体)",
                               font=("Arial", 8), fg="gray")
        example_label.pack(anchor="w", pady=(5, 0))

        error_label = tk.Label(range_frame, text="", fg="red", font=("Arial", 9))
        error_label.pack(anchor="w", pady=(5, 0))

        def on_range_entry_change(*args):
            if range_var.get().strip():
                translate_all_var.set(False)
                range_checkbox_var.set(True)
                error_label.config(text="")
            else:
                range_checkbox_var.set(False)

        def on_all_checkbox_change():
            if translate_all_var.get():
                range_checkbox_var.set(False)
                range_var.set("")
                error_label.config(text="")

        def on_range_checkbox_change():
            if range_checkbox_var.get():
                translate_all_var.set(False)
                range_entry.focus()
            else:
                range_var.set("")

        range_var.trace("w", on_range_entry_change)
        all_checkbox.config(command=on_all_checkbox_change)
        range_checkbox.config(command=on_range_checkbox_change)

        tk.Frame(main_frame, height=2, bg="lightgray").pack(fill="x", pady=20)

        # 言語選択セクション
        lang_frame = tk.Frame(main_frame)
        lang_frame.pack(fill="x", pady=(0, 20))

        tk.Label(lang_frame, text="翻訳先言語を選択してください",
                font=("Arial", 12, "bold")).pack(anchor="w")

        languages = {
            "英語 (English)": "English",
            "中国語簡体字 (Chinese Simplified)": "Chinese Simplified",
            "中国語繁体字 (Chinese Traditional)": "Chinese Traditional",
            "日本語 (Japanese)": "Japanese"
        }

        lang_var = tk.StringVar(sheet_win)
        lang_var.set("英語 (English)")
        lang_menu = tk.OptionMenu(lang_frame, lang_var, *languages.keys())
        lang_menu.config(font=("Arial", 10), width=25)
        lang_menu.pack(anchor="w", pady=(5, 0))

        def start_translation():
            """翻訳開始処理"""
            selected_language = languages[lang_var.get()]

            if not translate_all_var.get():
                valid, error_msg = validate_range(range_var.get())
                if not valid:
                    error_label.config(text=error_msg)
                    return

            translation_range = None if translate_all_var.get() else range_var.get().strip()

            selected_indices = sheet_listbox.curselection()
            if not selected_indices:
                messagebox.showwarning("警告", "シートを1つ以上選択してください。")
                return

            selected_sheets = [sheets[i] for i in selected_indices]
            sheet_win.destroy()
            translate_excel_parallel(path, selected_sheets, selected_language, 3, translation_range)

        # ボタンフレーム
        button_frame = tk.Frame(main_frame)
        button_frame.pack(pady=20)

        tk.Button(button_frame, text="翻訳開始", command=start_translation,
                 bg="#4CAF50", fg="white", padx=25, pady=12,
                 font=("Arial", 12, "bold")).pack(side=tk.LEFT, padx=15)
        tk.Button(button_frame, text="キャンセル", command=sheet_win.destroy,
                 padx=25, pady=12, font=("Arial", 12)).pack(side=tk.LEFT, padx=15)

    except Exception as e:
        messagebox.showerror("エラー", f"ファイルの読み込み中にエラーが発生しました:\n{str(e)}")


def check_api_key():
    """Gemini 認証情報の存在確認"""
    if not gemini_credentials_available():
        messagebox.showerror("エラー",
                           "Gemini認証情報が設定されていません。\n"
                           "以下のいずれかを設定してください:\n"
                           "- 環境変数 GEMINI_API_KEY （直接接続用）\n"
                           "- 環境変数 GEMINI_PROXY_URL （自宅PCプロキシ経由用）\n\n"
                           "※ setx で設定した場合は、コマンドプロンプトを\n"
                           "　 開き直してから起動してください。")
        return False
    return True


# GUI初期設定
if __name__ == "__main__":
    # 依存関係チェック
    if not check_dependencies():
        sys.exit(1)
    
    # API キーをチェック
    if not check_api_key():
        sys.exit(1)
    
    try:
        root = tk.Tk()
        root.title(f"Excel翻訳ツール (v{VERSION})")
        root.geometry("500x250")
        root.resizable(False, False)
        
        main_frame = tk.Frame(root)
        main_frame.pack(expand=True, fill='both', padx=20, pady=20)
        
        # タイトル
        title_label = tk.Label(main_frame, 
                              text="Excel翻訳ツール", 
                              font=("Arial", 16, "bold"))
        title_label.pack(pady=(10, 2))
        
        # バージョン表示
        version_label = tk.Label(main_frame, 
                                text=f"Version {VERSION}", 
                                font=("Arial", 9), fg="gray")
        version_label.pack(pady=(0, 10))
        
        # 説明
        desc_label = tk.Label(main_frame, 
                             text="Excelファイルを選択して翻訳します\n（並列処理対応・高速版・範囲指定対応）",
                             font=("Arial", 10))
        desc_label.pack(pady=5)
        
        # ファイル選択ボタン
        select_button = tk.Button(main_frame, 
                                 text="ファイル選択", 
                                 command=select_file,
                                 font=("Arial", 12),
                                 bg="#4CAF50",
                                 fg="white",
                                 padx=20,
                                 pady=10)
        select_button.pack(pady=15)
        
        # 注意事項
        note_label = tk.Label(main_frame, 
                             text="注意: 翻訳対象のExcelファイルは事前に閉じてください",
                             font=("Arial", 9),
                             fg="red")
        note_label.pack(pady=5)
        
        root.mainloop()
        
    except Exception as e:
        print(f"GUI初期化エラー: {e}")
        messagebox.showerror("起動エラー", f"アプリケーションの起動に失敗しました:\n{str(e)}")
        sys.exit(1)