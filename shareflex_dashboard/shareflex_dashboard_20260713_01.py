"""
Shareflex Document Dashboard

Nexus Document Management System (SharePoint / Shareflex) からエクスポートした
ドキュメント一覧(Excel)を読み込み、組織軸(Department)と業務プロセス軸(Top Level
Process)の2系統でドキュメント件数を集計した静的HTMLダッシュボードを生成する。

使い方:
    python shareflex_dashboard_20260713_01.py <export.xlsx> [-o output.html]
"""

import argparse
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as e:
    print(f"\n[エラー] 必要なライブラリ '{e.name}' が見つかりません。\n"
          "次のコマンドを実行してから再度実行してください:\n\n"
          "    pip install -r requirements.txt\n")
    sys.exit(1)


DEPARTMENT_LEVELS = ["Department", "Sub Team 1", "Sub Team 2"]
PROCESS_LEVELS = ["Top Level Process", "Sub Process 1", "Sub Process 2", "Sub Process 3"]
BREAKDOWN_COLUMNS = ["Document Type", "Document Status", "Confidentiality"]

EMPTY_LABEL = "(未設定)"


# ==========================================
# データ読み込み
# ==========================================

def load_documents(xlsx_path):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    header_index = {}
    documents = []

    for row in ws.iter_rows(values_only=True):
        if not header_index:
            if row and any(cell is not None and str(cell).strip() == "Document Number" for cell in row):
                header_index = {
                    str(cell).strip(): i
                    for i, cell in enumerate(row)
                    if cell is not None and str(cell).strip()
                }
            continue

        if row is None or all(cell is None for cell in row):
            continue

        documents.append({
            name: (row[idx] if idx < len(row) else None)
            for name, idx in header_index.items()
        })

    if not header_index:
        raise ValueError(
            "ヘッダー行(「Document Number」列を含む行)が見つかりませんでした。"
            "エクスポート形式を確認してください。"
        )

    return documents


# ==========================================
# 集計
# ==========================================

def clean(value):
    if value is None:
        return EMPTY_LABEL
    text = str(value).strip()
    return text if text else EMPTY_LABEL


def build_tree(documents, levels):
    root = {"count": 0, "children": {}}
    for doc in documents:
        node = root
        node["count"] += 1
        for level in levels:
            key = clean(doc.get(level))
            node = node["children"].setdefault(key, {"count": 0, "children": {}})
            node["count"] += 1
    return root


def count_by(documents, column):
    counts = defaultdict(int)
    for doc in documents:
        counts[clean(doc.get(column))] += 1
    return sorted(counts.items(), key=lambda kv: -kv[1])


# ==========================================
# HTML生成
# ==========================================

def esc(text):
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def render_tree(node, max_count, open_top=True):
    parts = []
    for label in sorted(node["children"], key=lambda k: -node["children"][k]["count"]):
        child = node["children"][label]
        width = round(child["count"] / max_count * 100, 1) if max_count else 0
        has_children = bool(child["children"])
        parts.append('<details class="tree-node"' + (" open" if open_top else "") + '>')
        parts.append(
            '<summary><span class="tree-label">%s</span>'
            '<span class="tree-bar"><span class="tree-bar-fill" style="width:%s%%"></span></span>'
            '<span class="tree-count">%d</span></summary>'
            % (esc(label), width, child["count"])
        )
        if has_children:
            parts.append('<div class="tree-children">')
            parts.append(render_tree(child, max_count, open_top=False))
            parts.append('</div>')
        parts.append('</details>')
    return "".join(parts)


def render_breakdown_table(counts, total):
    max_count = counts[0][1] if counts else 0
    rows = []
    for label, count in counts:
        width = round(count / max_count * 100, 1) if max_count else 0
        pct = round(count / total * 100, 1) if total else 0
        rows.append(
            '<tr><td>%s</td>'
            '<td class="bar-cell"><span class="bar-fill" style="width:%s%%"></span></td>'
            '<td class="num">%d</td><td class="num">%s%%</td></tr>'
            % (esc(label), width, count, pct)
        )
    return "".join(rows)


HTML_TEMPLATE = """<!DOCTYPE html><html lang="ja"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Shareflex Document Dashboard</title>
<style>
  body { font-family: "Segoe UI", "Meiryo", sans-serif; margin: 0; background: #f4f5f7; color: #1f2328; }
  header { background: #1f2937; color: #fff; padding: 20px 32px; }
  header h1 { margin: 0 0 4px; font-size: 20px; }
  header p { margin: 0; font-size: 13px; color: #c7cbd1; }
  main { max-width: 1080px; margin: 24px auto; padding: 0 16px 48px; }
  .cards { display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 24px; }
  .card { background: #fff; border: 1px solid #e2e4e8; border-radius: 8px; padding: 16px 20px; min-width: 150px; flex: 1; }
  .card .label { font-size: 12px; color: #6b7280; }
  .card .value { font-size: 28px; font-weight: 600; margin-top: 4px; }
  section { background: #fff; border: 1px solid #e2e4e8; border-radius: 8px; padding: 20px 24px; margin-bottom: 24px; }
  section h2 { margin-top: 0; font-size: 16px; }
  .tabs input[type=radio] { display: none; }
  .tab-labels { display: flex; gap: 4px; margin-bottom: 16px; }
  .tab-labels label { padding: 8px 16px; border-radius: 6px 6px 0 0; background: #eceef1; cursor: pointer; font-size: 13px; }
  .tab-panel { display: none; }
  #tab-dept:checked ~ .tab-labels label[for=tab-dept] { background: #1f2937; color: #fff; }
  #tab-proc:checked ~ .tab-labels label[for=tab-proc] { background: #1f2937; color: #fff; }
  #tab-dept:checked ~ .panels #panel-dept { display: block; }
  #tab-proc:checked ~ .panels #panel-proc { display: block; }
  .tree-node { margin: 2px 0 2px 0; }
  .tree-node > summary { list-style: none; cursor: pointer; display: flex; align-items: center; gap: 10px; padding: 4px 0; }
  .tree-node > summary::-webkit-details-marker { display: none; }
  .tree-node > summary::before { content: "▸"; color: #9aa0a8; width: 12px; }
  .tree-node[open] > summary::before { content: "▾"; }
  .tree-label { flex: 0 0 260px; font-size: 13px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
  .tree-bar { flex: 1; height: 8px; background: #eceef1; border-radius: 4px; overflow: hidden; }
  .tree-bar-fill { display: block; height: 100%; background: #3b82f6; }
  .tree-count { flex: 0 0 44px; text-align: right; font-size: 12px; color: #6b7280; }
  .tree-children { margin-left: 22px; border-left: 1px solid #eceef1; padding-left: 12px; }
  table { width: 100%; border-collapse: collapse; font-size: 13px; }
  table td, table th { padding: 6px 8px; border-bottom: 1px solid #eceef1; text-align: left; }
  table .num { text-align: right; white-space: nowrap; width: 60px; }
  .bar-cell { width: 40%; }
  .bar-fill { display: block; height: 8px; background: #3b82f6; border-radius: 4px; }
  footer { text-align: center; font-size: 12px; color: #9aa0a8; padding: 16px; }
</style></head><body>
<header>
  <h1>Shareflex Document Dashboard</h1>
  <p>Source: @@SOURCE@@ / Generated: @@GENERATED_AT@@ / Total Documents: @@TOTAL@@</p>
</header>
<main>
  <div class="cards">
    <div class="card"><div class="label">総ドキュメント数</div><div class="value">@@TOTAL@@</div></div>
    <div class="card"><div class="label">Department数</div><div class="value">@@DEPT_COUNT@@</div></div>
    <div class="card"><div class="label">Top Level Process数</div><div class="value">@@PROC_COUNT@@</div></div>
    <div class="card"><div class="label">Document Type数</div><div class="value">@@DOCTYPE_COUNT@@</div></div>
  </div>

  <section>
    <h2>階層別ドキュメント件数</h2>
    <div class="tabs">
      <input type="radio" name="tabs" id="tab-dept" checked>
      <input type="radio" name="tabs" id="tab-proc">
      <div class="tab-labels">
        <label for="tab-dept">組織軸 (Department)</label>
        <label for="tab-proc">プロセス軸 (Top Level Process)</label>
      </div>
      <div class="panels">
        <div class="tab-panel" id="panel-dept">@@DEPT_TREE@@</div>
        <div class="tab-panel" id="panel-proc">@@PROC_TREE@@</div>
      </div>
    </div>
  </section>

  <section>
    <h2>Document Type別</h2>
    <table><tbody>@@DOCTYPE_ROWS@@</tbody></table>
  </section>

  <section>
    <h2>Document Status別</h2>
    <table><tbody>@@STATUS_ROWS@@</tbody></table>
  </section>

  <section>
    <h2>Confidentiality別</h2>
    <table><tbody>@@CONF_ROWS@@</tbody></table>
  </section>
</main>
<footer>Generated by shareflex_dashboard_20260713_01.py</footer>
</body></html>"""


def generate_html(documents, source_name):
    total = len(documents)

    dept_tree = build_tree(documents, DEPARTMENT_LEVELS)
    proc_tree = build_tree(documents, PROCESS_LEVELS)

    doctype_counts = count_by(documents, "Document Type")
    status_counts = count_by(documents, "Document Status")
    conf_counts = count_by(documents, "Confidentiality")

    replacements = {
        "@@SOURCE@@": esc(source_name),
        "@@GENERATED_AT@@": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "@@TOTAL@@": str(total),
        "@@DEPT_COUNT@@": str(len({clean(d.get("Department")) for d in documents})),
        "@@PROC_COUNT@@": str(len({clean(d.get("Top Level Process")) for d in documents})),
        "@@DOCTYPE_COUNT@@": str(len(doctype_counts)),
        "@@DEPT_TREE@@": render_tree(dept_tree, dept_tree["count"] or 1),
        "@@PROC_TREE@@": render_tree(proc_tree, proc_tree["count"] or 1),
        "@@DOCTYPE_ROWS@@": render_breakdown_table(doctype_counts, total),
        "@@STATUS_ROWS@@": render_breakdown_table(status_counts, total),
        "@@CONF_ROWS@@": render_breakdown_table(conf_counts, total),
    }

    html = HTML_TEMPLATE
    for token, value in replacements.items():
        html = html.replace(token, value)
    return html


# ==========================================
# エントリポイント
# ==========================================

def main():
    parser = argparse.ArgumentParser(description="Shareflexドキュメント一覧からダッシュボードHTMLを生成する")
    parser.add_argument("input", help="SharePointからエクスポートしたExcelファイル(.xlsx)")
    parser.add_argument("-o", "--output", help="出力HTMLファイルパス(省略時は入力ファイルと同じ場所に *_dashboard.html)")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"[エラー] ファイルが見つかりません: {input_path}")
        sys.exit(1)

    output_path = Path(args.output) if args.output else input_path.with_name(input_path.stem + "_dashboard.html")

    documents = load_documents(input_path)
    html = generate_html(documents, input_path.name)

    output_path.write_text(html, encoding="utf-8")
    print(f"[完了] {len(documents)}件のドキュメントを集計しました。")
    print(f"[出力] {output_path}")


if __name__ == "__main__":
    main()
