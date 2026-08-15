# -*- coding: utf-8 -*-
"""
storage.py
学びジャーナル - Excel(SharePoint同期フォルダ)読み書きモジュール
Version: 0.11.0
"""

import os
import time
import shutil
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook

# ============================================================
# 設定値（要編集）
# ------------------------------------------------------------
# TODO: 越智さんのSharePoint(OneDrive)同期フォルダの実際のパスに
#       書き換えてください。例:
#       r"C:\Users\nx023836\Nexperia\JP Site - Journal\journal_data.xlsx"
# ============================================================
EXCEL_PATH = r"C:\Users\nx023836\Documents\PythonScripts\Journal\journal_data.xlsx"

ENTRIES_SHEET = "Entries"
TAGMASTER_SHEET = "TagMaster"
TIMELOG_SHEET = "TimeLog"
SUBITEM_SHEET = "SubItemMaster"
ACTIONS_SHEET = "Actions"
HIREBI_SHEET = "Hirebi"

ENTRIES_HEADER = ["日付", "タグ", "L", "K", "P", "T"]
TAGMASTER_HEADER = ["タグ名", "色コード"]
TIMELOG_HEADER = ["開始", "終了", "タグ", "中項目"]
SUBITEM_HEADER = ["タグ名", "中項目名", "表示順"]
# 「魂のひれぶり」信号。1タップ＝1件の記録とし、選ぶ・入力するを増やさない。
# 同じ瞬間に複数の感情がある場合は、その分だけタップすれば複数行として残る
# （1件に複数の感情フラグを持たせるより、後から集計しやすいため）
HIREBI_HEADER = ["日時", "タグ", "感情"]
EMOTION_LABELS = ["喜", "怒", "哀", "楽"]
# 完了・未完了という状態を持つため、追記のみの他シートとは別に管理する。
# 行番号自体を識別子として使う（削除・並べ替えを行わないため安定する）
ACTIONS_HEADER = ["作成日時", "タグ", "内容", "由来", "ステータス", "完了日時", "優先"]
ACTION_STATUS_PENDING = "未着手"
ACTION_STATUS_DONE = "完了"

# 前回チェックポイントから長時間経過していた場合の安全弁（この時間で打ち切る）
MAX_TIMELOG_GAP_HOURS = 2

# その日の最初のチェックインの起点となる勤務開始時刻。
# 以前は当日最初のチェックインを「0分の基準点」としていたため、8時から働いて
# 9時に初めて記録した場合、その1時間がどこにも残らなかった。この時刻を起点に
# することで、朝の分が自動的に埋まる（この時刻より前に記録した場合は従来通り
# 0分の基準点として扱う）
WORK_START_TIME = "08:00"

# MAX_TIMELOG_GAP_HOURSを超えて空いた分に付けるタグ。
# 以前は超過分を黙って捨てていたため、9:00→15:00のチェックインでは
# 13:00-15:00しか残らず4時間が消滅していた。捨てずにこのタグの行として
# 残すことで、ダッシュボード上で「ここが埋まっていない」と見えるようにする。
# TagMasterには登録しない（＝ポップアップの選択肢には出さない）
UNRECORDED_TAG = "未記録"

DEFAULT_TAGS = [
    ("R19", "#d9ae23"),
    ("JP Site", "#2fa84f"),
    ("NPI", "#c2399e"),
    ("その他", "#e08830"),
]

# タグ選択と同時に「何をしていたか」を1タップで選べる中項目の既定値。
# タグごとに5個以内（うち1つは自由記述の逃げ道）を目安に、SubItemMaster
# シートをExcelで直接編集すれば自由に追加・変更できる
OTHER_SUBITEM_LABEL = "Other"
DEFAULT_SUBITEMS = [
    ("R19", "メール返信", 1),
    ("R19", "データ整理", 2),
    ("R19", "会議", 3),
    ("R19", OTHER_SUBITEM_LABEL, 4),
    ("JP Site", "メール返信", 1),
    ("JP Site", "データ整理", 2),
    ("JP Site", "会議", 3),
    ("JP Site", OTHER_SUBITEM_LABEL, 4),
    ("NPI", "Caracal", 1),
    ("NPI", "Wheeling", 2),
    ("NPI", "GT", 3),
    ("NPI", OTHER_SUBITEM_LABEL, 4),
    # 「その他」タグ自体が既に受け皿のため、中項目は自由記述のみにしておく
    # （細分化したい場合はSubItemMasterシートに行を追加すればよい）
    ("その他", OTHER_SUBITEM_LABEL, 1),
]

MAX_RETRIES = 3
RETRY_WAIT_SEC = 1.5


def ensure_workbook_exists(path: str = EXCEL_PATH) -> None:
    """
    Excelブックが存在しない場合、Entries/TagMasterシートを
    初期タグ付きで新規作成する。
    """
    if os.path.exists(path):
        print(f"📘 既存ブックを検出しました: {path}")
        return

    os.makedirs(os.path.dirname(path), exist_ok=True)

    wb = Workbook()

    ws_entries = wb.active
    ws_entries.title = ENTRIES_SHEET
    ws_entries.append(ENTRIES_HEADER)

    ws_tags = wb.create_sheet(TAGMASTER_SHEET)
    ws_tags.append(TAGMASTER_HEADER)
    for tag_name, color_code in DEFAULT_TAGS:
        ws_tags.append([tag_name, color_code])

    ws_timelog = wb.create_sheet(TIMELOG_SHEET)
    ws_timelog.append(TIMELOG_HEADER)

    ws_subitems = wb.create_sheet(SUBITEM_SHEET)
    ws_subitems.append(SUBITEM_HEADER)
    for tag_name, subitem_name, order in DEFAULT_SUBITEMS:
        ws_subitems.append([tag_name, subitem_name, order])

    ws_actions = wb.create_sheet(ACTIONS_SHEET)
    ws_actions.append(ACTIONS_HEADER)

    ws_hirebi = wb.create_sheet(HIREBI_SHEET)
    ws_hirebi.append(HIREBI_HEADER)

    wb.save(path)
    print(f"✅ 新規ブックを作成しました: {path}")


def _ensure_timelog_sheet(wb) -> None:
    """
    TimeLogシートが無ければ作成する（既存のjournal_data.xlsxには
    自動で追加されないため、読み書き前にこの関数で自己修復する）。
    """
    if TIMELOG_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(TIMELOG_SHEET)
        ws.append(TIMELOG_HEADER)
        print(f"🔧 既存ブックに'{TIMELOG_SHEET}'シートを追加しました。")


def _ensure_timelog_subitem_column(wb) -> None:
    """
    TimeLogシートが旧形式（開始/終了/タグの3列）のままなら、
    中項目列を追加する自己修復を行う。
    """
    ws = wb[TIMELOG_SHEET]
    header = [cell.value for cell in ws[1]]
    if len(header) == 3 and header[2] == "タグ":
        ws.cell(row=1, column=4, value="中項目")
        print(f"🔧 '{TIMELOG_SHEET}'シートに中項目列を追加しました。")


def _ensure_subitem_sheet(wb) -> None:
    """
    SubItemMasterシートが無ければ、既定の中項目付きで作成する
    （既存のjournal_data.xlsxには自動で追加されないため、書き込み前に
    この関数で自己修復する）。
    """
    if SUBITEM_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(SUBITEM_SHEET)
        ws.append(SUBITEM_HEADER)
        for tag_name, subitem_name, order in DEFAULT_SUBITEMS:
            ws.append([tag_name, subitem_name, order])
        print(f"🔧 既存ブックに'{SUBITEM_SHEET}'シートを追加しました。")


def _ensure_actions_sheet(wb) -> None:
    """
    Actionsシートが無ければ作成する（既存のjournal_data.xlsxには
    自動で追加されないため、書き込み前にこの関数で自己修復する）。
    """
    if ACTIONS_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(ACTIONS_SHEET)
        ws.append(ACTIONS_HEADER)
        print(f"🔧 既存ブックに'{ACTIONS_SHEET}'シートを追加しました。")


def _ensure_actions_priority_column(wb) -> None:
    """
    Actionsシートが優先(★)列の無い旧形式（6列まで）のままなら、
    7列目に優先列を追加する自己修復を行う。
    """
    ws = wb[ACTIONS_SHEET]
    header = [cell.value for cell in ws[1]]
    if len(header) < 7:
        ws.cell(row=1, column=7, value="優先")
        print(f"🔧 '{ACTIONS_SHEET}'シートに優先列を追加しました。")


def _ensure_hirebi_sheet(wb) -> None:
    """
    Hirebiシートが無ければ作成する（既存のjournal_data.xlsxには
    自動で追加されないため、書き込み前にこの関数で自己修復する）。
    """
    if HIREBI_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(HIREBI_SHEET)
        ws.append(HIREBI_HEADER)
        print(f"🔧 既存ブックに'{HIREBI_SHEET}'シートを追加しました。")


def _ensure_entries_sheet(wb) -> None:
    """
    Entriesシートが無ければ作成する。
    _ensure_lkpt_columns()はwb[ENTRIES_SHEET]を無条件で参照するため、
    その前にこれを呼ばないとシートが無いブックでKeyErrorになる。
    """
    if ENTRIES_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(ENTRIES_SHEET)
        ws.append(ENTRIES_HEADER)
        print(f"🔧 既存ブックに'{ENTRIES_SHEET}'シートを追加しました。")


def _parse_excel_dt(value):
    """
    openpyxlが返す値（datetime、または"%Y-%m-%d %H:%M"形式の文字列）を
    datetimeに正規化する。解析できない場合はNoneを返す。
    同じイディオムが複数箇所に散らばっていたため共通化した
    （行の同一性を照合する楽観ロックでは、両者を同じ規則で比較する必要がある）。
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    try:
        return datetime.strptime(str(value), "%Y-%m-%d %H:%M")
    except ValueError:
        return None


def _ensure_lkpt_columns(wb) -> None:
    """
    Entriesシートが旧形式（日付/タグ/メモの3列）のままなら、
    メモ列をL列として読み替え、K/P/T列を追加する自己修復を行う。
    既存行の値は列の移動をせずそのまま残るため、既存のメモは
    自動的にLのデータとして統合される。
    """
    ws = wb[ENTRIES_SHEET]
    header = [cell.value for cell in ws[1]]
    if len(header) >= 3 and header[2] == "メモ":
        ws.cell(row=1, column=3, value="L")
        ws.insert_cols(4, amount=3)
        ws.cell(row=1, column=4, value="K")
        ws.cell(row=1, column=5, value="P")
        ws.cell(row=1, column=6, value="T")
        print("🔧 EntriesシートをLKPT形式（L/K/P/T）に移行しました。")


def _load_with_retry(path: str, max_retries: int = MAX_RETRIES,
                      retry_wait: float = RETRY_WAIT_SEC):
    """
    OneDrive同期等によるファイルロックを考慮し、
    リトライ付きでワークブックを読み込む内部関数。
    """
    last_error = None
    for attempt in range(1, max_retries + 1):
        try:
            wb = load_workbook(path)
            return wb
        except PermissionError as e:
            last_error = e
            print(f"⏳ ファイルがロック中です。{retry_wait}秒後に再試行します "
                  f"({attempt}/{max_retries})")
            time.sleep(retry_wait)
    print(f"❌ ファイルを開けませんでした（ロック解除待ちタイムアウト）: {last_error}")
    raise last_error


def _save_with_retry(wb, path: str, max_retries: int = MAX_RETRIES,
                      retry_wait: float = RETRY_WAIT_SEC) -> bool:
    """
    保存時のロック競合に対してリトライを行う内部関数。
    最終的に失敗した場合はローカル一時ファイルへ退避保存する。
    """
    last_error = None
    for attempt in range(1, max_retries + 1):
        try:
            wb.save(path)
            print(f"💾 保存に成功しました: {path}")
            return True
        except PermissionError as e:
            last_error = e
            print(f"⏳ 保存時にロック中です。{retry_wait}秒後に再試行します "
                  f"({attempt}/{max_retries})")
            time.sleep(retry_wait)

    # 最終フォールバック: ローカル退避ファイルに保存し、後で手動マージできるようにする
    fallback_dir = os.path.join(os.path.dirname(path), "_unsynced_fallback")
    os.makedirs(fallback_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    fallback_path = os.path.join(
        fallback_dir, f"fallback_{timestamp}.xlsx"
    )
    try:
        wb.save(fallback_path)
        print(f"⚠️ 本保存に失敗したため、退避ファイルに保存しました: {fallback_path}")
        print("⚠️ 後ほど手動でマージしてください。")
    except Exception as fallback_error:
        print(f"❌ 退避保存にも失敗しました: {fallback_error}")
        raise fallback_error from last_error

    return False


def append_entry(date_str: str, tag: str, memo: str,
                  path: str = EXCEL_PATH) -> bool:
    """
    Entriesシートに1行追記する。memoはL列（列C）にそのまま入る
    （C列は元は「メモ」列だったが、LKPT形式移行後は「L」列として
    読み替えられている）。K/P/T列は空のままになる。

    Args:
        date_str: "YYYY-MM-DD HH:MM" 形式の日時文字列
        tag: タグ名（TagMasterに存在するものを推奨）
        memo: 1行メモ（L列に記録される）
        path: Excelファイルパス

    Returns:
        bool: 保存に成功した場合True
    """
    ensure_workbook_exists(path)
    wb = _load_with_retry(path)

    if ENTRIES_SHEET not in wb.sheetnames:
        print(f"❌ シート'{ENTRIES_SHEET}'が見つかりません。")
        return False

    ws = wb[ENTRIES_SHEET]
    ws.append([date_str, tag, memo])

    success = _save_with_retry(wb, path)
    if success:
        print(f"📝 記録しました → [{date_str}] [{tag}] {memo}")
    return success


def _find_last_checkpoint(ws_timelog, now: datetime):
    """
    本日のTimeLogのうち、終了時刻が最も新しいものを返す（無ければNone）。
    昨日以前の行は対象外のため、日をまたいだ分は引き継がれない。
    """
    today = now.date()
    last_checkpoint = None
    for row in ws_timelog.iter_rows(min_row=2, values_only=True):
        end_dt = _parse_excel_dt(row[1])
        if end_dt is None:
            continue
        if end_dt.date() == today and (last_checkpoint is None or end_dt > last_checkpoint):
            last_checkpoint = end_dt
    return last_checkpoint


def _compute_time_range(ws_timelog, now: datetime) -> tuple:
    """
    【旧仕様・変更しない】本日のTimeLogの直近チェックポイントから、
    今回のチェックインで記録する作業時間の範囲を計算する。

    record_check_in()とpeek_next_time_range()専用。過去の
    daily_journal_*.pyが依存しているため、挙動を変えてはいけない。
    新しい入力経路はappend_time_log()（内部で_compute_time_range_v2）を使う。

    Returns:
        tuple[str, datetime, datetime]: (kind, start, end)
            kind: "anchor"（本日最初のチェックイン。start==end==nowで
                  作業時間としては記録されない）または"interval"
    """
    last_checkpoint = _find_last_checkpoint(ws_timelog, now)
    if last_checkpoint is None:
        return "anchor", now, now

    start = last_checkpoint
    max_gap = timedelta(hours=MAX_TIMELOG_GAP_HOURS)
    if now - start > max_gap:
        start = now - max_gap
    return "interval", start, now


def _work_start_of(now: datetime) -> datetime:
    """nowと同じ日付のWORK_START_TIMEをdatetimeで返す。"""
    hour, minute = (int(part) for part in WORK_START_TIME.split(":"))
    return now.replace(hour=hour, minute=minute, second=0, microsecond=0)


def _compute_time_range_v2(ws_timelog, now: datetime) -> tuple:
    """
    【新仕様】今回のチェックインで記録する作業時間の範囲を計算する。
    _compute_time_range()との違いは2点:

    1. 本日まだ記録が無い場合、勤務開始時刻(WORK_START_TIME)を起点にする。
       これにより朝の分が自動的に埋まる。勤務開始時刻より前に記録した場合
       だけ、従来通り0分の基準点("anchor")として扱う。
    2. 前回から MAX_TIMELOG_GAP_HOURS を超えて空いた場合、超過分を黙って
       捨てず "gap" として返す。呼び出し側(append_time_log)が
       「未記録」の行と実作業の行の2行に分けて記録する。

    Returns:
        tuple[str, datetime, datetime]: (kind, start, end)
            kind: "anchor"  … start==end==now（作業時間としては記録されない）
                  "interval"… start〜endを1行で記録する
                  "gap"     … start〜(end-上限)を「未記録」、残りを実作業として
                              2行に分けて記録する
    """
    last_checkpoint = _find_last_checkpoint(ws_timelog, now)

    if last_checkpoint is None:
        work_start = _work_start_of(now)
        if now <= work_start:
            # 勤務開始時刻より前の記録。埋めるべき朝の時間がまだ無い
            return "anchor", now, now
        start = work_start
    else:
        start = last_checkpoint

    if now - start > timedelta(hours=MAX_TIMELOG_GAP_HOURS):
        return "gap", start, now
    return "interval", start, now


def peek_next_time_range(path: str = EXCEL_PATH, now: datetime = None) -> tuple:
    """
    次にrecord_check_in()を呼んだ場合に記録される作業時間の範囲を、
    実際には書き込まずに事前計算する。タグ選択時にポップアップUIへ
    「何時から何時までの作業として記録されるか」をプレビュー表示するために使う。

    Args:
        path: Excelファイルパス
        now: 基準時刻（省略時はdatetime.now()）

    Returns:
        tuple[str, datetime, datetime]: (kind, start, end)
            kind: "anchor"（本日最初のチェックイン。作業時間としては
                  記録されない）または"interval"（start〜endが作業時間として
                  記録される）
    """
    if now is None:
        now = datetime.now()
    now = now.replace(second=0, microsecond=0)

    ensure_workbook_exists(path)
    wb = _load_with_retry(path)
    if TIMELOG_SHEET not in wb.sheetnames:
        return "anchor", now, now
    return _compute_time_range(wb[TIMELOG_SHEET], now)


def peek_next_time_range_v2(path: str = EXCEL_PATH, now: datetime = None) -> tuple:
    """
    次にappend_time_log()を呼んだ場合に記録される作業時間の範囲を、
    実際には書き込まずに事前計算する。ポップアップUIに「いま押すと何時から
    何時までになるか」を、タグを押す前から常時プレビュー表示するために使う。

    peek_next_time_range()は_compute_time_range()（record_check_in()向けの
    旧仕様。WORK_START_TIME起点や"gap"分割が無い）を使っているため、
    現在の主経路であるappend_time_log()の挙動とはズレる。record_check_in()の
    契約（旧仕様のまま変更しない）は壊さず、append_time_log()と同じ
    _compute_time_range_v2()を使う別関数として追加した。

    Args:
        path: Excelファイルパス
        now: 基準時刻（省略時はdatetime.now()）

    Returns:
        tuple[str, datetime, datetime]: (kind, start, end)
            kind: "anchor" | "interval" | "gap"（append_time_log()と同じ）。
            "gap"の場合、startは上限適用前の生の開始時刻。実際に
            「未記録」と「実作業」の2行に分かれる境界（now - MAX_TIMELOG_GAP_HOURS）
            は、append_time_log()の書き込みロジックと同じ計算を呼び出し側で
            行うこと（表示用の値をここで先読みで確定させない）。
    """
    if now is None:
        now = datetime.now()
    now = now.replace(second=0, microsecond=0)

    ensure_workbook_exists(path)
    wb = _load_with_retry(path)
    if TIMELOG_SHEET not in wb.sheetnames:
        return "anchor", now, now
    return _compute_time_range_v2(wb[TIMELOG_SHEET], now)


def record_check_in(tag: str, sub_item: str, l: str, k: str, p: str, t: str,
                     path: str = EXCEL_PATH, now: datetime = None) -> tuple:
    """
    タイムログのチェックインを記録する。

    本日まだTimeLogに記録が無ければ、これを「基準点」として
    開始=終了=nowの0分行を記録するだけに留める（作業記録は作らない）。
    2回目以降のチェックインでは、本日分の直近の終了時刻から今までを
    1件の作業記録として記録する（経過がMAX_TIMELOG_GAP_HOURSを超える
    場合は打ち切る）。
    l/k/p/tのいずれかが空でなければ、既存のEntriesシートにもLKPTとして
    記録する。TimeLog・Entriesへの追記は1回のload/saveにまとめ、
    部分成功を防ぐ。

    Args:
        tag: タグ名
        sub_item: 中項目名（SubItemMasterの値、または自由記述。空文字列可）
        l: Learned（学び）1行メモ（空文字列可）
        k: Keep（継続すべき事）1行メモ（空文字列可）
        p: Problem（課題）1行メモ（空文字列可）
        t: Try（次に挑戦したい事）1行メモ（空文字列可）
        path: Excelファイルパス
        now: 基準時刻（省略時はdatetime.now()）

    Returns:
        tuple[bool, str, int | None]:
            (保存に成功したか, "anchor"（基準点のみ）または"interval"（作業記録あり）,
             "interval"の場合の経過分数)
    """
    if now is None:
        now = datetime.now()
    now = now.replace(second=0, microsecond=0)

    ensure_workbook_exists(path)
    wb = _load_with_retry(path)
    _ensure_timelog_sheet(wb)
    _ensure_timelog_subitem_column(wb)
    _ensure_lkpt_columns(wb)
    _ensure_subitem_sheet(wb)
    ws_timelog = wb[TIMELOG_SHEET]

    kind, start, end = _compute_time_range(ws_timelog, now)
    if kind == "anchor":
        ws_timelog.append([now, now, tag, sub_item])
        minutes = None
    else:
        ws_timelog.append([start, end, tag, sub_item])
        minutes = int((end - start).total_seconds() // 60)

    has_lkpt = bool(l or k or p or t)
    if has_lkpt:
        if ENTRIES_SHEET not in wb.sheetnames:
            print(f"❌ シート'{ENTRIES_SHEET}'が見つかりません。")
        else:
            ws_entries = wb[ENTRIES_SHEET]
            ws_entries.append([now.strftime("%Y-%m-%d %H:%M"), tag, l, k, p, t])

    success = _save_with_retry(wb, path)
    if success:
        if kind == "anchor":
            print(f"⏱️ 本日の基準点を記録しました → {now.strftime('%Y-%m-%d %H:%M')}")
        else:
            print(
                f"⏱️ 作業記録: [{start.strftime('%H:%M')}-{now.strftime('%H:%M')}] "
                f"{tag} / {sub_item} ({minutes}分)"
            )
        if has_lkpt:
            print(f"📝 LKPTを記録しました → [{now.strftime('%Y-%m-%d %H:%M')}] [{tag}] "
                  f"L={l} K={k} P={p} T={t}")
    return success, kind, minutes


# ============================================================
# 新しい入力経路（TIME / LKPT を分離して記録する）
# ------------------------------------------------------------
# record_check_in()はTimeLogとEntriesを1回の保存でまとめて書くため、
# 「タグを選ばないとLKPTが書けない」という制約と一体になっていた。
# 目的の違う3つの記録（時間・振り返り・タスク）を独立させるため、
# TimeLog専用・Entries専用の関数をそれぞれ用意する。
# record_check_in()は過去のdaily_journal_*.pyが依存しているため残す。
# ============================================================


def append_time_log(tag: str, sub_item: str, path: str = EXCEL_PATH,
                     now: datetime = None) -> tuple:
    """
    TimeLogにのみ作業時間を記録する（Entriesには一切触れない）。

    前回チェックポイント（無ければ勤務開始時刻）から今までを1行として書く。
    上限を超えて空いていた場合は、超過分を「未記録」タグの行として先に書き、
    直近の上限時間ぶんだけを実作業として書く（合計2行）。

    Args:
        tag: タグ名
        sub_item: 中項目名（空文字列可。後からupdate_time_log_row()で補える）
        path: Excelファイルパス
        now: 基準時刻（省略時はdatetime.now()）

    Returns:
        tuple[bool, str, int | None, int | None]:
            (保存に成功したか,
             "anchor" | "interval" | "gap",
             実作業として記録した分数（anchorの場合はNone）,
             実作業を書き込んだExcelの行番号（失敗時はNone）)

        行番号は、直後の押し間違いを追記ではなく上書きで直すために返す
        （update_time_log_row() / delete_time_log_row() に渡す）。
    """
    if now is None:
        now = datetime.now()
    now = now.replace(second=0, microsecond=0)

    ensure_workbook_exists(path)
    wb = _load_with_retry(path)
    _ensure_timelog_sheet(wb)
    _ensure_timelog_subitem_column(wb)
    _ensure_subitem_sheet(wb)
    ws = wb[TIMELOG_SHEET]

    kind, start, end = _compute_time_range_v2(ws, now)

    if kind == "anchor":
        ws.append([now, now, tag, sub_item])
        minutes = None
    else:
        if kind == "gap":
            # 上限を超えた分を「未記録」として先に残し、どこが埋まって
            # いないのかを後から見えるようにする
            cutoff = now - timedelta(hours=MAX_TIMELOG_GAP_HOURS)
            ws.append([start, cutoff, UNRECORDED_TAG, ""])
            start = cutoff
        ws.append([start, end, tag, sub_item])
        minutes = int((end - start).total_seconds() // 60)

    row_idx = ws.max_row
    success = _save_with_retry(wb, path)
    if not success:
        return False, kind, minutes, None

    if kind == "anchor":
        print(f"⏱️ 本日の基準点を記録しました → {now.strftime('%Y-%m-%d %H:%M')}")
    else:
        if kind == "gap":
            print(f"⚠️ {MAX_TIMELOG_GAP_HOURS}時間以上空いていたため、超過分を"
                  f"「{UNRECORDED_TAG}」として記録しました。")
        print(
            f"⏱️ 作業記録: [{start.strftime('%H:%M')}-{end.strftime('%H:%M')}] "
            f"{tag} / {sub_item} ({minutes}分)"
        )
    return True, kind, minutes, row_idx


def _timelog_row_matches(ws, row: int, expect_start, expect_end) -> bool:
    """
    指定行の開始・終了が期待値と一致するかを確認する（楽観ロック）。
    行番号を識別子として持ち回る間にユーザーがExcelで行を挿入・削除すると、
    同じ行番号が別の記録を指してしまうため、書き換える前に必ず照合する。
    期待値がNoneの場合は照合を省略する。
    """
    if row < 2 or row > ws.max_row:
        return False
    if expect_start is None and expect_end is None:
        return True
    actual_start = _parse_excel_dt(ws.cell(row=row, column=1).value)
    actual_end = _parse_excel_dt(ws.cell(row=row, column=2).value)
    if expect_start is not None and actual_start != expect_start:
        return False
    if expect_end is not None and actual_end != expect_end:
        return False
    return True


def update_time_log_row(row: int, tag: str, sub_item: str,
                         expect_start: datetime = None, expect_end: datetime = None,
                         path: str = EXCEL_PATH) -> bool:
    """
    既存のTimeLog行のタグ（3列目）と中項目（4列目）だけを上書きする。
    開始・終了は動かさない（押し間違いの訂正であって、時間の訂正ではないため）。

    expect_start / expect_end を渡すと、その行の実際の開始・終了が一致した
    場合だけ書き換える。一致しない場合は何も書き換えずFalseを返すので、
    呼び出し側は「行番号が当てにならなくなった」と判断して追記に切り替えられる。

    Returns:
        bool: 上書きして保存できた場合True
    """
    ensure_workbook_exists(path)
    wb = _load_with_retry(path)
    if TIMELOG_SHEET not in wb.sheetnames:
        print(f"❌ シート'{TIMELOG_SHEET}'が見つかりません。")
        return False
    _ensure_timelog_subitem_column(wb)
    ws = wb[TIMELOG_SHEET]

    if not _timelog_row_matches(ws, row, expect_start, expect_end):
        print(f"⚠️ TimeLogの行{row}が想定と異なるため、上書きを中止しました。")
        return False

    ws.cell(row=row, column=3, value=tag)
    ws.cell(row=row, column=4, value=sub_item)

    success = _save_with_retry(wb, path)
    if success:
        print(f"✏️ 直前の作業記録を修正しました（行{row}）→ {tag} / {sub_item}")
    return success


def delete_time_log_row(row: int, expect_start: datetime = None,
                         expect_end: datetime = None,
                         path: str = EXCEL_PATH) -> bool:
    """
    指定したTimeLog行を削除する（直前の記録の取り消し用）。
    update_time_log_row()と同じ楽観ロックで照合してから消す。

    注意: 行を削除するとそれ以降の行番号がすべて繰り上がる。呼び出し側は
    削除に成功したら、保持している行番号を必ず破棄すること
    （Actionsシートが「行番号＝ID」のために行削除を禁じているのとは対照的に、
    TimeLogはこの関数でのみ削除を許す）。

    Returns:
        bool: 削除して保存できた場合True
    """
    ensure_workbook_exists(path)
    wb = _load_with_retry(path)
    if TIMELOG_SHEET not in wb.sheetnames:
        print(f"❌ シート'{TIMELOG_SHEET}'が見つかりません。")
        return False
    ws = wb[TIMELOG_SHEET]

    if not _timelog_row_matches(ws, row, expect_start, expect_end):
        print(f"⚠️ TimeLogの行{row}が想定と異なるため、削除を中止しました。")
        return False

    ws.delete_rows(row)
    success = _save_with_retry(wb, path)
    if success:
        print(f"🗑️ 直前の作業記録を取り消しました（行{row}）")
    return success


def append_lkpt_entry(tag: str, l: str, k: str, p: str, t: str,
                       path: str = EXCEL_PATH, now: datetime = None) -> tuple:
    """
    LKPT（振り返り）をEntriesシートにのみ1件記録する。
    TimeLogには一切触れないため、タグを選ばずに振り返りだけを書ける。

    Args:
        tag: 関連タグ（空文字列可）
        l, k, p, t: 各欄の内容。すべて空の場合は何も書かずに(False, None)を返す
        path: Excelファイルパス
        now: 記録時刻（省略時はdatetime.now()）

    Returns:
        tuple[bool, datetime | None]: (保存に成功したか, 記録した時刻)
    """
    l, k, p, t = (v.strip() if v else "" for v in (l, k, p, t))
    if not (l or k or p or t):
        return False, None

    if now is None:
        now = datetime.now()
    now = now.replace(second=0, microsecond=0)

    ensure_workbook_exists(path)
    wb = _load_with_retry(path)
    # 旧形式（日付/タグ/メモの3列）からの移行は、これまでrecord_check_in()
    # からしか呼ばれていなかった。新しい経路でも必ず通す必要がある
    _ensure_entries_sheet(wb)
    _ensure_lkpt_columns(wb)
    ws = wb[ENTRIES_SHEET]
    ws.append([now.strftime("%Y-%m-%d %H:%M"), tag, l, k, p, t])

    success = _save_with_retry(wb, path)
    if not success:
        return False, None

    print(f"📝 LKPTを記録しました → [{now.strftime('%Y-%m-%d %H:%M')}] [{tag}] "
          f"L={l} K={k} P={p} T={t}")
    return True, now


def append_hirebi(tag: str, emotion: str, path: str = EXCEL_PATH,
                   now: datetime = None) -> tuple:
    """
    「魂のひれぶり」を1件記録する（喜・怒・哀・楽のいずれか1タップ）。
    人生の羅針盤（2025年末）が終端指標に置いた「墓場に持っていけるのは
    人生の思い出。その思い出とは、喜怒哀楽の総量」を、日々のチェックインの
    中で実測するための最小の計器。テキスト入力を要らないので、
    LKPTを書かない日でもタップだけで残せる。

    Args:
        tag: 関連タグ（空文字列可）
        emotion: EMOTION_LABELSのいずれか
        path: Excelファイルパス
        now: 記録時刻（省略時はdatetime.now()）

    Returns:
        tuple[bool, datetime | None]: (保存に成功したか, 記録した時刻)
    """
    if emotion not in EMOTION_LABELS:
        print(f"❌ 未知の感情ラベルです: {emotion}")
        return False, None

    if now is None:
        now = datetime.now()
    now = now.replace(second=0, microsecond=0)

    ensure_workbook_exists(path)
    wb = _load_with_retry(path)
    _ensure_hirebi_sheet(wb)
    ws = wb[HIREBI_SHEET]
    ws.append([now.strftime("%Y-%m-%d %H:%M"), tag, emotion])

    success = _save_with_retry(wb, path)
    if not success:
        return False, None

    print(f"🎭 ひれぶりを記録しました → [{now.strftime('%Y-%m-%d %H:%M')}] [{tag}] {emotion}")
    return True, now


def undo_last_entry(path: str = EXCEL_PATH) -> bool:
    """
    Entriesシートの最終行（直近の記録1件）を削除する。

    Returns:
        bool: 削除に成功した場合True。削除対象が無い場合False。
    """
    ensure_workbook_exists(path)
    wb = _load_with_retry(path)
    ws = wb[ENTRIES_SHEET]

    if ws.max_row <= 1:
        print("ℹ️ 取り消せる記録がありません。")
        return False

    ws.delete_rows(ws.max_row)
    success = _save_with_retry(wb, path)
    if success:
        print("↩️ 直近の記録を取り消しました。")
    return success


def get_tag_master(path: str = EXCEL_PATH) -> list:
    """
    TagMasterシートからタグ一覧を取得する。

    Returns:
        list[tuple[str, str]]: (タグ名, 色コード) のリスト
    """
    ensure_workbook_exists(path)
    wb = _load_with_retry(path)
    ws = wb[TAGMASTER_SHEET]

    tags = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            tags.append((row[0], row[1]))

    print(f"🏷️ タグ一覧を取得しました（{len(tags)}件）")
    return tags


def get_sub_item_master(path: str = EXCEL_PATH) -> dict:
    """
    SubItemMasterシートから、タグ別の中項目一覧を取得する（表示順でソート済み）。

    読み取り専用の関数のため、SubItemMasterシートが無い旧形式のブックでも
    書き込みは行わず、既定値をそのまま返すだけに留める（実際の自己修復・
    永続化はrecord_check_in()が呼ばれた時に行われる）。

    Returns:
        dict[str, list[str]]: タグ名 -> 中項目名のリスト
    """
    ensure_workbook_exists(path)
    wb = _load_with_retry(path)

    if SUBITEM_SHEET not in wb.sheetnames:
        rows = DEFAULT_SUBITEMS
    else:
        ws = wb[SUBITEM_SHEET]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                order = row[2] if len(row) > 2 and row[2] is not None else 0
                rows.append((row[0], row[1], order))
        rows.sort(key=lambda r: (r[0], r[2]))

    result = {}
    for tag_name, subitem_name, _ in rows:
        result.setdefault(tag_name, []).append(subitem_name)

    print(f"📋 中項目マスタを取得しました（{len(rows)}件）")
    return result


def get_last_other_comment(tag: str, path: str = EXCEL_PATH) -> str:
    """
    指定タグで直近に使われた「Other」（中項目の自由記述）の内容を返す。
    無ければ空文字列。

    「同じ作業を続けているなら書き換えなくていい、変えたいときだけ書き換える」
    という運用にするため、ポップアップで「Other」を選んだ時にこの値を
    初期表示するのに使う（読み取り専用。書き込みは行わない）。

    TimeLogシートを新しい行から遡り、そのタグのプリセット中項目（"Other"
    というラベル自体を含む）のどれにも一致しない値を「実際に自由記述された
    内容だった」とみなす。自由記述が一度も無ければ空文字列を返す。

    Args:
        tag: 対象タグ名
        path: Excelファイルパス

    Returns:
        str: 直近の自由記述内容（無ければ空文字列）
    """
    ensure_workbook_exists(path)
    wb = _load_with_retry(path)
    if TIMELOG_SHEET not in wb.sheetnames:
        return ""

    preset_labels = set(get_sub_item_master(path).get(tag, []))

    ws = wb[TIMELOG_SHEET]
    for row in reversed(list(ws.iter_rows(min_row=2, values_only=True))):
        row_tag = row[2] if len(row) > 2 else None
        sub_item = row[3] if len(row) > 3 else None
        if row_tag != tag or not sub_item:
            continue
        if sub_item in preset_labels:
            continue
        return sub_item
    return ""


def add_action(content: str, tag: str = "", origin: str = "manual",
                path: str = EXCEL_PATH, now: datetime = None) -> bool:
    """
    アクションアイテムを1件追加する（ステータスは常に「未着手」で開始する）。

    Args:
        content: アクション内容（空文字列・空白のみの場合は何もせずFalseを返す）
        tag: 関連タグ（空文字列可。ポップアップでタグ未選択のまま追加した場合など）
        origin: "manual"（ポップアップの＋アクション欄から）または
                "P"（LKPTのP欄からのチェックによる自動作成）
        path: Excelファイルパス
        now: 作成日時（省略時はdatetime.now()）

    Returns:
        bool: 保存に成功した場合True
    """
    content = content.strip()
    if not content:
        return False
    if now is None:
        now = datetime.now()
    now = now.replace(second=0, microsecond=0)

    ensure_workbook_exists(path)
    wb = _load_with_retry(path)
    _ensure_actions_sheet(wb)
    _ensure_actions_priority_column(wb)
    ws = wb[ACTIONS_SHEET]
    ws.append([
        now.strftime("%Y-%m-%d %H:%M"), tag, content, origin,
        ACTION_STATUS_PENDING, "", False,
    ])

    success = _save_with_retry(wb, path)
    if success:
        print(f"📌 アクションを追加しました → {content}")
    return success


def get_actions(path: str = EXCEL_PATH) -> list:
    """
    Actionsシートの全アクションを取得する。

    Returns:
        list[dict]: 各要素は{"row", "created_at", "tag", "content", "origin",
        "status", "completed_at", "starred"}。"row"はcomplete_action()・
        set_action_priority()に渡すExcelの実行番号（IDを別列で持たず、
        行番号をそのまま識別子として使う。このシートは行の削除・並べ替えを
        行わない前提のため安定する）。"starred"は優先(★)列が無い旧形式の
        行ではFalse扱いになる
    """
    ensure_workbook_exists(path)
    wb = _load_with_retry(path)
    if ACTIONS_SHEET not in wb.sheetnames:
        return []
    ws = wb[ACTIONS_SHEET]

    actions = []
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True), start=2
    ):
        created_value = row[0]
        if not created_value:
            continue
        try:
            created_at = (
                created_value if isinstance(created_value, datetime)
                else datetime.strptime(str(created_value), "%Y-%m-%d %H:%M")
            )
        except ValueError:
            print(f"⚠️ 日時の解析に失敗した行をスキップしました: {created_value}")
            continue

        completed_value = row[5] if len(row) > 5 else None
        completed_at = None
        if completed_value:
            try:
                completed_at = (
                    completed_value if isinstance(completed_value, datetime)
                    else datetime.strptime(str(completed_value), "%Y-%m-%d %H:%M")
                )
            except ValueError:
                completed_at = None

        actions.append({
            "row": row_idx,
            "created_at": created_at,
            "tag": row[1] or "",
            "content": row[2] or "",
            "origin": row[3] or "",
            "status": row[4] or ACTION_STATUS_PENDING,
            "completed_at": completed_at,
            "starred": bool(row[6]) if len(row) > 6 and row[6] else False,
        })

    print(f"📌 アクション一覧を取得しました（{len(actions)}件）")
    return actions


def complete_action(row: int, path: str = EXCEL_PATH, now: datetime = None) -> bool:
    """
    指定した行のアクションを完了にする。

    Args:
        row: get_actions()が返す各要素の"row"（Excelの実行番号）
        path: Excelファイルパス
        now: 完了日時（省略時はdatetime.now()）

    Returns:
        bool: 保存に成功した場合True
    """
    if now is None:
        now = datetime.now()
    now = now.replace(second=0, microsecond=0)

    ensure_workbook_exists(path)
    wb = _load_with_retry(path)
    if ACTIONS_SHEET not in wb.sheetnames:
        print(f"❌ シート'{ACTIONS_SHEET}'が見つかりません。")
        return False
    ws = wb[ACTIONS_SHEET]

    ws.cell(row=row, column=5, value=ACTION_STATUS_DONE)
    ws.cell(row=row, column=6, value=now.strftime("%Y-%m-%d %H:%M"))

    success = _save_with_retry(wb, path)
    if success:
        print(f"✅ アクションを完了にしました（行{row}）")
    return success


def set_action_priority(row: int, starred: bool, path: str = EXCEL_PATH) -> bool:
    """
    指定した行のアクションの優先フラグ（★）を設定する。
    ポップアップのアクション一覧で★をクリックした時に呼ばれ、一覧の並び順
    （★付きを先頭に表示）を左右する。

    Args:
        row: get_actions()が返す各要素の"row"（Excelの実行番号）
        starred: True=優先（★）にする、False=解除する
        path: Excelファイルパス

    Returns:
        bool: 保存に成功した場合True
    """
    ensure_workbook_exists(path)
    wb = _load_with_retry(path)
    if ACTIONS_SHEET not in wb.sheetnames:
        print(f"❌ シート'{ACTIONS_SHEET}'が見つかりません。")
        return False
    _ensure_actions_priority_column(wb)
    ws = wb[ACTIONS_SHEET]
    ws.cell(row=row, column=7, value=starred)

    success = _save_with_retry(wb, path)
    if success:
        mark = "★" if starred else "☆"
        print(f"{mark} アクションの優先度を更新しました（行{row}）")
    return success


def add_tag(tag_name: str, color_code: str, path: str = EXCEL_PATH) -> bool:
    """
    TagMasterシートに新規タグを追加する。既存の場合は追加しない。
    """
    ensure_workbook_exists(path)
    wb = _load_with_retry(path)
    ws = wb[TAGMASTER_SHEET]

    existing = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
    if tag_name in existing:
        print(f"⚠️ タグ'{tag_name}'は既に存在します。追加をスキップしました。")
        return False

    ws.append([tag_name, color_code])
    success = _save_with_retry(wb, path)
    if success:
        print(f"➕ タグ'{tag_name}'を追加しました。")
    return success


def remove_tag(tag_name: str, path: str = EXCEL_PATH) -> bool:
    """
    TagMasterシートから指定タグを削除する。
    """
    ensure_workbook_exists(path)
    wb = _load_with_retry(path)
    ws = wb[TAGMASTER_SHEET]

    target_row = None
    for idx, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True), start=2
    ):
        if row[0] == tag_name:
            target_row = idx
            break

    if target_row is None:
        print(f"⚠️ タグ'{tag_name}'が見つかりませんでした。")
        return False

    ws.delete_rows(target_row)
    success = _save_with_retry(wb, path)
    if success:
        print(f"➖ タグ'{tag_name}'を削除しました。")
    return success


if __name__ == "__main__":
    print("🔧 storage.py 単体動作確認を開始します。")
    ensure_workbook_exists()
    tags = get_tag_master()
    for name, color in tags:
        print(f"  ・{name} ({color})")
    print("✅ 単体動作確認が完了しました。")
