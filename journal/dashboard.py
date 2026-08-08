# -*- coding: utf-8 -*-
"""
dashboard.py
LKPT Dashboard - 日次〜年次の集計ダッシュボード
Version: 0.18.0
"""

import json
import math
import os
import queue
import threading
import tkinter as tk
import urllib.request
from tkinter import font as tkfont
from datetime import datetime, timedelta

from openpyxl import load_workbook

from storage import (
    EXCEL_PATH, ENTRIES_SHEET, TAGMASTER_SHEET, TIMELOG_SHEET,
    ACTION_STATUS_DONE, ACTION_STATUS_PENDING, UNRECORDED_TAG,
    complete_action, get_actions,
)

# COACHビュー（LKPTの期間要約）用。追加ライブラリを増やさない方針のため、
# SDKは使わずurllib（標準ライブラリ）でGemini APIのREST版を直接呼ぶ。
GEMINI_MODEL = "gemini-2.5-flash"
GEMINI_API_URL = (
    f"https://generativelanguage.googleapis.com/v1beta/models/"
    f"{GEMINI_MODEL}:generateContent"
)

# ============================================================
# UI設定（ダークテーマ）
# ------------------------------------------------------------
# タグ4色とACCENT_COLORは入力ポップアップ側と共通の確定配色。
# カード面・罫線・本文色は、一覧を「ログの羅列」から脱却させるために
# 追加した中間色（背景#1a1a2eから少しずつ持ち上げた無彩色寄りの階調）。
# ============================================================
BG_COLOR = "#1a1a2e"
ACCENT_COLOR = "#f4a6b8"
TEXT_COLOR = "#ffffff"
BUTTON_TEXT_COLOR = "#2b2b40"
PERIOD_BTN_BG = "#dfeaf5"

CARD_BG = "#232338"        # 一覧のカード面
TRACK_BG = "#2b2b42"       # 作業時間バーの下地
LINE_COLOR = "#383850"     # 区切り線・タイムラインの軸
DIM_TEXT = "#a8a6bd"       # 時刻など補助情報
BODY_TEXT = "#ded9ec"      # 一覧の本文
BADGE_BG = "#454360"       # L/K/P/Tバッジの地色
BADGE_FG = "#e8e6f2"

PERIODS = ["D", "W", "M", "Q", "Y"]
MODES = ["LKPT", "TIME", "ACTION"]
# HISTORYの並べ方。CARD=記録単位、FLOW=時間軸、GROUP=カテゴリ軸、
# COACH=Gemini APIによるLKPT要約・アドバイスで並べる。
HISTORY_VIEWS = ["CARD", "FLOW", "GROUP", "COACH"]

# (バッジ文字, entry辞書のキー, グループ表示時の見出し)
LKPT_FIELDS = [
    ("L", "l", "LEARNED"),
    ("K", "k", "KEEP"),
    ("P", "p", "PROBLEM"),
    ("T", "t", "TRY"),
]

# LKPTモードの棒グラフはタグではなくL/K/P/Tで分類するため、タグ色とは
# 別の固定配色を使う（タグ4色と混同しないよう明確に離す）
LKPT_FIELD_COLORS = {
    "L": "#4fc3f7",
    "K": "#66bb6a",
    "P": "#ef5350",
    "T": "#ba68c8",
}

# ACTIONモードの棒グラフ・チェック印の色。PROBLEM(P)と同じ赤系=未対応、
# KEEP(K)と同じ緑系=解消済み、という意味の重なりをそのまま踏襲する
ACTION_STATUS_COLORS = {
    ACTION_STATUS_PENDING: "#ef5350",
    ACTION_STATUS_DONE: "#66bb6a",
}

# COACHビューの見出し（build_coach_prompt()が指示する4見出しと対応）に
# 付けるアイコン。見出し文字列そのものは表示上「## 」を取り除いて使う。
COACH_HEADING_ICONS = {
    "要約": "📋",
    "パターンと停滞ポイント": "🔍",
    "コーチングアドバイス": "💬",
    "次の一歩": "✅",
}

# 円グラフで扇の内側にパーセントを描く最小構成比（これ未満は外側に描く）
PIE_INSIDE_LABEL_MIN_PCT = 8.0

# グラフ領域の高さ。LKPTモードは棒グラフ1つだけなので低くし、
# その分をHISTORY（読むことが目的の領域）に回す。
CHART_HEIGHT_LKPT = 180
CHART_HEIGHT_TIME = 250


def _format_duration(minutes: int) -> str:
    """分数を"1h05m"形式の文字列に整形する。"""
    hours, mins = divmod(minutes, 60)
    if hours:
        return f"{hours}h{mins:02d}m"
    return f"{mins}m"


def _blend_color(hex_color: str, target_hex: str, factor: float) -> str:
    """hex_colorをtarget_hexの方向にfactor(0~1)だけ混ぜた色を返す。
    タグ絞り込み中、非選択タグの棒を背景色に寄せて目立たなくするのに使う。"""
    c1 = tuple(int(hex_color[i:i + 2], 16) for i in (1, 3, 5))
    c2 = tuple(int(target_hex[i:i + 2], 16) for i in (1, 3, 5))
    blended = tuple(int(c1[j] + (c2[j] - c1[j]) * factor) for j in range(3))
    return "#{:02x}{:02x}{:02x}".format(*blended)


def load_entries(path: str = EXCEL_PATH) -> list:
    """
    Entriesシートから全記録を読み込み、日時付き辞書のリストとして返す。
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[ENTRIES_SHEET]

    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_value, tag = row[0], row[1]
        if not date_value:
            continue
        try:
            if isinstance(date_value, datetime):
                dt = date_value
            else:
                dt = datetime.strptime(str(date_value), "%Y-%m-%d %H:%M")
        except ValueError:
            print(f"⚠️ 日時の解析に失敗した行をスキップしました: {date_value}")
            continue
        # 旧形式(メモ1列)のファイルをLKPT移行前に開いても壊れないよう、
        # 存在しない列は空文字列として扱う
        l = row[2] if len(row) > 2 and row[2] else ""
        k = row[3] if len(row) > 3 and row[3] else ""
        p = row[4] if len(row) > 4 and row[4] else ""
        t = row[5] if len(row) > 5 and row[5] else ""
        # LKPTはタグ無しでも記録できるようになったため、必ず正規化する。
        # openpyxlは空セルをNoneで返すので、ここで潰さないとf文字列に
        # 文字列"None"が混入する（特にCOACHのプロンプト）
        entries.append({
            "datetime": dt, "tag": tag or "", "l": l, "k": k, "p": p, "t": t,
        })

    print(f"📊 記録を読み込みました（{len(entries)}件）")
    return entries


def load_tag_colors(path: str = EXCEL_PATH) -> dict:
    """
    TagMasterシートからタグ名→色コードの辞書を取得する。
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[TAGMASTER_SHEET]

    colors = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            colors[row[0]] = row[1] or "#888888"
    # 「未記録」は、上限を超えて空いた時間に自動で付く印であってユーザーが
    # 選ぶタグではないため、TagMasterには登録していない（＝ポップアップの
    # 選択肢に出ない）。その代わり、ここで灰色を与えて棒グラフ・円グラフ・
    # チップに「埋まっていない時間」として表示できるようにする
    colors.setdefault(UNRECORDED_TAG, "#5a5a72")
    return colors


def load_time_log_entries(path: str = EXCEL_PATH) -> list:
    """
    TimeLogシートから全記録を読み込み、開始・終了・タグ付きの辞書のリストとして返す。
    "datetime"キーは終了時刻のエイリアスで、filter_entries_by_period()を
    そのまま再利用できるようにするためのもの。
    """
    wb = load_workbook(path, data_only=True)
    if TIMELOG_SHEET not in wb.sheetnames:
        return []
    ws = wb[TIMELOG_SHEET]

    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        start_value, end_value, tag = row[0], row[1], row[2]
        # 中項目（4列目）。これまで書き込むだけで一度も読んでいなかったため、
        # 「R19の中で何をしていたか」が画面から確認できなかった
        sub_item = (row[3] if len(row) > 3 else "") or ""
        if not start_value or not end_value:
            continue
        try:
            start_dt = (
                start_value if isinstance(start_value, datetime)
                else datetime.strptime(str(start_value), "%Y-%m-%d %H:%M")
            )
            end_dt = (
                end_value if isinstance(end_value, datetime)
                else datetime.strptime(str(end_value), "%Y-%m-%d %H:%M")
            )
        except ValueError:
            print(f"⚠️ 日時の解析に失敗した行をスキップしました: {start_value} - {end_value}")
            continue
        entries.append({
            "start": start_dt, "end": end_dt, "tag": tag or "",
            "sub_item": sub_item, "datetime": end_dt,
        })

    print(f"⏱️ 作業記録を読み込みました（{len(entries)}件）")
    return entries


def _period_range(period: str, reference: datetime = None) -> tuple:
    """
    指定された期間種別（D/W/M/Q/Y）に対応する開始日時・終了日時
    （終了は含まない）を返す。
    """
    reference = reference or datetime.now()
    today = reference.replace(hour=0, minute=0, second=0, microsecond=0)

    if period == "D":
        start = today
        end = today + timedelta(days=1)

    elif period == "W":
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=7)

    elif period == "M":
        start = today.replace(day=1)
        if start.month == 12:
            end = start.replace(year=start.year + 1, month=1)
        else:
            end = start.replace(month=start.month + 1)

    elif period == "Q":
        quarter_start_month = ((today.month - 1) // 3) * 3 + 1
        start = today.replace(month=quarter_start_month, day=1)
        end_month = quarter_start_month + 3
        if end_month > 12:
            end = start.replace(year=start.year + 1, month=end_month - 12)
        else:
            end = start.replace(month=end_month)

    elif period == "Y":
        start = today.replace(month=1, day=1)
        end = start.replace(year=start.year + 1)

    else:
        raise ValueError(f"不明な期間種別です: {period}")

    return start, end


def format_period_label(period: str, reference: datetime = None) -> str:
    """
    期間種別に応じて、実際にどの範囲を見ているかを表す文字列を返す。
    （D/W/M/Q/Yという略号だけでは範囲が分からないため、サマリー行に添える）
    """
    start, end = _period_range(period, reference)
    last = end - timedelta(days=1)

    if period == "D":
        return start.strftime("%Y-%m-%d")
    if period == "W":
        return f"{start.strftime('%Y-%m-%d')} – {last.strftime('%m-%d')}"
    if period == "M":
        return start.strftime("%Y-%m")
    if period == "Q":
        return f"{start.year} Q{(start.month - 1) // 3 + 1}"
    return str(start.year)


def filter_entries_by_period(entries: list, period: str,
                              reference: datetime = None) -> list:
    """
    指定期間種別に該当する記録のみを抽出する。
    """
    start, end = _period_range(period, reference)
    return [e for e in entries if start <= e["datetime"] < end]


def build_coach_prompt(entries: list, period_label: str) -> str:
    """
    COACHビュー用のプロンプトを組み立てる。期間内のLKPT記録を時系列で
    列挙したうえで、以下の方向でブラッシュアップしたプロンプトを渡す。
    - 出力フォーマットを見出し単位で固定する（生成のたびに体裁がぶれないため）
    - 記録間のパターン（繰り返す課題、放置されたP等）を明示的に探させる
    - 一般論・励まし系の紋切り型を禁止し、記録の具体的な引用を必須にする
    - コーチングスタイルは「PDCA語彙で状況を診断しつつ、率直に踏み込む」
      （小さく早くPDCAを回して成長したいマネージャー向けに選定した組み合わせ）
    - 末尾に、今週着手できる具体的な次の一歩を必須にする
    """
    lines = []
    for e in entries:
        stamp = e["datetime"].strftime("%m-%d %H:%M")
        body_lines = [
            f"  {letter}: {e[key]}" for letter, key, _ in LKPT_FIELDS if e[key]
        ]
        if body_lines:
            # タグ無しのLKPTがあり得るため、末尾の余白を落として
            # 「[08-07 10:07]」だけの行にする。「（タグなし）」のような
            # 代替語は入れない（AIが実在のタグ名だと誤解するため）
            lines.append(f"[{stamp}] {e['tag']}".rstrip())
            lines.extend(body_lines)
    records_text = "\n".join(lines) if lines else "（この期間に記録された内容はありません）"

    return (
        "あなたはシニアエグゼクティブコーチングアドバイザーです。\n"
        "コーチングを受けるのは「小さく早くPDCAを回して成長したいマネージャー」です。"
        "PDCA（Plan/Do/Check/Act）の語彙で状況を診断しつつ、遠慮せず率直に踏み込んで"
        "ください。慰めや一般論は不要です。\n\n"
        f"以下は、ユーザーが{period_label}の期間に記録したLKPT"
        "（L=学び, K=継続すべきこと, P=課題, T=次に挑戦したいこと）の記録です。\n\n"
        f"{records_text}\n\n"
        "次の4つの見出しで、日本語の文章を作成してください（見出しはそのまま使うこと）。\n\n"
        "## 要約\n"
        "この期間の記録内容を事実ベースで簡潔にまとめる（箇条書き可）。\n\n"
        "## パターンと停滞ポイント\n"
        "同じ課題（P）が形を変えて繰り返し記録されていないか、Pに対応するT"
        "（次のアクション）が無いまま放置されていないかなど、記録間のつながりから"
        "見えるPDCAの停滞箇所を具体的に指摘する。無ければ「停滞は見られない」と"
        "明記する。\n\n"
        "## コーチングアドバイス\n"
        "シニアエグゼクティブコーチとして、率直に踏み込んで助言する。一般論・精神論・"
        "励ましの決まり文句は禁止。必ず記録内容を最低1つ具体的に引用しながら、"
        "PDCAのどこが止まっているか、次にどう考えるべきかを指摘する。\n\n"
        "## 次の一歩\n"
        "今週中に着手できる、小さく具体的なアクションを1〜2個、箇条書きで挙げる。"
        "「頑張る」「意識する」のような曖昧な行動ではなく、今日〜明日に着手できる"
        "粒度にする。\n\n"
        "全体で600字程度を目安にしてください。"
    )


def call_gemini_summary(prompt: str) -> str:
    """
    Gemini API（REST）を呼び出して要約テキストを取得する。
    GEMINI_API_KEY環境変数が必須。ネットワーク呼び出しはブロッキングするため、
    呼び出し側は必ず別スレッドから呼ぶこと（Tkinterのメインループを止めないため）。
    """
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        raise RuntimeError("環境変数 GEMINI_API_KEY が設定されていません。")

    body = json.dumps({"contents": [{"parts": [{"text": prompt}]}]}).encode("utf-8")
    req = urllib.request.Request(
        f"{GEMINI_API_URL}?key={api_key}",
        data=body,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=45) as res:
        payload = json.loads(res.read().decode("utf-8"))

    try:
        return payload["candidates"][0]["content"]["parts"][0]["text"].strip()
    except (KeyError, IndexError) as e:
        raise RuntimeError(f"Gemini APIの応答形式が想定と異なります: {payload}") from e


def aggregate_lkpt_field_counts(entries: list) -> dict:
    """
    LKPTモードの棒グラフ用。タグ別ではなく、L/K/P/Tそれぞれに値が
    入っている件数を集計する（L→K→P→Tの順で固定、0件の項目は除く）。
    """
    counts = {}
    for letter, key, _ in LKPT_FIELDS:
        n = sum(1 for e in entries if e[key])
        if n:
            counts[letter] = n
    return counts


def aggregate_time_by_tag(entries: list) -> dict:
    """
    タグ別の作業時間合計（分）を集計する。所要時間0分の基準点マーカー行は
    合計に寄与しないため、特別なフィルタは不要。
    """
    totals = {}
    for e in entries:
        minutes = int((e["end"] - e["start"]).total_seconds() // 60)
        if minutes <= 0:
            continue
        totals[e["tag"]] = totals.get(e["tag"], 0) + minutes
    return totals


class DashboardWindow:
    """LKPT／作業時間の集計ダッシュボードウィンドウ。"""

    def __init__(self, root: tk.Tk, initial_mode: str = "LKPT"):
        self.root = root
        self.current_period = "D"
        self.current_mode = initial_mode
        self.current_view = "CARD"
        self.reference_date = datetime.now()
        # 棒グラフクリックによる絞り込み。LKPTモードではL/K/P/Tの文字、
        # TIMEモードではタグ名が入る（分類軸がモードで異なるため）
        self.chart_filter = None
        self.tag_colors = {}
        self.entries = []
        self.time_entries = []
        self.actions = []
        self._history_rows = []
        self.period_buttons = {}
        self.mode_buttons = {}
        self.view_buttons = {}

        # COACHビュー（Gemini要約）のキャッシュ。(period, start, end)をキーにし、
        # 同じ期間を見返すたびにAPIを叩き直さないようにする。
        self._coach_cache = {}
        self._coach_inflight_key = None
        self._coach_request_id = 0
        # Tkinterはメインスレッド以外からのウィジェット操作が禁止されているため、
        # ワーカースレッドからは直接rootを触らず、このキューに結果を積むだけにし、
        # メインスレッド側でafter()による定期ポーリングで取り出す
        self._coach_result_queue = queue.Queue()

        self.root.title("LKPT Dashboard")
        self.root.configure(bg=BG_COLOR)
        self.root.geometry("760x720")
        self.root.minsize(620, 520)

        self._init_fonts()
        self._build_ui()
        # _build_ui()はモードに関わらず一律の見た目で組み立てるため、
        # ハイライト・VIEW行の表示切替・絞り込み解除・再描画をまとめて
        # 行う_on_mode_change()を呼び、initial_modeに応じた画面に揃える
        # （initial_mode="LKPT"の場合、従来のrefresh()単独呼び出しと同じ結果になる）
        self._on_mode_change(self.current_mode)

    # ------------------------------------------------------------------
    # 構築
    # ------------------------------------------------------------------
    def _init_fonts(self) -> None:
        """画面全体で使うフォントを一箇所で定義する。"""
        self.f_title = tkfont.Font(family="Yu Gothic UI", size=16, weight="bold")
        self.f_button = tkfont.Font(family="Yu Gothic UI", size=11, weight="bold")
        self.f_summary = tkfont.Font(family="Yu Gothic UI", size=12, weight="bold")
        self.f_section = tkfont.Font(family="Yu Gothic UI", size=12, weight="bold")
        self.f_body = tkfont.Font(family="Yu Gothic UI", size=11)
        self.f_body_bold = tkfont.Font(family="Yu Gothic UI", size=11, weight="bold")
        self.f_coach_heading = tkfont.Font(family="Yu Gothic UI", size=13, weight="bold")
        self.f_chip = tkfont.Font(family="Yu Gothic UI", size=9, weight="bold")
        self.f_chart_label = tkfont.Font(family="Yu Gothic UI", size=10)
        self.f_chart_value = tkfont.Font(family="Consolas", size=11, weight="bold")
        self.f_pie_pct = tkfont.Font(family="Consolas", size=12, weight="bold")
        self.f_time = tkfont.Font(family="Consolas", size=10)
        self.f_badge = tkfont.Font(family="Consolas", size=9, weight="bold")
        self.f_empty = tkfont.Font(family="Yu Gothic UI", size=11)
        self.f_dot = tkfont.Font(family="Yu Gothic UI", size=11)

    def _build_ui(self) -> None:
        title = tk.Label(
            self.root,
            text="📊 LKPT Dashboard",
            bg=BG_COLOR,
            fg=TEXT_COLOR,
            font=self.f_title,
        )
        title.pack(pady=(14, 8))

        self.mode_frame = tk.Frame(self.root, bg=BG_COLOR)
        self.mode_frame.pack(pady=(0, 5))
        for mode in MODES:
            btn = tk.Button(
                self.mode_frame,
                text=mode,
                bg=PERIOD_BTN_BG,
                fg=BUTTON_TEXT_COLOR,
                activebackground=ACCENT_COLOR,
                relief="flat",
                font=self.f_button,
                width=8,
                command=lambda m=mode: self._on_mode_change(m),
            )
            btn.pack(side="left", padx=4)
            self.mode_buttons[mode] = btn

        self.period_frame = tk.Frame(self.root, bg=BG_COLOR)
        self.period_frame.pack(pady=5)

        prev_btn = tk.Button(
            self.period_frame,
            text="◀",
            bg=BG_COLOR,
            fg=DIM_TEXT,
            activebackground=ACCENT_COLOR,
            relief="flat",
            font=self.f_button,
            width=2,
            command=lambda: self._shift_period(-1),
        )
        prev_btn.pack(side="left", padx=(0, 6))

        for period in PERIODS:
            btn = tk.Button(
                self.period_frame,
                text=period,
                bg=PERIOD_BTN_BG,
                fg=BUTTON_TEXT_COLOR,
                activebackground=ACCENT_COLOR,
                relief="flat",
                font=self.f_button,
                width=4,
                command=lambda p=period: self._on_period_change(p),
            )
            btn.pack(side="left", padx=4)
            self.period_buttons[period] = btn

        next_btn = tk.Button(
            self.period_frame,
            text="▶",
            bg=BG_COLOR,
            fg=DIM_TEXT,
            activebackground=ACCENT_COLOR,
            relief="flat",
            font=self.f_button,
            width=2,
            command=lambda: self._shift_period(1),
        )
        next_btn.pack(side="left", padx=(6, 0))

        # 今日を含まない期間を見ている時だけ右上に浮かせて出す「今日に戻る」ボタン。
        # period_frame の中に置くと、出現時にボタン分だけ◀D W M Q Y▶の中心が
        # ずれてしまうため、あえてplace()でウィンドウ右上に独立配置する
        # （中央揃えの行レイアウトに影響を与えないため）
        self.today_btn = tk.Button(
            self.root,
            text="● TODAY",
            bg=ACCENT_COLOR,
            fg=BUTTON_TEXT_COLOR,
            activebackground=TEXT_COLOR,
            activeforeground=BUTTON_TEXT_COLOR,
            relief="flat",
            bd=0,
            font=self.f_chip,
            padx=10,
            pady=3,
            cursor="hand2",
            command=self._go_to_today,
        )

        # 一覧の見せ方の切替。TIMEモードでは一覧の形が固定のため、この行は隠す。
        # "VIEW"のキャプションをボタン行と同じ行に置くと、キャプション分だけ
        # ボタンの中心がD/W/M/Q/Y行の中心とずれてしまうため、キャプションは
        # 別行にして、ボタン行だけを期間ボタンと同じ組み方（左右対称のpadx）で
        # 中央揃えする。
        self.view_caption = tk.Label(
            self.root, text="VIEW", bg=BG_COLOR, fg=DIM_TEXT, font=self.f_chip,
        )
        self.view_caption.pack(pady=(4, 2))

        self.view_frame = tk.Frame(self.root, bg=BG_COLOR)
        self.view_frame.pack(pady=(0, 0))
        for view in HISTORY_VIEWS:
            btn = tk.Button(
                self.view_frame,
                text=view,
                bg=PERIOD_BTN_BG,
                fg=BUTTON_TEXT_COLOR,
                activebackground=ACCENT_COLOR,
                relief="flat",
                font=self.f_button,
                width=7,
                command=lambda v=view: self._on_view_change(v),
            )
            btn.pack(side="left", padx=3)
            self.view_buttons[view] = btn

        self.summary_label = tk.Label(
            self.root, text="", bg=BG_COLOR, fg=ACCENT_COLOR, font=self.f_summary,
        )
        self.summary_label.pack(pady=(10, 4))

        # 棒グラフでタグ絞り込み中の状態表示。絞り込み中のタグが次の期間に
        # 存在しない（＝棒自体が描かれない）と解除する手段が無くなるため、
        # 常設のクリック可能な表示として用意する。
        self.filter_label = tk.Label(
            self.root, text="", bg=BG_COLOR, fg=ACCENT_COLOR, font=self.f_chip,
            cursor="hand2",
        )
        self.filter_label.pack(pady=(0, 4))
        self.filter_label.bind("<Button-1>", self._clear_filter)

        self.canvas = tk.Canvas(
            self.root, width=700, height=CHART_HEIGHT_LKPT, bg=BG_COLOR,
            highlightthickness=0,
        )
        self.canvas.pack(pady=(0, 10))

        head = tk.Frame(self.root, bg=BG_COLOR)
        head.pack(fill="x", padx=16, pady=(0, 6))
        tk.Frame(head, bg=LINE_COLOR, height=1).pack(fill="x", pady=(0, 8))
        head_row = tk.Frame(head, bg=BG_COLOR)
        head_row.pack(fill="x")
        tk.Label(
            head_row, text="H I S T O R Y", bg=BG_COLOR, fg=TEXT_COLOR,
            font=self.f_section,
        ).pack(side="left")
        self.hist_count_label = tk.Label(
            head_row, text="", bg=BG_COLOR, fg=DIM_TEXT, font=self.f_time,
        )
        self.hist_count_label.pack(side="left", padx=(12, 0))

        self._build_history_area()

        self._highlight_period_button()
        self._highlight_mode_button()
        self._highlight_view_button()

    def _build_history_area(self) -> None:
        """
        スクロール可能な一覧領域を作る。tk.Listboxは1行1色のプレーンテキストしか
        描けないため、Canvas上に置いたFrameへ1件ずつウィジェットを積む方式にした
        （入力ポップアップのタグ行と同じ作り方）。
        """
        outer = tk.Frame(self.root, bg=BG_COLOR)
        outer.pack(fill="both", expand=True, padx=16, pady=(0, 14))

        self.hist_canvas = tk.Canvas(outer, bg=BG_COLOR, highlightthickness=0)
        scrollbar = tk.Scrollbar(
            outer, orient="vertical", command=self.hist_canvas.yview,
        )
        self.hist_canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        self.hist_canvas.pack(side="left", fill="both", expand=True)

        self.hist_inner = tk.Frame(self.hist_canvas, bg=BG_COLOR)
        self._hist_window = self.hist_canvas.create_window(
            (0, 0), window=self.hist_inner, anchor="nw",
        )

        self.hist_inner.bind(
            "<Configure>",
            lambda e: self.hist_canvas.configure(
                scrollregion=self.hist_canvas.bbox("all")
            ),
        )
        self.hist_canvas.bind(
            "<Configure>",
            lambda e: self.hist_canvas.itemconfig(self._hist_window, width=e.width),
        )
        # ホイールは子ウィジェット上でも効かせたいので、領域に入った時だけ
        # bind_allで拾い、出たら解除する
        self.hist_canvas.bind(
            "<Enter>",
            lambda e: self.hist_canvas.bind_all("<MouseWheel>", self._on_mousewheel),
        )
        self.hist_canvas.bind(
            "<Leave>", lambda e: self.hist_canvas.unbind_all("<MouseWheel>"),
        )

    def _on_mousewheel(self, event) -> None:
        self.hist_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    # ------------------------------------------------------------------
    # ボタンの状態
    # ------------------------------------------------------------------
    def _highlight_buttons(self, buttons: dict, current: str) -> None:
        for key, btn in buttons.items():
            btn.config(
                bg=ACCENT_COLOR if key == current else PERIOD_BTN_BG,
                fg=BUTTON_TEXT_COLOR,
            )

    def _highlight_period_button(self) -> None:
        self._highlight_buttons(self.period_buttons, self.current_period)

    def _highlight_mode_button(self) -> None:
        self._highlight_buttons(self.mode_buttons, self.current_mode)

    def _highlight_view_button(self) -> None:
        self._highlight_buttons(self.view_buttons, self.current_view)

    def _on_mode_change(self, mode: str) -> None:
        self.current_mode = mode
        self._highlight_mode_button()
        # LKPTモードはL/K/P/T、TIMEモードはタグと、棒グラフの分類軸が
        # 違うため、モードを切り替えたら絞り込みは意味を持たなくなる
        self.chart_filter = None
        if mode == "LKPT":
            self.view_caption.pack(after=self.period_frame, pady=(4, 2))
            self.view_frame.pack(after=self.view_caption, pady=(0, 0))
        else:
            self.view_caption.pack_forget()
            self.view_frame.pack_forget()
        self.refresh()

    def _on_period_change(self, period: str) -> None:
        self.current_period = period
        self._highlight_period_button()
        self.refresh()

    def _shift_period(self, direction: int) -> None:
        """
        表示中の期間を1つ前後にずらす。_period_range()のendは常に
        「次の期間の開始時刻」を表すため、基準日をendに合わせれば次の期間へ、
        startの1日前に合わせれば前の期間へ移動できる
        （D/W/M/Q/Yいずれも、この境界の付け替えだけで正しく動く）。
        """
        start, end = _period_range(self.current_period, self.reference_date)
        self.reference_date = end if direction > 0 else start - timedelta(days=1)
        self.refresh()

    def _showing_today(self) -> bool:
        start, end = _period_range(self.current_period, self.reference_date)
        return start <= datetime.now() < end

    def _update_today_button(self) -> None:
        """今日を含む期間を見ている時はTODAYボタンを隠す（押しても意味が無いため）。"""
        if self._showing_today():
            self.today_btn.place_forget()
        else:
            self.today_btn.place(relx=1.0, x=-16, y=94, anchor="ne")

    def _go_to_today(self) -> None:
        self.reference_date = datetime.now()
        self.refresh()

    def _on_view_change(self, view: str) -> None:
        self.current_view = view
        self._highlight_view_button()
        self.refresh()

    def _on_bar_click(self, category: str) -> None:
        """棒グラフのクリックで絞り込みをトグルする（LKPTモードはL/K/P/T、
        TIMEモードはタグ）。同じ棒をもう一度クリックすると全件表示に戻る。"""
        self.chart_filter = None if self.chart_filter == category else category
        self.refresh()

    def _clear_filter(self, event=None) -> None:
        """絞り込み解除。絞り込み中の分類が次の期間には無く棒が描かれない
        場合でも解除できるよう、常設のラベルからも解除できるようにする。"""
        if self.chart_filter:
            self.chart_filter = None
            self.refresh()

    # ------------------------------------------------------------------
    # データ読込と再描画
    # ------------------------------------------------------------------
    def refresh(self) -> None:
        """Excelからデータを再読込し、グラフ・一覧を再描画する。"""
        self._update_today_button()
        try:
            self.tag_colors = load_tag_colors()
            if self.current_mode == "LKPT":
                all_entries = load_entries()
            elif self.current_mode == "TIME":
                all_entries = load_time_log_entries()
            else:
                all_entries = None  # ACTIONモードはget_actions()で別途読む
                self.actions = get_actions()
        except Exception as e:
            self.summary_label.config(text=f"❌ 読込エラー: {e}")
            print(f"❌ ダッシュボードのデータ読込に失敗しました: {e}")
            return

        period_label = format_period_label(self.current_period, self.reference_date)

        if self.chart_filter and self.current_mode == "LKPT":
            heading = next(h for l, k, h in LKPT_FIELDS if l == self.chart_filter)
            filter_text = f"🔍 {self.chart_filter}（{heading}）のみ表示中（タップで解除）"
        elif self.chart_filter:
            filter_text = f"🔍 {self.chart_filter} のみ表示中（タップで解除）"
        else:
            filter_text = ""
        self.filter_label.config(text=filter_text)

        if self.current_mode == "LKPT":
            filtered = filter_entries_by_period(
                all_entries, self.current_period, reference=self.reference_date,
            )
            self.entries = sorted(filtered, key=lambda e: e["datetime"])
            # LKPTモードの棒グラフはタグではなくL/K/P/Tで分類する
            field_counts = aggregate_lkpt_field_counts(self.entries)
            all_rows = self._flatten_fields(self.entries)
            self._history_rows = [
                row for row in all_rows
                if not self.chart_filter or row[1] == self.chart_filter
            ]
            self.summary_label.config(
                text=f"{period_label} ／ {len(self.entries)} entries"
            )
            self.hist_count_label.config(text=f"{len(self._history_rows)} entries")
            self._draw_count_chart(field_counts)
            self._render_lkpt_history()
        elif self.current_mode == "TIME":
            filtered = filter_entries_by_period(
                all_entries, self.current_period, reference=self.reference_date,
            )
            self.time_entries = sorted(filtered, key=lambda e: e["start"])
            # TIMEモードの棒グラフは今まで通りタグで分類する
            minutes_by_tag = aggregate_time_by_tag(self.time_entries)
            total_minutes = sum(minutes_by_tag.values())
            blocks = [
                e for e in self.time_entries
                if (e["end"] - e["start"]).total_seconds() > 0
                and (not self.chart_filter or e["tag"] == self.chart_filter)
            ]
            self.summary_label.config(
                text=f"{period_label} ／ 合計 {_format_duration(total_minutes)}"
            )
            self.hist_count_label.config(text=f"{len(blocks)} blocks")
            self._draw_time_charts(minutes_by_tag)
            self._render_time_history(blocks)
        else:
            self._refresh_action_mode(period_label)

    def _refresh_action_mode(self, period_label: str) -> None:
        """
        ACTIONモード: 未着手は期間に関わらず常に全件、完了は表示中の期間で
        絞り込む（未完了のタスクが期間フィルタで見えなくなるのを防ぐため）。
        """
        start, end = _period_range(self.current_period, self.reference_date)
        pending = [a for a in self.actions if a["status"] == ACTION_STATUS_PENDING]
        pending.sort(key=lambda a: a["created_at"])
        completed = [
            a for a in self.actions
            if a["status"] == ACTION_STATUS_DONE and a["completed_at"]
            and start <= a["completed_at"] < end
        ]
        completed.sort(key=lambda a: a["completed_at"], reverse=True)

        counts = {}
        if pending:
            counts[ACTION_STATUS_PENDING] = len(pending)
        if completed:
            counts[ACTION_STATUS_DONE] = len(completed)

        show_pending = (
            pending if not self.chart_filter or self.chart_filter == ACTION_STATUS_PENDING
            else []
        )
        show_completed = (
            completed if not self.chart_filter or self.chart_filter == ACTION_STATUS_DONE
            else []
        )

        self.summary_label.config(
            text=f"{period_label} ／ 未着手 {len(pending)}件 ・ 完了(期間内) {len(completed)}件"
        )
        self.hist_count_label.config(
            text=f"{len(show_pending) + len(show_completed)} items"
        )
        self._draw_action_chart(counts)
        self._render_action_list(show_pending, show_completed)

    # ------------------------------------------------------------------
    # グラフ
    # ------------------------------------------------------------------
    def _chart_height(self) -> int:
        return CHART_HEIGHT_TIME if self.current_mode == "TIME" else CHART_HEIGHT_LKPT

    def _draw_empty_chart(self) -> None:
        self.canvas.create_text(
            350, self._chart_height() / 2, text="記録がありません",
            fill=DIM_TEXT, font=self.f_empty,
        )

    def _draw_bars(self, values: dict, x0: int, width: int,
                   value_formatter, color_map: dict) -> None:
        """
        分類別の値を棒グラフとして描く（LKPTモードはL/K/P/T、TIMEモードは
        タグで共用）。棒はクリックでその分類の絞り込みトグルになる。
        絞り込み中は非選択の棒を背景色寄りに沈めて選択中の棒を引き立てる。
        """
        max_value = max(values.values())
        base_y = self._chart_height() - 40   # 40 = 下のラベル用の余白
        max_bar_height = base_y - 45         # 45 = 上の数値ラベル用の余白
        slot = width / max(len(values), 1)
        bar_width = min(52, slot * 0.56)

        for i, (category, value) in enumerate(values.items()):
            base_color = color_map.get(category, "#888888")
            dimmed = bool(self.chart_filter) and category != self.chart_filter
            color = _blend_color(base_color, BG_COLOR, 0.7) if dimmed else base_color
            cx = x0 + slot * (i + 0.5)
            bar_height = (
                int((value / max_value) * max_bar_height) if max_value else 0
            )
            rect = self.canvas.create_rectangle(
                cx - bar_width / 2, base_y - bar_height,
                cx + bar_width / 2, base_y,
                fill=color, outline="",
            )
            value_text = self.canvas.create_text(
                cx, base_y - bar_height - 14,
                text=value_formatter(value),
                fill=DIM_TEXT if dimmed else TEXT_COLOR,
                font=self.f_chart_value,
            )
            label_text = self.canvas.create_text(
                cx, base_y + 18, text=category, fill=DIM_TEXT,
                font=self.f_chart_label,
            )

            # 棒・数値・ラベルのどこをクリックしても同じ絞り込みが働くようにする
            for item_id in (rect, value_text, label_text):
                self.canvas.tag_bind(
                    item_id, "<Button-1>",
                    lambda e, c=category: self._on_bar_click(c),
                )
                self.canvas.tag_bind(
                    item_id, "<Enter>", lambda e: self.canvas.config(cursor="hand2"),
                )
                self.canvas.tag_bind(
                    item_id, "<Leave>", lambda e: self.canvas.config(cursor=""),
                )

    def _draw_count_chart(self, field_counts: dict) -> None:
        """LKPTモード: L/K/P/T別の記録件数を棒グラフで描く。"""
        self.canvas.config(height=CHART_HEIGHT_LKPT)
        self.canvas.delete("all")
        if not field_counts:
            self._draw_empty_chart()
            return
        self._draw_bars(
            field_counts, x0=60, width=580, value_formatter=str,
            color_map=LKPT_FIELD_COLORS,
        )

    def _draw_time_charts(self, minutes_by_tag: dict) -> None:
        """TIMEモード: 左にタグ別作業時間の棒グラフ、右に構成比の円グラフを描く。"""
        self.canvas.config(height=CHART_HEIGHT_TIME)
        self.canvas.delete("all")
        if not minutes_by_tag:
            self._draw_empty_chart()
            return

        self._draw_bars(
            minutes_by_tag, x0=20, width=380, value_formatter=_format_duration,
            color_map=self.tag_colors,
        )
        self._draw_pie(minutes_by_tag, cx=545, cy=118, r=84)

    def _draw_action_chart(self, counts: dict) -> None:
        """ACTIONモード: 未着手/完了(期間内)の件数を棒グラフで描く。"""
        self.canvas.config(height=CHART_HEIGHT_LKPT)
        self.canvas.delete("all")
        if not counts:
            self._draw_empty_chart()
            return
        self._draw_bars(
            counts, x0=60, width=580, value_formatter=str,
            color_map=ACTION_STATUS_COLORS,
        )

    def _draw_pie(self, minutes_by_tag: dict, cx: int, cy: int, r: int) -> None:
        """
        構成比の円グラフ。各扇にパーセントを描く。
        扇が細いと内側に文字が収まらないため、一定の割合を下回る場合は
        円の外側へ逃がす。
        """
        total = sum(minutes_by_tag.values())
        if not total:
            return

        start_angle = 90.0
        for tag, minutes in minutes_by_tag.items():
            color = self.tag_colors.get(tag, "#888888")
            pct = 100.0 * minutes / total
            extent = -360.0 * (minutes / total)
            self.canvas.create_arc(
                cx - r, cy - r, cx + r, cy + r,
                start=start_angle, extent=extent,
                fill=color, outline=BG_COLOR, width=2,
            )

            # 扇の中心角の方向にラベルを置く
            mid_deg = start_angle + extent / 2.0
            mid_rad = math.radians(mid_deg)
            if pct >= PIE_INSIDE_LABEL_MIN_PCT:
                label_r = r * 0.62
                fill = TEXT_COLOR
            else:
                label_r = r + 22
                fill = DIM_TEXT
            lx = cx + label_r * math.cos(mid_rad)
            ly = cy - label_r * math.sin(mid_rad)
            self.canvas.create_text(
                lx, ly, text=f"{pct:.1f}%", fill=fill, font=self.f_pie_pct,
            )

            start_angle += extent

    # ------------------------------------------------------------------
    # 一覧（HISTORY）
    # ------------------------------------------------------------------
    def _clear_history(self) -> None:
        for child in self.hist_inner.winfo_children():
            child.destroy()
        self.hist_canvas.yview_moveto(0)

    def _wrap_length(self, indent: int) -> int:
        """
        本文ラベルの表示可能幅（px）。ウィジェット未実現時は既定値でしのぎ、
        以降のrefreshで実寸に追従する。
        """
        width = self.hist_canvas.winfo_width()
        if width <= 1:
            width = 700
        return max(220, width - indent)

    def _flatten_fields(self, entries: list) -> list:
        """
        L/K/P/Tは1件のチェックインにまとめず、値がある項目ごとに1行として
        並べる（1フィールド=1行にするため）。1件に複数フィールドがあっても、
        L→K→P→Tの順に別々の行として展開する。
        """
        rows = []
        for entry in entries:
            for letter, key, _ in LKPT_FIELDS:
                if entry[key]:
                    rows.append((entry, letter, key))
        return rows

    def _truncate_to_width(self, text: str, font: tkfont.Font, max_width: int) -> str:
        """本文を1行に収めるため、指定ピクセル幅に収まるまで末尾を省略記号で切る。"""
        if font.measure(text) <= max_width:
            return text
        lo, hi = 0, len(text)
        while lo < hi:
            mid = (lo + hi + 1) // 2
            if font.measure(text[:mid] + "…") <= max_width:
                lo = mid
            else:
                hi = mid - 1
        return text[:lo] + "…" if lo > 0 else "…"

    def _show_history_empty(self) -> None:
        tk.Label(
            self.hist_inner, text="この期間の記録はまだありません",
            bg=BG_COLOR, fg=DIM_TEXT, font=self.f_empty,
        ).pack(anchor="w", pady=14)

    def _tag_chip(self, parent, tag: str):
        """タグ名を、そのタグ固有の色を地色にしたチップとして表示する（TIMEモード用）。"""
        return tk.Label(
            parent, text=f" {tag} ", bg=self.tag_colors.get(tag, "#888888"),
            fg=TEXT_COLOR, font=self.f_chip,
        )

    def _tag_text(self, parent, tag: str):
        """
        タグ名をプレーンテキストで表示する（LKPTモードのHISTORY用）。
        LKPTモードの色の主役はL/K/P/Tなので、タグ色のチップは使わず
        タグ名は補助情報として控えめに出す。
        tagがNoneでも文字列"None"を描画しないよう、ここでも正規化しておく
        （呼び出し側でも空判定しているが、構造的に起こり得なくしておく）。
        """
        return tk.Label(
            parent, text=tag or "", bg=parent.cget("bg"), fg=DIM_TEXT, font=self.f_time,
        )

    def _badge(self, parent, letter: str):
        """L/K/P/Tの1文字バッジ。棒グラフと揃えたL/K/P/T固有色を地色にする。"""
        return tk.Label(
            parent, text=letter, bg=LKPT_FIELD_COLORS.get(letter, BADGE_BG),
            fg=TEXT_COLOR, font=self.f_badge, width=2,
        )

    def _render_lkpt_history(self) -> None:
        self._clear_history()

        # COACHは棒グラフの絞り込み（L/K/P/T単体）に関わらず、期間全体を
        # 対象に要約する方が意味があるため、self.entries（未絞り込み）で
        # 空判定する（他の3ビューはself._history_rowsで判定）
        if self.current_view == "COACH":
            if not self.entries:
                self._show_history_empty()
                return
            self._render_view_coach()
            return

        if not self._history_rows:
            self._show_history_empty()
            return

        if self.current_view == "CARD":
            self._render_view_cards()
        elif self.current_view == "FLOW":
            self._render_view_timeline()
        else:
            self._render_view_grouped()

    # --- CARD: L/K/P/Tを1件ずつ、1行のカードとして並べる ---
    def _render_view_cards(self) -> None:
        indent = 210  # 色帯+時刻+タグ+バッジの目安幅（本文の省略計算に使う）
        max_width = self._wrap_length(indent)
        for entry, letter, key in self._history_rows:
            color = LKPT_FIELD_COLORS.get(letter, "#888888")

            card = tk.Frame(self.hist_inner, bg=CARD_BG)
            card.pack(fill="x", pady=2)

            tk.Frame(card, bg=color, width=4).pack(side="left", fill="y")

            row = tk.Frame(card, bg=CARD_BG)
            row.pack(side="left", fill="x", expand=True, padx=10, pady=6)
            tk.Label(
                row, text=entry["datetime"].strftime("%m-%d %H:%M"),
                bg=CARD_BG, fg=DIM_TEXT, font=self.f_time,
            ).pack(side="left")
            self._badge(row, letter).pack(side="left", padx=(8, 8))
            if entry["tag"]:
                self._tag_text(row, entry["tag"]).pack(side="left", padx=(0, 8))
            tk.Label(
                row, text=self._truncate_to_width(entry[key], self.f_body, max_width),
                bg=CARD_BG, fg=BODY_TEXT, font=self.f_body, anchor="w",
            ).pack(side="left", fill="x", expand=True)

    # --- FLOW: 左に時刻軸を立てて時間の流れで追う（L/K/P/Tは1行ずつ） ---
    def _render_view_timeline(self) -> None:
        indent = 250
        max_width = self._wrap_length(indent)
        current_day = None

        for entry, letter, key in self._history_rows:
            day = entry["datetime"].date()
            if day != current_day:
                current_day = day
                tk.Label(
                    self.hist_inner,
                    text=entry["datetime"].strftime("%Y-%m-%d (%a)").upper(),
                    bg=BG_COLOR, fg=DIM_TEXT, font=self.f_time,
                ).pack(anchor="w", pady=(12, 4))

            color = LKPT_FIELD_COLORS.get(letter, "#888888")

            item = tk.Frame(self.hist_inner, bg=BG_COLOR)
            item.pack(fill="x", pady=1)

            tk.Label(
                item, text=entry["datetime"].strftime("%H:%M"),
                bg=BG_COLOR, fg=DIM_TEXT, font=self.f_time, width=6, anchor="e",
            ).pack(side="left")

            tk.Label(item, text="●", bg=BG_COLOR, fg=color, font=self.f_dot).pack(
                side="left", padx=(8, 8),
            )

            self._badge(item, letter).pack(side="left", padx=(0, 8))
            if entry["tag"]:
                self._tag_text(item, entry["tag"]).pack(side="left", padx=(0, 8))
            tk.Label(
                item, text=self._truncate_to_width(entry[key], self.f_body, max_width),
                bg=BG_COLOR, fg=BODY_TEXT, font=self.f_body, anchor="w",
            ).pack(side="left", fill="x", expand=True)

    # --- GROUP: L/K/P/Tごとに束ねて横断で見る（1件=1行） ---
    def _render_view_grouped(self) -> None:
        indent = 220
        max_width = self._wrap_length(indent)

        for letter, key, heading in LKPT_FIELDS:
            items = [entry for entry, l, k in self._history_rows if l == letter]
            if not items:
                continue

            section = tk.Frame(self.hist_inner, bg=BG_COLOR)
            section.pack(fill="x", pady=(10, 0))

            head = tk.Frame(section, bg=BG_COLOR)
            head.pack(fill="x")
            self._badge(head, letter).pack(side="left")
            tk.Label(
                head, text=heading, bg=BG_COLOR, fg=TEXT_COLOR, font=self.f_chip,
            ).pack(side="left", padx=(10, 0))
            tk.Label(
                head, text=f"({len(items)})", bg=BG_COLOR, fg=DIM_TEXT,
                font=self.f_time,
            ).pack(side="left", padx=(8, 0))
            tk.Frame(section, bg=LINE_COLOR, height=1).pack(fill="x", pady=(6, 2))

            for entry in items:
                row = tk.Frame(section, bg=BG_COLOR)
                row.pack(fill="x", pady=1)
                tk.Label(
                    row, text=entry["datetime"].strftime("%m-%d %H:%M"),
                    bg=BG_COLOR, fg=DIM_TEXT, font=self.f_time,
                ).pack(side="left")
                if entry["tag"]:
                    self._tag_text(row, entry["tag"]).pack(side="left", padx=(10, 0))
                tk.Label(
                    row,
                    text=self._truncate_to_width(entry[key], self.f_body, max_width),
                    bg=BG_COLOR, fg=BODY_TEXT, font=self.f_body, anchor="w",
                ).pack(side="left", padx=(10, 0), fill="x", expand=True)

    # --- COACH: Gemini APIによる期間要約＋コーチングアドバイス ---
    def _coach_cache_key(self) -> tuple:
        start, end = _period_range(self.current_period, self.reference_date)
        return (self.current_period, start, end)

    def _render_view_coach(self) -> None:
        key = self._coach_cache_key()
        cached = self._coach_cache.get(key)

        header = tk.Frame(self.hist_inner, bg=BG_COLOR)
        header.pack(fill="x", pady=(4, 8))
        tk.Label(
            header, text="🧭 COACH", bg=BG_COLOR, fg=TEXT_COLOR, font=self.f_section,
        ).pack(side="left")
        self.coach_regenerate_btn = tk.Button(
            header, text="🔄 再生成", bg=CARD_BG, fg=TEXT_COLOR, relief="flat",
            font=self.f_chip, padx=8, pady=3, bd=0, cursor="hand2",
            command=self._start_coach_generation,
        )
        self.coach_regenerate_btn.pack(side="right")

        self.coach_status_label = tk.Label(
            self.hist_inner, text="", bg=BG_COLOR, fg=DIM_TEXT, font=self.f_time,
            anchor="w", justify="left",
        )
        self.coach_status_label.pack(fill="x", pady=(0, 6))

        body = tk.Frame(self.hist_inner, bg=CARD_BG)
        body.pack(fill="both", expand=True, pady=(0, 10))
        # LabelのwraplengthはウィンドウをリサイズしてもTk側が自動で
        # 追従しない（固定px値のため）。wrap="word"のTextウィジェットは
        # 自身の実描画幅に合わせて自動で折り返すため、こちらを使う
        self.coach_text = tk.Text(
            body, bg=CARD_BG, fg=TEXT_COLOR, font=self.f_body, wrap="word",
            relief="flat", bd=0, padx=14, pady=12, highlightthickness=0,
            cursor="arrow", height=1,
        )
        self.coach_text.tag_configure(
            "heading", font=self.f_coach_heading, foreground=ACCENT_COLOR,
            spacing1=14, spacing3=6,
        )
        self.coach_text.tag_configure(
            "bullet", font=self.f_body, foreground=TEXT_COLOR,
            lmargin1=18, lmargin2=34, spacing1=2,
        )
        self.coach_text.tag_configure(
            "body", font=self.f_body, foreground=TEXT_COLOR, spacing1=2,
        )
        # bold用タグは最後に設定し、重なった時に最優先で効くようにする
        # （Tkinterのタグ優先度は「後から設定したタグほど高い」）
        self.coach_text.tag_configure("bold", font=self.f_body_bold)
        self.coach_text.config(state="disabled")
        self.coach_text.pack(fill="both", expand=True)
        self.coach_text.bind(
            "<Configure>", lambda e: self._autofit_coach_text_height(),
        )

        if cached:
            self._show_coach_result(cached["text"], cached["generated_at"])
        elif key == self._coach_inflight_key:
            # 別の期間・ビューを経由して戻ってきた場合、生成中のリクエストが
            # 既にあるなら新しく叩き直さず、そのまま待つ
            self.coach_regenerate_btn.config(state="disabled")
            self.coach_status_label.config(text="🤖 Geminiで生成中…（数秒〜数十秒かかります）")
        else:
            self._start_coach_generation()

    def _start_coach_generation(self) -> None:
        if self.current_view != "COACH":
            return
        self._coach_request_id += 1
        request_id = self._coach_request_id
        key = self._coach_cache_key()
        entries = list(self.entries)
        period_label = format_period_label(self.current_period, self.reference_date)

        self._coach_inflight_key = key
        self.coach_regenerate_btn.config(state="disabled")
        self.coach_status_label.config(text="🤖 Geminiで生成中…（数秒〜数十秒かかります）")
        self._clear_coach_text()

        def worker():
            # Tkinterのウィジェットにはメインスレッドからしか触れないため、
            # ここではrootに一切触らず、結果をキューに積むだけにする
            try:
                prompt = build_coach_prompt(entries, period_label)
                text = call_gemini_summary(prompt)
            except Exception as e:
                self._coach_result_queue.put(("error", request_id, key, e))
            else:
                self._coach_result_queue.put(("success", request_id, key, text))

        threading.Thread(target=worker, daemon=True).start()
        self.root.after(150, self._poll_coach_queue)

    def _poll_coach_queue(self) -> None:
        """メインスレッドからafter()経由で定期的にキューを覗き、結果が
        届いていれば処理する。ワーカースレッドからTkに直接触らないための仕組み。"""
        if not self.root.winfo_exists():
            return  # ダッシュボードが閉じられていたらポーリングを止める
        try:
            kind, request_id, key, payload = self._coach_result_queue.get_nowait()
        except queue.Empty:
            self.root.after(150, self._poll_coach_queue)
            return

        if kind == "success":
            self._on_coach_success(request_id, key, payload)
        else:
            self._on_coach_error(request_id, key, payload)

    def _on_coach_success(self, request_id: int, key: tuple, text: str) -> None:
        if key == self._coach_inflight_key:
            self._coach_inflight_key = None
        if request_id != self._coach_request_id or self.current_view != "COACH":
            return  # その間に別の期間・ビューへ移動していたら結果は捨てる
        generated_at = datetime.now()
        self._coach_cache[key] = {"text": text, "generated_at": generated_at}
        self._show_coach_result(text, generated_at)

    def _on_coach_error(self, request_id: int, key: tuple, error: Exception) -> None:
        if key == self._coach_inflight_key:
            self._coach_inflight_key = None
        if request_id != self._coach_request_id or self.current_view != "COACH":
            return
        self.coach_regenerate_btn.config(state="normal")
        self.coach_status_label.config(text=f"❌ 生成に失敗しました: {error}")

    def _show_coach_result(self, text: str, generated_at: datetime) -> None:
        self.coach_regenerate_btn.config(state="normal")
        self.coach_status_label.config(
            text=f"生成日時: {generated_at.strftime('%m-%d %H:%M')}"
        )
        self._render_coach_text(text)

    def _clear_coach_text(self) -> None:
        if hasattr(self, "coach_text") and self.coach_text.winfo_exists():
            self.coach_text.config(state="normal")
            self.coach_text.delete("1.0", "end")
            self.coach_text.config(state="disabled")

    def _render_coach_text(self, text: str) -> None:
        """
        Geminiが返す軽量マークダウン（## 見出し／- 箇条書き／**太字**）を
        パースし、生の記号を見せずにTextウィジェットのタグとして描画する。
        """
        widget = self.coach_text
        widget.config(state="normal")
        widget.delete("1.0", "end")

        for raw_line in text.split("\n"):
            line = raw_line.rstrip()
            if line.startswith("## "):
                heading_text = line[3:].strip()
                icon = COACH_HEADING_ICONS.get(heading_text, "▸")
                widget.insert("end", f"{icon} ", ("heading",))
                self._insert_coach_inline(widget, heading_text, ("heading",))
                widget.insert("end", "\n")
            elif line.startswith("- ") or line.startswith("・"):
                bullet_text = line[2:] if line.startswith("- ") else line[1:]
                widget.insert("end", "•  ", ("bullet",))
                self._insert_coach_inline(widget, bullet_text.strip(), ("bullet",))
                widget.insert("end", "\n")
            elif line == "":
                widget.insert("end", "\n")
            else:
                self._insert_coach_inline(widget, line, ("body",))
                widget.insert("end", "\n")

        widget.config(state="disabled")
        self.root.after_idle(self._autofit_coach_text_height)

    def _insert_coach_inline(self, widget, line: str, base_tags: tuple) -> None:
        """行内の**太字**をパースして挿入する（base_tagsは段落種別のタグ）。"""
        parts = line.split("**")
        for i, part in enumerate(parts):
            if not part:
                continue
            tags = base_tags + ("bold",) if i % 2 == 1 else base_tags
            widget.insert("end", part, tags)

    def _autofit_coach_text_height(self) -> None:
        """
        Textウィジェットは高さ(行数)を自動で内容に合わせないため、実際に
        折り返された表示行数を数えて高さを設定する。この関数はTextの
        <Configure>からも呼ばれるため、ウィンドウ幅が変わって折り返しが
        変化した時も高さが追従する。
        """
        if not hasattr(self, "coach_text") or not self.coach_text.winfo_exists():
            return
        self.coach_text.update_idletasks()
        try:
            num_lines = int(
                self.coach_text.count("1.0", "end-1c", "displaylines")[0]
            )
        except (TypeError, IndexError):
            num_lines = 1
        new_height = max(1, num_lines)
        if self.coach_text.cget("height") != new_height:
            self.coach_text.config(height=new_height)

    # --- TIMEモードの一覧（VIEWの選択に関わらず共通） ---
    def _render_time_history(self, blocks: list) -> None:
        self._clear_history()
        if not blocks:
            self._show_history_empty()
            return

        max_minutes = max(
            int((e["end"] - e["start"]).total_seconds() // 60) for e in blocks
        )

        for entry in blocks:
            minutes = int((entry["end"] - entry["start"]).total_seconds() // 60)
            color = self.tag_colors.get(entry["tag"], "#888888")

            row = tk.Frame(self.hist_inner, bg=BG_COLOR)
            row.pack(fill="x", pady=4)

            span = (
                f"{entry['start'].strftime('%m-%d %H:%M')}"
                f" – {entry['end'].strftime('%H:%M')}"
            )
            tk.Label(
                row, text=span, bg=BG_COLOR, fg=DIM_TEXT, font=self.f_time,
            ).pack(side="left")
            self._tag_chip(row, entry["tag"]).pack(side="left", padx=(10, 0))
            # 中項目（何をしていたか）。これまで書き込むだけで表示していな
            # かったため、「R19の中で何をしていたか」が画面から分からなかった
            if entry.get("sub_item"):
                tk.Label(
                    row, text=entry["sub_item"], bg=BG_COLOR, fg=DIM_TEXT,
                    font=self.f_time,
                ).pack(side="left", padx=(6, 0))
            tk.Label(
                row, text=_format_duration(minutes), bg=BG_COLOR, fg=TEXT_COLOR,
                font=self.f_chart_value, width=7, anchor="e",
            ).pack(side="right")

            # 所要時間そのものを横棒の長さにする（数字を読まなくても塊が見える）
            track = tk.Frame(row, bg=TRACK_BG, height=8)
            track.pack(side="left", fill="x", expand=True, padx=12)
            track.pack_propagate(False)
            ratio = minutes / max_minutes if max_minutes else 0
            tk.Frame(track, bg=color).place(
                x=0, y=0, relwidth=ratio, relheight=1.0,
            )

    # --- ACTIONモードの一覧（VIEWの選択に関わらず共通） ---
    def _render_action_list(self, pending: list, completed: list) -> None:
        """
        未着手・完了(期間内)を別セクションで表示する。未着手の行はクリックで
        即座に完了にできる（dashboard.py初めての書き込み操作）。
        """
        self._clear_history()
        if not pending and not completed:
            self._show_history_empty()
            return

        max_width = self._wrap_length(70)

        if pending:
            section = tk.Frame(self.hist_inner, bg=BG_COLOR)
            section.pack(fill="x", pady=(4, 0))
            head = tk.Frame(section, bg=BG_COLOR)
            head.pack(fill="x")
            tk.Label(
                head, text="🔲 未着手", bg=BG_COLOR, fg=TEXT_COLOR, font=self.f_chip,
            ).pack(side="left")
            tk.Label(
                head, text=f"({len(pending)})", bg=BG_COLOR, fg=DIM_TEXT,
                font=self.f_time,
            ).pack(side="left", padx=(8, 0))
            tk.Frame(section, bg=LINE_COLOR, height=1).pack(fill="x", pady=(6, 2))

            for action in pending:
                row = tk.Frame(section, bg=CARD_BG, cursor="hand2")
                row.pack(fill="x", pady=2)

                check = tk.Label(
                    row, text="☐", bg=CARD_BG,
                    fg=ACTION_STATUS_COLORS[ACTION_STATUS_PENDING], font=self.f_body,
                )
                check.pack(side="left", padx=(10, 8), pady=8)

                body = tk.Frame(row, bg=CARD_BG)
                body.pack(side="left", fill="both", expand=True, pady=8, padx=(0, 10))
                tk.Label(
                    body, text=action["content"], bg=CARD_BG, fg=BODY_TEXT,
                    font=self.f_body, anchor="w", justify="left", wraplength=max_width,
                ).pack(fill="x")

                meta = tk.Frame(body, bg=CARD_BG)
                meta.pack(fill="x", pady=(2, 0))
                tk.Label(
                    meta, text=action["created_at"].strftime("%m-%d %H:%M"),
                    bg=CARD_BG, fg=DIM_TEXT, font=self.f_time,
                ).pack(side="left")
                if action["tag"]:
                    self._tag_text(meta, action["tag"]).pack(side="left", padx=(8, 0))
                if action["origin"] == "P":
                    tk.Label(
                        meta, text="from P", bg=CARD_BG, fg=DIM_TEXT, font=self.f_time,
                    ).pack(side="left", padx=(8, 0))

                for widget in (row, check, body):
                    widget.bind(
                        "<Button-1>",
                        lambda e, r=action["row"]: self._complete_action(r),
                    )

        if completed:
            section = tk.Frame(self.hist_inner, bg=BG_COLOR)
            section.pack(fill="x", pady=(14, 0))
            head = tk.Frame(section, bg=BG_COLOR)
            head.pack(fill="x")
            tk.Label(
                head, text="☑ 完了", bg=BG_COLOR, fg=TEXT_COLOR, font=self.f_chip,
            ).pack(side="left")
            tk.Label(
                head, text=f"({len(completed)})", bg=BG_COLOR, fg=DIM_TEXT,
                font=self.f_time,
            ).pack(side="left", padx=(8, 0))
            tk.Frame(section, bg=LINE_COLOR, height=1).pack(fill="x", pady=(6, 2))

            for action in completed:
                row = tk.Frame(section, bg=BG_COLOR)
                row.pack(fill="x", pady=2)
                tk.Label(
                    row, text="☑", bg=BG_COLOR,
                    fg=ACTION_STATUS_COLORS[ACTION_STATUS_DONE], font=self.f_body,
                ).pack(side="left", padx=(10, 8))

                body = tk.Frame(row, bg=BG_COLOR)
                body.pack(side="left", fill="both", expand=True)
                tk.Label(
                    body, text=action["content"], bg=BG_COLOR, fg=DIM_TEXT,
                    font=self.f_body, anchor="w", justify="left", wraplength=max_width,
                ).pack(fill="x")
                tk.Label(
                    body, text=action["completed_at"].strftime("%m-%d %H:%M"),
                    bg=BG_COLOR, fg=DIM_TEXT, font=self.f_time,
                ).pack(anchor="w", pady=(2, 0))

    def _complete_action(self, row: int) -> None:
        """未着手のアクションをクリックした時、その場で完了にする。"""
        try:
            complete_action(row)
        except Exception as e:
            print(f"❌ アクションの完了処理に失敗しました: {e}")
            return
        self.refresh()


def run() -> None:
    """dashboard.py 単体起動用のエントリポイント。"""
    root = tk.Tk()
    DashboardWindow(root)
    print("🚀 LKPT Dashboardを起動しました。")
    root.mainloop()


if __name__ == "__main__":
    run()
