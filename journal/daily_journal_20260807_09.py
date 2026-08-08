# -*- coding: utf-8 -*-
"""
daily_journal_20260807_09.py
学びジャーナル - ホットキー起動の入力ポップアップUI
Version: 0.29.0
"""

import ctypes
import os
import queue
import threading
import tkinter as tk
from ctypes import wintypes
from datetime import datetime, timedelta
from tkinter import font as tkfont

from storage import (
    ACTION_STATUS_PENDING,
    OTHER_SUBITEM_LABEL,
    add_action,
    append_lkpt_entry,
    append_time_log,
    complete_action,
    delete_time_log_row,
    get_actions,
    get_sub_item_master,
    get_tag_master,
    set_action_priority,
    update_time_log_row,
)

# ============================================================
# UI設定（ダークテーマ）
# ============================================================
BG_COLOR = "#1a1a2e"
ACCENT_COLOR = "#f4a6b8"
TEXT_COLOR = "#ffffff"
ENTRY_BG = "#0f3460"
PLACEHOLDER_COLOR = "#7c86a8"  # 入力前のプレースホルダ文字色（本文の白と区別する）
BUTTON_TEXT_COLOR = "#2b2b40"
CANCEL_BTN_BG = "#d8d8e6"
DASHBOARD_BTN_BG = "#dfeaf5"  # dashboard.pyのPERIOD_BTN_BGと揃えた配色
LKPT_TOGGLE_BG = "#cbb8f5"  # 「振り返りを書く」トグルの地色。CANCEL/DBのボタン
                             # 群とは異なる色相にして、視認性のある独立した
                             # チップとして目立たせる（他ボタンと被らないラベンダー）

# たまっているアクション一覧（MS To Do風の1行カード）の配色。
# ○チェックは左端・★優先は右端という参考画像のレイアウトに合わせて選定した
ACTION_CARD_BG = "#2a2a44"        # 各行の地色（BG_COLORより一段明るいダーク）
ACTION_CARD_DONE_FLASH = "#3a5a45"  # チェック直後に一瞬光らせる完了フィードバック色
ACTION_STAR_ON_COLOR = "#4a90e2"    # ★（優先）点灯時の色
ACTION_STAR_OFF_COLOR = "#7c86a8"   # ☆（未優先）の色
ACTION_CHECK_COLOR = "#9aa4c8"      # ○チェックアイコンの色
FORECAST_BTN_BG = "#cfe6dd"         # 雨（予測）ボタンの地色。DB(青系)・
                                     # 閉じる(灰)と区別が付く淡い緑寄り

# L/K/P/Tが何を意味するか毎回忘れてしまう問題への対応。
# 入力欄には薄字のプレースホルダを常設し、ラベル文字にホバーすると
# 正式な意味をツールチップで出す（両方あることで、操作しなくても
# ある程度察せるし、正確な定義も確認できる）
FIELD_PLACEHOLDERS = {
    "L": "今日わかったこと",
    "K": "続けたいこと",
    "P": "気になっている課題",
    "T": "次に挑戦したいこと",
}
FIELD_TOOLTIPS = {
    "L": "Learned - 今日の学び・気づき",
    "K": "Keep - 続けたいこと・良かったこと",
    "P": "Problem - 気になっている課題・懸念",
    "T": "Try - 次に挑戦したいこと",
}

QUICK_ACTION_PLACEHOLDER = "Enterで登録"
SUBITEM_FREE_PLACEHOLDER = "Enterで確定"

# タグを押した瞬間に記録するようになったため、押し間違いを追記ではなく
# 上書きで直せる猶予を設ける。この時間を過ぎたタグクリックは「訂正」ではなく
# 「今から別の作業を始めた」の意味として新しい行を追加する
AMEND_WINDOW_MINUTES = 5

# 記録後、何も操作しなければ自動的に閉じるまでの時間。
# 「閉じるために毎回クリックが要る」のを避けつつ、訂正の余地も残すための値。
# タグ/中項目を操作するたびに数え直し、LKPT欄やタスク欄に触れたら恒久的に止める
AUTOCLOSE_DELAY_MS = 3500
P_TO_ACTION_LABEL = "→ アクションにする"


class _Tooltip:
    """
    ウィジェットにホバーすると、少し遅延してから説明を吹き出しで出す
    簡易ツールチップ。Tkinter標準には無いため自前で実装する。
    """

    def __init__(self, widget: tk.Widget, text: str, delay_ms: int = 400):
        self.widget = widget
        self.text = text
        self.delay_ms = delay_ms
        self._after_id = None
        self._tip_window = None
        widget.bind("<Enter>", self._schedule, add="+")
        widget.bind("<Leave>", self._hide, add="+")

    def _schedule(self, event=None) -> None:
        self._after_id = self.widget.after(self.delay_ms, self._show)

    def _show(self) -> None:
        if self._tip_window is not None or not self.widget.winfo_exists():
            return
        x = self.widget.winfo_rootx() + self.widget.winfo_width() + 8
        y = self.widget.winfo_rooty()
        self._tip_window = tk.Toplevel(self.widget)
        self._tip_window.wm_overrideredirect(True)
        self._tip_window.wm_geometry(f"+{x}+{y}")
        self._tip_window.attributes("-topmost", True)
        tk.Label(
            self._tip_window, text=self.text, bg="#111122", fg=TEXT_COLOR,
            font=tkfont.Font(family="Yu Gothic UI", size=9),
            padx=8, pady=4, relief="solid", borderwidth=1,
        ).pack()

    def _hide(self, event=None) -> None:
        if self._after_id is not None:
            self.widget.after_cancel(self._after_id)
            self._after_id = None
        if self._tip_window is not None:
            self._tip_window.destroy()
            self._tip_window = None


def _lighten_color(hex_color: str, factor: float = 0.78) -> str:
    """hex_colorを白方向にfactor(0〜1)だけ明るくした16進色を返す。"""
    hex_color = hex_color.lstrip("#")
    if len(hex_color) == 3:
        hex_color = "".join(c * 2 for c in hex_color)
    factor = max(0.0, min(1.0, factor))
    r, g, b = (int(hex_color[i:i + 2], 16) for i in (0, 2, 4))
    r, g, b = (round(c + (255 - c) * factor) for c in (r, g, b))
    return f"#{r:02x}{g:02x}{b:02x}"


def _blend_toward(hex_color: str, target_hex: str, ratio: float) -> str:
    """
    hex_colorをtarget_hex方向にratio(0〜1)だけ混ぜた16進色を返す。
    大項目タグチップの未選択時の地色（タグの色をうっすら地色BG_COLORに
    混ぜたもの）を作るのに使う。サブ項目ボタン（常に無彩色）との
    見分けが付くよう、未選択の間もタグ固有の色味を残しておくため
    """
    hex_color = hex_color.lstrip("#")
    target_hex = target_hex.lstrip("#")
    if len(hex_color) == 3:
        hex_color = "".join(c * 2 for c in hex_color)
    ratio = max(0.0, min(1.0, ratio))
    r1, g1, b1 = (int(hex_color[i:i + 2], 16) for i in (0, 2, 4))
    r2, g2, b2 = (int(target_hex[i:i + 2], 16) for i in (0, 2, 4))
    r = round(r1 + (r2 - r1) * ratio)
    g = round(g1 + (g2 - g1) * ratio)
    b = round(b1 + (b2 - b1) * ratio)
    return f"#{r:02x}{g:02x}{b:02x}"


def _relative_luminance(hex_color: str) -> float:
    """WCAGの相対輝度を計算する（文字色のコントラスト判定に使う）。"""
    hex_color = hex_color.lstrip("#")
    if len(hex_color) == 3:
        hex_color = "".join(c * 2 for c in hex_color)

    def channel(c: int) -> float:
        c = c / 255
        return c / 12.92 if c <= 0.03928 else ((c + 0.055) / 1.055) ** 2.4

    r, g, b = (int(hex_color[i:i + 2], 16) for i in (0, 2, 4))
    return 0.2126 * channel(r) + 0.7152 * channel(g) + 0.0722 * channel(b)


def _contrast_ratio(hex1: str, hex2: str) -> float:
    """WCAGのコントラスト比（1〜21、大きいほど読みやすい）を計算する。"""
    l1, l2 = _relative_luminance(hex1), _relative_luminance(hex2)
    lighter, darker = max(l1, l2), min(l1, l2)
    return (lighter + 0.05) / (darker + 0.05)


def _readable_text_color(bg_hex: str) -> str:
    """
    背景色に対して、白(TEXT_COLOR)と濃紺(BUTTON_TEXT_COLOR)のどちらが
    コントラスト比が高いかを判定して返す。
    選択中バッジの文字色を、常に白固定にしていると、タグ色が明るい
    （黄・オレンジ系）場合に読みづらくなっていたための対応。
    タグの色が今後増えても、明るさによらず自動的に読みやすい方を選ぶ。
    """
    white_ratio = _contrast_ratio(bg_hex, TEXT_COLOR)
    dark_ratio = _contrast_ratio(bg_hex, BUTTON_TEXT_COLOR)
    return TEXT_COLOR if white_ratio >= dark_ratio else BUTTON_TEXT_COLOR


def _format_duration(minutes: int) -> str:
    """分数を"1h05m"形式の文字列に整形する。"""
    hours, mins = divmod(minutes, 60)
    if hours:
        return f"{hours}h{mins:02d}m"
    return f"{mins}m"


HOTKEY = "ctrl+shift+j"
VERSION = "0.29.0"

# ファイル名（daily_journal_yyyymmdd_NN.py）そのものがバージョン識別子を
# 兼ねる運用のため、ここに手で書いた文字列を置くと更新を忘れて古いまま
# 残る恐れがある。__file__から動的に取り出すことで、ファイルを
# リネームするだけで表示側も自動的に最新化される。
_FILE_STEM = os.path.splitext(os.path.basename(__file__))[0]
FILE_VERSION_LABEL = (
    _FILE_STEM[len("daily_journal_"):]
    if _FILE_STEM.startswith("daily_journal_")
    else _FILE_STEM
)

_trigger_queue = queue.Queue()

# ============================================================
# グローバルホットキー（Windows RegisterHotKey API）
# ------------------------------------------------------------
# keyboardライブラリの低レベルフックは、長時間稼働中に監視スレッドが
# 静かに停止してしまう不具合が疑われたため、OS標準のRegisterHotKey APIに
# 置き換えた。他アプリが同じ組み合わせを登録済みの場合は明示的に
# 失敗が分かるため、原因切り分けもしやすくなる。
# ============================================================
_MOD_ALT = 0x0001
_MOD_CONTROL = 0x0002
_MOD_SHIFT = 0x0004
_MOD_WIN = 0x0008
_MOD_NOREPEAT = 0x4000  # キー押しっぱなしでの連続発火を防ぐ
_WM_HOTKEY = 0x0312
_HOTKEY_ID = 1
_ERROR_HOTKEY_ALREADY_REGISTERED = 1409

_MODIFIER_FLAGS = {
    "ctrl": _MOD_CONTROL,
    "alt": _MOD_ALT,
    "shift": _MOD_SHIFT,
    "win": _MOD_WIN,
}


def _parse_hotkey(hotkey: str) -> tuple:
    """"ctrl+shift+j"のような文字列を(修飾キーフラグ, 仮想キーコード)に変換する。"""
    *modifier_names, key = [part.strip().lower() for part in hotkey.split("+")]
    modifiers = 0
    for name in modifier_names:
        modifiers |= _MODIFIER_FLAGS[name]
    return modifiers, ord(key.upper())


def queue_popup_trigger() -> None:
    """
    外部モジュール（scheduler.py等）からポップアップ表示をトリガーするための
    公開関数。ホットキー検知と同じキューを使うため、スレッドセーフに扱える。
    """
    _trigger_queue.put(True)


def _hotkey_listener_loop(hotkey: str) -> None:
    """
    専用スレッドでRegisterHotKeyを登録し、WM_HOTKEYメッセージを待ち受ける。
    hWnd=NoneでRegisterHotKeyを呼ぶ場合、登録した同一スレッドでGetMessage
    ループを回す必要があるため、登録とメッセージ待受を同じ関数内で行う。
    """
    user32 = ctypes.windll.user32
    modifiers, vk = _parse_hotkey(hotkey)

    if not user32.RegisterHotKey(None, _HOTKEY_ID, modifiers | _MOD_NOREPEAT, vk):
        error_code = ctypes.GetLastError()
        hint = (
            "（他のアプリが同じキーの組み合わせを登録済みの可能性があります）"
            if error_code == _ERROR_HOTKEY_ALREADY_REGISTERED
            else ""
        )
        print(f"❌ ホットキー登録に失敗しました({hotkey})。エラーコード: {error_code} {hint}")
        return

    print(f"🔗 グローバルホットキーを登録しました: {hotkey}")

    msg = wintypes.MSG()
    try:
        while user32.GetMessageW(ctypes.byref(msg), None, 0, 0) != 0:
            if msg.message == _WM_HOTKEY:
                _trigger_queue.put(True)
                print("⌨️ ホットキーを検知しました。")
    finally:
        user32.UnregisterHotKey(None, _HOTKEY_ID)


def register_global_hotkey(hotkey: str = HOTKEY) -> None:
    """グローバルホットキーを登録する（専用スレッドでメッセージループを待受）。"""
    thread = threading.Thread(target=_hotkey_listener_loop, args=(hotkey,), daemon=True)
    thread.start()


class PopupWindow:
    """タグ選択＋中項目＋（任意で）LKPT入力のポップアップウィンドウ。"""

    def __init__(self, root: tk.Tk):
        self.root = root
        self.window = None
        self.dashboard_window = None
        self.selected_tag = None
        self.selected_subitem = None
        self.subitem_master = {}
        self.subitem_buttons = {}
        self.subitem_free_var = tk.StringVar()
        self.subitem_free_entry = None
        self.lkpt_expanded = False
        self.l_var = tk.StringVar()
        self.k_var = tk.StringVar()
        self.p_var = tk.StringVar()
        self.t_var = tk.StringVar()
        self.tag_chips = {}
        self.field_entries = {}
        self.selected_label = None
        self.pending_actions = []
        self.action_list_canvas = None
        self.action_list_inner = None
        self._popup_width = 300
        self.collapsed_height = None
        self.expanded_height = None
        self._themed_bg_widgets = []
        self._themed_fg_widgets = []

        # 直前に書き込んだTimeLog行を覚えておくための一式。
        # PopupWindowはアプリ起動中ずっと同じインスタンスが使われるため、
        # これらをshow()でリセットしないことで「一度閉じた後に開き直しても
        # AMEND_WINDOW_MINUTES以内なら直前の記録を修正できる」を実現する
        self.timelog_row = None          # int | None  Excelの行番号
        self.timelog_start = None        # datetime    楽観ロックの照合用
        self.timelog_end = None          # datetime    同上
        self.timelog_kind = None         # "anchor" | "interval" | "gap"
        self.timelog_minutes = None      # int | None
        self.timelog_tag = None          # str  修正時に引き継ぐ
        self.timelog_sub_item = ""       # str  同上
        self.timelog_committed_at = None # datetime  修正可能時間の起点
        self.timelog_busy = False        # 保存中の多重クリック抑止

        self.status_label = None         # 記録結果を出す行（旧time_preview_label）
        self.autoclose_after_id = None   # 自動クローズのタイマーID
        self.autoclose_cancelled = False # LKPT/タスクに触れたらTrue（恒久停止）
        self.lkpt_tag_var = None         # LKPTに選択中タグを引き継ぐか
        self.lkpt_tag_check = None       # 上記のチェックボックス本体
        self.forecast_btn = None         # 雨（予測）への入口ボタン
        self.forecast_seed_label = None  # タスク登録直後の「雨にする」導線

    def _register_themed(self, widget, bg: bool = True, fg: bool = False) -> None:
        """タグ選択時の背景/文字色切替に追従させるウィジェットを登録する。"""
        if bg:
            self._themed_bg_widgets.append(widget)
        if fg:
            self._themed_fg_widgets.append(widget)

    def _apply_theme(self, bg: str, fg: str) -> None:
        """登録済みウィジェットの背景色・文字色をまとめて切り替える。"""
        for widget in self._themed_bg_widgets:
            widget.configure(bg=bg)
        for widget in self._themed_fg_widgets:
            widget.configure(fg=fg)

    def show(self) -> None:
        """ポップアップを表示する。既に表示中の場合は前面化のみ行う。"""
        print("🪟 show()呼び出しを検知しました。")
        if self.window is not None and self.window.winfo_exists():
            self.window.lift()
            self.window.focus_force()
            return

        self.selected_tag = None
        self.selected_subitem = None
        self.subitem_buttons = {}
        self.subitem_free_var.set("")
        self.subitem_free_entry = None
        self.lkpt_expanded = False
        for var in (self.l_var, self.k_var, self.p_var, self.t_var):
            var.set("")
        self.tag_chips = {}
        self.field_entries = {}
        self._themed_bg_widgets = []
        self._themed_fg_widgets = []
        # 自動クローズの状態だけは毎回まっさらにする。
        # timelog_* は意図的にリセットしない（閉じた後の訂正を許すため）
        self.autoclose_after_id = None
        self.autoclose_cancelled = False
        self.timelog_busy = False

        try:
            tags = get_tag_master()
        except Exception as e:
            print(f"❌ タグ取得に失敗しました（Excelロック等の可能性）: {e}")
            tags = []

        try:
            self.subitem_master = get_sub_item_master()
        except Exception as e:
            print(f"❌ 中項目マスタ取得に失敗しました: {e}")
            self.subitem_master = {}

        self.pending_actions = self._fetch_pending_actions()

        self.window = tk.Toplevel(self.root)
        self.window.title(f"LKPT - {FILE_VERSION_LABEL}")
        self.window.configure(bg=BG_COLOR)
        self.window.attributes("-topmost", True)
        self.window.resizable(True, True)
        self._register_themed(self.window)

        row_height = 34
        # タイトル・selected_label(太字バッジ)・中項目ボタン・時間帯プレビュー・
        # LKPTトグルボタン・登録/キャンセル/DBボタン・＋アクション欄・
        # たまっているアクション一覧（固定4行分）・余白の合計目安
        # （LKPT欄自体は畳んだ状態の高さで、展開時はlkpt_block_heightを追加する）
        chrome_height = 650
        lkpt_block_height = 4 * 40 + 20
        self._popup_width = 300
        self.collapsed_height = max(360, chrome_height + row_height * len(tags))
        self.collapsed_height = min(
            self.collapsed_height, self.window.winfo_screenheight() - 100,
        )
        # collapsed_height同様、画面の縦幅を超えないようクランプする。
        # クランプにより画面が低い環境ではLKPT欄の下部が見切れる場合が
        # あるが、ウィンドウはresizable=Trueのため手動リサイズで対応できる
        self.expanded_height = min(
            self.collapsed_height + lkpt_block_height,
            self.window.winfo_screenheight() - 40,
        )

        screen_w = self.window.winfo_screenwidth()
        screen_h = self.window.winfo_screenheight()
        x = screen_w - self._popup_width - 20
        # 上端(y)は、畳んだ状態の高さではなく「展開した時の高さ」を基準に
        # 決める。畳んだ状態基準のまま置くと、LKPT欄を開いた時に画面下端を
        # はみ出さないよう上端を押し上げる必要が生じ、「開くと位置がずれる」
        # 原因になっていたため。あらかじめ展開後の高さぶんの余白を
        # 確保しておくことで、開閉時に上端を動かさずに済むようにする
        y = max(10, screen_h - self.expanded_height - 80)
        self.window.geometry(f"{self._popup_width}x{self.collapsed_height}+{x}+{y}")
        min_width = int(self._popup_width * 2 / 3)
        self.window.minsize(min_width, self.collapsed_height)

        title_font = tkfont.Font(family="Yu Gothic UI", size=12, weight="bold")
        label = tk.Label(
            self.window,
            text="📝 LKPT",
            bg=BG_COLOR,
            fg=TEXT_COLOR,
            font=title_font,
        )
        label.pack(pady=(12, 0))
        self._register_themed(label, bg=True, fg=True)

        # 起動しているファイルが最新版かどうかを、開くたびに一目で
        # 確認できるようにする（ファイル名から動的に取り出すため、
        # 新しいバージョンのファイルを追加するだけで自動的に更新される）
        version_font = tkfont.Font(family="Consolas", size=8)
        version_label = tk.Label(
            self.window, text=FILE_VERSION_LABEL, bg=BG_COLOR, fg=PLACEHOLDER_COLOR,
            font=version_font,
        )
        version_label.pack(pady=(0, 6))
        self._register_themed(version_label, bg=True)

        # 「選択中」バッジはタグ一覧の上に固定表示する（以前は下に表示して
        # いたが、タグ一覧を見てから選択結果を確認するまでの視線移動を
        # 減らすため、選ぶ前から見えている位置に変更した）
        selected_font = tkfont.Font(family="Yu Gothic UI", size=13, weight="bold")
        self.selected_label = tk.Label(
            self.window, text="未選択", bg=BG_COLOR, fg=ACCENT_COLOR,
            font=selected_font, padx=14, pady=6,
        )
        self.selected_label.pack(pady=(4, 4))
        # 選択中バッジは一括テーマ切替(_apply_theme)の対象外にし、_select_tag内で
        # タグの生の色を直接背景に当てて目立たせる（意図的に_register_themedを呼ばない）。
        # 文字色は固定せず、_select_tag内でタグの色に応じて白/濃紺を選ぶ。

        # タグ一覧：縦積みの行ではなく、2列のチップボタンで横に並べる
        # （「タグを横に並べませんか」との要望）。チップは未選択の間も
        # タグの色をうっすら地色に混ぜ、同色の縁取りを付ける。これにより、
        # 常に無彩色（CANCEL_BTN_BG）のサブ項目ボタンとの見分けが一目で
        # 付くようにしている（当初の「サブ項目と区別が付かない」という
        # 指摘への対応）
        self.tag_frame = tk.Frame(self.window, bg=BG_COLOR)
        self.tag_frame.pack(pady=(0, 0), fill="x", padx=20)
        self._register_themed(self.tag_frame)

        chip_font = tkfont.Font(family="Yu Gothic UI", size=10, weight="bold")
        dot_font = tkfont.Font(family="Yu Gothic UI", size=11)
        tag_cols = 2

        # 以前はチップを「行ごとのFrame」に分けてpack(side=left, expand=True)
        # していたが、タグ名の文字幅が行ごとに違う（"R19"と"JP Site"等）ため、
        # 行ごとに独立してexpandする形だと列の境界が行によってズレて
        # しまっていた（「タグのエリアが一つづつずれている」不具合）。
        # grid()＋columnconfigure(uniform=...)にして、全行・全列を横断して
        # 完全に同じ列幅になるよう強制する（登録/キャンセル/DBの3ボタンで
        # 使っているuniform="btn"と同じ仕組み）
        for col in range(tag_cols):
            self.tag_frame.columnconfigure(col, weight=1, uniform="tagcol")

        for i, (tag_name, color_code) in enumerate(tags):
            row, col = divmod(i, tag_cols)

            chip = tk.Frame(
                self.tag_frame, bg=BG_COLOR, cursor="hand2",
                highlightthickness=1,
            )
            chip.grid(row=row, column=col, sticky="ew", padx=3, pady=3)
            inner = tk.Frame(chip, bg=BG_COLOR)
            inner.pack(pady=7)
            dot = tk.Label(inner, text="●", font=dot_font, fg=color_code, bg=BG_COLOR)
            dot.pack(side="left", padx=(10, 4))
            name_label = tk.Label(
                inner, text=tag_name, font=chip_font, bg=BG_COLOR, fg=TEXT_COLOR,
            )
            name_label.pack(side="left", padx=(0, 10))

            for widget in (chip, inner, dot, name_label):
                widget.bind(
                    "<Button-1>",
                    lambda e, t=tag_name, c=color_code: self._on_tag_clicked(t, c),
                )

            # チップ自身の配色はテーマ一括切替(_apply_theme)の対象外にし、
            # _select_tag内で個別に塗り分ける（タグごとに異なる色を保つため、
            # 全ウィジェット一律の pale_bg/BUTTON_TEXT_COLOR には乗せない）
            self.tag_chips[tag_name] = {
                "chip": chip, "inner": inner, "dot": dot, "name": name_label,
                "color": color_code,
            }
            self._recolor_tag_chip(tag_name, selected=False)

        # 中項目（何をしていたか）ボタンと時間帯プレビューは、タググループ
        # 全体（2列チップの並び）の下にまとめて表示する。以前は選んだタグの
        # 行の直後に割り込ませていたが、「大項目タグをまとめて上に揃えた
        # 方がスッキリする」との指摘を受け、タグ一覧を最後まで表示してから
        # その下に固定表示する構成に変更した
        self.tag_detail_frame = tk.Frame(self.window, bg=BG_COLOR)
        self._register_themed(self.tag_detail_frame)

        self.subitem_frame = tk.Frame(self.tag_detail_frame, bg=BG_COLOR)
        self.subitem_frame.pack(pady=(6, 0), fill="x")
        self._register_themed(self.subitem_frame)
        self.subitem_font = tkfont.Font(family="Yu Gothic UI", size=10)
        self.entry_font = tkfont.Font(family="Yu Gothic UI", size=10)

        # 記録結果を表示する行。以前は「〜として記録されます」という未来形の
        # プレビューだったが、タグを押した時点で実際に記録されるようになった
        # ため、過去形＋「修正できる」ことの案内に変えた。
        # tag_detail_frameの中ではなくウィンドウ直下に置く。閉じた直後に開き
        # 直した場合、タグ未選択でも「直前の記録」を示す必要があるため。
        # tag_detail_frameはafter=tag_frameでpackされるので、この行より上に入る
        status_font = tkfont.Font(family="Consolas", size=9)
        self.status_label = tk.Label(
            self.window, text="", bg=BG_COLOR, fg=PLACEHOLDER_COLOR,
            font=status_font, wraplength=self._popup_width - 30, justify="center",
        )
        self.status_label.pack(pady=(4, 2), padx=15, fill="x")
        self._register_themed(self.status_label, bg=True)
        self._update_status_label()

        # LKPT（振り返り）は既定では畳んでおき、書きたい時だけ開く。
        # 押すたびに開閉し、ウィンドウの高さもそれに合わせて変える。
        # 以前は背景をウィンドウと同色にした「テキストリンク風」だったが、
        # ①タグ選択後の淡い背景に対してコントラスト不足になる、②地味すぎて
        # 目立たない、の両方の指摘を受けたため、独立した地色（LKPT_TOGGLE_BG）を
        # 持つ丸みのあるチップ状のボタンに変更した。タグの選択状態（テーマ）に
        # 関わらず常に同じ配色で見えるため、①の根本対応にもなっている
        toggle_font = tkfont.Font(family="Yu Gothic UI", size=10, weight="bold")
        self.lkpt_toggle_btn = tk.Button(
            self.window, text="LKPT",
            bg=LKPT_TOGGLE_BG, fg=BUTTON_TEXT_COLOR,
            activebackground=ACCENT_COLOR, activeforeground=BUTTON_TEXT_COLOR,
            relief="flat", bd=0, highlightthickness=0, font=toggle_font, cursor="hand2",
            padx=14, pady=5,
            command=self._toggle_lkpt,
        )
        # 直前の要素（時間帯プレビュー）との間に固定の間隔を空けて配置する。
        # 横幅は他のボタン列（登録/キャンセル/DB）やタグ一覧・入力欄と
        # 揃うよう、中央寄せの小さなピルではなくfill="x"で全幅にする
        self.lkpt_toggle_btn.pack(pady=(14, 6), padx=20, fill="x")

        # LKPT欄自体はここで組み立てておくが、既定では畳んだ状態にするため
        # pack()はまだ呼ばない（_toggle_lkpt()で開いた時に初めてpackする）
        self._lkpt_frame = tk.Frame(self.window, bg=BG_COLOR)
        self._register_themed(self._lkpt_frame)

        lkpt_label_font = tkfont.Font(family="Yu Gothic UI", size=10, weight="bold")
        for field_label, var in (
            ("L", self.l_var), ("K", self.k_var), ("P", self.p_var), ("T", self.t_var),
        ):
            field_row = tk.Frame(self._lkpt_frame, bg=BG_COLOR)
            field_row.pack(fill="x", pady=2)
            self._register_themed(field_row)

            field_label_widget = tk.Label(
                field_row, text=field_label, font=lkpt_label_font,
                bg=BG_COLOR, fg=TEXT_COLOR, width=2,
            )
            field_label_widget.pack(side="left")
            self._register_themed(field_label_widget, bg=True, fg=True)
            _Tooltip(field_label_widget, FIELD_TOOLTIPS[field_label])

            field_entry = tk.Entry(
                field_row,
                bg=ENTRY_BG,
                fg=PLACEHOLDER_COLOR,
                insertbackground=TEXT_COLOR,
                relief="flat",
                font=self.entry_font,
                textvariable=var,
            )
            field_entry.pack(side="left", ipady=3, fill="x", expand=True)
            field_entry.insert(0, FIELD_PLACEHOLDERS[field_label])
            field_entry.bind(
                "<FocusIn>",
                lambda e, ent=field_entry, ph=FIELD_PLACEHOLDERS[field_label]:
                    self._on_lkpt_focus_in(ent, ph),
            )
            field_entry.bind(
                "<FocusOut>",
                lambda e, ent=field_entry, ph=FIELD_PLACEHOLDERS[field_label]:
                    self._restore_placeholder(ent, ph),
            )
            # どの欄でEnterを押しても、L/K/P/T全体を1件として記録する
            # （4件に分かれないよう、まとめて1行にする）
            field_entry.bind("<Return>", lambda e: self._commit_lkpt())
            self.field_entries[field_label] = field_entry

            # P（課題）だけ、そのままアクションアイテムとしても登録できる
            # チェックを添える。登録時にPの内容が空でなければ、チェックが
            # 入っている場合に限りActionsシートへも1件追加する
            if field_label == "P":
                self.p_to_action_var = tk.BooleanVar(value=False)
                check_row = tk.Frame(self._lkpt_frame, bg=BG_COLOR)
                check_row.pack(fill="x", pady=(0, 4))
                self._register_themed(check_row)
                p_to_action_check = tk.Checkbutton(
                    check_row, text=P_TO_ACTION_LABEL, variable=self.p_to_action_var,
                    bg=BG_COLOR, fg=PLACEHOLDER_COLOR, selectcolor=ENTRY_BG,
                    activebackground=BG_COLOR, activeforeground=PLACEHOLDER_COLOR,
                    font=tkfont.Font(family="Yu Gothic UI", size=9), bd=0,
                    highlightthickness=0, cursor="hand2",
                )
                p_to_action_check.pack(side="left", padx=(28, 0))
                self._register_themed(p_to_action_check, bg=True, fg=True)

        # LKPTはタグ無しでも記録できるが、TIMEのタグを選択中なら、それを
        # 引き継いでおくと後からタグ別に振り返れる。ただし黙って引き継ぐと
        # 「15時に書いた朝の会議の振り返り」に15時のタグが付く取り違えが
        # 起きるため、引き継ぐかどうかを明示して切り替えられるようにする
        self.lkpt_tag_var = tk.BooleanVar(value=True)
        self.lkpt_tag_check = tk.Checkbutton(
            self._lkpt_frame, text="", variable=self.lkpt_tag_var,
            bg=BG_COLOR, fg=PLACEHOLDER_COLOR, selectcolor=ENTRY_BG,
            activebackground=BG_COLOR, activeforeground=PLACEHOLDER_COLOR,
            font=tkfont.Font(family="Yu Gothic UI", size=9), bd=0,
            highlightthickness=0, cursor="hand2", anchor="w",
        )
        self.lkpt_tag_check.pack(fill="x", pady=(2, 0), padx=(28, 0))
        self._register_themed(self.lkpt_tag_check, bg=True, fg=True)
        self._update_lkpt_tag_check()

        self.btn_frame = tk.Frame(self.window, bg=BG_COLOR)
        self.btn_frame.pack(pady=6, padx=20, fill="x")
        self._register_themed(self.btn_frame)
        # 「登録」ボタンは廃止した。TIMEはタグ/中項目を押した時点で、
        # LKPTとタスクはEnterで、それぞれ即記録されるようになったため、
        # このボタンに残された仕事が無くなった。
        # 代わりに雨（予測）への入口を足して3列にする
        for col in range(3):
            # uniform="btn"を指定して均等分割することで、
            # ボタンのテキスト長に関わらず幅を完全に揃える
            # （pack()のexpand=Trueだけでは自然サイズの差が残ってしまうため）
            self.btn_frame.columnconfigure(col, weight=1, uniform="btn")

        # 3つとも同じfont・pady（=同じボタンの箱の大きさ）に統一する。
        # 以前DBだけフォントを大きくして揃えようとしたが、行の高さが一番
        # 大きいDBに引っ張られてかえってバランスが崩れたため撤回した経緯がある
        icon_btn_font = tkfont.Font(family="Yu Gothic UI", size=13)

        # 記録済みの内容は取り消さないため「キャンセル」ではなく「閉じる」。
        # TIMEを押した時点で既に書き込まれているので、キャンセルと名乗るのは
        # 嘘になる（取り消したい場合は同じタグをもう一度押す）
        close_btn = tk.Button(
            self.btn_frame,
            text="✖",
            font=icon_btn_font,
            bg=CANCEL_BTN_BG,
            fg=BUTTON_TEXT_COLOR,
            activebackground=ACCENT_COLOR,
            relief="flat",
            pady=4,
            command=self._close,
        )
        close_btn.grid(row=0, column=0, sticky="ew", padx=3)

        # 雨（予測）の一覧・入力への入口。期日が来た未記入がある間は
        # 件数を出して、こちらから気づけるようにする
        self.forecast_btn = tk.Button(
            self.btn_frame,
            text="🌧️",
            font=icon_btn_font,
            bg=FORECAST_BTN_BG,
            fg=BUTTON_TEXT_COLOR,
            activebackground=ACCENT_COLOR,
            relief="flat",
            pady=4,
            command=self._open_forecasts,
        )
        self.forecast_btn.grid(row=0, column=1, sticky="ew", padx=3)
        self._update_forecast_button()

        dashboard_btn = tk.Button(
            self.btn_frame,
            text="📊",
            font=icon_btn_font,
            bg=DASHBOARD_BTN_BG,
            fg=BUTTON_TEXT_COLOR,
            activebackground=ACCENT_COLOR,
            relief="flat",
            pady=4,
            command=self._open_dashboard,
        )
        dashboard_btn.grid(row=0, column=2, sticky="ew", padx=3)

        # たまっているアクション一覧。MS To Doの参考画像に合わせ、○チェックは
        # 左端・★優先は右端、行は横幅いっぱいの1行カードとして縦に積む。
        # ポップアップの高さがアクション件数に比例して伸び続けないよう、
        # visible_rows件分の高さで固定し、それ以上は縦スクロールで見せる。
        # アクションを入力してもこのUI自体は閉じない前提のため、複数件
        # 書き留めた直後でもこの一覧がその場で増えていく様子が見える
        action_section_label = tk.Label(
            self.window, text="タスク", bg=BG_COLOR,
            fg=PLACEHOLDER_COLOR, font=tkfont.Font(family="Yu Gothic UI", size=9, weight="bold"),
            anchor="w",
        )
        action_section_label.pack(pady=(8, 2), padx=20, fill="x")
        self._register_themed(action_section_label, bg=True)

        # 「今すぐ書き留めたい」思いつきを、タグ選択の有無に関わらずいつでも
        # 記録できる常設欄。登録ボタンとは独立しており、追加してもポップアップは
        # 閉じない（1回開いている間に複数件メモできるようにするため）。
        # 「タスク」の見出しの直下、一覧の直前に配置する
        quick_action_font = tkfont.Font(family="Yu Gothic UI", size=10)
        quick_action_frame = tk.Frame(self.window, bg=BG_COLOR)
        quick_action_frame.pack(pady=(0, 2), padx=20, fill="x")
        self._register_themed(quick_action_frame)

        self.quick_action_var = tk.StringVar()
        self.quick_action_entry = tk.Entry(
            quick_action_frame, textvariable=self.quick_action_var, bg=ENTRY_BG,
            fg=PLACEHOLDER_COLOR, insertbackground=TEXT_COLOR, relief="flat",
            font=quick_action_font,
        )
        self.quick_action_entry.pack(side="left", ipady=3, fill="x", expand=True)
        self.quick_action_entry.insert(0, QUICK_ACTION_PLACEHOLDER)
        self.quick_action_entry.bind(
            "<FocusIn>",
            lambda e: (
                self._cancel_autoclose(permanent=True),
                self._clear_placeholder(
                    self.quick_action_entry, QUICK_ACTION_PLACEHOLDER,
                ),
            ),
        )
        self.quick_action_entry.bind(
            "<FocusOut>",
            lambda e: self._restore_placeholder(
                self.quick_action_entry, QUICK_ACTION_PLACEHOLDER,
            ),
        )
        # 登録は＋ボタンではなく、基本的にEnterキーで行う仕様にする
        # （＋ボタンは廃止。プレースホルダに「Enterで登録」と明示することで
        # 押せるボタンが無くても操作方法が伝わるようにする）
        self.quick_action_entry.bind("<Return>", lambda e: self._add_quick_action())

        feedback_font = tkfont.Font(family="Yu Gothic UI", size=9)
        self.quick_action_feedback_label = tk.Label(
            self.window, text="", bg=BG_COLOR, fg=ACCENT_COLOR, font=feedback_font,
        )
        self.quick_action_feedback_label.pack(pady=(0, 2))
        self._register_themed(self.quick_action_feedback_label, bg=True, fg=True)

        # タスク登録直後にだけ出す「雨にする」導線（既定では非表示）
        self.forecast_seed_label = tk.Label(
            self.window, text="", bg=BG_COLOR, fg=ACCENT_COLOR,
            font=tkfont.Font(family="Yu Gothic UI", size=9, underline=True),
            cursor="hand2",
        )
        self._register_themed(self.forecast_seed_label, bg=True, fg=True)

        action_scroll_outer = tk.Frame(self.window, bg=BG_COLOR)
        action_scroll_outer.pack(pady=(0, 10), padx=20, fill="both", expand=True)
        self._register_themed(action_scroll_outer)

        visible_rows = 4
        self.action_row_height = 40
        self.action_list_canvas = tk.Canvas(
            action_scroll_outer, bg=BG_COLOR,
            height=self.action_row_height * visible_rows, highlightthickness=0,
        )
        action_vbar = tk.Scrollbar(
            action_scroll_outer, orient="vertical", command=self.action_list_canvas.yview,
        )
        self.action_list_canvas.configure(yscrollcommand=action_vbar.set)
        # スクロールバーを先に、幅固定(fill="y"のみ)でpackしてから、
        # 残りの領域いっぱいにCanvas(fill="both", expand=True)をpackする。
        # 逆順（Canvasを先にexpand=Trueでpack）だと、Canvasがcavityを
        # 全て使い切ってしまい、後からpackするスクロールバーの幅が
        # 確保されず実質見えなくなる不具合があったため、この順序にしている
        action_vbar.pack(side="right", fill="y")
        self.action_list_canvas.pack(side="left", fill="both", expand=True)
        self._register_themed(self.action_list_canvas, bg=True)

        self.action_list_inner = tk.Frame(self.action_list_canvas, bg=BG_COLOR)
        self._action_list_window = self.action_list_canvas.create_window(
            (0, 0), window=self.action_list_inner, anchor="nw",
        )
        self.action_list_inner.bind(
            "<Configure>",
            lambda e: self.action_list_canvas.configure(
                scrollregion=self.action_list_canvas.bbox("all"),
            ),
        )
        self.action_list_canvas.bind(
            "<Configure>",
            lambda e: self.action_list_canvas.itemconfig(self._action_list_window, width=e.width),
        )
        self._register_themed(self.action_list_inner, bg=True)

        # マウスホイールでもスクロールできるようにする（スクロールバーを
        # 見つけてドラッグしなくても、一覧の上にカーソルを置いてホイールを
        # 回すだけで見られるようにするための補助）。個々の行にも
        # _render_action_rows()内で同じハンドラを紐づけている
        self.action_list_canvas.bind("<MouseWheel>", self._on_action_list_mousewheel)
        self.action_list_inner.bind("<MouseWheel>", self._on_action_list_mousewheel)

        self._render_action_rows()

        self.window.bind("<Escape>", lambda e: self._close())
        # Windowsが Alt キー解放をシステムメニュー呼び出しと誤認識し、
        # ウィンドウが一瞬で背面に回る不具合を防ぐための対策
        self.window.bind("<Alt-KeyPress>", lambda e: "break")
        self.window.bind("<Alt-KeyRelease>", lambda e: "break")

        # Windowsは「直前にユーザー入力の無いプロセス」からの前面化要求を
        # ブロックすることがある（タイマー発火時に顕著）。topmost属性の
        # 解除→再設定を挟むことでこの制限を回避する。
        self.window.attributes("-topmost", True)
        self.window.deiconify()
        self.window.lift()
        self.window.after(50, self._force_foreground)

    def _force_foreground(self) -> None:
        """Windowsのフォアグラウンド制限を回避しつつウィンドウを前面化する。"""
        try:
            self.window.attributes("-topmost", False)
            self.window.attributes("-topmost", True)
            self.window.lift()
            self.window.focus_force()
            print("🪟 ポップアップを前面表示しました。")
        except tk.TclError as e:
            print(f"⚠️ 前面表示に失敗しました: {e}")

    def _resize_window(self, height: int) -> None:
        """
        LKPT欄の開閉に合わせてウィンドウの高さを変える。幅は保つ。
        以前は下端を画面下端付近に固定していたため、開閉のたびに上端(y)が
        動き、「アコーディオンを開くとLKPTの位置がずれる」という指摘が
        あった。上端(y)は現在の位置のまま変えず、下端だけが動く（開けば
        下に伸び、閉じれば下から縮む）ようにする。ただし、それによって
        下端が画面からはみ出す場合に限り、はみ出さない範囲でyを上にずらす
        （画面が低い環境向けのフォールバック）。
        """
        x = self.window.winfo_x()
        y = self.window.winfo_y()
        screen_h = self.window.winfo_screenheight()
        if y + height > screen_h - 10:
            y = max(10, screen_h - height - 10)
        self.window.geometry(f"{self._popup_width}x{height}+{x}+{y}")
        self.window.minsize(int(self._popup_width * 2 / 3), height)

    def _toggle_lkpt(self) -> None:
        """LKPT（振り返り）欄の開閉を切り替える。"""
        # 振り返りを書き始める＝しばらく操作が続くということなので、
        # 自動クローズは恒久的に止める（書いている途中で閉じないように）
        self._cancel_autoclose(permanent=True)
        self.lkpt_expanded = not self.lkpt_expanded
        if self.lkpt_expanded:
            self._lkpt_frame.pack(pady=(4, 8), padx=20, fill="x", before=self.btn_frame)
            self.lkpt_toggle_btn.config(text="LKPT")
            self._resize_window(self.expanded_height)
        else:
            self._lkpt_frame.pack_forget()
            self.lkpt_toggle_btn.config(text="LKPT")
            self._resize_window(self.collapsed_height)

    def _recolor_tag_chip(self, tag_name: str, selected: bool) -> None:
        """
        タグチップ1個分の配色を選択状態に応じて塗り分ける。
        選択中: タグの色をそのまま地色に、縁取りも同色で少し太くする。
        未選択: タグの色をうっすら地色BG_COLORに混ぜたものを地色にし、
        縁取りは同色のまま細くする（サブ項目ボタンの無彩色と区別するため、
        未選択でも色味を完全には消さない）。
        """
        widgets = self.tag_chips[tag_name]
        color = widgets["color"]
        if selected:
            bg = color
            fg = _readable_text_color(color)
            # 選択中はチップの地色がタグの生の色そのものになるため、
            # ドットの文字色も生の色のままだと地色と完全に同化して
            # 見えなくなってしまっていた（バグ）。文字色と同じ
            # 読みやすい色に切り替えて、選択中も視認できるようにする
            dot_fg = fg
            border_w = 2
        else:
            bg = _blend_toward(color, BG_COLOR, 0.72)
            fg = TEXT_COLOR
            # 未選択時は地色が暗く濁った色になるため、タグの生の色のままでも
            # 十分にコントラストが付く
            dot_fg = color
            border_w = 1
        widgets["chip"].config(bg=bg, highlightbackground=color, highlightcolor=color, highlightthickness=border_w)
        widgets["inner"].config(bg=bg)
        widgets["dot"].config(bg=bg, fg=dot_fg)
        widgets["name"].config(bg=bg, fg=fg)

    # ------------------------------------------------------------------
    # TIMEの記録（タグを押した時点で即記録し、以降のクリックは上書き修正）
    # ------------------------------------------------------------------

    def _timelog_token_live(self) -> bool:
        """
        直前に書いたTimeLog行を「押し間違いの修正」として上書きしてよいかを返す。
        AMEND_WINDOW_MINUTES以内であることを条件にする。ウィンドウを開いたまま
        放置して1時間後にタグを押した場合、それは訂正ではなく「今から別の作業を
        始めた」の意味なので、期限を切らないと古い記録を書き換えてしまう。
        """
        if self.timelog_row is None or self.timelog_committed_at is None:
            return False
        elapsed = datetime.now() - self.timelog_committed_at
        return elapsed <= timedelta(minutes=AMEND_WINDOW_MINUTES)

    def _forget_timelog_token(self) -> None:
        """保持している行番号を破棄する（以降は追記に切り替わる）。"""
        self.timelog_row = None
        self.timelog_start = None
        self.timelog_end = None
        self.timelog_kind = None
        self.timelog_minutes = None
        self.timelog_tag = None
        self.timelog_sub_item = ""
        self.timelog_committed_at = None

    def _commit_or_amend_timelog(self, tag: str, sub_item: str) -> None:
        """
        TIMEを記録する。直前に書いた行が修正可能なうちは追記せず上書きする。
        これが無いと、誤ったタグをクリック→正しいタグをクリック とした時に
        「誤ったタグが実作業時間を保持し、訂正のクリックは0分の行になる」
        というデータの取り違えが起きる。
        """
        if self.timelog_busy:
            return
        self.timelog_busy = True
        try:
            # 保存はOneDriveのロック待ちで数秒固まることがあるため、
            # 何が起きているかを先に出しておく
            self.status_label.config(text="⏱ 記録中…")
            self.status_label.update_idletasks()

            if self._timelog_token_live():
                ok = update_time_log_row(
                    self.timelog_row, tag, sub_item,
                    expect_start=self.timelog_start, expect_end=self.timelog_end,
                )
                if ok:
                    self.timelog_tag = tag
                    self.timelog_sub_item = sub_item
                    self._update_status_label()
                    self._schedule_autoclose()
                    return
                # 行番号が当てにならなくなっている（Excelを直接編集した等）。
                # 破棄して追記に切り替える
                self._forget_timelog_token()

            ok, kind, minutes, row = append_time_log(tag, sub_item)
            if not ok:
                self.status_label.config(
                    text="⚠️ 記録に失敗しました（退避保存を確認してください）"
                )
                return
            self.timelog_row = row
            self.timelog_kind = kind
            self.timelog_minutes = minutes
            self.timelog_tag = tag
            self.timelog_sub_item = sub_item
            self.timelog_committed_at = datetime.now()
            # 楽観ロックの照合用に、書き込んだ行の開始・終了を控える
            self.timelog_start, self.timelog_end = self._read_back_timelog_bounds(row)
            self._update_status_label()
            self._schedule_autoclose()
        except Exception as e:
            print(f"❌ 作業時間の記録に失敗しました: {e}")
            self.status_label.config(text="⚠️ 記録に失敗しました")
        finally:
            self.timelog_busy = False

    def _read_back_timelog_bounds(self, row: int):
        """
        書き込んだTimeLog行の開始・終了を読み戻す（楽観ロックの照合用）。
        取得に失敗した場合は(None, None)を返し、以降の照合を省略させる。
        """
        try:
            from openpyxl import load_workbook
            from storage import EXCEL_PATH, TIMELOG_SHEET, _parse_excel_dt
            wb = load_workbook(EXCEL_PATH)
            ws = wb[TIMELOG_SHEET]
            return (
                _parse_excel_dt(ws.cell(row=row, column=1).value),
                _parse_excel_dt(ws.cell(row=row, column=2).value),
            )
        except Exception as e:
            print(f"⚠️ 記録した行の読み戻しに失敗しました（照合を省略します）: {e}")
            return None, None

    def _undo_timelog(self) -> bool:
        """直前に書いたTimeLog行を削除する。削除できた場合Trueを返す。"""
        if not self._timelog_token_live():
            return False
        try:
            ok = delete_time_log_row(
                self.timelog_row,
                expect_start=self.timelog_start, expect_end=self.timelog_end,
            )
        except Exception as e:
            print(f"❌ 作業記録の取り消しに失敗しました: {e}")
            ok = False
        if ok:
            # 行を消すと以降の行番号が繰り上がるため、必ず破棄する
            self._forget_timelog_token()
        return ok

    def _update_status_label(self) -> None:
        """記録結果と、修正できる旨の案内を表示する。"""
        if self.status_label is None:
            return
        if self.timelog_row is None:
            self.status_label.config(text="")
            return

        if self.timelog_kind == "anchor":
            body = "⏱ 本日の基準点として記録しました"
        else:
            span = ""
            if self.timelog_start and self.timelog_end:
                span = (f"{self.timelog_start.strftime('%H:%M')}–"
                        f"{self.timelog_end.strftime('%H:%M')} ")
            detail = self.timelog_tag or ""
            if self.timelog_sub_item:
                detail += f" / {self.timelog_sub_item}"
            minutes = (f" ({_format_duration(self.timelog_minutes)})"
                       if self.timelog_minutes else "")
            body = f"⏱ {span}{detail} として記録しました{minutes}"
            if self.timelog_kind == "gap":
                body += "\n※ 空き時間は「未記録」として残しました"

        # 修正できなくなったら案内も消す（できない操作を案内しないため）
        if self._timelog_token_live():
            body += "\n別のタグ/中項目で修正・同じタグをもう一度押すと取り消し"
        self.status_label.config(text=body)

    # ------------------------------------------------------------------
    # 自動クローズ（記録したら勝手に閉じる。ただし操作したら止める）
    # ------------------------------------------------------------------

    def _schedule_autoclose(self) -> None:
        """
        記録後、何も操作しなければ自動的に閉じるタイマーを開始/再開始する。
        「閉じるために毎回クリックが必要」を避けつつ、訂正の余地も残すための仕組み。
        """
        if self.autoclose_cancelled or self.window is None:
            return
        if self.autoclose_after_id is not None:
            self.window.after_cancel(self.autoclose_after_id)
        self.autoclose_after_id = self.window.after(
            AUTOCLOSE_DELAY_MS, self._autoclose_fire,
        )

    def _cancel_autoclose(self, permanent: bool = False) -> None:
        """
        自動クローズを止める。permanent=Trueにすると、以降このウィンドウでは
        二度とタイマーを開始しない（LKPTやタスクを書いている最中に、
        文章の途中で勝手に閉じられるのを防ぐため）。
        """
        if self.autoclose_after_id is not None and self.window is not None:
            try:
                self.window.after_cancel(self.autoclose_after_id)
            except Exception:
                pass
            self.autoclose_after_id = None
        if permanent:
            self.autoclose_cancelled = True
            self._update_status_label()

    def _autoclose_fire(self) -> None:
        """タイマー満了。記録済みなのでそのまま閉じる。"""
        self.autoclose_after_id = None
        if self.autoclose_cancelled:
            return
        print("🪟 記録後、操作が無かったため自動的に閉じます。")
        self._close()

    # ------------------------------------------------------------------

    def _on_tag_clicked(self, tag_name: str, color_code: str) -> None:
        """
        タグチップのクリックを振り分ける。
        選択中のタグをもう一度クリックした場合は、記録そのものを取り消して
        未選択に戻す（修正可能な時間内のみ。期限切れ後は選択解除だけ行う）。
        """
        if self.timelog_busy:
            return
        if self.selected_tag == tag_name:
            if self._timelog_token_live():
                self._undo_timelog()
            self._deselect_all_tags()
            self._update_status_label()
            self._schedule_autoclose()
            return
        self._select_tag(tag_name, color_code)
        self._commit_or_amend_timelog(tag_name, "")

    def _deselect_all_tags(self) -> None:
        """タグの選択表示を未選択状態に戻す（記録の取り消しは呼び出し側の責任）。"""
        self.selected_tag = None
        self.selected_label.config(text="未選択", bg=BG_COLOR, fg=ACCENT_COLOR)
        self._apply_theme(BG_COLOR, TEXT_COLOR)
        for name in self.tag_chips:
            self._recolor_tag_chip(name, selected=False)
        self.tag_detail_frame.pack_forget()
        self.selected_subitem = None
        self.subitem_buttons = {}
        self._update_lkpt_tag_check()
        print("🏷️ タグ選択を解除しました。")

    def _select_tag(self, tag_name: str, color_code: str) -> None:
        self.selected_tag = tag_name
        color_code = str(color_code).strip()
        badge_fg = _readable_text_color(color_code)
        self.selected_label.config(text=tag_name, bg=color_code, fg=badge_fg)
        try:
            pale_bg = _lighten_color(color_code, 0.78)
        except (ValueError, IndexError):
            pale_bg = "#3a3a55"
        self._apply_theme(pale_bg, BUTTON_TEXT_COLOR)
        print(f"🏷️ タグ選択: {tag_name}")

        # 選ばれたタグのチップだけを選択色にし、他のチップは未選択配色に戻す
        for name in self.tag_chips:
            self._recolor_tag_chip(name, selected=(name == tag_name))

        # 中項目のブロックは、タググループ全体の下の固定位置に表示する。
        # 初回選択時にpackし、以降タグを切り替えても位置は動かさず、
        # 中身（_build_subitem_buttons）だけを差し替える
        if not self.tag_detail_frame.winfo_ismapped():
            self.tag_detail_frame.pack(after=self.tag_frame, fill="x", pady=(6, 0))

        self._build_subitem_buttons(tag_name)
        self._update_lkpt_tag_check()

    def _update_lkpt_tag_check(self) -> None:
        """LKPTのタグ引き継ぎチェックの文言・有効無効を選択状態に合わせる。"""
        if self.lkpt_tag_check is None:
            return
        if self.selected_tag:
            self.lkpt_tag_check.config(
                text=f"{self.selected_tag} として記録", state="normal",
            )
        else:
            self.lkpt_tag_check.config(text="タグなしで記録", state="disabled")

    def _build_subitem_buttons(self, tag_name: str) -> None:
        """
        選んだタグの中項目（何をしていたか）をボタンで選べるようにする。
        タグを切り替えるたびに作り直す（中項目の一覧自体がタグ依存のため）。
        """
        for child in self.subitem_frame.winfo_children():
            child.destroy()
        self.subitem_buttons = {}
        self.selected_subitem = None
        self.subitem_free_var.set("")
        self.subitem_free_entry = None

        items = self.subitem_master.get(tag_name, [])
        if not items:
            return

        # タグ選択によりsubitem_frame自体の背景は既にpale_bgへ切り替わって
        # いるため、新しく作る枠もそれに合わせる（テーマ管理対象外の子要素）
        current_bg = self.subitem_frame.cget("bg")

        grid = tk.Frame(self.subitem_frame, bg=current_bg)
        grid.pack(fill="x")
        for col in range(2):
            grid.columnconfigure(col, weight=1, uniform="subitem")

        for i, item in enumerate(items):
            btn = tk.Button(
                grid, text=item, bg=CANCEL_BTN_BG, fg=BUTTON_TEXT_COLOR,
                activebackground=ACCENT_COLOR, relief="flat", font=self.subitem_font,
                command=lambda it=item: self._select_subitem(it),
            )
            btn.grid(row=i // 2, column=i % 2, sticky="ew", padx=3, pady=3)
            self.subitem_buttons[item] = btn

        # Otherを選んだ時だけ表示する自由記述欄。他の入力欄と揃えてENTRY_BGにする。
        # Enterでも、他をクリックしてフォーカスが外れた時でも確定する
        # （書いたまま放置しても消えないように）
        self.subitem_free_entry = tk.Entry(
            self.subitem_frame, textvariable=self.subitem_free_var, bg=ENTRY_BG,
            fg=TEXT_COLOR, insertbackground=TEXT_COLOR, relief="flat",
            font=self.entry_font,
        )
        self.subitem_free_entry.bind("<Return>", lambda e: self._commit_free_subitem())
        self.subitem_free_entry.bind("<FocusOut>", lambda e: self._commit_free_subitem())

    def _select_subitem(self, item: str) -> None:
        self.selected_subitem = item
        for name, btn in self.subitem_buttons.items():
            btn.config(bg=ACCENT_COLOR if name == item else CANCEL_BTN_BG)
        if item == OTHER_SUBITEM_LABEL:
            self.subitem_free_entry.pack(fill="x", pady=(6, 0), ipady=3)
            self.subitem_free_entry.focus_set()
        else:
            self.subitem_free_entry.pack_forget()
        print(f"🗂️ 中項目選択: {item}")
        # 中項目を選んだ時点で、直前に書いたTimeLog行を上書き更新する。
        # Otherの場合はまず"Other"として確定させ、自由記述が入った時点で
        # あらためて上書きする（記録が未確定のまま宙に浮く瞬間を作らない）
        self._commit_or_amend_timelog(self.selected_tag, self._resolve_subitem())

    def _commit_free_subitem(self) -> None:
        """Other の自由記述欄の内容を中項目として確定する。"""
        if self.selected_subitem != OTHER_SUBITEM_LABEL:
            return
        resolved = self._resolve_subitem()
        # 前回書き込んだ内容と同じなら、ブックを開き直す意味が無いので何もしない
        # （FocusOutのたびに保存が走るのを防ぐ）
        if resolved == self.timelog_sub_item:
            return
        self._commit_or_amend_timelog(self.selected_tag, resolved)

    def _resolve_subitem(self) -> str:
        """
        選択中の中項目から、TimeLogに書き込む実際の文字列を決める。
        「Other」の場合は自由記述欄の内容（未入力ならラベルそのまま）を使う。
        """
        if self.selected_subitem is None:
            return ""
        if self.selected_subitem == OTHER_SUBITEM_LABEL:
            free_text = self.subitem_free_var.get().strip()
            return free_text if free_text else OTHER_SUBITEM_LABEL
        return self.selected_subitem

    def _on_lkpt_focus_in(self, entry: tk.Entry, placeholder: str) -> None:
        """
        LKPT欄にフォーカスが入った時。プレースホルダを消すのに加えて、
        自動クローズを恒久的に止める（文章を書いている途中で閉じないように）。
        """
        self._cancel_autoclose(permanent=True)
        self._clear_placeholder(entry, placeholder)

    def _clear_placeholder(self, entry: tk.Entry, placeholder: str) -> None:
        """フォーカスされた時、プレースホルダを表示中であれば消して入力可能にする。"""
        if str(entry.cget("fg")) == PLACEHOLDER_COLOR and entry.get() == placeholder:
            entry.delete(0, "end")
            entry.config(fg=TEXT_COLOR)

    def _restore_placeholder(self, entry: tk.Entry, placeholder: str) -> None:
        """フォーカスが外れた時、何も入力されていなければプレースホルダに戻す。"""
        if not entry.get():
            entry.insert(0, placeholder)
            entry.config(fg=PLACEHOLDER_COLOR)

    def _field_value(self, letter: str) -> str:
        """
        入力欄の実際の値を取得する。プレースホルダを表示中の場合は
        未入力（空文字）として扱う。
        """
        entry = self.field_entries[letter]
        if str(entry.cget("fg")) == PLACEHOLDER_COLOR:
            return ""
        return entry.get().strip()

    def _quick_action_value(self) -> str:
        """＋アクション欄の実際の値を取得する（プレースホルダ表示中は空扱い）。"""
        if str(self.quick_action_entry.cget("fg")) == PLACEHOLDER_COLOR:
            return ""
        return self.quick_action_entry.get().strip()

    def _fetch_pending_actions(self) -> list:
        """
        未着手アクションの一覧を、★優先を先頭・その中は登録順（古い順）で
        取得する（取得失敗時は空リストを返す）。
        """
        try:
            pending = [
                a for a in get_actions() if a["status"] == ACTION_STATUS_PENDING
            ]
        except Exception as e:
            print(f"⚠️ 未着手アクション一覧の取得に失敗しました: {e}")
            return []
        pending.sort(key=lambda a: (not a["starred"], a["created_at"]))
        return pending

    def _on_action_list_mousewheel(self, event) -> None:
        """
        アクション一覧の上でマウスホイールを回した時にスクロールする。
        Windowsの<MouseWheel>はevent.deltaが120単位で来るため、120で割って
        1行分の単位に変換する。
        一覧を眺め始めた＝まだ用があるということなので、自動クローズも止める。
        """
        self._cancel_autoclose(permanent=True)
        self.action_list_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _render_action_rows(self) -> None:
        """
        たまっているアクション一覧を、MS To Do風の1行カードとして
        縦に積んで描画し直す。★優先の切替・チェックでの完了のたびに呼ばれる。
        """
        if self.action_list_inner is None:
            return
        for child in self.action_list_inner.winfo_children():
            child.destroy()

        icon_font = tkfont.Font(family="Yu Gothic UI", size=13)
        text_font = tkfont.Font(family="Yu Gothic UI", size=10)

        for action in self.pending_actions:
            row_id = action["row"]
            row = tk.Frame(self.action_list_inner, bg=ACTION_CARD_BG)
            row.pack(side="top", fill="x", pady=(0, 3))
            row.bind("<MouseWheel>", self._on_action_list_mousewheel)

            # ○チェック（左端）。クリックで即完了＝一覧から消える
            check_btn = tk.Label(
                row, text="○", bg=ACTION_CARD_BG, fg=ACTION_CHECK_COLOR,
                font=icon_font, cursor="hand2", padx=10, pady=8,
            )
            check_btn.pack(side="left")
            check_btn.bind(
                "<Button-1>", lambda e, r=row_id: self._complete_pending_action(r),
            )
            check_btn.bind("<MouseWheel>", self._on_action_list_mousewheel)

            content = action["content"]
            label = tk.Label(
                row, text=content, bg=ACTION_CARD_BG, fg=TEXT_COLOR, font=text_font,
                anchor="w",
            )
            label.pack(side="left", fill="x", expand=True)
            # 固定文字数での省略をやめ、実際にこのラベルへ割り当てられた
            # ピクセル幅に収まる分だけ表示し、収まらない分だけ省略記号(…)を
            # 付ける。ラベルはfill="x", expand=Trueで親の残り幅いっぱいに
            # 広がるため、ウィンドウを横に広げるほど<Configure>で通知される
            # 幅が増え、表示できる文字数も連動して増える
            label.bind(
                "<Configure>",
                lambda e, lbl=label, full=content, fnt=text_font:
                    self._fit_action_row_text(lbl, full, fnt, e.width),
            )
            label.bind("<MouseWheel>", self._on_action_list_mousewheel)

            # ★優先（右端）。オンで青塗り、オフで灰アウトライン
            star_color = ACTION_STAR_ON_COLOR if action["starred"] else ACTION_STAR_OFF_COLOR
            star_char = "★" if action["starred"] else "☆"
            star_btn = tk.Label(
                row, text=star_char, bg=ACTION_CARD_BG, fg=star_color, font=icon_font,
                cursor="hand2", padx=10,
            )
            star_btn.pack(side="right")
            star_btn.bind(
                "<Button-1>", lambda e, r=row_id: self._toggle_action_star(r),
            )
            star_btn.bind("<MouseWheel>", self._on_action_list_mousewheel)

    def _fit_action_row_text(
        self, label: tk.Label, full_text: str, font: tkfont.Font, width: int,
    ) -> None:
        """
        アクション一覧の行テキストを、実際に割り当てられたピクセル幅
        （<Configure>のe.width）に収まる最大の長さまで表示し、収まらない
        場合だけ末尾に省略記号(…)を付ける。二分探索で「full_text[:n]+…」が
        widthに収まる最大のnを求める。
        """
        if width <= 4:
            return
        ellipsis = "…"
        if font.measure(full_text) <= width:
            label.config(text=full_text)
            return
        lo, hi = 0, len(full_text)
        best = ellipsis
        while lo <= hi:
            mid = (lo + hi) // 2
            candidate = full_text[:mid] + ellipsis
            if font.measure(candidate) <= width:
                best = candidate
                lo = mid + 1
            else:
                hi = mid - 1
        label.config(text=best)

    def _complete_pending_action(self, row: int) -> None:
        """
        ○チェックをクリックしたアクションを完了にする。ポップアップ自体は
        閉じず、一覧からその行だけがその場で消える。
        """
        # タスクを触り始めたら、自動クローズは止める
        self._cancel_autoclose(permanent=True)
        try:
            ok = complete_action(row)
        except Exception as e:
            print(f"❌ アクションの完了処理に失敗しました: {e}")
            ok = False
        if not ok:
            return
        self.pending_actions = [a for a in self.pending_actions if a["row"] != row]
        self._render_action_rows()

    def _toggle_action_star(self, row: int) -> None:
        """
        ★優先マークをクリックしたアクションの優先度を切り替える。
        保存に成功した場合のみ、★付きが先頭に来るよう並び替えて再描画する。
        """
        self._cancel_autoclose(permanent=True)
        target = next((a for a in self.pending_actions if a["row"] == row), None)
        if target is None:
            return
        new_starred = not target["starred"]
        try:
            ok = set_action_priority(row, new_starred)
        except Exception as e:
            print(f"❌ 優先度の更新に失敗しました: {e}")
            ok = False
        if not ok:
            return
        target["starred"] = new_starred
        self.pending_actions.sort(key=lambda a: (not a["starred"], a["created_at"]))
        self._render_action_rows()

    def _add_quick_action(self) -> None:
        """
        ＋アクション欄の内容をActionsシートに追加する。登録ボタンとは独立して
        おり、追加してもポップアップは閉じない（続けて何件でもメモできる）。
        追加後は一覧を取得し直し、その場でカードが増えた様子を見せる。
        """
        content = self._quick_action_value()
        if not content:
            return
        try:
            ok = add_action(content, tag=self.selected_tag or "", origin="manual")
        except Exception as e:
            print(f"❌ アクションの追加に失敗しました: {e}")
            ok = False

        if ok:
            # 追加直後にプレースホルダ文字とPLACEHOLDER_COLORへ戻していたが、
            # このときEntryはまだフォーカスを持ったままなので<FocusIn>が
            # 発火せず、_clear_placeholder()が呼ばれないままfgだけが
            # PLACEHOLDER_COLORに戻ってしまっていた。続けて2件目を入力して
            # Enter/＋を押しても_quick_action_value()がfg==PLACEHOLDER_COLORを
            # 「プレースホルダ表示中＝未入力」と誤判定し、常に空文字を返して
            # 何も追加されないバグになっていた（連続入力を想定した機能なのに
            # 1件目までしか効かない状態）。
            # 空にするだけにとどめ、プレースホルダの再表示はフォーカスが
            # 外れた時の_restore_placeholder()に任せることで解消する
            self.quick_action_entry.delete(0, "end")
            self.quick_action_entry.config(fg=TEXT_COLOR)
            self._flash_quick_action_feedback("✅ アクションに追加しました")
            self.pending_actions = self._fetch_pending_actions()
            self._render_action_rows()
            # いま書いたタスクを雨の種として拾えるようにする
            self._offer_forecast_seed(content)
        else:
            self._flash_quick_action_feedback("⚠️ 追加に失敗しました")

    def _flash_quick_action_feedback(self, message: str) -> None:
        """＋アクション欄の下に一瞬だけメッセージを出し、少し待って消す。"""
        self.quick_action_feedback_label.config(text=message)
        self.window.after(
            1800, lambda: self.quick_action_feedback_label.config(text=""),
        )

    def _offer_forecast_seed(self, task_text: str) -> None:
        """
        タスクを登録した直後に「なぜ今これを？ → 雨にする」を出す。

        タスクの裏には必ず予測があるので、そこを捕まえれば新しい習慣を
        作らずに雨を書き始められる。押さなければ数秒で消えるだけなので、
        「打つ→Enter」の速い動線は一切変わらない。
        """
        if self.forecast_seed_label is None:
            return
        self.forecast_seed_label.config(text="なぜ今これを？ → 雨にする")
        # after= を付けないと、実行時packのためウィンドウ最下部
        # （タスク一覧より下）に出てしまい、いま打った欄から離れて気づけない
        self.forecast_seed_label.pack(
            pady=(0, 2), after=self.quick_action_feedback_label,
        )
        self.forecast_seed_label.bind(
            "<Button-1>", lambda e, t=task_text: self._seed_forecast_from_task(t),
        )
        self.window.after(6000, self._hide_forecast_seed)

    def _hide_forecast_seed(self) -> None:
        if self.forecast_seed_label is not None:
            self.forecast_seed_label.pack_forget()

    def _commit_lkpt(self) -> None:
        """
        LKPT（振り返り）を1件記録する。L/K/P/Tのどの欄でEnterを押しても、
        入っている内容をまとめて1行として記録する（4件に分かれない）。
        タグの選択は不要で、TimeLogには一切触れない。
        記録してもポップアップは閉じず、続けて何件でも書ける。
        """
        self._cancel_autoclose(permanent=True)

        l = self._field_value("L")
        k = self._field_value("K")
        p = self._field_value("P")
        t = self._field_value("T")
        if not (l or k or p or t):
            self._flash_quick_action_feedback("⚠️ 入力が空です")
            return

        # タグを引き継ぐかは明示的なチェックで決める（黙って引き継ぐと、
        # 15時に書いた朝の会議の振り返りに15時のタグが付いてしまう）
        tag = self.selected_tag if (self.selected_tag and self.lkpt_tag_var.get()) else ""

        try:
            ok, stamp = append_lkpt_entry(tag, l, k, p, t)
        except Exception as e:
            print(f"❌ LKPTの記録に失敗しました: {e}")
            ok, stamp = False, None
        if not ok:
            self._flash_quick_action_feedback("⚠️ 記録に失敗しました")
            return

        # Pの内容を「→ アクションにする」にチェックしていれば、Actionsにも
        # 1件追加する。LKPTの記録が成功した後に行うため、こちらが失敗しても
        # 振り返り自体は失われない
        if p and self.p_to_action_var.get():
            try:
                add_action(p, tag=tag, origin="P")
                # 窓が開いたままになったため、追加したタスクを一覧に即反映する
                self.pending_actions = self._fetch_pending_actions()
                self._render_action_rows()
            except Exception as e:
                print(f"❌ Pのアクション化に失敗しました: {e}")

        self._clear_lkpt_fields()
        self.p_to_action_var.set(False)
        self._flash_quick_action_feedback(
            f"✅ {stamp.strftime('%H:%M')} にLKPTを記録しました"
        )

    def _clear_lkpt_fields(self) -> None:
        """
        記録後にL/K/P/Tの入力欄を空に戻す。

        フォーカスが残っている欄にプレースホルダを入れてはいけない。
        入れるとfgがPLACEHOLDER_COLORのままになり、次に書いた内容を
        _field_value()が「プレースホルダ表示中＝未入力」と誤判定して、
        2件目の振り返りが黙って消える（タスク欄で実際に起きたv1.22.1のバグ）。
        """
        focused = self.window.focus_get() if self.window is not None else None
        for letter, entry in self.field_entries.items():
            entry.delete(0, "end")
            if entry is focused:
                entry.config(fg=TEXT_COLOR)
            else:
                entry.insert(0, FIELD_PLACEHOLDERS[letter])
                entry.config(fg=PLACEHOLDER_COLOR)

    def _close(self) -> None:
        """
        ポップアップを閉じる。記録済みの内容は取り消さない
        （TIMEはタグを押した時点で、LKPT・タスクはEnterの時点で保存済み）。
        """
        self._cancel_autoclose()
        print("🪟 ポップアップを閉じました（記録済みの内容はそのまま残ります）。")
        self.window.destroy()

    def _open_dashboard(self, mode: str = "LKPT") -> None:
        """
        ダッシュボードウィンドウを開く。既に開いていれば前面化のみ行う
        （この場合、既存ウィンドウのモードはそのまま維持する）。
        modeを指定すると、その表示モードで直接開く（未着手アクション
        バッジからの1クリック遷移用。既定は従来通りLKPT）。
        """
        # ダッシュボードを開くのは明確な操作なので、その裏でポップアップが
        # 勝手に閉じないようにする
        self._cancel_autoclose(permanent=True)
        if self.dashboard_window is not None and self.dashboard_window.winfo_exists():
            self.dashboard_window.lift()
            self.dashboard_window.focus_force()
            return

        # popup_ui側とdashboard側で同名の定数（BG_COLOR等）を持つため、
        # モジュールレベルでの衝突を避けて遅延importする
        import dashboard

        self.dashboard_window = tk.Toplevel(self.root)
        dashboard.DashboardWindow(self.dashboard_window, initial_mode=mode)
        print(f"📊 ダッシュボードを開きました（{mode}モード）。")

    # ------------------------------------------------------------------
    # 雨（予測）
    # ------------------------------------------------------------------

    def _update_forecast_button(self) -> None:
        """確認期日が来た未記入の雨があれば、ボタンに件数を出す。"""
        if self.forecast_btn is None:
            return
        try:
            import forecast
            due = forecast.count_due()
        except Exception as e:
            print(f"⚠️ 雨の件数取得に失敗しました: {e}")
            due = 0
        self.forecast_btn.config(text=f"🌧️{due}" if due else "🌧️")

    def _open_forecasts(self) -> None:
        """雨（予測）の一覧を開く。"""
        self._cancel_autoclose(permanent=True)
        try:
            import forecast_ui
            forecast_ui.open_forecast_list(self.root)
        except Exception as e:
            print(f"❌ 雨の一覧を開けませんでした: {e}")

    def _seed_forecast_from_task(self, task_text: str) -> None:
        """
        いま書いたタスクを傘として、雨の入力を開く。

        タスクを書く時点で頭の中では必ず何かを予測している
        （「ベンダーに催促する」の裏には「放っておくと間に合わない」がある）。
        ゼロから予測を捻り出すのは難しいので、その予測を捕まえる入口にする。
        速い動線（打つ→Enter）は一切変えず、書きたい時だけ1クリックで進める
        """
        self._cancel_autoclose(permanent=True)
        try:
            import forecast_ui
            forecast_ui.open_forecast_entry(self.root, seed_umbrella=task_text)
        except Exception as e:
            print(f"❌ 雨の入力を開けませんでした: {e}")


def poll_trigger_queue(root: tk.Tk, popup: "PopupWindow") -> None:
    """
    トリガーキューを定期的に確認し、通知があればポップアップを表示する。
    Tkinterのメインスレッド上でroot.after()により定期実行される。
    例外が発生してもループ自体は必ず継続する（自己修復設計）。
    """
    try:
        _trigger_queue.get_nowait()
        # 修飾キー（Ctrl/Alt/L）が完全に離されるのを待ってから表示する。
        # 直後に表示すると、Windowsが Alt キー解放をシステムメニュー呼び出しと
        # 誤認識し、ウィンドウが一瞬で背面に回る不具合が発生するため。
        root.after(150, popup.show)
    except queue.Empty:
        pass
    except Exception as e:
        print(f"❌ poll_trigger_queueで例外が発生しました: {e}")
    finally:
        root.after(200, poll_trigger_queue, root, popup)


def run() -> None:
    """このファイル単体起動用のメインループ。"""
    print(f"📦 {os.path.basename(__file__)} version: {VERSION}")
    root = tk.Tk()
    root.withdraw()  # メインウィンドウ本体は非表示にする

    popup = PopupWindow(root)
    register_global_hotkey()

    root.after(200, poll_trigger_queue, root, popup)

    # 循環importおよびモジュール二重読み込みを避けるため、
    # run()内で遅延importし、自モジュールの関数をコールバックとして渡す
    from scheduler import start_scheduler_loop
    start_scheduler_loop(root, queue_popup_trigger)

    # 起動直後に一度ポップアップを表示し、本日最初のチェックイン（基準点）を
    # すぐに促す。Windowsスタートアップから自動起動された場合、次の定時
    # リマインドまで何も表示されず待たされてしまうのを防ぐため
    queue_popup_trigger()

    # 確認期日の来た雨があれば答え合わせを促す。人は自分から見返さないので、
    # 向こうから出てくる必要がある。ロジックはforecast_ui側に置いてあるので、
    # このバージョン付きファイルを更新しても移植の必要は無い
    def _prompt_forecasts():
        try:
            import forecast_ui
            forecast_ui.prompt_due_forecasts_if_needed(root)
        except Exception as e:
            print(f"⚠️ 雨の督促を実行できませんでした: {e}")

    # チェックインのポップアップより後に出して、答え合わせ画面が前面に来るようにする
    root.after(1200, _prompt_forecasts)

    print(f"🚀 ポップアップUIを起動しました。ホットキー({HOTKEY})待受中...")
    root.mainloop()


if __name__ == "__main__":
    run()
