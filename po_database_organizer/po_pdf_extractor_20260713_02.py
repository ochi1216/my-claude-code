#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PO PDF Extractor（調査/プロトタイプ版）

version : 20260713_02
purpose : PO本体PDF（SAPが発行するPurchase Order）を開き、
          1) 1ページ目冒頭の「PO番号」（国コード+拠点コード+PO番号、例 HK648210235980）
          2) 明細テーブル（Line / Material no.(12NC) / Order quantity / Unit /
             Unit price / Price Unit / Total price / Description）
          を抽出する。フォルダ内のPDFを一括処理し、抽出結果をExcelサマリーに
          まとめる調査用ツール。

前提（実PO 12件（JP/HK/CN/US）で確認した書式）:
    - SAPが生成するネイティブテキストPDF（画像スキャンではない）で、地域（HK/CN/JP/US等）が
      違っても完全に同一レイアウト。
    - PO番号は1ページ目の2行目（"Purchase Order"の次の行）に単独で記載されており、
      末尾の数字部分がファイル名の"PO<番号>"と一致する（一致しない場合はファイル名の
      誤り・付け間違いの可能性があるため要確認）。
    - 明細テーブルは罫線があり、pdfplumberのテーブル検出で「見出し行」と「明細ブロック」を
      分離できる。明細ブロックは各行が5桁のLine番号（00010, 00020...）で始まり、
      その後に数量/単位/単価/通貨/金額が続き、以降の複数行がDescription（自由記述）になる。
    - 明細が多い場合、テーブルが複数ページにまたがる（例: 4行目以降が次ページに続く）。
      全ページを走査して連結することで対応済み（v02で修正）。

使い方:
    pip install -r requirements.txt
    python po_pdf_extractor_20260713_02.py <PDFが入ったフォルダ> [-o summary.xlsx]
"""

import argparse
import re
import sys
from pathlib import Path

try:
    import pdfplumber
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ModuleNotFoundError as e:
    print(f"\n[エラー] 必要なライブラリ '{e.name}' が見つかりません。\n"
          "次のコマンドを実行してから再度実行してください:\n\n"
          "    pip install -r requirements.txt\n")
    sys.exit(1)


HEADER_PO_RE = re.compile(r"^([A-Z]{2})(\d{2})(\d+)$")
LINE_ITEM_START_RE = re.compile(r"^(\d{5})\s+(.*)$")
TOTAL_AMOUNT_RE = re.compile(
    r"Total Order Amount.*?:\s*([\d,\.]+)\s*(\S+)", re.IGNORECASE
)
ORDER_AMOUNT_RE = re.compile(r"^([\d,\.]+)\s+([A-Z]{3})$")


def _split_item_fields(first_line: str) -> dict:
    """
    明細1行目（例 "1 PU 4,168.68 USD 1 PU 4,168.68 USD"、
    または材料番号ありの場合 "935123456 1 PU ..."）をトークン分割する。
    末尾から Total価格+通貨、Price Unit数量+単位、Unit価格+通貨 を取り、
    残りを 材料番号(あれば) + Order数量 + Order単位 とみなす。
    """
    tokens = first_line.split()
    result = {
        "material_no": "", "order_qty": "", "order_unit": "",
        "unit_price": "", "unit_price_currency": "",
        "price_unit_qty": "", "price_unit_unit": "",
        "total_price": "", "total_price_currency": "",
        "_raw_tokens": tokens,
    }
    if len(tokens) < 7:
        return result  # 想定より短い→呼び出し側で要確認フラグを立てる

    total_currency  = tokens[-1]
    total_price     = tokens[-2]
    pu_unit         = tokens[-3]
    pu_qty          = tokens[-4]
    unit_currency   = tokens[-5]
    unit_price      = tokens[-6]
    head            = tokens[:-6]  # [material_no?, order_qty, order_unit]

    result.update({
        "total_price_currency": total_currency,
        "total_price":          total_price,
        "price_unit_unit":      pu_unit,
        "price_unit_qty":       pu_qty,
        "unit_price_currency":  unit_currency,
        "unit_price":           unit_price,
    })

    if len(head) == 2:
        result["order_qty"], result["order_unit"] = head
    elif len(head) == 3:
        result["material_no"], result["order_qty"], result["order_unit"] = head
    else:
        result["_unparsed_head"] = head

    return result


def extract_line_items(item_block_text: str) -> list:
    """
    明細ブロックの生テキストから、Line番号(5桁)で始まる各明細を切り出す。
    "( details )" 以降（内訳の自由記述）は明細としては扱わない。
    """
    if "( details )" in item_block_text:
        item_block_text = item_block_text.split("( details )")[0]

    lines = item_block_text.split("\n")
    items = []
    current = None

    for line in lines:
        m = LINE_ITEM_START_RE.match(line.strip())
        if m:
            if current is not None:
                items.append(current)
            current = {
                "line_no": m.group(1),
                "first_line": m.group(2).strip(),
                "description_lines": [],
            }
        elif current is not None:
            stripped = line.strip()
            if stripped:
                if TOTAL_AMOUNT_RE.search(stripped):
                    continue  # Total Order Amount行は明細と別扱い
                current["description_lines"].append(stripped)

    if current is not None:
        items.append(current)

    result = []
    for it in items:
        fields = _split_item_fields(it["first_line"])
        result.append({
            "line_no":              it["line_no"],
            "material_no":          fields["material_no"],
            "order_qty":            fields["order_qty"],
            "order_unit":           fields["order_unit"],
            "unit_price":           fields["unit_price"],
            "unit_price_currency":  fields["unit_price_currency"],
            "price_unit_qty":       fields["price_unit_qty"],
            "price_unit_unit":      fields["price_unit_unit"],
            "total_price":          fields["total_price"],
            "total_price_currency": fields["total_price_currency"],
            "description":          " / ".join(it["description_lines"]),
            "_field_count_ok":      len(fields["_raw_tokens"]) >= 7,
        })
    return result


def extract(pdf_path: Path) -> dict:
    """1件のPO PDFからヘッダーPO番号・明細行を抽出する。失敗理由はerrorsに積む。"""
    result = {
        "filename":            pdf_path.name,
        "filename_po_number":  "",
        "header_po_number":    "",
        "header_po_prefix":    "",
        "po_number_match":     None,
        "order_amount":        "",
        "order_amount_currency": "",
        "line_items":          [],
        "errors":              [],
    }

    m = re.search(r"PO(\d+)", pdf_path.stem, re.IGNORECASE)
    if m:
        result["filename_po_number"] = m.group(1)

    try:
        with pdfplumber.open(pdf_path) as pdf:
            page1_text = pdf.pages[0].get_text() if hasattr(pdf.pages[0], "get_text") \
                else pdf.pages[0].extract_text() or ""
            lines = [l.strip() for l in page1_text.split("\n") if l.strip()]

            if len(lines) >= 2 and lines[0].lower() == "purchase order":
                header_line = lines[1]
                hm = HEADER_PO_RE.match(header_line)
                if hm:
                    result["header_po_prefix"] = hm.group(1) + hm.group(2)
                    result["header_po_number"] = hm.group(3)
                else:
                    result["errors"].append(
                        f"ヘッダーPO番号のパターン不一致: {header_line!r}"
                    )
            else:
                result["errors"].append(
                    "1ページ目冒頭が 'Purchase Order' で始まっていない"
                )

            if result["filename_po_number"] and result["header_po_number"]:
                result["po_number_match"] = (
                    result["filename_po_number"] == result["header_po_number"]
                )

            for i, line in enumerate(lines):
                if not line.lower().startswith("order amount:"):
                    continue
                # 同一行に金額まで含む場合（例: pdfplumberのレイアウト）と、
                # 見出し行のみで次行に金額がある場合（例: PyMuPDFのレイアウト）の両方に対応する
                same_line_amount = line.split(":", 1)[1].strip()
                am = ORDER_AMOUNT_RE.match(same_line_amount) if same_line_amount else None
                if not am and i + 1 < len(lines):
                    am = ORDER_AMOUNT_RE.match(lines[i + 1])
                if am:
                    result["order_amount"] = am.group(1)
                    result["order_amount_currency"] = am.group(2)
                else:
                    result["errors"].append(
                        f"発注金額のパターン不一致: {line!r}"
                    )
                break

            # 明細テーブルは複数ページにまたがる場合がある（例: 4行目以降が次ページに
            # 続く）ため、最初に見つかったページで打ち切らず全ページを走査して連結する。
            item_block_texts = []
            for page in pdf.pages:
                tables = page.find_tables()
                for t in tables:
                    rows = t.extract()
                    header_idx = None
                    for idx, row in enumerate(rows):
                        if row and row[0] and str(row[0]).strip() == "Line":
                            header_idx = idx
                            break
                    if header_idx is None:
                        continue
                    for row in rows[header_idx + 1:]:
                        cell0 = (row[0] or "").strip()
                        cell1 = (row[1] or "").strip() if len(row) > 1 else ""
                        if cell0 == "" and cell1 == "Description":
                            continue
                        if cell0 and LINE_ITEM_START_RE.match(cell0.split("\n")[0].strip()):
                            item_block_texts.append(cell0)
                            continue
                        joined = " ".join(str(c) for c in row if c)
                        if LINE_ITEM_START_RE.match(joined.split("\n")[0].strip()):
                            item_block_texts.append(joined)

            item_block_text = "\n".join(item_block_texts) if item_block_texts else None

            if item_block_text:
                result["line_items"] = extract_line_items(item_block_text)
            else:
                result["errors"].append("明細テーブル（Line見出し）が見つからない")

    except Exception as e:
        result["errors"].append(f"PDF処理エラー: {e}")

    return result


def build_summary_workbook(results: list, out_path: Path) -> None:
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    warn_fill   = PatternFill("solid", fgColor="FCA5A5")

    def _header(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

    ws1 = wb.active
    ws1.title = "サマリー"
    _header(ws1, [
        "ファイル名", "ファイル名PO番号", "ヘッダーPO番号", "ヘッダー接頭辞",
        "番号一致", "発注金額", "通貨", "明細行数", "エラー",
    ])
    for r in results:
        ws1.append([
            r["filename"], r["filename_po_number"], r["header_po_number"],
            r["header_po_prefix"],
            "" if r["po_number_match"] is None else ("OK" if r["po_number_match"] else "NG"),
            r["order_amount"], r["order_amount_currency"],
            len(r["line_items"]), "; ".join(r["errors"]),
        ])
        if r["errors"] or r["po_number_match"] is False:
            for cell in ws1[ws1.max_row]:
                cell.fill = warn_fill

    ws2 = wb.create_sheet("明細")
    _header(ws2, [
        "ファイル名", "Line", "Material No.", "Order Qty", "Unit",
        "Unit Price", "通貨", "Price Unit Qty", "Price Unit",
        "Total Price", "通貨", "Description", "フィールド数チェック",
    ])
    for r in results:
        for it in r["line_items"]:
            ws2.append([
                r["filename"], it["line_no"], it["material_no"],
                it["order_qty"], it["order_unit"],
                it["unit_price"], it["unit_price_currency"],
                it["price_unit_qty"], it["price_unit_unit"],
                it["total_price"], it["total_price_currency"],
                it["description"],
                "OK" if it["_field_count_ok"] else "要確認",
            ])
            if not it["_field_count_ok"]:
                for cell in ws2[ws2.max_row]:
                    cell.fill = warn_fill

    for ws in (ws1, ws2):
        for col in ws.columns:
            length = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(length + 2, 60)

    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser(description="PO PDFからPO番号・明細行を抽出しExcelサマリーを作る（調査用）")
    parser.add_argument("input_dir", help="PO PDFが入ったフォルダ")
    parser.add_argument("-o", "--output", default="po_pdf_extract_summary.xlsx",
                         help="出力Excelファイル名（既定: po_pdf_extract_summary.xlsx）")
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    if not input_dir.is_dir():
        print(f"[エラー] フォルダが見つかりません: {input_dir}")
        sys.exit(1)

    pdf_paths = sorted(input_dir.glob("*.pdf"))
    if not pdf_paths:
        print(f"[エラー] PDFファイルが見つかりません: {input_dir}")
        sys.exit(1)

    print(f"[開始] {len(pdf_paths)} 件のPDFを処理します...")
    results = []
    for p in pdf_paths:
        r = extract(p)
        status = "OK" if not r["errors"] else "要確認: " + "; ".join(r["errors"])
        print(f"  {p.name}: 明細{len(r['line_items'])}件 / {status}")
        results.append(r)

    out_path = Path(args.output)
    build_summary_workbook(results, out_path)
    print(f"[完了] サマリーを出力しました: {out_path}")


if __name__ == "__main__":
    main()
