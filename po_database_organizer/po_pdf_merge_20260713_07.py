#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PO PDF Merge Integrator

version : 20260713_07
purpose : po_database_organizer が出力した「PO一覧」シート付きのExcelを読み込み、
          各行のPO本体PDFをSharePointから直接ダウンロードして
          （ヘッダーPO番号・発注金額・明細行）を抽出し、
          同じ「PO一覧」シートに列として統合する。

ブラウザを一切使わない理由:
    PDFのリンクをブラウザで開くと、Microsoft Defender for Cloud Apps の
    「Access to Microsoft SharePoint Online is monitored」という確認画面が
    毎回はさまり、Chromeでは手動でボタンを押す必要がある。
    本スクリプトは po_database_organizer と同じ Graph API（MSAL認証・Bearerトークン）
    でファイル本体をAPI経由で直接取得するため、ブラウザもMCASの確認画面も経由しない。
    「PO一覧」シートの Project / Vendor / PO関連フォルダ / ファイル名 列（プレーンテキスト）
    から相対パスを組み立て、Graph APIのパス指定ダウンロード
    （/drives/{drive-id}/root:/{path}:/content）でPDFを取得する。

進め方:
    - 全行を順番に処理し、[現在の件数/全件数] を画面に表示する。
    - 10件処理するごとに一度停止し、
        [1] 次の10件へ進む
        [2] 最後まで一気に進める（以降は確認なし）
        [3] ここで中断し、ここまでの結果を保存して終了する
      を選択できる。
    - 各行の処理結果はメモリ上に保持し、最後（または中断時）にExcelへ書き戻す。
      個別ファイルの取得・抽出エラーはその行にエラーメッセージを記録し、
      全体は止めずに次の行へ進む。

使い方:
    1. config.json を用意する（po_database_organizer と同じもので良い。
       tenant_id/client_id/site_host/site_path/library_name が必要）
    2. pip install -r requirements.txt
    3. python po_pdf_merge_20260713_07.py [PO一覧Excelファイル] [-o output.xlsx] [--start-row N]
       入力ファイルを引数で指定しない場合は、起動時にファイル選択ダイアログが
       開くのでExcelファイルを選ぶ。
       出力先を省略した場合は "<入力ファイル名>_detail.xlsx" に保存する
       （元のExcelは上書きしない）。

v03での修正:
    - 「PDFヘッダーPO番号」が数字部分のみになっていたのを、国コード+拠点コード
      （例 HK64, CN69, JP68）を別列「ヘッダー接頭辞」として復活。
    - Vendor列とPO関連フォルダ列が同名になっている行でファイルが見つからない
      場合、フォルダの重複セグメントを1つ省いたパスでも再試行するフォールバックを追加。

v04での修正:
    - 1ページ目冒頭が "Purchase Order" ではなく "Changed Purchase Order"
      （変更発注書）の場合でも処理できるよう対応。実際に見つかったタイトル文字列を
      「PDF種別」列・「PO明細(PDF抽出)」シートの「ヘッダー接頭辞」列として記録する。

v05での修正:
    - po_database_organizer v03でのExcelフォーマット変更（PO一覧が1書類=1行、
      列名「代表ファイル」→「ファイル名」）に対応。
    - 「PO一覧」の新規列の並びを ヘッダー接頭辞 → PDFヘッダーPO番号 の順に変更。
    - 「PO一覧」のPDFヘッダーPO番号セルから「PO明細(PDF抽出)」の該当PO番号・
      Line 00010行のA列へジャンプするハイパーリンクを追加。
    - 「PO明細(PDF抽出)」のPO番号セルから「PO一覧」の該当行のPDFヘッダーPO番号
      セルへジャンプするハイパーリンクを追加（双方向ジャンプ）。

v06での修正:
    - 処理速度の改善。入力Excelの「未分類書類」等、本ツールが触れない他シート
      （数千行規模になりがち）を読み込み直後に破棄するようにした。従来は10件ごとの
      中間保存のたびにファイル全体（未分類書類を含む）を書き直しており、行数が
      多いとこれが処理時間の大半を占めていた（実測でシート有無により保存1回あたり
      約18倍の差）。ハイパーリンク自体は軽い処理で速度への影響はごく僅か。
    - 「PO一覧」の新規列の並びを v05以前の PDFヘッダーPO番号 → ヘッダー接頭辞 の順に戻した
      （v05でヘッダー接頭辞を先頭にしたが、視認性の観点から差し戻し）。

v07での修正:
    - アクセストークンを起動時に1回だけ取得して使い回していたのを、リクエストの
      たびに取得し直すよう修正（MSALはキャッシュが有効なら即座に返すため速度への
      影響はない）。処理に時間がかかりトークンの有効期限（通常1時間程度）が切れた
      場合、"401 Unauthorized" でダウンロードが失敗していた問題への対応。
    - 途中から再開できる機能を追加。
        - "--start-row N" で開始行を明示的に指定できる（例: --start-row 107）。
        - 既に「PDFヘッダーPO番号」が入力済みの行（前回成功した行）は自動的に
          スキップする。前回出力した "_detail.xlsx" をそのまま入力に指定すれば、
          エラーで止まった行や未処理の行だけ再処理できる。
        - 「PO明細(PDF抽出)」シートが既にある場合は前回分を保持したまま追記する。
"""

import argparse
import io
import json
import re
import subprocess
import sys
import time
from pathlib import Path
from typing import Optional
from urllib.parse import quote

try:
    import msal
    import requests as http_req
    import pdfplumber
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as e:
    print(f"\n[エラー] 必要なライブラリ '{e.name}' が見つかりません。\n"
          "次のコマンドを実行してから再度実行してください:\n\n"
          "    pip install -r requirements.txt\n")
    sys.exit(1)


# ─────────────────────────────────────────────────────────────
# 設定ファイル読み込み（po_database_organizer と共通形式）
# ─────────────────────────────────────────────────────────────
def _load_config(path: str = "config.json") -> dict:
    config_path = Path(path)
    if not config_path.exists():
        print(f"[エラー] {path} が見つかりません。"
              f" po_database_organizer と同じ config.json を用意してください。")
        raise SystemExit(1)
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def _check_placeholder(value: str, key: str) -> None:
    if not value or value.startswith("<") or value.endswith(">"):
        print(f"[エラー] config.json の \"{key}\" がプレースホルダーのままです"
              f"（現在の値: {value!r}）。実際の値に書き換えてから再実行してください。")
        raise SystemExit(1)


_CFG         = _load_config()
TENANT_ID    = _CFG["tenant_id"]
CLIENT_ID    = _CFG["client_id"]
_check_placeholder(TENANT_ID, "tenant_id")
_check_placeholder(CLIENT_ID, "client_id")
TOKEN_CACHE_PATH = _CFG.get("token_cache_path", "token_cache.json")
SITE_HOST    = _CFG["site_host"]
SITE_PATH    = _CFG["site_path"]
LIBRARY_NAME = _CFG.get("library_name", "")

SCOPES   = ["https://graph.microsoft.com/Sites.Read.All"]
GRAPH_V1 = "https://graph.microsoft.com/v1.0"

CHECKPOINT_INTERVAL = 10


# ─────────────────────────────────────────────────────────────
# Class: GraphAuthManager（po_database_organizer と同一）
# ─────────────────────────────────────────────────────────────
class GraphAuthManager:
    """MSAL Device Code Flow による認証管理クラス"""

    def __init__(self, tenant_id: str, client_id: str, cache_path: str):
        self.tenant_id  = tenant_id
        self.client_id  = client_id
        self.cache_path = cache_path
        self.cache      = msal.SerializableTokenCache()
        self._load_cache()
        self.app = msal.PublicClientApplication(
            client_id=self.client_id,
            authority=f"https://login.microsoftonline.com/{self.tenant_id}",
            token_cache=self.cache,
        )

    def _load_cache(self) -> None:
        if Path(self.cache_path).exists():
            with open(self.cache_path, "r", encoding="utf-8") as f:
                self.cache.deserialize(f.read())

    def _save_cache(self) -> None:
        if self.cache.has_state_changed:
            with open(self.cache_path, "w", encoding="utf-8") as f:
                f.write(self.cache.serialize())

    def get_token(self, scopes: list) -> str:
        accounts = self.app.get_accounts()
        result   = None
        if accounts:
            result = self.app.acquire_token_silent(scopes, account=accounts[0])
        if not result:
            flow = self.app.initiate_device_flow(scopes=scopes)
            if "user_code" not in flow:
                raise RuntimeError(f"Device Flow 開始失敗: {flow.get('error_description')}")
            print("\n" + "=" * 60)
            print(flow["message"])
            print("=" * 60 + "\n")
            result = self.app.acquire_token_by_device_flow(flow)
        if "access_token" not in result:
            raise RuntimeError(f"トークン取得失敗: {result.get('error_description', '不明')}")
        self._save_cache()
        return result["access_token"]


# ─────────────────────────────────────────────────────────────
# Class: SharePointDownloadClient
# ─────────────────────────────────────────────────────────────
class SharePointDownloadClient:
    """
    Graph API 経由でサイト・ドライブを特定し、パス指定でファイル本体を取得する。
    アクセストークンは有効期限が切れる（通常1時間程度）ため、リクエストのたびに
    auth.get_token() を呼び直して最新のトークンを使う。MSALは有効なキャッシュが
    あればネットワーク通信なしで即座に返すため、毎回呼び出しても速度への影響はない。
    """

    def __init__(self, auth: "GraphAuthManager", site_host: str, site_path: str):
        self.auth = auth
        self.site_host = site_host
        self.site_path = site_path

    def _headers(self) -> dict:
        token = self.auth.get_token(SCOPES)
        return {"Authorization": f"Bearer {token}", "Accept": "application/json"}

    def _request_json(self, url: str) -> dict:
        MAX_RETRY = 3
        last_error = None
        for attempt in range(1, MAX_RETRY + 1):
            try:
                resp = http_req.get(url, headers=self._headers(), timeout=30)
                if resp.status_code == 403:
                    raise PermissionError("403 Forbidden: Sites.Read.All の Admin Consent を確認してください。")
                resp.raise_for_status()
                return resp.json()
            except PermissionError:
                raise
            except Exception as e:
                last_error = e
                print(f"  [リトライ {attempt}/{MAX_RETRY}] {e}")
                time.sleep(1)
        raise RuntimeError(f"リクエスト失敗（{MAX_RETRY}回試行）: {last_error}")

    def get_site_id(self) -> str:
        url = f"{GRAPH_V1}/sites/{self.site_host}:{self.site_path}"
        data = self._request_json(url)
        site_id = data.get("id", "")
        if not site_id:
            raise RuntimeError(f"site-id 取得失敗: {data}")
        return site_id

    def get_drive_id(self, site_id: str, library_name: str = "") -> str:
        url = f"{GRAPH_V1}/sites/{site_id}/drives"
        data = self._request_json(url)
        drives = data.get("value", [])
        if library_name:
            for d in drives:
                if d.get("name", "") == library_name:
                    return d["id"]
        for priority_name in ["Documents", "Shared Documents"]:
            for d in drives:
                if d.get("name", "") == priority_name:
                    return d["id"]
        for d in drives:
            if d.get("driveType", "") == "documentLibrary":
                return d["id"]
        raise RuntimeError("ドキュメントライブラリが見つかりません。")

    def download_by_path(self, drive_id: str, relative_path: str) -> bytes:
        """
        Project/Vendor/PO関連フォルダ/ファイル名 の相対パスを指定してファイル本体を取得する。
        Graph API のパス指定アドレッシング（root:/{path}:/content）を使うため、
        ブラウザ・MCASの確認画面を経由しない。
        """
        encoded = quote(relative_path, safe="/")
        url = f"{GRAPH_V1}/drives/{drive_id}/root:/{encoded}:/content"
        MAX_RETRY = 3
        last_error = None
        for attempt in range(1, MAX_RETRY + 1):
            try:
                resp = http_req.get(
                    url, headers=self._headers(),
                    allow_redirects=False, timeout=30, stream=True,
                )
                if resp.status_code in (301, 302, 303, 307, 308):
                    redirect_url = resp.headers.get("Location", "")
                    resp.close()
                    if not redirect_url:
                        raise RuntimeError("リダイレクト先URLが取得できません")
                    resp = http_req.get(redirect_url, timeout=60, stream=True)
                if resp.status_code == 404:
                    raise FileNotFoundError(f"ファイルが見つかりません: {relative_path}")
                resp.raise_for_status()
                content = resp.content
                resp.close()
                return content
            except FileNotFoundError:
                raise
            except Exception as e:
                last_error = e
                print(f"    [リトライ {attempt}/{MAX_RETRY}] {e}")
                time.sleep(1)
        raise RuntimeError(f"ダウンロード失敗（{MAX_RETRY}回試行）: {last_error}")


# ─────────────────────────────────────────────────────────────
# PDF抽出ロジック（po_pdf_extractor_20260713_04.py と同一のロジックを移植）
# ─────────────────────────────────────────────────────────────
HEADER_PO_RE = re.compile(r"^([A-Z]{2})(\d{2})(\d+)$")
LINE_ITEM_START_RE = re.compile(r"^(\d{5})\s+(.*)$")
TOTAL_AMOUNT_RE = re.compile(r"Total Order Amount.*?:\s*([\d,\.]+)\s*(\S+)", re.IGNORECASE)
ORDER_AMOUNT_RE = re.compile(r"^([\d,\.]+)\s+([A-Z]{3})$")


def _split_item_fields(first_line: str) -> dict:
    tokens = first_line.split()
    result = {
        "material_no": "", "order_qty": "", "order_unit": "",
        "unit_price": "", "unit_price_currency": "",
        "price_unit_qty": "", "price_unit_unit": "",
        "total_price": "", "total_price_currency": "",
        "_raw_tokens": tokens,
    }
    if len(tokens) < 7:
        return result

    total_currency = tokens[-1]
    total_price    = tokens[-2]
    pu_unit        = tokens[-3]
    pu_qty         = tokens[-4]
    unit_currency  = tokens[-5]
    unit_price     = tokens[-6]
    head           = tokens[:-6]

    result.update({
        "total_price_currency": total_currency,
        "total_price":          total_price,
        "price_unit_unit":      pu_unit,
        "price_unit_qty":       pu_qty,
        "unit_price_currency":  unit_currency,
        "unit_price":           unit_price,
    })
    if len(head) == 2:
        result["order_qty"], result["order_unit"] = head
    elif len(head) == 3:
        result["material_no"], result["order_qty"], result["order_unit"] = head
    return result


def extract_line_items(item_block_text: str) -> list:
    if "( details )" in item_block_text:
        item_block_text = item_block_text.split("( details )")[0]

    lines = item_block_text.split("\n")
    items = []
    current = None
    for line in lines:
        m = LINE_ITEM_START_RE.match(line.strip())
        if m:
            if current is not None:
                items.append(current)
            current = {"line_no": m.group(1), "first_line": m.group(2).strip(), "description_lines": []}
        elif current is not None:
            stripped = line.strip()
            if stripped:
                if TOTAL_AMOUNT_RE.search(stripped):
                    continue
                current["description_lines"].append(stripped)
    if current is not None:
        items.append(current)

    result = []
    for it in items:
        fields = _split_item_fields(it["first_line"])
        result.append({
            "line_no": it["line_no"], "material_no": fields["material_no"],
            "order_qty": fields["order_qty"], "order_unit": fields["order_unit"],
            "unit_price": fields["unit_price"], "unit_price_currency": fields["unit_price_currency"],
            "price_unit_qty": fields["price_unit_qty"], "price_unit_unit": fields["price_unit_unit"],
            "total_price": fields["total_price"], "total_price_currency": fields["total_price_currency"],
            "description": " / ".join(it["description_lines"]),
        })
    return result


def extract_from_bytes(pdf_bytes: bytes, display_filename: str) -> dict:
    """ダウンロード済みPDFのバイト列からヘッダーPO番号・発注金額・明細行を抽出する。"""
    result = {
        "header_po_number": "", "header_po_prefix": "",
        "document_type": "",
        "order_amount": "", "order_amount_currency": "",
        "line_items": [], "errors": [],
    }
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            page1_text = pdf.pages[0].extract_text() or ""
            lines = [l.strip() for l in page1_text.split("\n") if l.strip()]

            # "Purchase Order" のほか、変更発注書 "Changed Purchase Order" にも対応する
            if len(lines) >= 2 and "purchase order" in lines[0].lower():
                result["document_type"] = lines[0]
                hm = HEADER_PO_RE.match(lines[1])
                if hm:
                    result["header_po_prefix"] = hm.group(1) + hm.group(2)
                    result["header_po_number"] = hm.group(3)
                else:
                    result["errors"].append(f"ヘッダーPO番号のパターン不一致: {lines[1]!r}")
            else:
                result["errors"].append("1ページ目冒頭が 'Purchase Order' 系のタイトルで始まっていない")

            for i, line in enumerate(lines):
                if not line.lower().startswith("order amount:"):
                    continue
                same_line_amount = line.split(":", 1)[1].strip()
                am = ORDER_AMOUNT_RE.match(same_line_amount) if same_line_amount else None
                if not am and i + 1 < len(lines):
                    am = ORDER_AMOUNT_RE.match(lines[i + 1])
                if am:
                    result["order_amount"] = am.group(1)
                    result["order_amount_currency"] = am.group(2)
                break

            item_block_texts = []
            for page in pdf.pages:
                for t in page.find_tables():
                    rows = t.extract()
                    header_idx = None
                    for idx, row in enumerate(rows):
                        if row and row[0] and str(row[0]).strip() == "Line":
                            header_idx = idx
                            break
                    if header_idx is None:
                        continue
                    for row in rows[header_idx + 1:]:
                        cell0 = (row[0] or "").strip()
                        cell1 = (row[1] or "").strip() if len(row) > 1 else ""
                        if cell0 == "" and cell1 == "Description":
                            continue
                        if cell0 and LINE_ITEM_START_RE.match(cell0.split("\n")[0].strip()):
                            item_block_texts.append(cell0)
                            continue
                        joined = " ".join(str(c) for c in row if c)
                        if LINE_ITEM_START_RE.match(joined.split("\n")[0].strip()):
                            item_block_texts.append(joined)

            item_block_text = "\n".join(item_block_texts) if item_block_texts else None
            if item_block_text:
                result["line_items"] = extract_line_items(item_block_text)
            else:
                result["errors"].append("明細テーブル（Line見出し）が見つからない")

    except Exception as e:
        result["errors"].append(f"PDF処理エラー: {e}")

    return result


# ─────────────────────────────────────────────────────────────
# メイン処理
# ─────────────────────────────────────────────────────────────
def _prompt_checkpoint(done: int, total: int) -> str:
    """10件ごとの確認プロンプト。'continue' / 'auto' / 'stop' のいずれかを返す。"""
    while True:
        print(f"\n--- {done}/{total} 件処理しました ---")
        print("  [1] 次の10件へ進む")
        print("  [2] 最後まで一気に進める（以降は確認なし）")
        print("  [3] ここで中断し、ここまでの結果を保存して終了する")
        choice = input("選択 (1/2/3): ").strip()
        if choice == "1":
            return "continue"
        if choice == "2":
            return "auto"
        if choice == "3":
            return "stop"
        print("  1〜3のいずれかを入力してください。")


def _pick_excel_file() -> Optional[str]:
    """
    サブプロセスでtkinterのファイル選択ダイアログを起動し、選択されたパスを返す。
    tkinterはpipでインストールできない標準ライブラリのため、サブプロセス側だけで
    importすることでメインスクリプトの依存関係チェックに影響させない
    （po_database_organizer の route_pick_folder と同じ方式）。
    """
    # サブプロセスのstdoutをUTF-8に固定する（パスに日本語が含まれる場合の文字化け防止）
    script = (
        "import sys,tkinter as tk; from tkinter import filedialog; "
        "root=tk.Tk(); root.withdraw(); root.lift(); "
        "root.attributes('-topmost',True); "
        "chosen=filedialog.askopenfilename("
        "title='PO一覧シートがある"
        "Excelファイルを選択して"
        "ください',"
        "filetypes=[('Excel ファイル','*.xlsx'),"
        "('すべてのファイル','*.*')]); "
        "sys.stdout.reconfigure(encoding='utf-8'); "
        "print(chosen if chosen else '',end='')"
    )
    try:
        result = subprocess.run(
            [sys.executable, "-c", script],
            capture_output=True, text=True, encoding="utf-8", timeout=120,
        )
    except Exception as e:
        print(f"[エラー] ファイル選択ダイアログの起動に失敗しました: {e}")
        return None
    chosen = result.stdout.strip()
    return chosen or None


def main():
    parser = argparse.ArgumentParser(
        description="PO一覧Excelの各行のPDFをSharePointから取得し、抽出結果を統合する"
    )
    parser.add_argument("input_xlsx", nargs="?", default=None,
                         help="po_database_organizer が出力したExcelファイル"
                              "（省略時はファイル選択ダイアログが開く）")
    parser.add_argument("-o", "--output",
                         help="出力Excelファイル（省略時は '<入力ファイル名>_detail.xlsx'。元ファイルは上書きしない）")
    parser.add_argument("--start-row", type=int, default=None,
                         help="処理を開始するExcelの行番号（例: 107 なら107行目から）。"
                              "省略時は2行目（先頭データ行）から開始する。"
                              "入力に '_detail.xlsx' 等、既にPDFヘッダーPO番号が入っている行が"
                              "あれば、start-rowの指定に関わらず自動的にスキップして再開する。")
    args = parser.parse_args()

    input_arg = args.input_xlsx
    if not input_arg:
        print("[選択] ファイル選択ダイアログを開いています...")
        input_arg = _pick_excel_file()
        if not input_arg:
            print("[中止] ファイルが選択されませんでした。")
            sys.exit(1)

    input_path = Path(input_arg)
    if not input_path.exists():
        print(f"[エラー] ファイルが見つかりません: {input_path}")
        sys.exit(1)
    output_path = Path(args.output) if args.output else input_path.with_name(input_path.stem + "_detail.xlsx")

    print("=" * 60)
    print(" PO PDF Merge Integrator")
    print("=" * 60)
    print(f"  対象サイト: {SITE_HOST}{SITE_PATH}")
    if LIBRARY_NAME:
        print(f"  ライブラリ: {LIBRARY_NAME}")
    print()

    auth = GraphAuthManager(TENANT_ID, CLIENT_ID, TOKEN_CACHE_PATH)
    print("[認証] トークンを取得しています...")
    auth.get_token(SCOPES)
    print("[OK] 認証完了\n")

    client = SharePointDownloadClient(auth, SITE_HOST, SITE_PATH)
    site_id = client.get_site_id()
    drive_id = client.get_drive_id(site_id, LIBRARY_NAME)
    print(f"[OK] drive_id 取得完了\n")

    wb = load_workbook(input_path)
    if "PO一覧" not in wb.sheetnames:
        print("[エラー] 'PO一覧' シートが見つかりません。")
        sys.exit(1)
    ws = wb["PO一覧"]

    DETAIL_SHEET_NAME = "PO明細(PDF抽出)"
    resume_detail_sheet = DETAIL_SHEET_NAME in wb.sheetnames

    # 「未分類書類」等、本ツールが触れない他シートは巨大なことが多く（数千行）、
    # 10件ごとの中間保存のたびにファイル全体を書き直すコストが跳ね上がるため、
    # 処理対象外のシートは出力から除外する（元のExcelは別ファイルなので情報は失われない）。
    # 既に「PO明細(PDF抽出)」がある場合（前回の続きから再開する場合）はそのまま保持する。
    for name in list(wb.sheetnames):
        if name not in ("PO一覧", DETAIL_SHEET_NAME):
            del wb[name]

    header = [c.value for c in ws[1]]
    try:
        col_project = header.index("Project")
        col_vendor  = header.index("Vendor")
        col_pofolder = header.index("PO関連フォルダ")
        col_ponumber = header.index("PO番号")
        col_repfile  = header.index("ファイル名") if "ファイル名" in header else header.index("代表ファイル")
    except ValueError as e:
        print(f"[エラー] 想定する列が見つかりません: {e}")
        sys.exit(1)

    hyperlink_font = Font(color="0563C1", underline="single")

    # PDFヘッダーPO番号を先頭に、その隣にヘッダー接頭辞を配置する
    new_col_names = ["PDFヘッダーPO番号", "ヘッダー接頭辞", "PDF種別", "PO番号一致", "発注金額", "通貨", "明細行数", "抽出エラー"]
    if new_col_names[0] in header:
        # 再開: 前回付与された列をそのまま使う（新規に列を追加しない）
        col_header_ponumber = header.index("PDFヘッダーPO番号") + 1
        col_header_prefix   = header.index("ヘッダー接頭辞") + 1
        col_doctype         = header.index("PDF種別") + 1
        col_match           = header.index("PO番号一致") + 1
        col_amount          = header.index("発注金額") + 1
        col_currency        = header.index("通貨") + 1
        col_linecount       = header.index("明細行数") + 1
        col_error           = header.index("抽出エラー") + 1
    else:
        base = len(header)
        (col_header_ponumber, col_header_prefix, col_doctype, col_match,
         col_amount, col_currency, col_linecount, col_error) = range(base + 1, base + 9)
        for offset, name in enumerate(new_col_names):
            ws.cell(row=1, column=base + 1 + offset, value=name)

    start_row = args.start_row if args.start_row and args.start_row >= 2 else 2
    if start_row > ws.max_row:
        print(f"[エラー] --start-row {start_row} がデータ範囲を超えています（最終行: {ws.max_row}）。")
        sys.exit(1)

    total = ws.max_row - start_row + 1
    if start_row > 2:
        print(f"[開始] PO一覧 {start_row}行目〜{ws.max_row}行目（{total}件）を処理します。\n")
    else:
        print(f"[開始] PO一覧 全 {total} 件を処理します。\n")
    print("      （PDFヘッダーPO番号が既に入っている行は自動的にスキップします）\n")

    all_line_items = []  # (po一覧のrow_num, ヘッダーPO番号, ヘッダー接頭辞, line_item dict)
    auto_continue = False
    processed = 0
    attempted = 0  # スキップした行を除く、実際に処理を試みた件数（チェックポイント判定用）

    for row_num in range(start_row, ws.max_row + 1):
        processed += 1
        project_name = ws.cell(row=row_num, column=col_project + 1).value or ""
        vendor_name  = ws.cell(row=row_num, column=col_vendor + 1).value or ""
        po_folder    = ws.cell(row=row_num, column=col_pofolder + 1).value or ""
        po_number    = ws.cell(row=row_num, column=col_ponumber + 1).value or ""
        rep_file     = ws.cell(row=row_num, column=col_repfile + 1).value or ""

        if ws.cell(row=row_num, column=col_header_ponumber).value:
            print(f"[{processed}/{total}] スキップ（処理済み）: PO{po_number} ({vendor_name} / {rep_file})")
            continue

        print(f"[{processed}/{total}] 処理中: PO{po_number} ({vendor_name} / {rep_file})")

        error_msg = ""
        extracted = None
        try:
            if not po_folder or not rep_file:
                raise ValueError("PO関連フォルダ または ファイル名 が空欄です")
            relative_path = f"{project_name}/{vendor_name}/{po_folder}/{rep_file}"
            try:
                pdf_bytes = client.download_by_path(drive_id, relative_path)
            except FileNotFoundError:
                # Vendorフォルダ＝PO関連フォルダで、実際にはPOフォルダの階層が
                # もう一段少ないケースへのフォールバック（重複セグメントを1つ省く）
                if po_folder and po_folder == vendor_name:
                    fallback_path = f"{project_name}/{vendor_name}/{rep_file}"
                    pdf_bytes = client.download_by_path(drive_id, fallback_path)
                else:
                    raise
            extracted = extract_from_bytes(pdf_bytes, rep_file)
            if extracted["errors"]:
                error_msg = "; ".join(extracted["errors"])
        except Exception as e:
            error_msg = str(e)

        if extracted:
            match = ""
            if extracted["header_po_number"]:
                match = "OK" if str(po_number) == extracted["header_po_number"] else "NG"
            ws.cell(row=row_num, column=col_header_ponumber, value=extracted["header_po_number"])
            ws.cell(row=row_num, column=col_header_prefix, value=extracted["header_po_prefix"])
            ws.cell(row=row_num, column=col_doctype, value=extracted["document_type"])
            ws.cell(row=row_num, column=col_match, value=match)
            ws.cell(row=row_num, column=col_amount, value=extracted["order_amount"])
            ws.cell(row=row_num, column=col_currency, value=extracted["order_amount_currency"])
            ws.cell(row=row_num, column=col_linecount, value=len(extracted["line_items"]))
            for it in extracted["line_items"]:
                all_line_items.append((row_num, extracted["header_po_number"],
                                        extracted["header_po_prefix"], it))
        ws.cell(row=row_num, column=col_error, value=error_msg)

        if error_msg:
            print(f"    [要確認] {error_msg}")

        attempted += 1
        if (attempted % CHECKPOINT_INTERVAL == 0) and processed != total and not auto_continue:
            wb.save(output_path)
            print(f"  [中間保存] {output_path}")
            action = _prompt_checkpoint(processed, total)
            if action == "auto":
                auto_continue = True
            elif action == "stop":
                print(f"\n[中断] {processed}/{total} 件で中断しました。")
                break

    if all_line_items:
        if resume_detail_sheet:
            # 再開時: 前回の内容を保持したまま追記する（ヘッダー行は追加しない）
            ws_detail = wb[DETAIL_SHEET_NAME]
        else:
            ws_detail = wb.create_sheet(DETAIL_SHEET_NAME)
            ws_detail.append([
                "PO番号", "ヘッダー接頭辞", "Line", "Material No.", "Order Qty", "Unit",
                "Unit Price", "通貨", "Total Price", "通貨", "Description",
            ])

        po_ponumber_letter = get_column_letter(col_header_ponumber)
        first_line_detail_row = {}  # PO一覧のrow_num -> PO明細でLine=00010が入った行番号

        for src_row_num, header_po_number, header_prefix, it in all_line_items:
            ws_detail.append([
                header_po_number, header_prefix, it["line_no"], it["material_no"],
                it["order_qty"], it["order_unit"],
                it["unit_price"], it["unit_price_currency"],
                it["total_price"], it["total_price_currency"],
                it["description"],
            ])
            detail_row = ws_detail.max_row
            if it["line_no"] == "00010" and src_row_num not in first_line_detail_row:
                first_line_detail_row[src_row_num] = detail_row

            # 逆リンク: PO明細のPO番号セル → PO一覧の該当行のPDFヘッダーPO番号セル
            back_cell = ws_detail.cell(row=detail_row, column=1)
            back_cell.hyperlink = f"#'PO一覧'!{po_ponumber_letter}{src_row_num}"
            back_cell.font = hyperlink_font

        # 順リンク: PO一覧のPDFヘッダーPO番号セル → PO明細の該当Line 00010行のA列
        for src_row_num, detail_row in first_line_detail_row.items():
            fwd_cell = ws.cell(row=src_row_num, column=col_header_ponumber)
            fwd_cell.hyperlink = f"#'{DETAIL_SHEET_NAME}'!A{detail_row}"
            fwd_cell.font = hyperlink_font

    wb.save(output_path)
    print(f"\n[完了] {processed}/{total} 件処理し、結果を保存しました: {output_path}")


if __name__ == "__main__":
    main()
