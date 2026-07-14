#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PO Query Import

version : 20260714_01
purpose : SharePointの「クエリ」エクスポート（Name / Status / Requester /
          Item Type / Path の5列を持つExcel）を読み込み、Graph APIに一切
          アクセスせず、po_database_organizer と同じ形式の「PO一覧」
          「未分類書類」Excelをオフラインで高速に再構築する。

背景:
    ライブスキャン（po_database_organizer）はSharePointの全フォルダ・全ファイルを
    Graph APIで1件ずつ辿るため、大規模サイトでは時間がかかり、トークン失効や
    スロットリングのリスクもある。SharePoint側で「PO」ライブラリの全アイテムを
    一括エクスポートできるなら、そのクエリ結果から同じ出力を組み立てる方が
    圧倒的に速い（数千件でも数秒）。

前提・制約:
    - クエリの Path 列は「サイトルートからの相対パス（親フォルダ）」。
      例: "sites/JapanDesign/PO/Caracal Study/DBH"
    - Project > Vendor > PO関連フォルダ > 書類 の3階層構造を前提に、
      Pathの区切り数からどの階層のアイテムかを判定する
      （po_database_organizerのスキャン設計と同じ深さの意味付け）。
    - クエリには「最終更新日」列が無いため、「最終更新日」列は空欄になる
      （ファイル名で管理する運用を前提とする）。
    - クエリには直接のリンクURLが含まれないため、Path+NameからSharePointの
      URLを再構築する（スペース→%20、#→%23、非ASCII文字→%XX、括弧・カンマ等は
      そのまま、という実際に観測されたエンコード規則に基づく）。極端に特殊な
      文字を含むパス名では、実際のURLと一致しない可能性がある。

使い方:
    1. config.json を用意する（po_database_organizer と同じもので良い。
       tenant_id/client_idは使わないが、site_host/site_path/library_name/
       po_number_patternをそのまま流用する）
    2. pip install -r requirements.txt
    3. python po_query_import_20260714_01.py [クエリExcelファイル] [-o output.xlsx]
       入力ファイルを引数で指定しない場合は、起動時にファイル選択ダイアログが開く。
       出力先を省略した場合は "<入力ファイル名>_imported.xlsx" に保存する。
"""

import argparse
import json
import re
import subprocess
import sys
from pathlib import Path
from typing import Optional
from urllib.parse import quote

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill
except ModuleNotFoundError as e:
    print(f"\n[エラー] 必要なライブラリ '{e.name}' が見つかりません。\n"
          "次のコマンドを実行してから再度実行してください:\n\n"
          "    pip install -r requirements.txt\n")
    sys.exit(1)


# ─────────────────────────────────────────────────────────────
# 設定ファイル読み込み（po_database_organizer と共通形式。認証情報は使わない）
# ─────────────────────────────────────────────────────────────
def _load_config(path: str = "config.json") -> dict:
    config_path = Path(path)
    if not config_path.exists():
        print(f"[エラー] {path} が見つかりません。"
              f" po_database_organizer と同じ config.json を用意してください。")
        raise SystemExit(1)
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


_CFG              = _load_config()
SITE_HOST         = _CFG["site_host"]
SITE_PATH         = _CFG["site_path"]
LIBRARY_NAME      = _CFG.get("library_name", "")
PO_NUMBER_PATTERN = _CFG.get("po_number_pattern", r"^PO[-_# ]?(\d{3,})")

if not LIBRARY_NAME:
    print("[エラー] config.json に library_name が設定されていません。"
          " クエリのPath列の起点を特定できないため必須です。")
    raise SystemExit(1)

ROOT_PREFIX = f"{SITE_PATH.strip('/')}/{LIBRARY_NAME}"
PROJECT_ROOT_VENDOR_NAME = "(プロジェクト直下)"

_PO_RE = re.compile(PO_NUMBER_PATTERN, re.IGNORECASE)

# SharePointの実URLで確認された「エンコードされない」記号（括弧・カンマ等）。
# 半角スペースと # は必ずエンコードされる（quote()の既定動作のまま）。
_URL_SAFE_CHARS = "/()',.&+;=!$"


def classify_filename(filename: str) -> tuple:
    """ファイル名からPO番号を抽出する。'PO'系の接頭辞で始まるファイルのみPO本体として認識する。"""
    stem = Path(filename).stem
    m = _PO_RE.match(stem)
    if m:
        return "PO本体", m.group(1)
    return "未分類", None


def build_url(path: str) -> str:
    """サイトルートからの相対パスをSharePointの実URLに近い形へ復元する。"""
    encoded = quote(path, safe=_URL_SAFE_CHARS)
    return f"https://{SITE_HOST}/{encoded}"


# ─────────────────────────────────────────────────────────────
# ファイル選択ダイアログ（po_pdf_merge と同じサブプロセス方式）
# ─────────────────────────────────────────────────────────────
def _pick_excel_file() -> Optional[str]:
    script = (
        "import sys,tkinter as tk; from tkinter import filedialog; "
        "root=tk.Tk(); root.withdraw(); root.lift(); "
        "root.attributes('-topmost',True); "
        "chosen=filedialog.askopenfilename("
        "title='SharePointのクエリエクスポート"
        "Excelファイルを選択してください',"
        "filetypes=[('Excel ファイル','*.xlsx'),"
        "('すべてのファイル','*.*')]); "
        "sys.stdout.reconfigure(encoding='utf-8'); "
        "print(chosen if chosen else '',end='')"
    )
    try:
        result = subprocess.run(
            [sys.executable, "-c", script],
            capture_output=True, text=True, encoding="utf-8", timeout=120,
        )
    except Exception as e:
        print(f"[エラー] ファイル選択ダイアログの起動に失敗しました: {e}")
        return None
    chosen = result.stdout.strip()
    return chosen or None


# ─────────────────────────────────────────────────────────────
# クエリExcelの読み込み・変換
# ─────────────────────────────────────────────────────────────
def load_query_rows(input_path: Path) -> list:
    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("クエリExcelにデータがありません。")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    required = ["Name", "Item Type", "Path"]
    missing = [c for c in required if c not in header]
    if missing:
        raise ValueError(f"想定する列が見つかりません: {missing}（見つかった列: {header}）")

    idx = {name: header.index(name) for name in header if name}
    result = []
    for row in rows[1:]:
        if row is None or all(c is None for c in row):
            continue
        result.append({
            "name":      row[idx["Name"]] if idx["Name"] < len(row) else None,
            "item_type": row[idx["Item Type"]] if idx["Item Type"] < len(row) else None,
            "path":      row[idx["Path"]] if idx["Path"] < len(row) else None,
        })
    return result


def build_documents(rows: list) -> dict:
    """クエリ行から documents / project_web_url / vendor_web_url を構築する。"""
    documents = []
    project_ids: dict = {}       # project_name -> project_id
    project_web_url: dict = {}   # project_id -> url
    vendor_ids: dict = {}        # (project_id, vendor_name) -> vendor_id
    vendor_web_url: dict = {}    # vendor_id -> url
    vendor_count_per_project: dict = {}  # project_id -> count

    skipped_outside_root = 0
    skipped_root_level   = 0

    def get_project_id(project_name: str) -> str:
        if project_name not in project_ids:
            pid = f"PRJ{len(project_ids) + 1:04d}"
            project_ids[project_name] = pid
            project_web_url[pid] = build_url(f"{ROOT_PREFIX}/{project_name}")
        return project_ids[project_name]

    def get_vendor_id(project_id: str, project_name: str, vendor_name: str) -> str:
        key = (project_id, vendor_name)
        if key not in vendor_ids:
            n = vendor_count_per_project.get(project_id, 0) + 1
            vendor_count_per_project[project_id] = n
            vid = f"{project_id}-V{n:04d}"
            vendor_ids[key] = vid
            if vendor_name != PROJECT_ROOT_VENDOR_NAME:
                vendor_web_url[vid] = build_url(f"{ROOT_PREFIX}/{project_name}/{vendor_name}")
        return vendor_ids[key]

    for row in rows:
        name, item_type, path = row["name"], row["item_type"], row["path"]
        if item_type != "アイテム" or not name or not path:
            continue
        if not path.startswith(ROOT_PREFIX):
            skipped_outside_root += 1
            continue

        rel = path[len(ROOT_PREFIX):].strip("/")
        segments = [s for s in rel.split("/") if s] if rel else []
        depth = len(segments)

        if depth == 0:
            # PO一覧のルート直下（Projectフォルダより上）にあるファイル。
            # 想定外のレイアウトのため、対応するProject/Vendorが定義できず対象外とする。
            skipped_root_level += 1
            continue

        project_name = segments[0]
        project_id   = get_project_id(project_name)

        if depth == 1:
            vendor_name       = PROJECT_ROOT_VENDOR_NAME
            vendor_id         = get_vendor_id(project_id, project_name, vendor_name)
            po_folder_name    = ""
            po_folder_web_url = ""
        else:
            vendor_name = segments[1]
            vendor_id   = get_vendor_id(project_id, project_name, vendor_name)
            if depth == 2:
                po_folder_name    = ""
                po_folder_web_url = ""
            else:
                po_folder_name    = segments[2]
                po_folder_web_url = build_url(f"{ROOT_PREFIX}/{project_name}/{vendor_name}/{po_folder_name}")

        doc_type, po_number = classify_filename(name)
        file_web_url = build_url(f"{path}/{name}")

        documents.append({
            "project_id":        project_id,
            "project_name":      project_name,
            "vendor_id":         vendor_id,
            "vendor_name":       vendor_name,
            "po_folder_name":    po_folder_name,
            "po_folder_web_url": po_folder_web_url,
            "po_number":         po_number,
            "doc_type":          doc_type,
            "filename":          name,
            "web_url":           file_web_url,
        })

    return {
        "documents":              documents,
        "project_web_url":        project_web_url,
        "vendor_web_url":         vendor_web_url,
        "skipped_outside_root":   skipped_outside_root,
        "skipped_root_level":     skipped_root_level,
    }


# ─────────────────────────────────────────────────────────────
# Excel出力（po_database_organizer と同じ列構成・スタイル）
# ─────────────────────────────────────────────────────────────
def export_excel(result: dict, out_path: Path) -> None:
    wb             = Workbook()
    header_font    = Font(bold=True, color="FFFFFF")
    header_fill    = PatternFill("solid", fgColor="1E3A5F")
    hyperlink_font = Font(color="0563C1", underline="single")

    def _sheet(ws, headers, rows):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for row in rows:
            values = [item[0] if isinstance(item, tuple) else item for item in row]
            ws.append(values)
            r = ws.max_row
            for col_idx, item in enumerate(row, start=1):
                if isinstance(item, tuple) and item[1]:
                    cell = ws.cell(row=r, column=col_idx)
                    cell.hyperlink = item[1]
                    cell.font = hyperlink_font
        for col in ws.columns:
            length = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(length + 2, 60)

    project_web_url = result["project_web_url"]
    vendor_web_url  = result["vendor_web_url"]
    documents       = result["documents"]

    ws1 = wb.active
    ws1.title = "PO一覧"
    _sheet(
        ws1,
        ["Project ID", "Project", "Vendor", "PO関連フォルダ", "PO番号", "ファイル名",
         "最終更新日", "Status(未定義)"],
        [
            [d["project_id"],
             (d["project_name"], project_web_url.get(d["project_id"])),
             (d["vendor_name"], vendor_web_url.get(d["vendor_id"])),
             (d["po_folder_name"], d["po_folder_web_url"]),
             d["po_number"],
             (d["filename"], d["web_url"]), "", ""]
            for d in documents if d["doc_type"] == "PO本体"
        ],
    )

    ws2 = wb.create_sheet("未分類書類")
    _sheet(
        ws2,
        ["Project ID", "Project", "Vendor", "PO関連フォルダ", "ファイル名", "最終更新日"],
        [
            [d["project_id"],
             (d["project_name"], project_web_url.get(d["project_id"])),
             (d["vendor_name"], vendor_web_url.get(d["vendor_id"])),
             (d["po_folder_name"], d["po_folder_web_url"]),
             (d["filename"], d["web_url"]), ""]
            for d in documents if d["doc_type"] != "PO本体"
        ],
    )

    wb.save(out_path)


# ─────────────────────────────────────────────────────────────
# エントリポイント
# ─────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="SharePointのクエリExcelからPO一覧/未分類書類を高速に再構築する"
    )
    parser.add_argument("input_xlsx", nargs="?", default=None,
                         help="クエリExcelファイル（省略時はファイル選択ダイアログが開く）")
    parser.add_argument("-o", "--output",
                         help="出力Excelファイル（省略時は '<入力ファイル名>_imported.xlsx'）")
    args = parser.parse_args()

    input_arg = args.input_xlsx
    if not input_arg:
        print("[選択] ファイル選択ダイアログを開いています...")
        input_arg = _pick_excel_file()
        if not input_arg:
            print("[中止] ファイルが選択されませんでした。")
            sys.exit(1)

    input_path = Path(input_arg)
    if not input_path.exists():
        print(f"[エラー] ファイルが見つかりません: {input_path}")
        sys.exit(1)
    output_path = Path(args.output) if args.output else input_path.with_name(input_path.stem + "_imported.xlsx")

    print("=" * 60)
    print(" PO Query Import")
    print("=" * 60)
    print(f"  対象ルート: https://{SITE_HOST}/{ROOT_PREFIX}")
    print()

    print("[読込] クエリExcelを読み込んでいます...")
    rows = load_query_rows(input_path)
    print(f"  [OK] {len(rows)} 行を読み込みました")

    print("[変換] Project/Vendor/PO関連フォルダの階層とPO番号を判定しています...")
    result = build_documents(rows)
    documents = result["documents"]
    po_docs = [d for d in documents if d["doc_type"] == "PO本体"]
    other_docs = [d for d in documents if d["doc_type"] != "PO本体"]

    print(f"  [OK] Project数: {len(result['project_web_url'])}")
    print(f"  [OK] Vendor数: {len(result['vendor_web_url'])}")
    print(f"  [OK] PO本体: {len(po_docs)} 件 / 未分類: {len(other_docs)} 件")
    if result["skipped_outside_root"]:
        print(f"  [警告] 対象ルート外のためスキップ: {result['skipped_outside_root']} 件"
              f"（config.jsonのsite_path/library_nameを確認してください）")
    if result["skipped_root_level"]:
        print(f"  [警告] Projectフォルダより上位階層のためスキップ: {result['skipped_root_level']} 件")

    export_excel(result, output_path)
    print(f"\n[完了] 結果を保存しました: {output_path}")


if __name__ == "__main__":
    main()
