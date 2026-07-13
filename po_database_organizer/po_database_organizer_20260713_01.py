#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PO Database Organizer

version : 20260713_01
purpose : SharePoint上のPOフォルダ（Project > Vendor > 書類）をスキャンし、
          PO番号を軸にしたカタログをExcel/JSONで出力する。

Phase 1のスコープ:
    - ファイル名が "PO" で始まるものだけを確実にPO本体として認識し、PO番号を抽出する
      （それ以外の命名規則はまだ確立していないため、無理に自動分類しない）
    - PO番号に紐付かない書類（メール履歴・エビデンス等）は Project/Vendor 単位までの
      情報だけを保持し、「未分類」として一覧化する
    - 発注/検収/請求などのステータス判定はここでは行わない。Excel出力の「PO一覧」シートに
      空の Status 列を用意し、他の管理Excelと PO番号 キーで結合して後付けできるようにする

使い方:
    1. config.example.json を config.json にコピーし、tenant_id/client_id/
       site_host/site_path/library_name を環境に合わせて設定する
    2. pip install -r requirements.txt
    3. python po_database_organizer_20260713_01.py
    4. 初回はターミナルにDevice Code Flowの認証コードが表示されるので、
       表示されたURLをブラウザで開いてサインインする
    5. http://127.0.0.1:5010 が自動で開くので、「スキャン開始」を押す
    6. 完了後、「Excel出力」または「JSON出力」でカタログを保存する
"""

import json
import re
import sys
import time
import queue
import threading
import webbrowser
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional

try:
    import msal
    import requests as http_req
    from flask import (
        Flask, jsonify, request as flask_req,
        Response, stream_with_context, send_file,
    )
except ModuleNotFoundError as e:
    print(f"\n[エラー] 必要なライブラリ '{e.name}' が見つかりません。\n"
          "次のコマンドを実行してから再度実行してください:\n\n"
          "    pip install -r requirements.txt\n")
    sys.exit(1)


# ─────────────────────────────────────────────────────────────
# 設定ファイル読み込み
# ─────────────────────────────────────────────────────────────
def _load_config(path: str = "config.json") -> dict:
    config_path = Path(path)
    if not config_path.exists():
        print(f"[エラー] {path} が見つかりません。"
              f" config.example.json をコピーして作成してください。")
        raise SystemExit(1)
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def _check_placeholder(value: str, key: str) -> None:
    """config.json の値がexample由来のプレースホルダーのままでないか確認する。"""
    if not value or value.startswith("<") or value.endswith(">"):
        print(f"[エラー] config.json の \"{key}\" がプレースホルダーのままです"
              f"（現在の値: {value!r}）。")
        print(f"         実際の {key} の値に書き換えてから再実行してください。")
        raise SystemExit(1)


_CFG              = _load_config()
TENANT_ID         = _CFG["tenant_id"]
CLIENT_ID         = _CFG["client_id"]
_check_placeholder(TENANT_ID, "tenant_id")
_check_placeholder(CLIENT_ID, "client_id")
TOKEN_CACHE_PATH  = _CFG.get("token_cache_path", "token_cache.json")
SITE_HOST         = _CFG["site_host"]
SITE_PATH         = _CFG["site_path"]
LIBRARY_NAME      = _CFG.get("library_name", "")
PO_NUMBER_PATTERN = _CFG.get("po_number_pattern", r"^PO[-_]?(\d{3,})")
MAX_DEPTH         = _CFG.get("max_depth", 6)
MAX_ITEMS         = _CFG.get("max_items_per_folder", 1000)
SKIP_FOLDER_NAMES = {n.lower() for n in _CFG.get("skip_folder_names", ["old"])}

SCOPES     = ["https://graph.microsoft.com/Sites.Read.All"]
GRAPH_V1   = "https://graph.microsoft.com/v1.0"
FLASK_PORT = 5010
CACHE_DIR  = Path("cache")
JST        = timezone(timedelta(hours=9))

_PO_RE = re.compile(PO_NUMBER_PATTERN, re.IGNORECASE)

_lock = threading.Lock()
_auth: Optional["GraphAuthManager"] = None
_scan_state: dict = {
    "running":   False,
    "cancelled": False,
    "count":     0,
}
_last_result: Optional[dict] = None

flask_app = Flask(__name__)


# ─────────────────────────────────────────────────────────────
# Class: GraphAuthManager
# ─────────────────────────────────────────────────────────────
class GraphAuthManager:
    """MSAL Device Code Flow による認証管理クラス"""

    def __init__(self, tenant_id: str, client_id: str, cache_path: str):
        self.tenant_id  = tenant_id
        self.client_id  = client_id
        self.cache_path = cache_path
        self.cache      = msal.SerializableTokenCache()
        self._load_cache()
        self.app = msal.PublicClientApplication(
            client_id=self.client_id,
            authority=f"https://login.microsoftonline.com/{self.tenant_id}",
            token_cache=self.cache,
        )

    def _load_cache(self) -> None:
        if Path(self.cache_path).exists():
            with open(self.cache_path, "r", encoding="utf-8") as f:
                self.cache.deserialize(f.read())

    def _save_cache(self) -> None:
        if self.cache.has_state_changed:
            with open(self.cache_path, "w", encoding="utf-8") as f:
                f.write(self.cache.serialize())

    def get_token(self, scopes: list) -> str:
        accounts = self.app.get_accounts()
        result   = None
        if accounts:
            result = self.app.acquire_token_silent(scopes, account=accounts[0])
        if not result:
            flow = self.app.initiate_device_flow(scopes=scopes)
            if "user_code" not in flow:
                raise RuntimeError(
                    f"Device Flow 開始失敗: {flow.get('error_description')}"
                )
            print("\n" + "=" * 60)
            print(flow["message"])
            print("=" * 60 + "\n")
            result = self.app.acquire_token_by_device_flow(flow)
        if "access_token" not in result:
            raise RuntimeError(
                f"トークン取得失敗: {result.get('error_description', '不明')}"
            )
        self._save_cache()
        return result["access_token"]


# ─────────────────────────────────────────────────────────────
# Class: SharePointClient
# ─────────────────────────────────────────────────────────────
class SharePointClient:
    """Graph API 経由でサイト・ドライブ・フォルダ内容を取得するクラス"""

    def __init__(self, token: str, site_host: str, site_path: str):
        self.headers = {
            "Authorization": f"Bearer {token}",
            "Accept":        "application/json",
        }
        self.site_host = site_host
        self.site_path = site_path

    def _request(self, url: str) -> dict:
        """GETリクエストを実行する。403は即raise、400/404は_not_found、他は3回リトライ。"""
        MAX_RETRY  = 3
        last_error = None
        for attempt in range(1, MAX_RETRY + 1):
            try:
                resp = http_req.get(url, headers=self.headers, timeout=30)
                if resp.status_code == 403:
                    raise PermissionError(
                        "403 Forbidden: Sites.Read.All の"
                        " Admin Consent を確認してください。"
                    )
                if resp.status_code in (400, 404):
                    return {"_not_found": True}
                resp.raise_for_status()
                return resp.json()
            except PermissionError:
                raise
            except Exception as e:
                last_error = e
                print(f"  [リトライ {attempt}/{MAX_RETRY}] {e}")
                time.sleep(1)
        raise RuntimeError(f"リクエスト失敗（{MAX_RETRY}回試行）: {last_error}")

    def get_site_id(self) -> str:
        url     = f"{GRAPH_V1}/sites/{self.site_host}:{self.site_path}"
        data    = self._request(url)
        site_id = data.get("id", "")
        if not site_id:
            raise RuntimeError(f"site-id 取得失敗: {data}")
        print(f"  [OK] site-id 取得: {site_id} ({self.site_path})")
        return site_id

    def get_drive_id(self, site_id: str, library_name: str = "") -> str:
        """
        drive-id を取得する。判定優先順位:
          1. library_name が指定されていればその名前と完全一致
          2. name == "Documents" / "Shared Documents"
          3. driveType == "documentLibrary" の先頭
        """
        url    = f"{GRAPH_V1}/sites/{site_id}/drives"
        data   = self._request(url)
        drives = data.get("value", [])

        if library_name:
            for d in drives:
                if d.get("name", "") == library_name:
                    print(f"  [OK] drive-id 取得: {d['id']} (name={d['name']})")
                    return d["id"]
            print(f"  [警告] ライブラリ名 '{library_name}' が見つかりません。候補:")
            for d in drives:
                print(f"    - {d.get('name', '')}")

        for priority_name in ["Documents", "Shared Documents"]:
            for d in drives:
                if d.get("name", "") == priority_name:
                    print(f"  [OK] drive-id 取得: {d['id']} (name={d['name']})")
                    return d["id"]
        for d in drives:
            if d.get("driveType", "") == "documentLibrary":
                print(f"  [OK] drive-id 取得（fallback）: {d['id']}")
                return d["id"]
        raise RuntimeError("ドキュメントライブラリが見つかりません。")

    def fetch_all_children(self, drive_id: str, item_id: str) -> list:
        """指定アイテムの子一覧をページネーション対応で全件取得する。"""
        if item_id == "root":
            url = f"{GRAPH_V1}/drives/{drive_id}/root/children?$top=200"
        else:
            url = f"{GRAPH_V1}/drives/{drive_id}/items/{item_id}/children?$top=200"
        items = []
        while url:
            data = self._request(url)
            if data.get("_not_found"):
                break
            items.extend(data.get("value", []))
            url = data.get("@odata.nextLink", None)
            if url:
                time.sleep(0.15)
        return items


# ─────────────────────────────────────────────────────────────
# Class: ScanCache
# ─────────────────────────────────────────────────────────────
class ScanCache:
    """Vendorフォルダ単位でスキャン結果を永続化する（中断・再開対応）"""

    def __init__(self, path: Path):
        self.path = path
        self.data: dict = {"vendors": {}}
        if self.path.exists():
            with open(self.path, "r", encoding="utf-8") as f:
                self.data = json.load(f)

    def get_vendor(self, vendor_path: str) -> "Optional[list]":
        return self.data.get("vendors", {}).get(vendor_path)

    def set_vendor(self, vendor_path: str, files: list) -> None:
        self.data.setdefault("vendors", {})[vendor_path] = files
        self._save()

    def reset(self) -> None:
        self.data = {"vendors": {}}
        if self.path.exists():
            self.path.unlink()

    def _save(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        with open(self.path, "w", encoding="utf-8") as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)


# ─────────────────────────────────────────────────────────────
# PO番号抽出
# ─────────────────────────────────────────────────────────────
def classify_filename(filename: str) -> tuple:
    """
    ファイル名からPO番号を抽出する。
    'PO'で始まるファイルのみ確実にPO本体として認識し、それ以外は未分類として扱う。
    """
    stem = Path(filename).stem
    m = _PO_RE.match(stem)
    if m:
        return "PO本体", m.group(1)
    return "未分類", None


# ─────────────────────────────────────────────────────────────
# Class: POScanner
# ─────────────────────────────────────────────────────────────
class POScanner:
    """Project > Vendor > 書類 の階層をスキャンし、PO単位のカタログを構築する"""

    def __init__(self, client: SharePointClient, cache: ScanCache):
        self.client = client
        self.cache  = cache

    def scan(self, drive_id: str, event_q: "queue.Queue", state: dict) -> dict:
        projects, vendors, pos, documents = [], [], [], []
        po_index: dict = {}  # (vendor_id, po_number) -> po_id

        root_children   = self.client.fetch_all_children(drive_id, "root")
        project_folders = [c for c in root_children if "folder" in c]

        for p_idx, pf in enumerate(project_folders):
            if state.get("cancelled"):
                break
            project_id   = f"PRJ{p_idx + 1:04d}"
            project_name = pf.get("name", "")
            projects.append({
                "project_id":   project_id,
                "project_name": project_name,
                "web_url":      pf.get("webUrl", ""),
            })
            event_q.put({"type": "exploring", "path": project_name})

            vendor_children = self.client.fetch_all_children(drive_id, pf["id"])
            vendor_folders  = [
                c for c in vendor_children
                if "folder" in c and c.get("name", "").lower() not in SKIP_FOLDER_NAMES
            ]

            for v_idx, vf in enumerate(vendor_folders):
                if state.get("cancelled"):
                    break
                vendor_id   = f"{project_id}-V{v_idx + 1:04d}"
                vendor_name = vf.get("name", "")
                vendors.append({
                    "vendor_id":   vendor_id,
                    "vendor_name": vendor_name,
                    "project_id":  project_id,
                    "web_url":     vf.get("webUrl", ""),
                })

                vendor_path = f"{project_name}/{vendor_name}"
                files = self.cache.get_vendor(vendor_path)
                if files is None:
                    files = self._collect_files(drive_id, vf["id"], "", 0)
                    self.cache.set_vendor(vendor_path, files)

                for f in files:
                    doc_type, po_number = classify_filename(f["name"])
                    doc_id = f"{vendor_id}-D{len(documents) + 1:04d}"
                    po_id  = None

                    if po_number:
                        key = (vendor_id, po_number)
                        if key not in po_index:
                            po_id = f"{vendor_id}-PO{len(pos) + 1:04d}"
                            po_index[key] = po_id
                            pos.append({
                                "po_id":               po_id,
                                "po_number":           po_number,
                                "project_id":          project_id,
                                "project_name":        project_name,
                                "vendor_id":           vendor_id,
                                "vendor_name":         vendor_name,
                                "representative_file": f["name"],
                                "web_url":             f["web_url"],
                                "last_modified":       f["last_modified"],
                                "doc_count":           0,
                            })
                        po_id = po_index[key]
                        po_rec = next(p for p in pos if p["po_id"] == po_id)
                        po_rec["doc_count"] += 1
                        if f["last_modified"] > po_rec["last_modified"]:
                            po_rec["last_modified"]       = f["last_modified"]
                            po_rec["representative_file"] = f["name"]
                            po_rec["web_url"]              = f["web_url"]

                    documents.append({
                        "doc_id":         doc_id,
                        "po_id":          po_id,
                        "project_id":     project_id,
                        "project_name":   project_name,
                        "vendor_id":      vendor_id,
                        "vendor_name":    vendor_name,
                        "filename":       f["name"],
                        "relative_path":  f["relative_path"],
                        "web_url":        f["web_url"],
                        "last_modified":  f["last_modified"],
                        "size":           f["size"],
                        "doc_type":       doc_type,
                    })

                state["count"] += 1
                event_q.put({
                    "type":  "progress",
                    "count": state["count"],
                    "path":  vendor_path,
                    "files": len(files),
                })

        return {
            "projects":   projects,
            "vendors":    vendors,
            "pos":        pos,
            "documents":  documents,
            "scanned_at": datetime.now(JST).strftime("%Y-%m-%d %H:%M:%S JST"),
        }

    def _collect_files(self, drive_id: str, folder_id: str,
                        rel_path: str, depth: int) -> list:
        """Vendorフォルダ配下を再帰的に走査し、ファイル一覧をフラットに集める。"""
        if depth >= MAX_DEPTH:
            return []
        children    = self.client.fetch_all_children(drive_id, folder_id)
        files_raw   = [c for c in children if "file" in c][:MAX_ITEMS]
        folders_raw = [
            c for c in children
            if "folder" in c and c.get("name", "").lower() not in SKIP_FOLDER_NAMES
        ]

        result = [
            {
                "name":          f.get("name", ""),
                "web_url":       f.get("webUrl", ""),
                "size":          f.get("size", 0),
                "last_modified": f.get("lastModifiedDateTime", ""),
                "relative_path": rel_path,
            }
            for f in files_raw
        ]
        for sub in folders_raw:
            sub_rel = f"{rel_path}/{sub.get('name', '')}" if rel_path else sub.get("name", "")
            result.extend(self._collect_files(drive_id, sub["id"], sub_rel, depth + 1))
        return result


# ─────────────────────────────────────────────────────────────
# Excel / JSON 出力
# ─────────────────────────────────────────────────────────────
def export_excel(result: dict, out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb             = Workbook()
    header_font    = Font(bold=True, color="FFFFFF")
    header_fill    = PatternFill("solid", fgColor="1E3A5F")
    hyperlink_font = Font(color="0563C1", underline="single")

    def _sheet(ws, headers, rows):
        """
        rows の各セルは通常の値、または (表示値, リンク先URL) のタプルを渡せる。
        タプルの場合はハイパーリンクを設定し、下線付きの青字にする。
        """
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for row in rows:
            values = [item[0] if isinstance(item, tuple) else item for item in row]
            ws.append(values)
            r = ws.max_row
            for col_idx, item in enumerate(row, start=1):
                if isinstance(item, tuple) and item[1]:
                    cell = ws.cell(row=r, column=col_idx)
                    cell.hyperlink = item[1]
                    cell.font = hyperlink_font
        for col in ws.columns:
            length = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(length + 2, 60)

    po_number_by_id  = {p["po_id"]: p["po_number"] for p in result["pos"]}
    project_web_url  = {p["project_id"]: p["web_url"] for p in result["projects"]}
    vendor_web_url   = {v["vendor_id"]: v["web_url"] for v in result["vendors"]}

    ws1 = wb.active
    ws1.title = "PO一覧"
    _sheet(
        ws1,
        ["Project", "Vendor", "PO番号", "代表ファイル", "関連書類数",
         "最終更新日", "Status(未定義)"],
        [
            [(po["project_name"], project_web_url.get(po["project_id"])),
             (po["vendor_name"], vendor_web_url.get(po["vendor_id"])),
             po["po_number"],
             (po["representative_file"], po["web_url"]),
             po["doc_count"], po["last_modified"], ""]
            for po in result["pos"]
        ],
    )

    ws2 = wb.create_sheet("関連書類")
    _sheet(
        ws2,
        ["Project", "Vendor", "PO番号", "ファイル名", "最終更新日"],
        [
            [(d["project_name"], project_web_url.get(d["project_id"])),
             (d["vendor_name"], vendor_web_url.get(d["vendor_id"])),
             po_number_by_id.get(d["po_id"], ""),
             (d["filename"], d["web_url"]), d["last_modified"]]
            for d in result["documents"] if d["po_id"]
        ],
    )

    ws3 = wb.create_sheet("未分類書類")
    _sheet(
        ws3,
        ["Project", "Vendor", "サブフォルダ", "ファイル名", "最終更新日"],
        [
            [(d["project_name"], project_web_url.get(d["project_id"])),
             (d["vendor_name"], vendor_web_url.get(d["vendor_id"])),
             d["relative_path"],
             (d["filename"], d["web_url"]), d["last_modified"]]
            for d in result["documents"] if not d["po_id"]
        ],
    )

    ws4 = wb.create_sheet("Projects")
    _sheet(
        ws4, ["Project ID", "Project名"],
        [[p["project_id"], (p["project_name"], p["web_url"])] for p in result["projects"]],
    )

    ws5 = wb.create_sheet("Vendors")
    _sheet(
        ws5, ["Vendor ID", "Vendor名", "Project ID"],
        [[v["vendor_id"], (v["vendor_name"], v["web_url"]), v["project_id"]]
         for v in result["vendors"]],
    )

    wb.save(out_path)


# ─────────────────────────────────────────────────────────────
# ユーティリティ
# ─────────────────────────────────────────────────────────────
def _get_result() -> Optional[dict]:
    global _last_result
    if _last_result is not None:
        return _last_result
    path = CACHE_DIR / "scan_result.json"
    if path.exists():
        with open(path, "r", encoding="utf-8") as f:
            _last_result = json.load(f)
    return _last_result


# ─────────────────────────────────────────────────────────────
# Flask ルート定義
# ─────────────────────────────────────────────────────────────
@flask_app.route("/")
def route_index():
    resp = flask_app.make_response(HTML_TEMPLATE)
    resp.headers["Cache-Control"] = "no-store"
    return resp


@flask_app.route("/api/scan")
def route_scan():
    """[SSE] Project/Vendorフォルダをスキャンする。?force=1でキャッシュ無視。"""
    if _scan_state.get("running"):
        def _busy():
            yield "data: " + json.dumps(
                {"type": "error", "message": "既にスキャン中です"}, ensure_ascii=False
            ) + "\n\n"
        return Response(stream_with_context(_busy()), mimetype="text/event-stream")

    force = flask_req.args.get("force", "0") == "1"

    with _lock:
        _scan_state["running"]   = True
        _scan_state["cancelled"] = False
        _scan_state["count"]     = 0

    event_q = queue.Queue()

    def _run():
        global _last_result
        try:
            cache = ScanCache(CACHE_DIR / "scan_cache.json")
            if force:
                cache.reset()

            with _lock:
                token = _auth.get_token(SCOPES)
            client   = SharePointClient(token, SITE_HOST, SITE_PATH)
            site_id  = client.get_site_id()
            drive_id = client.get_drive_id(site_id, LIBRARY_NAME)

            scanner = POScanner(client, cache)
            result  = scanner.scan(drive_id, event_q, _scan_state)

            _last_result = result
            CACHE_DIR.mkdir(exist_ok=True)
            with open(CACHE_DIR / "scan_result.json", "w", encoding="utf-8") as f:
                json.dump(result, f, ensure_ascii=False, indent=2)

            unclassified = len([d for d in result["documents"] if not d["po_id"]])
            event_q.put({
                "type":    "cancelled" if _scan_state.get("cancelled") else "done",
                "count":   _scan_state["count"],
                "summary": {
                    "projects":     len(result["projects"]),
                    "vendors":      len(result["vendors"]),
                    "pos":          len(result["pos"]),
                    "documents":    len(result["documents"]),
                    "unclassified": unclassified,
                },
            })
        except Exception as e:
            event_q.put({"type": "error", "message": str(e)})
        finally:
            _scan_state["running"] = False

    threading.Thread(target=_run, daemon=True).start()

    def _generate():
        while True:
            try:
                event = event_q.get(timeout=30)
                yield "data: " + json.dumps(event, ensure_ascii=False) + "\n\n"
                if event["type"] in ("done", "cancelled", "error"):
                    break
            except queue.Empty:
                yield ": heartbeat\n\n"
                if not _scan_state.get("running"):
                    break

    return Response(
        stream_with_context(_generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@flask_app.route("/api/scan/cancel", methods=["POST"])
def route_scan_cancel():
    if not _scan_state.get("running"):
        return jsonify({"ok": False, "message": "スキャン中ではありません"}), 400
    _scan_state["cancelled"] = True
    return jsonify({"ok": True})


@flask_app.route("/api/result")
def route_result():
    result = _get_result()
    if result is None:
        return jsonify({"projects": [], "vendors": [], "pos": [], "documents": []})
    return jsonify(result)


@flask_app.route("/api/export/excel")
def route_export_excel():
    result = _get_result()
    if result is None:
        return jsonify({"ok": False, "message": "先にスキャンを実行してください"}), 400
    CACHE_DIR.mkdir(exist_ok=True)
    out_path = CACHE_DIR / f"po_catalog_{datetime.now(JST).strftime('%Y%m%d_%H%M%S')}.xlsx"
    export_excel(result, out_path)
    return send_file(out_path, as_attachment=True, download_name=out_path.name)


@flask_app.route("/api/export/json")
def route_export_json():
    result = _get_result()
    if result is None:
        return jsonify({"ok": False, "message": "先にスキャンを実行してください"}), 400
    return jsonify(result)


# ─────────────────────────────────────────────────────────────
# HTML テンプレート
# ─────────────────────────────────────────────────────────────
HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>PO Database Organizer</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: 'Segoe UI', 'Noto Sans JP', sans-serif; background: #f3f4f6; color: #1f2937; }
  header { background: #1e3a5f; color: #fff; padding: 16px 24px; }
  header h1 { font-size: 18px; font-weight: 600; }
  header p { font-size: 12px; opacity: 0.75; margin-top: 4px; }
  main { max-width: 1100px; margin: 20px auto; padding: 0 16px 48px; }
  section { background: #fff; border-radius: 10px; box-shadow: 0 1px 6px rgba(0,0,0,0.07); padding: 18px 22px; margin-bottom: 18px; }
  section h2 { font-size: 15px; margin-bottom: 12px; }
  .btn { background: #1e3a5f; color: #fff; border: none; border-radius: 6px; padding: 9px 18px; font-size: 13px; cursor: pointer; margin-right: 8px; }
  .btn:hover { background: #16304d; }
  .btn.secondary { background: #6b7280; }
  .btn.secondary:hover { background: #545b66; }
  .btn:disabled { background: #9ca3af; cursor: not-allowed; }
  #log { margin-top: 12px; background: #0f172a; color: #cbd5e1; font-family: monospace; font-size: 12px; padding: 10px 12px; border-radius: 6px; height: 140px; overflow-y: auto; white-space: pre-wrap; }
  .summary { display: flex; gap: 12px; margin-top: 12px; flex-wrap: wrap; }
  .card { flex: 1; min-width: 110px; background: #f3f4f6; border-radius: 8px; padding: 10px 14px; }
  .card .label { font-size: 11px; color: #6b7280; }
  .card .value { font-size: 22px; font-weight: 600; }
  .filters { display: flex; gap: 8px; margin-bottom: 10px; flex-wrap: wrap; }
  .filters input, .filters select { padding: 6px 10px; border: 1px solid #d1d5db; border-radius: 6px; font-size: 13px; }
  table { width: 100%; border-collapse: collapse; font-size: 12.5px; }
  th, td { padding: 6px 8px; border-bottom: 1px solid #eceef1; text-align: left; }
  th { background: #f3f4f6; position: sticky; top: 0; }
  .table-wrap { max-height: 420px; overflow-y: auto; border: 1px solid #eceef1; border-radius: 6px; }
  a { color: #1e3a5f; }
  .badge { display: inline-block; background: #fca5a5; color: #7f1d1d; border-radius: 10px; padding: 1px 8px; font-size: 11px; margin-left: 6px; }
</style>
</head>
<body>
<header>
  <h1>PO Database Organizer</h1>
  <p>Project &gt; Vendor フォルダをスキャンし、PO番号を軸にしたカタログを作成します（ステータス判定は行いません）</p>
</header>
<main>
  <section>
    <h2>1. スキャン実行</h2>
    <button class="btn" id="btnScan" onclick="startScan(false)">スキャン開始</button>
    <button class="btn secondary" id="btnRescan" onclick="startScan(true)">キャッシュ無視で再スキャン</button>
    <button class="btn secondary" id="btnCancel" onclick="cancelScan()" disabled>キャンセル</button>
    <div id="log">スキャン待機中...</div>
    <div class="summary" id="summary"></div>
  </section>

  <section>
    <h2>2. PO一覧</h2>
    <div class="filters">
      <input type="text" id="fProject" placeholder="Projectで絞り込み" oninput="renderTable()">
      <input type="text" id="fVendor" placeholder="Vendorで絞り込み" oninput="renderTable()">
      <input type="text" id="fPo" placeholder="PO番号で絞り込み" oninput="renderTable()">
    </div>
    <div class="table-wrap">
      <table>
        <thead><tr><th>Project</th><th>Vendor</th><th>PO番号</th><th>代表ファイル</th><th>関連書類数</th><th>最終更新日</th><th>リンク</th></tr></thead>
        <tbody id="poTableBody"></tbody>
      </table>
    </div>
  </section>

  <section>
    <h2>3. エクスポート</h2>
    <button class="btn" onclick="location.href='/api/export/excel'">Excel出力</button>
    <button class="btn secondary" onclick="location.href='/api/export/json'">JSON出力</button>
  </section>
</main>

<script>
let scanData = { pos: [] };

function log(text) {
  const el = document.getElementById('log');
  el.textContent += "\\n" + text;
  el.scrollTop = el.scrollHeight;
}

function startScan(force) {
  document.getElementById('log').textContent = 'スキャンを開始します...';
  document.getElementById('btnScan').disabled = true;
  document.getElementById('btnRescan').disabled = true;
  document.getElementById('btnCancel').disabled = false;

  const es = new EventSource('/api/scan?force=' + (force ? '1' : '0'));
  es.onmessage = (ev) => {
    const data = JSON.parse(ev.data);
    if (data.type === 'exploring') {
      log('[Project] ' + data.path);
    } else if (data.type === 'progress') {
      log('  [Vendor] ' + data.path + ' (' + data.files + '件, 累計' + data.count + ')');
    } else if (data.type === 'done' || data.type === 'cancelled') {
      log(data.type === 'done' ? '完了しました。' : 'キャンセルされました。');
      renderSummary(data.summary);
      es.close();
      finishScan();
      loadResult();
    } else if (data.type === 'error') {
      log('[エラー] ' + data.message);
      es.close();
      finishScan();
    }
  };
  es.onerror = () => { es.close(); finishScan(); };
}

function finishScan() {
  document.getElementById('btnScan').disabled = false;
  document.getElementById('btnRescan').disabled = false;
  document.getElementById('btnCancel').disabled = true;
}

function cancelScan() {
  fetch('/api/scan/cancel', { method: 'POST' });
}

function renderSummary(summary) {
  if (!summary) return;
  const el = document.getElementById('summary');
  el.innerHTML = [
    ['Projects', summary.projects], ['Vendors', summary.vendors],
    ['POs', summary.pos], ['書類総数', summary.documents],
    ['未分類', summary.unclassified],
  ].map(([label, value]) =>
    `<div class="card"><div class="label">${label}</div><div class="value">${value}</div></div>`
  ).join('');
}

function loadResult() {
  fetch('/api/result').then(r => r.json()).then(data => {
    scanData = data;
    renderTable();
  });
}

function renderTable() {
  const fProject = document.getElementById('fProject').value.toLowerCase();
  const fVendor  = document.getElementById('fVendor').value.toLowerCase();
  const fPo      = document.getElementById('fPo').value.toLowerCase();

  const rows = scanData.pos.filter(po =>
    po.project_name.toLowerCase().includes(fProject) &&
    po.vendor_name.toLowerCase().includes(fVendor) &&
    po.po_number.toLowerCase().includes(fPo)
  );

  document.getElementById('poTableBody').innerHTML = rows.map(po => `
    <tr>
      <td>${po.project_name}</td>
      <td>${po.vendor_name}</td>
      <td>${po.po_number}</td>
      <td>${po.representative_file}</td>
      <td>${po.doc_count}</td>
      <td>${po.last_modified}</td>
      <td><a href="${po.web_url}" target="_blank">開く</a></td>
    </tr>
  `).join('');
}

loadResult();
</script>
</body>
</html>"""


# ─────────────────────────────────────────────────────────────
# エントリポイント
# ─────────────────────────────────────────────────────────────
def main():
    global _auth

    print("=" * 60)
    print(" PO Database Organizer")
    print("=" * 60)
    print(f"  対象サイト: {SITE_HOST}{SITE_PATH}")
    if LIBRARY_NAME:
        print(f"  ライブラリ: {LIBRARY_NAME}")
    print()

    CACHE_DIR.mkdir(exist_ok=True)

    _auth = GraphAuthManager(TENANT_ID, CLIENT_ID, TOKEN_CACHE_PATH)
    print("[認証] トークンを取得しています...")
    _auth.get_token(SCOPES)
    print("[OK] 認証完了\n")

    url = f"http://127.0.0.1:{FLASK_PORT}"
    print(f"[起動] {url} をブラウザで開きます...")
    threading.Timer(1.0, lambda: webbrowser.open(url)).start()
    flask_app.run(host="127.0.0.1", port=FLASK_PORT, debug=False)


if __name__ == "__main__":
    main()
